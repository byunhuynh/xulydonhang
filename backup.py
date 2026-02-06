import fitz  # PyMuPDF
import re
import openpyxl  # Đọc file Excel
from datetime import datetime,timedelta
import math
import os
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment
from PySide6.QtCore import Signal, QObject
from fuzzywuzzy import fuzz
import configparser
import pandas as pd
import pytesseract
from PIL import Image, ImageEnhance, ImageOps
from statistics import mode

import io

# Đọc thông tin từ file setting.ini
CONFIG_FILE = "settings.ini"
config = configparser.ConfigParser()

class ProcessHandler(QObject):
    table_signal = Signal(str, str, str,str, str, str,str)  # Tín hiệu mới để cập nhật bảng

    log_signal = Signal(str)  # Tín hiệu gửi log về giao diện

    


    def extract_sku(value):
        """Trích xuất mã SKU từ value (chuỗi có thể chứa ký tự chữ & số)."""
        match = re.search(r"\b[A-Z0-9]+\b", value)  # Tìm từ viết hoa + số
        return match.group(0) if match else None

    def check_value_in_sanpham(value):
        """Tìm tất cả mã SKU có trong value dựa trên danh sách SKU từ 'data.xlsx'."""
        file_path = "data.xlsx"
        sheet_name = "SanPham"

        if not value or not isinstance(value, str):
            return False

        try:
            # 📌 Tải danh sách SKU từ file Excel
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb[sheet_name]

            # Chuyển tất cả mã SKU thành chuỗi để tránh lỗi
            sku_list = {str(row[0]).strip() for row in sheet.iter_rows(min_col=1, max_col=1, values_only=True) if row[0]}

            # 🔍 Tìm tất cả mã SKU có trong chuỗi value
            pattern = r"\b(" + "|".join(map(re.escape, sku_list)) + r")\b"
            matches = re.findall(pattern, value)

            return matches if matches else False

        except Exception as e:
            print(f"⚠ Lỗi khi kiểm tra SKU: {e}")
            return False






    def extract_text_from_pdf(pdf_path):
        """Trích xuất nội dung từ PDF"""
        doc = fitz.open(pdf_path)
        text = "\n".join([page.get_text("text") for page in doc])
        return text


    def identify_vendor(text):
        """Xác định Vendor bất kể khoảng trắng hoặc xuống dòng"""
        cleaned_text = re.sub(r"\s+", " ", text).strip()  # Chuẩn hóa khoảng trắng
       
       

        if re.search(r"Vendor\s*[-:]\s*21569", cleaned_text):
            return "Coop"
        
        if re.search(r"3005382", cleaned_text) or re.search(r"CTY TNHH DV EB", cleaned_text, re.IGNORECASE):
            return "BigC"
        
        if re.search(r"0107889783\s*009333", cleaned_text):
            return "Lotte"
        
        if re.search(r"VD-00002345", cleaned_text):
            return "Satra"
        
        if re.search(r"CONG TY TNHH TMDV XNK HA THANH \(101017\)", cleaned_text):
            return "Emart"
        
        if re.search(r"0313403198", cleaned_text):
            return "Kingfood"

        if re.search(r"Nhà cung cấp \(Supplier\): 0002011398", cleaned_text):
            return "Winmart"
        

        if re.search(r"254000001538", cleaned_text):
            return "FujiMart"
        
        if re.search(r"Công Ty Cổ Phần Thương Mại Bách Hóa Xanh", cleaned_text):
            return "BHX"
        

        if re.search(r"Mã NCC:V0001626", cleaned_text):
            return "Farmer"

        return "Unknown"

    def extract_po_location(text):
        """Lấy P/O Location từ PDF"""
        match = re.search(r"P/O Location:\s*(.+)", text)
        return match.group(1).strip() if match else None

    def get_makhachhang(po_location):
        """Lấy Mã Khách Hàng từ file data.xlsx"""
        file_path = "data.xlsx"
        sheet_name = "MaKH"
        
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb[sheet_name]

        for row in sheet.iter_rows(min_row=2, values_only=True):  # Bỏ qua tiêu đề
            col_A, col_B, col_C = row[0], row[1], row[2]  # Cột A, B, C trong Excel

            if col_A and str(col_A).strip().upper() in ["COOP", "COOPFOOD"]:
                if col_C and str(col_C).strip().endswith(str(po_location)):  # Chuyển cả hai về chuỗi
                    return str(col_C)  # Trả về giá trị cột C dưới dạng chuỗi
        return "Không tìm thấy"
    def normalize_text(text):
        """Chuẩn hóa chuỗi: viết thường, loại bỏ dấu câu, khoảng trắng thừa."""
        text = text.lower()
        text = re.sub(r'[^a-z0-9\s]', '', text)  # Xóa ký tự đặc biệt
        text = re.sub(r'\s+', ' ', text).strip()  # Chuẩn hóa khoảng trắng
        return text
    





    def laymakhachhang_fujimart(po_location):
        """Tìm chính xác po_location trong cột B, trả về giá trị tương ứng ở cột C."""
        file_path = "data.xlsx"
        sheet_name = "MaKH"

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb[sheet_name]

            for row in sheet.iter_rows(min_row=2):
                ten_kh = row[1].value  # Cột B
                if ten_kh and str(ten_kh).strip().lower() == str(po_location).strip().lower():
                    return row[2].value  # Cột C
            return None

        except Exception as e:
            print("Lỗi:", e)
            return None
    
    def laymakhachhang_satra(po_location,hethong):
        """Lấy Mã Khách Hàng từ file data.xlsx với tìm kiếm gần đúng."""
        file_path = "data.xlsx"
        sheet_name = "MaKH"
        
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb[sheet_name]
        
        po_location_norm = ProcessHandler.normalize_text(po_location)  # Chuẩn hóa địa chỉ nhập vào
        best_match = None
        best_score = 0
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            col_A, col_B, col_C, col_D = row[:4]

            if col_A and str(col_A).strip().upper() in hethong:
                if col_D:
                    col_D_norm = ProcessHandler.normalize_text(str(col_D))
                    
                    score = fuzz.partial_ratio(po_location_norm, col_D_norm)  # Dùng partial_ratio để tăng độ chính xác
                    if score > best_score:
                        best_score = score
                        best_match = str(col_C)
        
        return best_match if best_score > 70 else None  # Giảm ngưỡng xuống 70

    def layhethong_COOP(makhachhang):
        """Lấy Mã Khách Hàng từ file data.xlsx"""
        file_path = "data.xlsx"
        sheet_name = "MaKH"
        
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb[sheet_name]

        for row in sheet.iter_rows(min_row=2, values_only=True):  # Bỏ qua tiêu đề
            col_A, col_B, col_C = row[0], row[1], row[2]  # Cột A, B, C trong Excel

            if col_C and str(col_C).strip() in [makhachhang]:
                return str(col_A)  # Trả về giá trị cột C dưới dạng chuỗi
        return None
    
    def get_makhachhang_lotte(storecode):
        """Lấy Mã Khách Hàng từ file data.xlsx"""
        file_path = "data.xlsx"
        sheet_name = "MaKH"
        
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb[sheet_name]

        for row in sheet.iter_rows(min_row=2, values_only=True):  # Bỏ qua tiêu đề
            col_A, col_B, col_C = row[0], row[1], row[2]  # Cột A, B, C trong Excel

            if col_A and str(col_A).strip().upper() in ["LOTTE"]:
                if col_C and str(col_C).strip().endswith(str(storecode)):  # Chuyển cả hai về chuỗi
                    return str(col_C)  # Trả về giá trị cột C dưới dạng chuỗi
        return None


    @staticmethod
    def clean_text(text):
        """Xóa dòng trống & gộp dòng liên quan để dễ xử lý"""
        lines = text.split("\n")
        cleaned_lines = [line.strip() for line in lines if line.strip()]  # Xóa dòng trống
        return " ".join(cleaned_lines)  # Gộp tất cả thành 1 đoạn

    @staticmethod
    def tachSP_text(text):
        """Trích xuất SKU, Net Buy Cost, Qty Ord/Pcs, Extended Cost từ văn bản"""
        text = process_handler.clean_text(text)  # Làm sạch và gộp nội dung


        # Regex tìm SKU
        sku_pattern = re.compile(r"(\d{6,}-\d+)")  # SKU có dạng 3573266-1 hoặc 3573267-6

        # Regex tìm giá trị số tiền (có thể có dấu phẩy và chấm)
        cost_pattern = re.compile(r"(\d{1,3}(?:,\d{3})*\.\d{2})")

        matches = sku_pattern.finditer(text)
        products = []

        for match in matches:
            sku = match.group(1)  # SKU
            start_idx = match.end()  # Vị trí kết thúc của SKU trong văn bản
            
            # Cắt phần văn bản sau SKU để tìm các số
            sub_text = text[start_idx:start_idx + 100]  # Lấy 100 ký tự sau SKU để tìm số liệu
            cost_values = cost_pattern.findall(sub_text)

            if len(cost_values) >= 3:
                net_buy_cost = cost_values[-3]  # Net Buy Cost
                qty_ord_pcs = cost_values[-2]  # Qty Ord/Pcs
                extended_cost = cost_values[-1]  # Extended Cost
                
                products.append({
                    "SKU Number": sku,
                    "Net Buy Cost": net_buy_cost,
                    "Qty Ord/Pcs": qty_ord_pcs,
                    "Extended Cost": extended_cost
                })
            else:
                print(f"⚠️ Không tìm đủ dữ liệu cho SKU: {sku}")

        return products



    @staticmethod
    def extract_products(text):
        """Trích xuất SKU Number, Qty Ord/Pcs, Extended Cost từ dữ liệu dạng dọc hoặc ngang với debug."""

        print("Bắt đầu trích xuất sản phẩm...\n")

        # Cắt phần dư thừa
        text = re.split(r"S\s*u\s*b\s*T\s*o\s*t\s*a\s*l", text, flags=re.IGNORECASE)[0]
        text = re.split(r"V\s*N\s*D\s*V\s*i\s*e\s*t\s*N\s*a\s*m\s*D\s*o\s*n\s*g", text, flags=re.IGNORECASE)[-1]

        # Chuẩn hóa dòng
        lines = [line.strip() for line in text.strip().split('\n') if line.strip()]
        print(f"DEBUG: Tổng số dòng sau chuẩn hóa -> {len(lines)}")

        # Tìm các vị trí có SKU
        sku_indices = [i for i, line in enumerate(lines) if re.search(r'\d{7}-\s*\d', line)]
        print(f"DEBUG: Chỉ số SKU được tìm thấy tại {sku_indices}\n")

        products = []

        for i in range(len(sku_indices)):
            start = sku_indices[i]
            end = sku_indices[i + 1] if i + 1 < len(sku_indices) else len(lines)
            block = lines[start:end]
            block = [line.replace(", ", ",").replace(". ", ".") for line in block]

            print(f"DEBUG: Xử lý Khối {i+1}:\n" + "\n".join(block) + "\n")

            item = {
                'SKU Number': None,
                'Qty Ord/Pcs': None,
                'Extended Cost': None
            }

            # SKU Number
            match = re.search(r'(\d{7}-\s*\d)', block[0])
            if match:
                item['SKU Number'] = match.group(1).replace(" ", "")
            print(f"DEBUG: Số SKU đã trích xuất -> {item['SKU Number']}")

            # Tách số tài chính từ block
            numeric_lines = [line for line in block if re.fullmatch(r'[\d,.\s]+$', line)]
            all_numbers = re.findall(r'\d[\d,]*\.\d+', " ".join(numeric_lines))
            print(f"DEBUG: Dòng số tài chính -> {numeric_lines}")
            print(f"DEBUG: Tất cả các số -> {all_numbers}")

            # Chuyển số thành float để phân tích
            numbers_float = [float(n.replace(",", "")) for n in all_numbers]

            # Xử lý block cuối hoặc block thường
            if i == len(sku_indices) - 1:
                print("DEBUG: 🔍 Phân tích thông minh block cuối có thể dính tổng đơn...")
                # Lọc các số tiền lớn (có dấu phẩy)
                large_numbers = [n for n in all_numbers if ',' in n]
                print(f"DEBUG: Các số tiền lớn -> {large_numbers}")

                if large_numbers:
                    # Sắp xếp các số lớn theo giá trị
                    large_numbers_float = [float(n.replace(",", "")) for n in large_numbers]
                    sorted_large_numbers = sorted(enumerate(large_numbers), key=lambda x: large_numbers_float[x[0]])

                    # Thử với số nhỏ nhất trước
                    for idx, cost in sorted_large_numbers:
                        cost_idx = all_numbers.index(cost)
                        if cost_idx > 0:
                            qty = all_numbers[cost_idx - 1]
                            qty_float = float(qty.replace(",", ""))
                            cost_float = float(cost.replace(",", ""))
                            unit_price = cost_float / qty_float
                            print(f"DEBUG: Thử Qty = {qty}, Cost = {cost}, Đơn giá = {unit_price:.2f}")

                            if 1000 < unit_price < 2000000:
                                item['Qty Ord/Pcs'] = qty
                                item['Extended Cost'] = cost
                                print(f"DEBUG: ✅ Chọn Qty = {qty}, Cost = {cost}")
                                break
                    if not item['Qty Ord/Pcs']:
                        print("DEBUG: ⚠️ Không tìm thấy cặp Qty và Cost hợp lý!")
                else:
                    print("DEBUG: ⚠️ Không tìm thấy số tiền lớn nào!")
            else:
                # Block thường: Lấy số thứ hai từ cuối và cuối cùng
                if len(all_numbers) >= 2:
                    item['Qty Ord/Pcs'] = all_numbers[-2]
                    item['Extended Cost'] = all_numbers[-1]
                else:
                    print("DEBUG: ⚠️ Không đủ số để trích xuất Qty và Cost!")

            print(f"DEBUG: Số lượng đặt hàng đã trích xuất -> {item['Qty Ord/Pcs']}")
            print(f"DEBUG: Chi phí mở rộng đã trích xuất -> {item['Extended Cost']}\n")

            # Thêm kiểm tra đơn giá (nếu đã có Qty và Cost)
            if item['Qty Ord/Pcs'] and item['Extended Cost']:
                qty = float(item['Qty Ord/Pcs'].replace(",", ""))
                cost = float(item['Extended Cost'].replace(",", ""))
                unit_price = cost / qty
                if not (1000 < unit_price < 2000000):
                    print(f"DEBUG: ⚠️ Đơn giá không hợp lý: {unit_price:.2f}")

            if item['SKU Number']:
                products.append(item)

        print(f"Tổng sản phẩm đã trích xuất: {len(products)}\n")

        return products

        



    def xoakhoangtrang(text):
        return re.sub(r'\s+', '', text)  # Xóa tất cả khoảng trắng trong chuỗi
    


    def remove_line_numbers(text):
        # Chia text thành các dòng
        lines = text.split('\n')
        # Danh sách để chứa các dòng đã xử lý
        cleaned_lines = []
        # Giới hạn số thứ tự cần xóa (1 đến 10)
        valid_numbers = set(str(i) for i in range(1, 11))  # Tập hợp "1" đến "10"
        
        for line in lines:
            # Kiểm tra nếu dòng bắt đầu bằng số
            if line.strip():
                first_part = line.split(maxsplit=1)[0]
                if first_part in valid_numbers:
                    # Nếu là số thứ tự từ 1-10, loại bỏ nó và giữ nội dung phía sau
                    if len(line.split(maxsplit=1)) > 1:  # Đảm bảo có nội dung phía sau
                        cleaned_line = line.split(maxsplit=1)[1]
                        cleaned_lines.append(cleaned_line)
                    # Nếu chỉ có số (không nội dung), bỏ qua dòng
                else:
                    # Giữ nguyên dòng nếu không phải số thứ tự từ 1-10
                    cleaned_lines.append(line)
            else:
                # Giữ nguyên dòng trống
                cleaned_lines.append(line)
        
        # Gộp lại các dòng thành text hoàn chỉnh
        return '\n'.join(cleaned_lines)

    def extract_info(text):
        """Trích xuất các thông tin quan trọng từ văn bản"""
        text = ProcessHandler.remove_line_numbers(text)
        text = re.sub(r"\s+", " ", text)
        text = ProcessHandler.xoakhoangtrang(text)
        text = text.split("Currency")[0] if "Currency" in text else text
        print(text)
        
        info = {
            "P/O Number": re.search(r"P\s*/\s*O\s*N\s*u\s*m\s*b\s*e\s*r\s*[:-]?\s*([\d-]+)", text),
            "P/O Location": re.search(r"P\s*/\s*O\s*L\s*o\s*c\s*a\s*t\s*i\s*o\s*n\s*:\s*(\d+)", text),
            "Entry Date": re.search(r"(?:Entry\s*Date|E\s*n\s*t\s*r\s*y\s*D\s*a\s*t\s*e)\s*-\s*([\d/]+)", text),
            "Cancel Date": re.search(r"Cancel\s*Date-\s*([\d/]+)", text)
        }

        if not info["P/O Location"]:
            info["P/O Location"] = re.search(r"Store-\s*(\d+)", text)

        extracted_info = {key: (match.group(1).strip() if match else "Không tìm thấy") for key, match in info.items()}

        return extracted_info
        
    def extract_discount(value):
        """Trích xuất số phần trăm giảm giá từ chuỗi."""
        if isinstance(value, list):  
            value = value[0] if value else ""

        if not isinstance(value, str):  
            value = str(value)  # Chuyển thành chuỗi nếu chưa phải chuỗi

        match = re.search(r"(\d+(\.\d+)?)%", value)  # Tìm số trước dấu "%"
        if match:
            return float(match.group(1))  # Chuyển thành số thực
        
        return 0  # Không có chiết khấu

    
        


    def extract_number(text):
        """Hàm trích xuất phần số từ chuỗi, nếu không có số thì trả về None"""
        match = re.search(r'\d+', str(text))
        return int(match.group()) if match else None
    

    def tachkhuyenmai_coop(text, hethong):
        if hethong == "COOP":
            match = re.search(r'CM[^()]*', text)  # Lấy nội dung sau "CM" đến trước "("
        elif hethong == "COOPFOOD":
            # Nếu CF nằm trong dấu ngoặc, lấy cụm trong ngoặc có chứa "CF"
            match = re.search(r'\(([^()]*(?=\sCF))[^()]*\)', text)
            if not match:
                # Nếu CF không nằm trong ngoặc, lấy toàn bộ nội dung từ "CF" đến trước "("
                match = re.search(r'CF[^()]*', text)
        else:
            return text  # Trả về toàn bộ văn bản nếu hệ thống không hợp lệ

        # Nếu tìm thấy kết quả, trả về đoạn text đã trích xuất, ngược lại trả về toàn bộ text
        return (match.group(1).strip() if match and match.lastindex == 1 else match.group().strip()) if match else text


    def timten_sanpham(search_value,file_path = "data.xlsx", sheet_name="SanPham"):
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name]

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2, values_only=True):
            key, value = row
            if key == search_value:
                return value  # Trả về giá trị cột B nếu tìm thấy

        return "Không thấy tên sản phẩm"  # Trả về None nếu không tìm thấy


    def write_to_dondathang(self,products, makhachhang, po_number, entry_date, cancle_date,STT_donhang,vendor,ghichu,shipto):


        """Ghi danh sách sản phẩm vào file dondathang.xlsx"""
        file_path = "dondathang.xlsx"
        sheet_name = "Don dat hang"
        
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]

        start_row = sheet.max_row + 1
        current_row = start_row
        # Lấy giá trị G1, đảm bảo nó là số
        
        print(f"📍 PO Number {po_number} ")
        STT_donhang_str = f"{STT_donhang:05d}"     
        hethong = ProcessHandler.layhethong_COOP(makhachhang)


        saigia = 0

        if makhachhang[:2] == "MB":
            khuvuc = "MT_MB"
            kho = "TP_HN_5"
            mien = "HN"
        else:
            kho = "LA_KHO1"
            khuvuc = "MT_MN"
            mien = "LA"


        sheet[f"A{current_row}"] = entry_date
        if current_row == 9:
            sheet[f"B{current_row}"] = f'="ĐĐH{vendor}"&TEXT(G1,"00000")'
        elif current_row > 9:
            sheet[f"B{current_row}"] = f'=IF(L{current_row}=L{current_row - 1},B{current_row - 1},"ĐĐH{vendor}"&TEXT(RIGHT(B{current_row - 1},4)+1,"00000"))'

            
            

        sheet[f"C{current_row}"] = "Chưa thực hiện"
        sheet[f"D{current_row}"] = cancle_date
        sheet[f"E{current_row}"] = shipto
        sheet[f"G{current_row}"] = makhachhang
        sheet[f"L{current_row}"] = f"{hethong} PO{po_number}" + (f" - {ghichu}" if ghichu else "")
        sheet[f"V{current_row}"] = kho
        sheet[f"AE{current_row}"] = 10
        sheet[f"AJ{current_row}"] = khuvuc
        sheet[f"AM{current_row}"] = mien
        sheet[f"U{current_row}"] = "Không"
        sheet[f"Z{current_row}"] = 0
        sheet[f"S{current_row}"] = f"{hethong} PO{po_number}"
        sheet[f"T{current_row}"] = "Có"
        sheet[f"X{current_row}"] = 0
        sheet[f"Y{current_row}"] = 0

        current_row += 1
            
        for product in products:
            sheet[f"A{current_row}"] = entry_date
            if current_row == 9:
                sheet[f"B{current_row}"] = f'="ĐĐH{vendor}"&TEXT(G1,"00000")'
            elif current_row > 9:
                sheet[f"B{current_row}"] = f'=IF(L{current_row}=L{current_row - 1},B{current_row - 1},"ĐĐH{vendor}"&TEXT(RIGHT(B{current_row - 1},4)+1,"00000"))'
        
            sheet[f"C{current_row}"] = "Chưa thực hiện"
            sheet[f"D{current_row}"] = cancle_date
            sheet[f"E{current_row}"] = shipto
            sheet[f"G{current_row}"] = makhachhang
            sheet[f"L{current_row}"] = f"{hethong} PO{po_number}" + (f" - {ghichu}" if ghichu else "")
            sheet[f"Q{current_row}"] = product["SKU Number"]
            sheet[f"V{current_row}"] = kho
            sheet[f"AE{current_row}"] = 10
            sheet[f"AJ{current_row}"] = khuvuc
            sheet[f"AM{current_row}"] = mien
            sheet[f"U{current_row}"] = "Không"
            sheet[f"Y{current_row}"] = 0
            sheet[f"T{current_row}"] = "Không"
            sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(product["SKU Number"])

            # ✅ Xử lý "Extended Cost"
            extended_cost = product["Extended Cost"]
            if isinstance(extended_cost, str):  # Nếu là chuỗi, loại bỏ dấu phẩy
                extended_cost = float(extended_cost.replace(",", ""))
            extended_cost = int(extended_cost) if extended_cost.is_integer() else extended_cost
            sheet[f"Z{current_row}"] = f"=Y{current_row}*X{current_row}"

            # ✅ Xử lý "Qty Ord/Pcs"
            qty_ord_pcs = product["Qty Ord/Pcs"]
            if isinstance(qty_ord_pcs, str):  # Nếu là chuỗi, loại bỏ dấu phẩy
                qty_ord_pcs = float(qty_ord_pcs.replace(",", ""))
            qty_ord_pcs = int(qty_ord_pcs) if qty_ord_pcs.is_integer() else qty_ord_pcs
            sheet[f"X{current_row}"] = qty_ord_pcs

            # ✅ Tính giá hóa đơn
            giahoadon = extended_cost / qty_ord_pcs, 2
            giathucte = ProcessHandler.find_price_by_sku(product["SKU Number"])
           
            # 🔍 **Tìm giá trị trong CTKM.xlsx theo SKU & thời gian**
            value = ProcessHandler.find_value_by_sku_and_time(product["SKU Number"], entry_date)

            # ✅ Nếu có giá trị, xử lý chiết khấu
            if value:
                value = ProcessHandler.tachkhuyenmai_coop(value,hethong)
                print(f"🔹 Đã lấy được giá trị: {value}, tiếp tục xử lý...")
                self.log_signal.emit(
    f'🎉 Mã hàng <b><span style="color: blue;">{product["SKU Number"]}</span></b> '
    f'có chương trình khuyến mãi <b><span style="color: green;">{value}</span></b> 🛒'
)

                # ✅ Kiểm tra nếu value chứa "%", trích xuất và áp dụng chiết khấu
                discount = ProcessHandler.extract_discount(value)
                print(f'Giảm giá: {discount}')
                print(f"🔍 Debug: type(discount) = {type(discount)}, value = {discount}")

                if discount != 0:

                    if isinstance(discount, str):  # Nếu discount là chuỗi
                        discount = discount.replace("%", "").strip()  # Loại bỏ dấu %
                        discount = float(discount) if discount else 0  # Chuyển về float, nếu rỗng thì là 0

                    # Chuyển giathucte về float nếu nó là chuỗi
                    if isinstance(giathucte, str):
                        giathucte = giathucte.replace(",", "").strip()  # Loại bỏ dấu phẩy (nếu có)
                        giathucte = float(giathucte) if giathucte else 0  # Chuyển về float, nếu rỗng thì là 0

                    # Giảm giá theo %    
                    giathucte = giathucte - (giathucte * float(discount) / 100)




            
            giahoadon = float(giahoadon[0]) if isinstance(giahoadon, tuple) else float(giahoadon)
            giathucte = float(giathucte[0]) if isinstance(giathucte, tuple) else float(giathucte)
            # ✅ So sánh giá sau khi xử lý chiết khấu
            if giathucte is not None and math.isclose(giahoadon, giathucte, rel_tol=1e-3):
                sheet[f"Y{current_row}"] = giathucte
                self.log_signal.emit(
    f'✅ Mã hàng <b><span style="color: blue;">{product["SKU Number"]}</span></b> '
    f'có giá chính xác! 🎯 '
    f'Giá hóa đơn: <span style="color: green; font-weight: bold;">{giahoadon}</span>')
            else:

                sheet[f"Y{current_row}"] = giathucte
                red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                sheet[f"Y{current_row}"].fill = red_fill
                comment_text = f"Kiểm tra lại giá mã này! - Giá hóa đơn: {giahoadon} - Chênh lệch: {giahoadon - giathucte}"
                comment = Comment(comment_text, "System")
                sheet[f"Y{current_row}"].comment = comment
                self.log_signal.emit(
    f"⚠️ Mã hàng <b>{product["SKU Number"]}</b> sai giá! 🛑 Giá hóa đơn: <span style='color: red;'><b>{giahoadon}</b></span> - "
    f"Giá đúng: <span style='color: green;'><b>{giathucte}</b></span>"               
    )           
                saigia += 1
              

            # 🔍 Kiểm tra SKU có trong 'SanPham'
            kiemtra = ProcessHandler.check_value_in_sanpham(value) 

            match = re.search(r"(\d+)\s*\+\s*1", value)  # Tìm số đầu tiên trong biểu thức X+1
            if match:
                x = int(match.group(1))  # Chuyển đổi thành số nguyên
                kiemtra = kiemtra or product["SKU Number"]
                
                if x >= 2:  # Chỉ chia khi x >= 2, còn 1+1 thì giữ nguyên
                    qty_ord_pcs = math.floor(qty_ord_pcs / x)  # Chia rồi làm tròn xuống

            if kiemtra:
                if isinstance(kiemtra, list):  # Nếu có nhiều mã SKU, nối thành chuỗi
                    kiemtra = ", ".join(kiemtra)
                current_row += 1
                self.log_signal.emit(
    f'🎁 Mã hàng <b><span style="color: blue;">{product["SKU Number"]}</span></b> '
    f'đã thêm hàng khuyến mãi <b><span style="color: green;">{kiemtra}</span></b> tặng kèm! 🎉'
)
                sheet[f"A{current_row}"] = entry_date
                if current_row == 9:
                    sheet[f"B{current_row}"] = f'="ĐĐH{vendor}"&TEXT(G1,"00000")'
                elif current_row > 9:
                    sheet[f"B{current_row}"] = f'=IF(L{current_row}=L{current_row - 1},B{current_row - 1},"ĐĐH{vendor}"&TEXT(RIGHT(B{current_row - 1},4)+1,"00000"))'
                sheet[f"C{current_row}"] = "Chưa thực hiện"
                sheet[f"D{current_row}"] = cancle_date
                sheet[f"E{current_row}"] = shipto
                sheet[f"G{current_row}"] = makhachhang
                sheet[f"L{current_row}"] = f"{hethong} PO{po_number}" + (f" - {ghichu}" if ghichu else "")
                sheet[f"Q{current_row}"] = kiemtra
                sheet[f"X{current_row}"] = qty_ord_pcs
                sheet[f"V{current_row}"] = kho
                sheet[f"Y{current_row}"] = 0
                sheet[f"Z{current_row}"] = 0
                sheet[f"U{current_row}"] = "Có"
                sheet[f"AJ{current_row}"] = khuvuc
                sheet[f"AE{current_row}"] = 10
                sheet[f"AM{current_row}"] = mien
                sheet[f"T{current_row}"] = "Không"
                sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(kiemtra)
            sheet["G1"] = STT_donhang
            current_row += 1

        # Lưu lại file Excel
        wb.save(file_path)
        print(f"✅ Đã ghi {len(products)} dòng vào '{file_path}', sheet '{sheet_name}'!")
        return saigia


    def write_to_dondathang_lotte(self,products, makhachhang, po_number, entry_date, cancle_date,STT_donhang,vendor,diachigiaohang):

        vendor = vendor.upper()
        """Ghi danh sách sản phẩm vào file dondathang.xlsx"""
        file_path = "dondathang.xlsx"
        sheet_name = "Don dat hang"
        
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]

        start_row = sheet.max_row + 1
        current_row = start_row
        # Lấy giá trị G1, đảm bảo nó là số
        
        print(f"📍 PO Number {po_number} ")
        STT_donhang_str = f"{STT_donhang:05d}"     
        sku_mapping = ProcessHandler.load_sku_mapping()  # Load mapping SKU
        products = ProcessHandler.replace_sku_numbers(products, sku_mapping)  # Thay SKU
        saigia = 0
        if makhachhang[:2] == "MB":
            khuvuc = "MT_MB"
            kho = "TP_HN_5"
            mien = "HN"
        else:
            kho = "LA_KHO1"
            khuvuc = "MT_MN"
            mien = "LA"




        sheet[f"A{current_row}"] = entry_date
        if current_row == 9:
            sheet[f"B{current_row}"] = f'="ĐĐH{vendor}"&TEXT(G1,"00000")'
        elif current_row > 9:
            sheet[f"B{current_row}"] = f'=IF(L{current_row}=L{current_row - 1},B{current_row - 1},"ĐĐH{vendor}"&TEXT(RIGHT(B{current_row - 1},4)+1,"00000"))'
        sheet[f"C{current_row}"] = "Chưa thực hiện"
        sheet[f"D{current_row}"] = cancle_date
        sheet[f"G{current_row}"] = makhachhang
        sheet[f"L{current_row}"] = f"{vendor} PO{po_number}"
        sheet[f"V{current_row}"] = kho
        sheet[f"AE{current_row}"] = 10
        sheet[f"AJ{current_row}"] = khuvuc
        sheet[f"AM{current_row}"] = mien
        sheet[f"U{current_row}"] = "Không"
        sheet[f"Z{current_row}"] = 0
        sheet[f"S{current_row}"] = f"{vendor} PO{po_number}"
        sheet[f"T{current_row}"] = "Có"
        sheet[f"X{current_row}"] = 0
        sheet[f"Y{current_row}"] = 0
        sheet[f"E{current_row}"] = diachigiaohang



        
        current_row += 1

        for product in products:
            
            sheet[f"A{current_row}"] = entry_date
            if current_row == 9:
                sheet[f"B{current_row}"] = f'="ĐĐH{vendor}"&TEXT(G1,"00000")'
            elif current_row > 9:
                sheet[f"B{current_row}"] = f'=IF(L{current_row}=L{current_row - 1},B{current_row - 1},"ĐĐH{vendor}"&TEXT(RIGHT(B{current_row - 1},4)+1,"00000"))'
            sheet[f"C{current_row}"] = "Chưa thực hiện"
            sheet[f"D{current_row}"] = cancle_date
            sheet[f"G{current_row}"] = makhachhang
            sheet[f"L{current_row}"] = f"{vendor} PO{po_number}"
            sheet[f"Q{current_row}"] = product["Barcode"]
            sheet[f"V{current_row}"] = kho
            sheet[f"AE{current_row}"] = 10
            sheet[f"AJ{current_row}"] = khuvuc
            sheet[f"AM{current_row}"] = mien
            sheet[f"U{current_row}"] = "Không"
            sheet[f"E{current_row}"] = diachigiaohang

            sheet[f"T{current_row}"] = "Không"
            sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(product["Barcode"])

            # ✅ Xử lý "Extended Cost"
            extended_cost = product["Total Price"]
            if isinstance(extended_cost, str):  # Nếu là chuỗi, loại bỏ dấu phẩy
                extended_cost = float(extended_cost.replace(",", ""))
            extended_cost = int(extended_cost) if extended_cost.is_integer() else extended_cost
            sheet[f"Z{current_row}"] = f"=Y{current_row}*X{current_row}"

            # ✅ Xử lý "Qty Ord/Pcs"
            qty_ord_pcs = product["Qty-Box"] * product["Box Quantity"]
            if isinstance(qty_ord_pcs, str):  # Nếu là chuỗi, loại bỏ dấu phẩy
                qty_ord_pcs = float(qty_ord_pcs.replace(",", ""))
            qty_ord_pcs = int(qty_ord_pcs) if qty_ord_pcs.is_integer() else qty_ord_pcs
            sheet[f"X{current_row}"] = qty_ord_pcs

            # ✅ Tính giá hóa đơn
            giahoadon = extended_cost / qty_ord_pcs, 2
            giathucte = ProcessHandler.find_price_by_sku(product["Barcode"],"Lotte")

            # 🔍 **Tìm giá trị trong CTKM.xlsx theo SKU & thời gian**
            value = ProcessHandler.find_value_by_sku_and_time(product["Barcode"], entry_date,"Lotte")

            # ✅ Nếu có giá trị, xử lý chiết khấu
            if value:
                print(f"🔹 Đã lấy được giá trị: {value}, tiếp tục xử lý...")
                self.log_signal.emit(
    f'🎉 Mã hàng <b><span style="color: blue;">{product["Barcode"]}</span></b> '
    f'có chương trình khuyến mãi <b><span style="color: green;">{value}</span></b> 🛒'
)

                # ✅ Kiểm tra nếu value chứa "%", trích xuất và áp dụng chiết khấu
                discount = ProcessHandler.extract_discount(value)
               
                
                if discount != 0:
                    if isinstance(discount, str):  # Nếu discount là chuỗi
                        discount = discount.replace("%", "").strip()  # Loại bỏ dấu %
                        discount = float(discount) if discount else 0  # Chuyển về float, nếu rỗng thì là 0

                    # Chuyển giathucte về float nếu nó là chuỗi
                    if isinstance(giathucte, str):
                        giathucte = giathucte.replace(",", "").strip()  # Loại bỏ dấu phẩy (nếu có)
                        giathucte = float(giathucte) if giathucte else 0  # Chuyển về float, nếu rỗng thì là 0

                    # Giảm giá theo %    
                    
                    giathucte = giathucte - (giathucte * float(discount) / 100)  # Tính giá trị thực tế


            
            # ✅ So sánh giá sau khi xử lý chiết khấu
            giahoadon = float(giahoadon[0]) if isinstance(giahoadon, tuple) else float(giahoadon)
            giathucte = float(giathucte[0]) if isinstance(giathucte, tuple) else float(giathucte)

            if giathucte is not None and math.isclose(giahoadon, giathucte, rel_tol=1e-3):
                sheet[f"Y{current_row}"] = giathucte
                self.log_signal.emit(
    f'✅ Mã hàng <b><span style="color: blue;">{product["Barcode"]}</span></b> '
    f'có giá chính xác! 🎯 '
    f'Giá hóa đơn: <span style="color: green; font-weight: bold;">{giahoadon}</span>')
            else:

                sheet[f"Y{current_row}"] = giathucte
                red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                sheet[f"Y{current_row}"].fill = red_fill
                comment_text = f"Kiểm tra lại giá mã này! - Giá hóa đơn: {giahoadon} - Chênh lệch: {giahoadon - giathucte}"
                comment = Comment(comment_text, "System")
                sheet[f"Y{current_row}"].comment = comment
                self.log_signal.emit(
    f"⚠️ Mã hàng <b>{product["Barcode"]}</b> sai giá! 🛑 Giá hóa đơn: <span style='color: red;'><b>{giahoadon}</b></span> - "
    f"Giá đúng: <span style='color: green;'><b>{giathucte}</b></span>"
)
                saigia += 1
                

            # 🔍 Kiểm tra SKU có trong 'SanPham'
            kiemtra = ProcessHandler.check_value_in_sanpham(value)
            match = re.search(r"(\d+)\s*\+\s*1", value)  # Tìm số đầu tiên trong biểu thức X+1
            if match:
                x = int(match.group(1))  # Chuyển đổi thành số nguyên
                kiemtra = kiemtra or product["Barcode"]
                
                if x >= 2:  # Chỉ chia khi x >= 2, còn 1+1 thì giữ nguyên
                    qty_ord_pcs = math.floor(qty_ord_pcs / x)  # Chia rồi làm tròn xuống

            
               
            if kiemtra:
                if isinstance(kiemtra, list):  # Nếu có nhiều mã SKU, nối thành chuỗi
                    kiemtra = ", ".join(kiemtra)
                current_row += 1
                self.log_signal.emit(
    f'🎁 Mã hàng <b><span style="color: blue;">{product["Barcode"]}</span></b> '
    f'đã thêm hàng khuyến mãi <b><span style="color: green;">{kiemtra}</span></b> tặng kèm! 🎉'
)
                sheet[f"A{current_row}"] = entry_date
                if current_row == 9:
                    sheet[f"B{current_row}"] = f'="ĐĐH{vendor}"&TEXT(G1,"00000")'
                elif current_row > 9:
                    sheet[f"B{current_row}"] = f'=IF(L{current_row}=L{current_row - 1},B{current_row - 1},"ĐĐH{vendor}"&TEXT(RIGHT(B{current_row - 1},4)+1,"00000"))'
                sheet[f"C{current_row}"] = "Chưa thực hiện"
                sheet[f"D{current_row}"] = cancle_date
                sheet[f"G{current_row}"] = makhachhang
                sheet[f"L{current_row}"] = f"{vendor} PO{po_number}"
                sheet[f"Q{current_row}"] = kiemtra
                sheet[f"X{current_row}"] = qty_ord_pcs
                sheet[f"V{current_row}"] = kho
                sheet[f"Y{current_row}"] = 0
                sheet[f"Z{current_row}"] = 0
                sheet[f"U{current_row}"] = "Có"
                sheet[f"AJ{current_row}"] = khuvuc
                sheet[f"AE{current_row}"] = 10
                sheet[f"AM{current_row}"] = mien
                sheet[f"T{current_row}"] = "Không"
                sheet[f"E{current_row}"] = diachigiaohang
                sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(kiemtra)
            sheet["G1"] = STT_donhang
            current_row += 1

        # Lưu lại file Excel
        wb.save(file_path)
        print(f"✅ Đã ghi {len(products)} dòng vào '{file_path}', sheet '{sheet_name}'!")
        return saigia
    





    

    


    def write_to_dondathang_satra(self,products, makhachhang, po_number, entry_date, cancle_date,STT_donhang,vendor,diachigiao):

        vendor = vendor.upper()
        """Ghi danh sách sản phẩm vào file dondathang.xlsx"""
        file_path = "dondathang.xlsx"
        sheet_name = "Don dat hang"
        
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]

        start_row = sheet.max_row + 1
        current_row = start_row
        # Lấy giá trị G1, đảm bảo nó là số
        print(f"📍 PO Number {po_number} ")
        STT_donhang_str = f"{STT_donhang:05d}"     
        saigia = 0
        if makhachhang[:2] == "MB":
            khuvuc = "MT_MB"
            kho = "TP_HN_5"
            mien = "HN"
        else:
            kho = "LA_KHO1"
            khuvuc = "MT_MN"
            mien = "LA"




        sheet[f"A{current_row}"] = entry_date
        if current_row == 9:
            sheet[f"B{current_row}"] = f'="ĐĐH{vendor}"&TEXT(G1,"00000")'
        elif current_row > 9:
            sheet[f"B{current_row}"] = f'=IF(L{current_row}=L{current_row - 1},B{current_row - 1},"ĐĐH{vendor}"&TEXT(RIGHT(B{current_row - 1},4)+1,"00000"))'
        sheet[f"C{current_row}"] = "Chưa thực hiện"
        sheet[f"D{current_row}"] = cancle_date
        sheet[f"G{current_row}"] = makhachhang
        sheet[f"L{current_row}"] = f"{vendor} {po_number}"
        sheet[f"V{current_row}"] = kho
        sheet[f"AE{current_row}"] = 10
        sheet[f"AJ{current_row}"] = khuvuc
        sheet[f"AM{current_row}"] = mien
        sheet[f"U{current_row}"] = "Không"
        sheet[f"Z{current_row}"] = 0
        sheet[f"S{current_row}"] = f"{vendor} PO{po_number}"
        sheet[f"T{current_row}"] = "Có"
        sheet[f"X{current_row}"] = 0
        sheet[f"Y{current_row}"] = 0
        sheet[f"E{current_row}"] = diachigiao
        current_row += 1

        for product in products:
            barcode = product["Article"]
            soluong = float(product["OU Qty"])
            
            dongia = float(product["Net Purchase Price"])
           
            sku_mapping = ProcessHandler.load_sku_mapping()  # Load mapping SKU
            products = ProcessHandler.replace_sku_numbers(products, sku_mapping)  # Thay SKU
            sheet[f"A{current_row}"] = entry_date
            if current_row == 9:
                sheet[f"B{current_row}"] = f'="ĐĐH{vendor}"&TEXT(G1,"00000")'
            elif current_row > 9:
                sheet[f"B{current_row}"] = f'=IF(L{current_row}=L{current_row - 1},B{current_row - 1},"ĐĐH{vendor}"&TEXT(RIGHT(B{current_row - 1},4)+1,"00000"))'
            sheet[f"C{current_row}"] = "Chưa thực hiện"
            sheet[f"D{current_row}"] = cancle_date
            sheet[f"G{current_row}"] = makhachhang
            sheet[f"L{current_row}"] = f"{vendor} {po_number}"
            sheet[f"Q{current_row}"] = product["Article"]
            sheet[f"V{current_row}"] = kho
            sheet[f"AE{current_row}"] = 10
            sheet[f"AJ{current_row}"] = khuvuc
            sheet[f"AM{current_row}"] = mien
            sheet[f"U{current_row}"] = "Không"
            sheet[f"E{current_row}"] = diachigiao
            sheet[f"T{current_row}"] = "Không"
            sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(product["Article"])

            # ✅ Xử lý "Extended Cost"
            extended_cost = soluong * dongia
            if isinstance(extended_cost, str):  # Nếu là chuỗi, loại bỏ dấu phẩy
                extended_cost = float(extended_cost.replace(",", ""))
            extended_cost = int(extended_cost) if extended_cost.is_integer() else extended_cost
            sheet[f"Z{current_row}"] = f"=Y{current_row}*X{current_row}"

            # ✅ Xử lý "Qty Ord/Pcs"
            qty_ord_pcs = product["OU Qty"]
            if isinstance(qty_ord_pcs, str):  # Nếu là chuỗi, loại bỏ dấu phẩy
                qty_ord_pcs = float(qty_ord_pcs.replace(",", ""))
            qty_ord_pcs = int(qty_ord_pcs) if qty_ord_pcs.is_integer() else qty_ord_pcs
            sheet[f"X{current_row}"] = qty_ord_pcs

            # ✅ Tính giá hóa đơn
            giahoadon = dongia
            giathucte = ProcessHandler.find_price_by_sku(barcode,"SATRA")

            # 🔍 **Tìm giá trị trong CTKM.xlsx theo SKU & thời gian**
            value = ProcessHandler.find_value_by_sku_and_time(barcode, entry_date,"SATRA")

            # ✅ Nếu có giá trị, xử lý chiết khấu
            if value:
                print(f"🔹 Đã lấy được giá trị: {value}, tiếp tục xử lý...")
                self.log_signal.emit(
    f'🎉 Mã hàng <b><span style="color: blue;">{barcode}</span></b> '
    f'có chương trình khuyến mãi <b><span style="color: green;">{value}</span></b> 🛒'
)

                # ✅ Kiểm tra nếu value chứa "%", trích xuất và áp dụng chiết khấu
                discount = ProcessHandler.extract_discount(value)
                
                if discount != 0:

                    if isinstance(discount, str):  # Nếu discount là chuỗi
                        discount = discount.replace("%", "").strip()  # Loại bỏ dấu %
                        discount = float(discount) if discount else 0  # Chuyển về float, nếu rỗng thì là 0

                    # Chuyển giathucte về float nếu nó là chuỗi
                    if isinstance(giathucte, str):
                        giathucte = giathucte.replace(",", "").strip()  # Loại bỏ dấu phẩy (nếu có)
                        giathucte = float(giathucte) if giathucte else 0  # Chuyển về float, nếu rỗng thì là 0

                    # Giảm giá theo %    
                    giathucte = giathucte - (giathucte * float(discount) / 100)
                    



            
            # ✅ So sánh giá sau khi xử lý chiết khấu
            giahoadon = float(giahoadon[0]) if isinstance(giahoadon, tuple) else float(giahoadon)
            giathucte = float(giathucte[0]) if isinstance(giathucte, tuple) else float(giathucte)

            if giathucte is not None and math.isclose(giahoadon, giathucte, rel_tol=1e-3):
                sheet[f"Y{current_row}"] = giathucte
                self.log_signal.emit(
    f'✅ Mã hàng <b><span style="color: blue;">{barcode}</span></b> '
    f'có giá chính xác! 🎯 '
    f'Giá hóa đơn: <span style="color: green; font-weight: bold;">{giahoadon}</span>')
            else:

                sheet[f"Y{current_row}"] = giathucte
                red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                sheet[f"Y{current_row}"].fill = red_fill
                comment_text = f"Kiểm tra lại giá mã này! - Giá hóa đơn: {giahoadon} - Chênh lệch: {giahoadon - giathucte}"
                comment = Comment(comment_text, "System")
                sheet[f"Y{current_row}"].comment = comment
                self.log_signal.emit(
    f"⚠️ Mã hàng <b>{barcode}</b> sai giá! 🛑 Giá hóa đơn: <span style='color: red;'><b>{giahoadon}</b></span> - "
    f"Giá đúng: <span style='color: green;'><b>{giathucte}</b></span>"
)
                saigia += 1
                

            # 🔍 Kiểm tra SKU có trong 'SanPham'
            kiemtra = ProcessHandler.check_value_in_sanpham(value)
            match = re.search(r"(\d+)\s*\+\s*1", value)  # Tìm số đầu tiên trong biểu thức X+1
            if match:
                x = int(match.group(1))  # Chuyển đổi thành số nguyên
                kiemtra = kiemtra or barcode
                
                if x >= 2:  # Chỉ chia khi x >= 2, còn 1+1 thì giữ nguyên
                    qty_ord_pcs = math.floor(qty_ord_pcs / x)  # Chia rồi làm tròn xuống

            
               
            if kiemtra:
                if isinstance(kiemtra, list):  # Nếu có nhiều mã SKU, nối thành chuỗi
                    kiemtra = ", ".join(kiemtra)
                current_row += 1
                self.log_signal.emit(
    f'🎁 Mã hàng <b><span style="color: blue;">{barcode}</span></b> '
    f'đã thêm hàng khuyến mãi <b><span style="color: green;">{kiemtra}</span></b> tặng kèm! 🎉'
)
                sheet[f"A{current_row}"] = entry_date
                if current_row == 9:
                    sheet[f"B{current_row}"] = f'="ĐĐH{vendor}"&TEXT(G1,"00000")'
                elif current_row > 9:
                    sheet[f"B{current_row}"] = f'=IF(L{current_row}=L{current_row - 1},B{current_row - 1},"ĐĐH{vendor}"&TEXT(RIGHT(B{current_row - 1},4)+1,"00000"))'
                sheet[f"C{current_row}"] = "Chưa thực hiện"
                sheet[f"D{current_row}"] = cancle_date
                sheet[f"G{current_row}"] = makhachhang
                sheet[f"L{current_row}"] = f"{vendor} {po_number}"
                sheet[f"Q{current_row}"] = kiemtra
                sheet[f"X{current_row}"] = qty_ord_pcs
                sheet[f"V{current_row}"] = kho
                sheet[f"Y{current_row}"] = 0
                sheet[f"Z{current_row}"] = 0
                sheet[f"U{current_row}"] = "Có"
                sheet[f"AJ{current_row}"] = khuvuc
                sheet[f"AE{current_row}"] = 10
                sheet[f"AM{current_row}"] = mien
                sheet[f"E{current_row}"] = diachigiao
                sheet[f"T{current_row}"] = "Không"
                sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(kiemtra)
            sheet["G1"] = STT_donhang
            current_row += 1

        # Lưu lại file Excel
        wb.save(file_path)
        print(f"✅ Đã ghi {len(products)} dòng vào '{file_path}', sheet '{sheet_name}'!")
        return saigia
    




    def write_to_dondathang_fujimart(self,products, makhachhang, po_number, entry_date, cancle_date,STT_donhang,vendor,diachigiao):

        vendor = vendor.upper()
        """Ghi danh sách sản phẩm vào file dondathang.xlsx"""
        file_path = "dondathang.xlsx"
        sheet_name = "Don dat hang"
        
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]

        start_row = sheet.max_row + 1
        current_row = start_row
        # Lấy giá trị G1, đảm bảo nó là số
        print(f"📍 PO Number {po_number} ")
        STT_donhang_str = f"{STT_donhang:05d}"     
        saigia = 0
        if makhachhang[:2] == "MB":
            khuvuc = "MT_MB"
            kho = "TP_HN_5"
            mien = "HN"
        else:
            kho = "LA_KHO1"
            khuvuc = "MT_MN"
            mien = "LA"




        sheet[f"A{current_row}"] = entry_date
        if current_row == 9:
            sheet[f"B{current_row}"] = f'="ĐĐH{vendor}"&TEXT(G1,"00000")'
        elif current_row > 9:
            sheet[f"B{current_row}"] = f'=IF(L{current_row}=L{current_row - 1},B{current_row - 1},"ĐĐH{vendor}"&TEXT(RIGHT(B{current_row - 1},4)+1,"00000"))'
        sheet[f"C{current_row}"] = "Chưa thực hiện"
        sheet[f"D{current_row}"] = cancle_date
        sheet[f"G{current_row}"] = makhachhang
        sheet[f"L{current_row}"] = f"{vendor} PO{po_number}"
        sheet[f"V{current_row}"] = kho
        sheet[f"AE{current_row}"] = 10
        sheet[f"AJ{current_row}"] = khuvuc
        sheet[f"AM{current_row}"] = mien
        sheet[f"U{current_row}"] = "Không"
        sheet[f"Z{current_row}"] = 0
        sheet[f"S{current_row}"] = f"{vendor} PO{po_number}"
        sheet[f"T{current_row}"] = "Có"
        sheet[f"X{current_row}"] = 0
        sheet[f"Y{current_row}"] = 0
        sheet[f"E{current_row}"] = diachigiao
        current_row += 1

        for product in products:
            barcode = product["Article"]
            soluong = float(product["OU Qty"])
            
            dongia = float(product["Net Purchase Price"])
           
            
            sheet[f"A{current_row}"] = entry_date
            if current_row == 9:
                sheet[f"B{current_row}"] = f'="ĐĐH{vendor}"&TEXT(G1,"00000")'
            elif current_row > 9:
                sheet[f"B{current_row}"] = f'=IF(L{current_row}=L{current_row - 1},B{current_row - 1},"ĐĐH{vendor}"&TEXT(RIGHT(B{current_row - 1},4)+1,"00000"))'
            sheet[f"C{current_row}"] = "Chưa thực hiện"
            sheet[f"D{current_row}"] = cancle_date
            sheet[f"G{current_row}"] = makhachhang
            sheet[f"L{current_row}"] = f"{vendor} PO{po_number}"
            sheet[f"Q{current_row}"] = product["Article"]
            sheet[f"V{current_row}"] = kho
            sheet[f"AE{current_row}"] = 10
            sheet[f"AJ{current_row}"] = khuvuc
            sheet[f"AM{current_row}"] = mien
            sheet[f"U{current_row}"] = "Không"
            sheet[f"E{current_row}"] = diachigiao
            sheet[f"T{current_row}"] = "Không"
            sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(product["Article"])

            # ✅ Xử lý "Extended Cost"
            extended_cost =  dongia
            if isinstance(extended_cost, str):  # Nếu là chuỗi, loại bỏ dấu phẩy
                extended_cost = float(extended_cost.replace(",", ""))
            extended_cost = int(extended_cost) if extended_cost.is_integer() else extended_cost
            sheet[f"Z{current_row}"] = f"=Y{current_row}*X{current_row}"

            # ✅ Xử lý "Qty Ord/Pcs"
            qty_ord_pcs = product["OU Qty"]
            if isinstance(qty_ord_pcs, str):  # Nếu là chuỗi, loại bỏ dấu phẩy
                qty_ord_pcs = float(qty_ord_pcs.replace(",", ""))
            qty_ord_pcs = int(qty_ord_pcs) if qty_ord_pcs.is_integer() else qty_ord_pcs
            sheet[f"X{current_row}"] = qty_ord_pcs

            # ✅ Tính giá hóa đơn
            giahoadon = dongia / qty_ord_pcs
            giathucte = ProcessHandler.find_price_by_sku(barcode,"FUJIMART")

            # 🔍 **Tìm giá trị trong CTKM.xlsx theo SKU & thời gian**
            value = ProcessHandler.find_value_by_sku_and_time(barcode, entry_date,"FUJIMART")

            # ✅ Nếu có giá trị, xử lý chiết khấu
            if value:
                print(f"🔹 Đã lấy được giá trị: {value}, tiếp tục xử lý...")
                self.log_signal.emit(
    f'🎉 Mã hàng <b><span style="color: blue;">{barcode}</span></b> '
    f'có chương trình khuyến mãi <b><span style="color: green;">{value}</span></b> 🛒'
)

                # ✅ Kiểm tra nếu value chứa "%", trích xuất và áp dụng chiết khấu
                discount = ProcessHandler.extract_discount(value)
                
                if discount != 0:

                    if isinstance(discount, str):  # Nếu discount là chuỗi
                        discount = discount.replace("%", "").strip()  # Loại bỏ dấu %
                        discount = float(discount) if discount else 0  # Chuyển về float, nếu rỗng thì là 0

                    # Chuyển giathucte về float nếu nó là chuỗi
                    if isinstance(giathucte, str):
                        giathucte = giathucte.replace(",", "").strip()  # Loại bỏ dấu phẩy (nếu có)
                        giathucte = float(giathucte) if giathucte else 0  # Chuyển về float, nếu rỗng thì là 0

                    # Giảm giá theo %    
                    giathucte = giathucte - (giathucte * float(discount) / 100)
                    



            
            # ✅ So sánh giá sau khi xử lý chiết khấu
            giahoadon = float(giahoadon[0]) if isinstance(giahoadon, tuple) else float(giahoadon)
            giathucte = float(giathucte[0]) if isinstance(giathucte, tuple) else float(giathucte)

            if giathucte is not None and math.isclose(giahoadon, giathucte, rel_tol=1e-3):
                sheet[f"Y{current_row}"] = giathucte
                self.log_signal.emit(
    f'✅ Mã hàng <b><span style="color: blue;">{barcode}</span></b> '
    f'có giá chính xác! 🎯 '
    f'Giá hóa đơn: <span style="color: green; font-weight: bold;">{giahoadon}</span>')
            else:

                sheet[f"Y{current_row}"] = giathucte
                red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                sheet[f"Y{current_row}"].fill = red_fill
                comment_text = f"Kiểm tra lại giá mã này! - Giá hóa đơn: {giahoadon} - Chênh lệch: {giahoadon - giathucte}"
                comment = Comment(comment_text, "System")
                sheet[f"Y{current_row}"].comment = comment
                self.log_signal.emit(
    f"⚠️ Mã hàng <b>{barcode}</b> sai giá! 🛑 Giá hóa đơn: <span style='color: red;'><b>{giahoadon}</b></span> - "
    f"Giá đúng: <span style='color: green;'><b>{giathucte}</b></span>"
)
                saigia += 1
                

            # 🔍 Kiểm tra SKU có trong 'SanPham'
            kiemtra = ProcessHandler.check_value_in_sanpham(value)
            match = re.search(r"(\d+)\s*\+\s*1", value)  # Tìm số đầu tiên trong biểu thức X+1
            if match:
                x = int(match.group(1))  # Chuyển đổi thành số nguyên
                kiemtra = kiemtra or barcode
                
                if x >= 2:  # Chỉ chia khi x >= 2, còn 1+1 thì giữ nguyên
                    qty_ord_pcs = math.floor(qty_ord_pcs / x)  # Chia rồi làm tròn xuống

            
               
            if kiemtra:
                if isinstance(kiemtra, list):  # Nếu có nhiều mã SKU, nối thành chuỗi
                    kiemtra = ", ".join(kiemtra)
                current_row += 1
                self.log_signal.emit(
    f'🎁 Mã hàng <b><span style="color: blue;">{barcode}</span></b> '
    f'đã thêm hàng khuyến mãi <b><span style="color: green;">{kiemtra}</span></b> tặng kèm! 🎉'
)
                sheet[f"A{current_row}"] = entry_date
                if current_row == 9:
                    sheet[f"B{current_row}"] = f'="ĐĐH{vendor}"&TEXT(G1,"00000")'
                elif current_row > 9:
                    sheet[f"B{current_row}"] = f'=IF(L{current_row}=L{current_row - 1},B{current_row - 1},"ĐĐH{vendor}"&TEXT(RIGHT(B{current_row - 1},4)+1,"00000"))'
                sheet[f"C{current_row}"] = "Chưa thực hiện"
                sheet[f"D{current_row}"] = cancle_date
                sheet[f"G{current_row}"] = makhachhang
                sheet[f"L{current_row}"] = f"{vendor} PO{po_number}"
                sheet[f"Q{current_row}"] = kiemtra
                sheet[f"X{current_row}"] = qty_ord_pcs
                sheet[f"V{current_row}"] = kho
                sheet[f"Y{current_row}"] = 0
                sheet[f"Z{current_row}"] = 0
                sheet[f"U{current_row}"] = "Có"
                sheet[f"AJ{current_row}"] = khuvuc
                sheet[f"AE{current_row}"] = 10
                sheet[f"AM{current_row}"] = mien
                sheet[f"E{current_row}"] = diachigiao
                sheet[f"T{current_row}"] = "Không"
                sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(kiemtra)
            sheet["G1"] = STT_donhang
            current_row += 1

        # Lưu lại file Excel
        wb.save(file_path)
        print(f"✅ Đã ghi {len(products)} dòng vào '{file_path}', sheet '{sheet_name}'!")
        return saigia
    


    def write_to_dondathang_farmer(self,products, makhachhang, po_number, entry_date, cancle_date,STT_donhang,vendor,diachigiao):

        vendor = vendor.upper()
        """Ghi danh sách sản phẩm vào file dondathang.xlsx"""
        file_path = "dondathang.xlsx"
        sheet_name = "Don dat hang"
        
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]

        start_row = sheet.max_row + 1
        current_row = start_row
        # Lấy giá trị G1, đảm bảo nó là số
        print(f"📍 PO Number {po_number} ")
        STT_donhang_str = f"{STT_donhang:05d}"     
        saigia = 0
        if makhachhang[:2] == "MB":
            khuvuc = "MT_MB"
            kho = "TP_HN_5"
            mien = "HN"
        else:
            kho = "LA_KHO1"
            khuvuc = "MT_MN"
            mien = "LA"




        sheet[f"A{current_row}"] = entry_date
        if current_row == 9:
            sheet[f"B{current_row}"] = f'="ĐĐH{vendor}"&TEXT(G1,"00000")'
        elif current_row > 9:
            sheet[f"B{current_row}"] = f'=IF(L{current_row}=L{current_row - 1},B{current_row - 1},"ĐĐH{vendor}"&TEXT(RIGHT(B{current_row - 1},4)+1,"00000"))'
        sheet[f"C{current_row}"] = "Chưa thực hiện"
        sheet[f"D{current_row}"] = cancle_date
        sheet[f"G{current_row}"] = makhachhang
        sheet[f"L{current_row}"] = f"{vendor} {po_number}"
        sheet[f"V{current_row}"] = kho
        sheet[f"AE{current_row}"] = 10
        sheet[f"AJ{current_row}"] = khuvuc
        sheet[f"AM{current_row}"] = mien
        sheet[f"U{current_row}"] = "Không"
        sheet[f"Z{current_row}"] = 0
        sheet[f"S{current_row}"] = f"{vendor} {po_number}"
        sheet[f"T{current_row}"] = "Có"
        sheet[f"X{current_row}"] = 0
        sheet[f"Y{current_row}"] = 0
        sheet[f"E{current_row}"] = diachigiao
        current_row += 1

        for product in products:
            barcode = product["Article"]
            soluong = float(product["OU Qty"])
            
            dongia = float(product["Net Purchase Price"])
           
            
            sheet[f"A{current_row}"] = entry_date
            if current_row == 9:
                sheet[f"B{current_row}"] = f'="ĐĐH{vendor}"&TEXT(G1,"00000")'
            elif current_row > 9:
                sheet[f"B{current_row}"] = f'=IF(L{current_row}=L{current_row - 1},B{current_row - 1},"ĐĐH{vendor}"&TEXT(RIGHT(B{current_row - 1},4)+1,"00000"))'
            sheet[f"C{current_row}"] = "Chưa thực hiện"
            sheet[f"D{current_row}"] = cancle_date
            sheet[f"G{current_row}"] = makhachhang
            sheet[f"L{current_row}"] = f"{vendor} {po_number}"
            sheet[f"Q{current_row}"] = product["Article"]
            sheet[f"V{current_row}"] = kho
            sheet[f"AE{current_row}"] = 10
            sheet[f"AJ{current_row}"] = khuvuc
            sheet[f"AM{current_row}"] = mien
            sheet[f"U{current_row}"] = "Không"
            sheet[f"E{current_row}"] = diachigiao
            sheet[f"T{current_row}"] = "Không"
            sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(product["Article"])

            # ✅ Xử lý "Extended Cost"
            extended_cost =  dongia
            if isinstance(extended_cost, str):  # Nếu là chuỗi, loại bỏ dấu phẩy
                extended_cost = float(extended_cost.replace(",", ""))
            extended_cost = int(extended_cost) if extended_cost.is_integer() else extended_cost
            sheet[f"Z{current_row}"] = f"=Y{current_row}*X{current_row}"

            # ✅ Xử lý "Qty Ord/Pcs"
            qty_ord_pcs = product["OU Qty"]
            if isinstance(qty_ord_pcs, str):  # Nếu là chuỗi, loại bỏ dấu phẩy
                qty_ord_pcs = float(qty_ord_pcs.replace(",", ""))
            qty_ord_pcs = int(qty_ord_pcs) if qty_ord_pcs.is_integer() else qty_ord_pcs
            sheet[f"X{current_row}"] = qty_ord_pcs

            # ✅ Tính giá hóa đơn
            giahoadon = dongia / qty_ord_pcs
            giathucte = ProcessHandler.find_price_by_sku(barcode,"Farmer")

            # 🔍 **Tìm giá trị trong CTKM.xlsx theo SKU & thời gian**
            value = ProcessHandler.find_value_by_sku_and_time(barcode, entry_date,"Farmer")

            # ✅ Nếu có giá trị, xử lý chiết khấu
            if value:
                print(f"🔹 Đã lấy được giá trị: {value}, tiếp tục xử lý...")
                self.log_signal.emit(
    f'🎉 Mã hàng <b><span style="color: blue;">{barcode}</span></b> '
    f'có chương trình khuyến mãi <b><span style="color: green;">{value}</span></b> 🛒'
)

                # ✅ Kiểm tra nếu value chứa "%", trích xuất và áp dụng chiết khấu
                discount = ProcessHandler.extract_discount(value)
                
                if discount != 0:

                    if isinstance(discount, str):  # Nếu discount là chuỗi
                        discount = discount.replace("%", "").strip()  # Loại bỏ dấu %
                        discount = float(discount) if discount else 0  # Chuyển về float, nếu rỗng thì là 0

                    # Chuyển giathucte về float nếu nó là chuỗi
                    if isinstance(giathucte, str):
                        giathucte = giathucte.replace(",", "").strip()  # Loại bỏ dấu phẩy (nếu có)
                        giathucte = float(giathucte) if giathucte else 0  # Chuyển về float, nếu rỗng thì là 0

                    # Giảm giá theo %    
                    giathucte = giathucte - (giathucte * float(discount) / 100)
                    



            
            # ✅ So sánh giá sau khi xử lý chiết khấu
            giahoadon = float(giahoadon[0]) if isinstance(giahoadon, tuple) else float(giahoadon)
            giathucte = float(giathucte[0]) if isinstance(giathucte, tuple) else float(giathucte)

            if giathucte is not None and math.isclose(giahoadon, giathucte, rel_tol=1e-3):
                sheet[f"Y{current_row}"] = giathucte
                self.log_signal.emit(
    f'✅ Mã hàng <b><span style="color: blue;">{barcode}</span></b> '
    f'có giá chính xác! 🎯 '
    f'Giá hóa đơn: <span style="color: green; font-weight: bold;">{giahoadon}</span>')
            else:

                sheet[f"Y{current_row}"] = giathucte
                red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                sheet[f"Y{current_row}"].fill = red_fill
                comment_text = f"Kiểm tra lại giá mã này! - Giá hóa đơn: {giahoadon} - Chênh lệch: {giahoadon - giathucte}"
                comment = Comment(comment_text, "System")
                sheet[f"Y{current_row}"].comment = comment
                self.log_signal.emit(
    f"⚠️ Mã hàng <b>{barcode}</b> sai giá! 🛑 Giá hóa đơn: <span style='color: red;'><b>{giahoadon}</b></span> - "
    f"Giá đúng: <span style='color: green;'><b>{giathucte}</b></span>"
)
                saigia += 1
                

            # 🔍 Kiểm tra SKU có trong 'SanPham'
            kiemtra = ProcessHandler.check_value_in_sanpham(value)
            match = re.search(r"(\d+)\s*\+\s*1", value)  # Tìm số đầu tiên trong biểu thức X+1
            if match:
                x = int(match.group(1))  # Chuyển đổi thành số nguyên
                kiemtra = kiemtra or barcode
                
                if x >= 2:  # Chỉ chia khi x >= 2, còn 1+1 thì giữ nguyên
                    qty_ord_pcs = math.floor(qty_ord_pcs / x)  # Chia rồi làm tròn xuống

            
               
            if kiemtra:
                if isinstance(kiemtra, list):  # Nếu có nhiều mã SKU, nối thành chuỗi
                    kiemtra = ", ".join(kiemtra)
                current_row += 1
                self.log_signal.emit(
    f'🎁 Mã hàng <b><span style="color: blue;">{barcode}</span></b> '
    f'đã thêm hàng khuyến mãi <b><span style="color: green;">{kiemtra}</span></b> tặng kèm! 🎉'
)
                sheet[f"A{current_row}"] = entry_date
                if current_row == 9:
                    sheet[f"B{current_row}"] = f'="ĐĐH{vendor}"&TEXT(G1,"00000")'
                elif current_row > 9:
                    sheet[f"B{current_row}"] = f'=IF(L{current_row}=L{current_row - 1},B{current_row - 1},"ĐĐH{vendor}"&TEXT(RIGHT(B{current_row - 1},4)+1,"00000"))'
                sheet[f"C{current_row}"] = "Chưa thực hiện"
                sheet[f"D{current_row}"] = cancle_date
                sheet[f"G{current_row}"] = makhachhang
                sheet[f"L{current_row}"] = f"{vendor} {po_number}"
                sheet[f"Q{current_row}"] = kiemtra
                sheet[f"X{current_row}"] = qty_ord_pcs
                sheet[f"V{current_row}"] = kho
                sheet[f"Y{current_row}"] = 0
                sheet[f"Z{current_row}"] = 0
                sheet[f"U{current_row}"] = "Có"
                sheet[f"AJ{current_row}"] = khuvuc
                sheet[f"AE{current_row}"] = 10
                sheet[f"AM{current_row}"] = mien
                sheet[f"E{current_row}"] = diachigiao
                sheet[f"T{current_row}"] = "Không"
                sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(kiemtra)
            sheet["G1"] = STT_donhang
            current_row += 1

        # Lưu lại file Excel
        wb.save(file_path)
        print(f"✅ Đã ghi {len(products)} dòng vào '{file_path}', sheet '{sheet_name}'!")
        return saigia
    

    def write_to_dondathang_bhx(self,products, makhachhang, po_number, entry_date, cancle_date,STT_donhang,vendor,diachigiao):

        vendor = vendor.upper()
        """Ghi danh sách sản phẩm vào file dondathang.xlsx"""
        file_path = "dondathang.xlsx"
        sheet_name = "Don dat hang"
        
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]

        start_row = sheet.max_row + 1
        current_row = start_row
        # Lấy giá trị G1, đảm bảo nó là số
        print(f"📍 PO Number {po_number} ")
        STT_donhang_str = f"{STT_donhang:05d}"     
        saigia = 0
        if makhachhang[:2] == "MB":
            khuvuc = "MT_MB"
            kho = "TP_HN_5"
            mien = "HN"
        else:
            kho = "LA_KHO1"
            khuvuc = "MT_MN"
            mien = "LA"




        sheet[f"A{current_row}"] = entry_date
        if current_row == 9:
            sheet[f"B{current_row}"] = f'="ĐĐH{vendor}"&TEXT(G1,"00000")'
        elif current_row > 9:
            sheet[f"B{current_row}"] = f'=IF(L{current_row}=L{current_row - 1},B{current_row - 1},"ĐĐH{vendor}"&TEXT(RIGHT(B{current_row - 1},4)+1,"00000"))'
        sheet[f"C{current_row}"] = "Chưa thực hiện"
        sheet[f"D{current_row}"] = cancle_date
        sheet[f"G{current_row}"] = makhachhang
        sheet[f"L{current_row}"] = f"{vendor} PO{po_number}"
        sheet[f"V{current_row}"] = kho
        sheet[f"AE{current_row}"] = 10
        sheet[f"AJ{current_row}"] = khuvuc
        sheet[f"AM{current_row}"] = mien
        sheet[f"U{current_row}"] = "Không"
        sheet[f"Z{current_row}"] = 0
        sheet[f"S{current_row}"] = f"{vendor} PO{po_number}"
        sheet[f"T{current_row}"] = "Có"
        sheet[f"X{current_row}"] = 0
        sheet[f"Y{current_row}"] = 0
        sheet[f"E{current_row}"] = diachigiao
        current_row += 1

        for product in products:
            barcode = product["Article"]
            soluong = float(product["OU Qty"])
            
            dongia = float(product["Net Purchase Price"])
           
            
            sheet[f"A{current_row}"] = entry_date
            if current_row == 9:
                sheet[f"B{current_row}"] = f'="ĐĐH{vendor}"&TEXT(G1,"00000")'
            elif current_row > 9:
                sheet[f"B{current_row}"] = f'=IF(L{current_row}=L{current_row - 1},B{current_row - 1},"ĐĐH{vendor}"&TEXT(RIGHT(B{current_row - 1},4)+1,"00000"))'
            sheet[f"C{current_row}"] = "Chưa thực hiện"
            sheet[f"D{current_row}"] = cancle_date
            sheet[f"G{current_row}"] = makhachhang
            sheet[f"L{current_row}"] = f"{vendor} PO{po_number}"
            sheet[f"Q{current_row}"] = product["Article"]
            sheet[f"V{current_row}"] = kho
            sheet[f"AE{current_row}"] = 10
            sheet[f"AJ{current_row}"] = khuvuc
            sheet[f"AM{current_row}"] = mien
            sheet[f"U{current_row}"] = "Không"
            sheet[f"E{current_row}"] = diachigiao
            sheet[f"T{current_row}"] = "Không"
            sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(product["Article"])

            # ✅ Xử lý "Extended Cost"
            extended_cost =  dongia
            if isinstance(extended_cost, str):  # Nếu là chuỗi, loại bỏ dấu phẩy
                extended_cost = float(extended_cost.replace(",", ""))
            extended_cost = int(extended_cost) if extended_cost.is_integer() else extended_cost
            sheet[f"Z{current_row}"] = f"=Y{current_row}*X{current_row}"

            # ✅ Xử lý "Qty Ord/Pcs"
            qty_ord_pcs = product["OU Qty"]
            if isinstance(qty_ord_pcs, str):  # Nếu là chuỗi, loại bỏ dấu phẩy
                qty_ord_pcs = float(qty_ord_pcs.replace(",", ""))
            qty_ord_pcs = int(qty_ord_pcs) if qty_ord_pcs.is_integer() else qty_ord_pcs
            sheet[f"X{current_row}"] = qty_ord_pcs

            # ✅ Tính giá hóa đơn
            giahoadon = dongia / qty_ord_pcs
            giathucte = ProcessHandler.find_price_by_sku(barcode,"BHX")

            # 🔍 **Tìm giá trị trong CTKM.xlsx theo SKU & thời gian**
            value = ProcessHandler.find_value_by_sku_and_time(barcode, entry_date,"BHX")

            # ✅ Nếu có giá trị, xử lý chiết khấu
            if value:
                print(f"🔹 Đã lấy được giá trị: {value}, tiếp tục xử lý...")
                self.log_signal.emit(
    f'🎉 Mã hàng <b><span style="color: blue;">{barcode}</span></b> '
    f'có chương trình khuyến mãi <b><span style="color: green;">{value}</span></b> 🛒'
)

                # ✅ Kiểm tra nếu value chứa "%", trích xuất và áp dụng chiết khấu
                discount = ProcessHandler.extract_discount(value)
                
                if discount != 0:

                    if isinstance(discount, str):  # Nếu discount là chuỗi
                        discount = discount.replace("%", "").strip()  # Loại bỏ dấu %
                        discount = float(discount) if discount else 0  # Chuyển về float, nếu rỗng thì là 0

                    # Chuyển giathucte về float nếu nó là chuỗi
                    if isinstance(giathucte, str):
                        giathucte = giathucte.replace(",", "").strip()  # Loại bỏ dấu phẩy (nếu có)
                        giathucte = float(giathucte) if giathucte else 0  # Chuyển về float, nếu rỗng thì là 0

                    # Giảm giá theo %    
                    giathucte = giathucte - (giathucte * float(discount) / 100)
                    



            
            # ✅ So sánh giá sau khi xử lý chiết khấu
            giahoadon = float(giahoadon[0]) if isinstance(giahoadon, tuple) else float(giahoadon)
            giathucte = float(giathucte[0]) if isinstance(giathucte, tuple) else float(giathucte)

            if giathucte is not None and math.isclose(giahoadon, giathucte, rel_tol=1e-3):
                sheet[f"Y{current_row}"] = giathucte
                self.log_signal.emit(
    f'✅ Mã hàng <b><span style="color: blue;">{barcode}</span></b> '
    f'có giá chính xác! 🎯 '
    f'Giá hóa đơn: <span style="color: green; font-weight: bold;">{giahoadon}</span>')
            else:

                sheet[f"Y{current_row}"] = giathucte
                red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                sheet[f"Y{current_row}"].fill = red_fill
                comment_text = f"Kiểm tra lại giá mã này! - Giá hóa đơn: {giahoadon} - Chênh lệch: {giahoadon - giathucte}"
                comment = Comment(comment_text, "System")
                sheet[f"Y{current_row}"].comment = comment
                self.log_signal.emit(
    f"⚠️ Mã hàng <b>{barcode}</b> sai giá! 🛑 Giá hóa đơn: <span style='color: red;'><b>{giahoadon}</b></span> - "
    f"Giá đúng: <span style='color: green;'><b>{giathucte}</b></span>"
)
                saigia += 1
                

            # 🔍 Kiểm tra SKU có trong 'SanPham'
            kiemtra = ProcessHandler.check_value_in_sanpham(value)
            match = re.search(r"(\d+)\s*\+\s*1", value)  # Tìm số đầu tiên trong biểu thức X+1
            if match:
                x = int(match.group(1))  # Chuyển đổi thành số nguyên
                kiemtra = kiemtra or barcode
                
                if x >= 2:  # Chỉ chia khi x >= 2, còn 1+1 thì giữ nguyên
                    qty_ord_pcs = math.floor(qty_ord_pcs / x)  # Chia rồi làm tròn xuống

            
               
            if kiemtra:
                if isinstance(kiemtra, list):  # Nếu có nhiều mã SKU, nối thành chuỗi
                    kiemtra = ", ".join(kiemtra)
                current_row += 1
                self.log_signal.emit(
    f'🎁 Mã hàng <b><span style="color: blue;">{barcode}</span></b> '
    f'đã thêm hàng khuyến mãi <b><span style="color: green;">{kiemtra}</span></b> tặng kèm! 🎉'
)
                sheet[f"A{current_row}"] = entry_date
                if current_row == 9:
                    sheet[f"B{current_row}"] = f'="ĐĐH{vendor}"&TEXT(G1,"00000")'
                elif current_row > 9:
                    sheet[f"B{current_row}"] = f'=IF(L{current_row}=L{current_row - 1},B{current_row - 1},"ĐĐH{vendor}"&TEXT(RIGHT(B{current_row - 1},4)+1,"00000"))'
                sheet[f"C{current_row}"] = "Chưa thực hiện"
                sheet[f"D{current_row}"] = cancle_date
                sheet[f"G{current_row}"] = makhachhang
                sheet[f"L{current_row}"] = f"{vendor} PO{po_number}"
                sheet[f"Q{current_row}"] = kiemtra
                sheet[f"X{current_row}"] = qty_ord_pcs
                sheet[f"V{current_row}"] = kho
                sheet[f"Y{current_row}"] = 0
                sheet[f"Z{current_row}"] = 0
                sheet[f"U{current_row}"] = "Có"
                sheet[f"AJ{current_row}"] = khuvuc
                sheet[f"AE{current_row}"] = 10
                sheet[f"AM{current_row}"] = mien
                sheet[f"E{current_row}"] = diachigiao
                sheet[f"T{current_row}"] = "Không"
                sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(kiemtra)
            sheet["G1"] = STT_donhang
            current_row += 1

        # Lưu lại file Excel
        wb.save(file_path)
        print(f"✅ Đã ghi {len(products)} dòng vào '{file_path}', sheet '{sheet_name}'!")
        return saigia





    def write_to_dondathang_kingfood(self,products, makhachhang, po_number, entry_date, cancle_date,STT_donhang,vendor):

        vendor = vendor.upper()
        """Ghi danh sách sản phẩm vào file dondathang.xlsx"""
        file_path = "dondathang.xlsx"
        sheet_name = "Don dat hang"
        
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]

        start_row = sheet.max_row + 1
        current_row = start_row
        # Lấy giá trị G1, đảm bảo nó là số
        print(f"📍 PO Number {po_number} ")
        STT_donhang_str = f"{STT_donhang:05d}"  
        saigia = 0   
        if makhachhang[:2] == "MB":
            khuvuc = "MT_MB"
            kho = "TP_HN_5"
            mien = "HN"
        else:
            kho = "LA_KHO1"
            khuvuc = "MT_MN"
            mien = "LA"


        sheet[f"A{current_row}"] = entry_date
        if current_row == 9:
            sheet[f"B{current_row}"] = f'="ĐĐH{vendor}"&TEXT(G1,"00000")'
        elif current_row > 9:
            sheet[f"B{current_row}"] = f'=IF(L{current_row}=L{current_row - 1},B{current_row - 1},"ĐĐH{vendor}"&TEXT(RIGHT(B{current_row - 1},4)+1,"00000"))'
        sheet[f"C{current_row}"] = "Chưa thực hiện"
        sheet[f"D{current_row}"] = cancle_date
        sheet[f"G{current_row}"] = makhachhang
        sheet[f"L{current_row}"] = f"{vendor} {po_number}"
        sheet[f"V{current_row}"] = kho
        sheet[f"AE{current_row}"] = 10
        sheet[f"AJ{current_row}"] = khuvuc
        sheet[f"AM{current_row}"] = mien
        sheet[f"U{current_row}"] = "Không"
        sheet[f"Z{current_row}"] = 0
        sheet[f"S{current_row}"] = f"{vendor} PO{po_number}"
        sheet[f"T{current_row}"] = "Có"
        sheet[f"X{current_row}"] = 0
        sheet[f"Y{current_row}"] = 0
        sheet[f"E{current_row}"] = "KHO SEEDLOG"


        
        current_row += 1


        for product in products:
            barcode = product["Article"]
            soluong = float(product["OU Qty"])
            dongia = float(product["Net Purchase Price"])
           
            sku_mapping = ProcessHandler.load_sku_mapping()  # Load mapping SKU
            products = ProcessHandler.replace_sku_numbers(products, sku_mapping)  # Thay SKU
            sheet[f"A{current_row}"] = entry_date
            if current_row == 9:
                sheet[f"B{current_row}"] = f'="ĐĐH{vendor}"&TEXT(G1,"00000")'
            elif current_row > 9:
                sheet[f"B{current_row}"] = f'=IF(L{current_row}=L{current_row - 1},B{current_row - 1},"ĐĐH{vendor}"&TEXT(RIGHT(B{current_row - 1},4)+1,"00000"))'
            sheet[f"C{current_row}"] = "Chưa thực hiện"
            sheet[f"D{current_row}"] = cancle_date
            sheet[f"G{current_row}"] = makhachhang
            sheet[f"L{current_row}"] = f"{vendor} {po_number}"
            sheet[f"Q{current_row}"] = product["Article"]
            sheet[f"V{current_row}"] = kho
            sheet[f"AE{current_row}"] = 10
            sheet[f"AJ{current_row}"] = khuvuc
            sheet[f"AM{current_row}"] = mien
            sheet[f"U{current_row}"] = "Không"
            sheet[f"E{current_row}"] = "KHO SEEDLOG"
            sheet[f"T{current_row}"] = "Không"
            sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(product["Article"])

            # ✅ Xử lý "Extended Cost"
            extended_cost = soluong * dongia
            if isinstance(extended_cost, str):  # Nếu là chuỗi, loại bỏ dấu phẩy
                extended_cost = float(extended_cost.replace(",", ""))
            extended_cost = int(extended_cost) if extended_cost.is_integer() else extended_cost
            sheet[f"Z{current_row}"] = f"=Y{current_row}*X{current_row}"

            # ✅ Xử lý "Qty Ord/Pcs"
            qty_ord_pcs = product["OU Qty"]
            if isinstance(qty_ord_pcs, str):  # Nếu là chuỗi, loại bỏ dấu phẩy
                qty_ord_pcs = float(qty_ord_pcs.replace(",", ""))
            qty_ord_pcs = int(qty_ord_pcs) if qty_ord_pcs.is_integer() else qty_ord_pcs
            sheet[f"X{current_row}"] = qty_ord_pcs

            # ✅ Tính giá hóa đơn
            giahoadon = dongia
            giathucte = ProcessHandler.find_price_by_sku(barcode,"KINGFOOD")

            # 🔍 **Tìm giá trị trong CTKM.xlsx theo SKU & thời gian**
            
            value = ProcessHandler.find_value_by_sku_and_time(barcode, entry_date,"KINGFOOD")

            # ✅ Nếu có giá trị, xử lý chiết khấu
            if value:
                print(f"🔹 Đã lấy được giá trị: {value}, tiếp tục xử lý...")
                self.log_signal.emit(
    f'🎉 Mã hàng <b><span style="color: blue;">{barcode}</span></b> '
    f'có chương trình khuyến mãi <b><span style="color: green;">{value}</span></b> 🛒'
)

                # ✅ Kiểm tra nếu value chứa "%", trích xuất và áp dụng chiết khấu
                discount = ProcessHandler.extract_discount(value)
                print(f'Giảm giá: {discount}')
                if discount != 0:
                    if isinstance(discount, str):  # Nếu discount là chuỗi
                        discount = discount.replace("%", "").strip()  # Loại bỏ dấu %
                        discount = float(discount) if discount else 0  # Chuyển về float, nếu rỗng thì là 0

                    # Chuyển giathucte về float nếu nó là chuỗi
                    if isinstance(giathucte, str):
                        giathucte = giathucte.replace(",", "").strip()  # Loại bỏ dấu phẩy (nếu có)
                        giathucte = float(giathucte) if giathucte else 0  # Chuyển về float, nếu rỗng thì là 0

                    # Giảm giá theo %    
                    giathucte = giathucte - (giathucte * float(discount) / 100)


            
            # ✅ So sánh giá sau khi xử lý chiết khấu
            giahoadon = float(giahoadon[0]) if isinstance(giahoadon, tuple) else float(giahoadon)
            giathucte = float(giathucte[0]) if isinstance(giathucte, tuple) else float(giathucte)

            if giathucte is not None and math.isclose(giahoadon, giathucte, rel_tol=1e-3):
                sheet[f"Y{current_row}"] = giathucte
                self.log_signal.emit(
    f'✅ Mã hàng <b><span style="color: blue;">{barcode}</span></b> '
    f'có giá chính xác! 🎯 '
    f'Giá hóa đơn: <span style="color: green; font-weight: bold;">{giahoadon}</span>')
            else:

                sheet[f"Y{current_row}"] = giathucte
                red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                sheet[f"Y{current_row}"].fill = red_fill
                comment_text = f"Kiểm tra lại giá mã này! - Giá hóa đơn: {giahoadon} - Chênh lệch: {giahoadon - giathucte}"
                comment = Comment(comment_text, "System")
                sheet[f"Y{current_row}"].comment = comment
                self.log_signal.emit(
    f"⚠️ Mã hàng <b>{barcode}</b> sai giá! 🛑 Giá hóa đơn: <span style='color: red;'><b>{giahoadon}</b></span> - "
    f"Giá đúng: <span style='color: green;'><b>{giathucte}</b></span>"
)
                saigia += 1
                

            # 🔍 Kiểm tra SKU có trong 'SanPham'
            kiemtra = ProcessHandler.check_value_in_sanpham(value)
            match = re.search(r"(\d+)\s*\+\s*1", value)  # Tìm số đầu tiên trong biểu thức X+1
            if match:
                x = int(match.group(1))  # Chuyển đổi thành số nguyên
                kiemtra = kiemtra or barcode
                
                if x >= 2:  # Chỉ chia khi x >= 2, còn 1+1 thì giữ nguyên
                    qty_ord_pcs = math.floor(qty_ord_pcs / x)  # Chia rồi làm tròn xuống
               
            if kiemtra:
                if isinstance(kiemtra, list):  # Nếu có nhiều mã SKU, nối thành chuỗi
                    kiemtra = ", ".join(kiemtra)
                current_row += 1
                self.log_signal.emit(
    f'🎁 Mã hàng <b><span style="color: blue;">{barcode}</span></b> '
    f'đã thêm hàng khuyến mãi <b><span style="color: green;">{kiemtra}</span></b> tặng kèm! 🎉'
)
                sheet[f"A{current_row}"] = entry_date
                if current_row == 9:
                    sheet[f"B{current_row}"] = f'="ĐĐH{vendor}"&TEXT(G1,"00000")'
                elif current_row > 9:
                    sheet[f"B{current_row}"] = f'=IF(L{current_row}=L{current_row - 1},B{current_row - 1},"ĐĐH{vendor}"&TEXT(RIGHT(B{current_row - 1},4)+1,"00000"))'
                sheet[f"C{current_row}"] = "Chưa thực hiện"
                sheet[f"D{current_row}"] = cancle_date
                sheet[f"G{current_row}"] = makhachhang
                sheet[f"L{current_row}"] = f"{vendor} {po_number}"
                sheet[f"Q{current_row}"] = kiemtra
                sheet[f"X{current_row}"] = qty_ord_pcs
                sheet[f"V{current_row}"] = kho
                sheet[f"Y{current_row}"] = 0
                sheet[f"Z{current_row}"] = 0
                sheet[f"U{current_row}"] = "Có"
                sheet[f"AJ{current_row}"] = khuvuc
                sheet[f"AE{current_row}"] = 10
                sheet[f"AM{current_row}"] = mien
                sheet[f"T{current_row}"] = "Không"
                sheet[f"E{current_row}"] = "KHO SEEDLOG"
                sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(kiemtra)
            sheet["G1"] = STT_donhang
            current_row += 1

        # Lưu lại file Excel
        wb.save(file_path)
        print(f"✅ Đã ghi {len(products)} dòng vào '{file_path}', sheet '{sheet_name}'!")
        return saigia






    def write_to_dondathang_winmart(self,products, makhachhang, po_number, entry_date, cancle_date,STT_donhang,vendor,diachigiaohang,ghichu):

        vendor = vendor.upper()
        """Ghi danh sách sản phẩm vào file dondathang.xlsx"""
        file_path = "dondathang.xlsx"
        sheet_name = "Don dat hang"
        
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]

        start_row = sheet.max_row + 1
        current_row = start_row
        # Lấy giá trị G1, đảm bảo nó là số
        print(f"📍 PO Number {po_number} ")
        STT_donhang_str = f"{STT_donhang:05d}"   
        saigia = 0  
        
        if makhachhang[:2] == "MB":
            khuvuc = "MT_MB"
            kho = "TP_HN_5"
            mien = "HN"
        else:
            kho = "LA_KHO1"
            khuvuc = "MT_MN"
            mien = "LA"

        if ghichu:
            po_number = f'{po_number} - {ghichu}'


        sheet[f"A{current_row}"] = entry_date
        if current_row == 9:
            sheet[f"B{current_row}"] = f'="ĐĐH{vendor}"&TEXT(G1,"00000")'
        elif current_row > 9:
            sheet[f"B{current_row}"] = f'=IF(L{current_row}=L{current_row - 1},B{current_row - 1},"ĐĐH{vendor}"&TEXT(RIGHT(B{current_row - 1},4)+1,"00000"))'
        sheet[f"C{current_row}"] = "Chưa thực hiện"
        sheet[f"D{current_row}"] = cancle_date
        sheet[f"G{current_row}"] = makhachhang
        sheet[f"L{current_row}"] = f"{vendor} PO{po_number}"
        sheet[f"V{current_row}"] = kho
        sheet[f"AE{current_row}"] = 10
        sheet[f"AJ{current_row}"] = khuvuc
        sheet[f"AM{current_row}"] = mien
        sheet[f"U{current_row}"] = "Không"
        sheet[f"Z{current_row}"] = 0
        sheet[f"S{current_row}"] = f"{vendor} PO{po_number.split('-')[0]}"
        sheet[f"T{current_row}"] = "Có"
        sheet[f"X{current_row}"] = 0
        sheet[f"Y{current_row}"] = 0
        sheet[f"E{current_row}"] = diachigiaohang
        current_row += 1


        for product in products:
            barcode = product["Article"]
            soluong = float(product["OU Qty"])
            dongia = float(product["Net Purchase Price"])
           
            sku_mapping = ProcessHandler.load_sku_mapping()  # Load mapping SKU
            products = ProcessHandler.replace_sku_numbers(products, sku_mapping)  # Thay SKU
            sheet[f"A{current_row}"] = entry_date
            if current_row == 9:
                sheet[f"B{current_row}"] = f'="ĐĐH{vendor}"&TEXT(G1,"00000")'
            elif current_row > 9:
                sheet[f"B{current_row}"] = f'=IF(L{current_row}=L{current_row - 1},B{current_row - 1},"ĐĐH{vendor}"&TEXT(RIGHT(B{current_row - 1},4)+1,"00000"))'
            sheet[f"C{current_row}"] = "Chưa thực hiện"
            sheet[f"D{current_row}"] = cancle_date
            sheet[f"G{current_row}"] = makhachhang
            sheet[f"L{current_row}"] = f"{vendor} PO{po_number}"
            sheet[f"Q{current_row}"] = product["Article"]
            sheet[f"V{current_row}"] = kho
            sheet[f"AE{current_row}"] = 10
            sheet[f"AJ{current_row}"] = khuvuc
            sheet[f"AM{current_row}"] = mien
            sheet[f"U{current_row}"] = "Không"
            sheet[f"T{current_row}"] = "Không"
            sheet[f"E{current_row}"] = diachigiaohang
            sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(product["Article"])

            # ✅ Xử lý "Extended Cost"
            extended_cost = product["Net Purchase Price"]
            if isinstance(extended_cost, str):  # Nếu là chuỗi, loại bỏ dấu phẩy
                extended_cost = float(extended_cost.replace(",", ""))
            extended_cost = int(extended_cost) if extended_cost.is_integer() else extended_cost
            sheet[f"Z{current_row}"] = f"=Y{current_row}*X{current_row}"

            # ✅ Xử lý "Qty Ord/Pcs"
            qty_ord_pcs = product["OU Qty"]
            if isinstance(qty_ord_pcs, str):  # Nếu là chuỗi, loại bỏ dấu phẩy
                qty_ord_pcs = float(qty_ord_pcs.replace(",", ""))
            qty_ord_pcs = int(qty_ord_pcs) if qty_ord_pcs.is_integer() else qty_ord_pcs
            sheet[f"X{current_row}"] = qty_ord_pcs

            # ✅ Tính giá hóa đơn
            giahoadon = extended_cost / qty_ord_pcs
            giathucte = ProcessHandler.find_price_by_sku(barcode,"WINMART")

            # 🔍 **Tìm giá trị trong CTKM.xlsx theo SKU & thời gian**
            
            value = ProcessHandler.find_value_by_sku_and_time(barcode, entry_date,"WINMART")

            # ✅ Nếu có giá trị, xử lý chiết khấu
            if value:
                print(f"🔹 Đã lấy được giá trị: {value}, tiếp tục xử lý...")
                self.log_signal.emit(
    f'🎉 Mã hàng <b><span style="color: blue;">{barcode}</span></b> '
    f'có chương trình khuyến mãi <b><span style="color: green;">{value}</span></b> 🛒'
)

                # ✅ Kiểm tra nếu value chứa "%", trích xuất và áp dụng chiết khấu
                discount = ProcessHandler.extract_discount(value)
                print(f'Giảm giá: {discount}')
                if discount != 0:
                    if isinstance(discount, str):  # Nếu discount là chuỗi
                        discount = discount.replace("%", "").strip()  # Loại bỏ dấu %
                        discount = float(discount) if discount else 0  # Chuyển về float, nếu rỗng thì là 0

                    # Chuyển giathucte về float nếu nó là chuỗi
                    if isinstance(giathucte, str):
                        giathucte = giathucte.replace(",", "").strip()  # Loại bỏ dấu phẩy (nếu có)
                        giathucte = float(giathucte) if giathucte else 0  # Chuyển về float, nếu rỗng thì là 0

                    # Giảm giá theo %    
                    giathucte = giathucte - (giathucte * float(discount) / 100)


            
            # ✅ So sánh giá sau khi xử lý chiết khấu
            giahoadon = float(giahoadon[0]) if isinstance(giahoadon, tuple) else float(giahoadon)
            giathucte = float(giathucte[0]) if isinstance(giathucte, tuple) else float(giathucte)

            if giathucte is not None and math.isclose(giahoadon, giathucte, rel_tol=1e-3):
                sheet[f"Y{current_row}"] = giathucte
                self.log_signal.emit(
    f'✅ Mã hàng <b><span style="color: blue;">{barcode}</span></b> '
    f'có giá chính xác! 🎯 '
    f'Giá hóa đơn: <span style="color: green; font-weight: bold;">{giahoadon}</span>')
            else:

                sheet[f"Y{current_row}"] = giathucte
                red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                sheet[f"Y{current_row}"].fill = red_fill
                comment_text = f"Kiểm tra lại giá mã này! - Giá hóa đơn: {giahoadon} - Chênh lệch: {giahoadon - giathucte}"
                comment = Comment(comment_text, "System")
                sheet[f"Y{current_row}"].comment = comment
                self.log_signal.emit(
    f"⚠️ Mã hàng <b>{barcode}</b> sai giá! 🛑 Giá hóa đơn: <span style='color: red;'><b>{giahoadon}</b></span> - "
    f"Giá đúng: <span style='color: green;'><b>{giathucte}</b></span>"
)
                saigia += 1
                

            # 🔍 Kiểm tra SKU có trong 'SanPham'
            kiemtra = ProcessHandler.check_value_in_sanpham(value)
            match = re.search(r"(\d+)\s*\+\s*1", value)  # Tìm số đầu tiên trong biểu thức X+1
            if match:
                x = int(match.group(1))  # Chuyển đổi thành số nguyên
                kiemtra = kiemtra or barcode
                
                if x >= 2:  # Chỉ chia khi x >= 2, còn 1+1 thì giữ nguyên
                    qty_ord_pcs = math.floor(qty_ord_pcs / x)  # Chia rồi làm tròn xuống
               
            if kiemtra:
                if isinstance(kiemtra, list):  # Nếu có nhiều mã SKU, nối thành chuỗi
                    kiemtra = ", ".join(kiemtra)
                current_row += 1
                self.log_signal.emit(
    f'🎁 Mã hàng <b><span style="color: blue;">{barcode}</span></b> '
    f'đã thêm hàng khuyến mãi <b><span style="color: green;">{kiemtra}</span></b> tặng kèm! 🎉'
)
                sheet[f"A{current_row}"] = entry_date
                if current_row == 9:
                    sheet[f"B{current_row}"] = f'="ĐĐH{vendor}"&TEXT(G1,"00000")'
                elif current_row > 9:
                    sheet[f"B{current_row}"] = f'=IF(L{current_row}=L{current_row - 1},B{current_row - 1},"ĐĐH{vendor}"&TEXT(RIGHT(B{current_row - 1},4)+1,"00000"))'
                sheet[f"C{current_row}"] = "Chưa thực hiện"
                sheet[f"D{current_row}"] = cancle_date
                sheet[f"G{current_row}"] = makhachhang
                sheet[f"L{current_row}"] = f"{vendor} PO{po_number}"
                sheet[f"Q{current_row}"] = kiemtra
                sheet[f"X{current_row}"] = qty_ord_pcs
                sheet[f"V{current_row}"] = kho
                sheet[f"Y{current_row}"] = 0
                sheet[f"Z{current_row}"] = 0
                sheet[f"U{current_row}"] = "Có"
                sheet[f"AJ{current_row}"] = khuvuc
                sheet[f"AE{current_row}"] = 10
                sheet[f"AM{current_row}"] = mien
                sheet[f"T{current_row}"] = "Không"
                sheet[f"E{current_row}"] = diachigiaohang
                sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(kiemtra)
            sheet["G1"] = STT_donhang
            current_row += 1

        # Lưu lại file Excel
        wb.save(file_path)
        print(f"✅ Đã ghi {len(products)} dòng vào '{file_path}', sheet '{sheet_name}'!")
        return saigia
        



        

    def tim_gia_tri_congtrinh(congtrinh):
        file_path = "data.xlsx"
        sheet_name = "MaKH"

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb[sheet_name]

            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=3, values_only=True):
                ma_ct, gia_tri = row  # Cột B và C
                if ma_ct == congtrinh:
                    return gia_tri  # Trả về giá trị cột C

            return congtrinh.replace(" ", "")  # Không tìm thấy

        except Exception as e:
            print(f"Lỗi khi mở file: {e}")
            return None





    def write_to_dondathang_bigc(self,products,items, po_number, entry_date, cancle_date, congtrinh,STT_donhang ,makhachhang, vendor,page_num,diachigiao):

        vendor = vendor.upper()

        """Ghi danh sách sản phẩm vào file dondathang.xlsx"""
        file_path = "dondathang.xlsx"
        sheet_name = "Don dat hang"
        
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]

        start_row = sheet.max_row + 1
        current_row = start_row
        congtrinh = ProcessHandler.tim_gia_tri_congtrinh(congtrinh)


        STT_donhang_str = f"{STT_donhang:05d}"    

        items = ProcessHandler.ghepgia_donhangbigc(items,products) 

        sku_mapping = ProcessHandler.load_sku_mapping()  # Load mapping SKU
        items = ProcessHandler.replace_sku_numbers(items, sku_mapping)  # Thay SKU
        saigia = 0
        if makhachhang[:2] == "MB":
            khuvuc = "MT_MB"
            kho = "TP_HN_5"
            mien = "HN"
        else:
            kho = "LA_KHO1"
            khuvuc = "MT_MN"
            mien = "LA"
        if page_num == 1:
            sheet[f"A{current_row}"] = entry_date
            if current_row == 9:
                sheet[f"B{current_row}"] = f'="ĐĐH{vendor}"&TEXT(G1,"00000")'
            elif current_row > 9:
                sheet[f"B{current_row}"] = f'=IF(L{current_row}=L{current_row - 1},B{current_row - 1},"ĐĐH{vendor}"&TEXT(RIGHT(B{current_row - 1},4)+1,"00000"))'
            sheet[f"C{current_row}"] = "Chưa thực hiện"
            sheet[f"D{current_row}"] = cancle_date
            sheet[f"G{current_row}"] = makhachhang
            sheet[f"L{current_row}"] = f"{vendor} PO{po_number}"
            sheet[f"V{current_row}"] = kho
            sheet[f"AE{current_row}"] = 10
            sheet[f"AJ{current_row}"] = khuvuc
            sheet[f"AM{current_row}"] = mien
            sheet[f"U{current_row}"] = "Không"
            sheet[f"Z{current_row}"] = 0
            sheet[f"S{current_row}"] = f"{vendor} PO{po_number}"
            sheet[f"T{current_row}"] = "Có"
            sheet[f"X{current_row}"] = 0
            sheet[f"Y{current_row}"] = 0
            sheet[f"AN{current_row}"] = "BIGCANLAC"
            sheet[f"E{current_row}"] = diachigiao
            current_row += 1

        for item in items:
            sheet[f"A{current_row}"] = entry_date
            if current_row == 9:
                sheet[f"B{current_row}"] = f'="ĐĐH{vendor}"&TEXT(G1,"00000")'
            elif current_row > 9:
                sheet[f"B{current_row}"] = f'=IF(L{current_row}=L{current_row - 1},B{current_row - 1},"ĐĐH{vendor}"&TEXT(RIGHT(B{current_row - 1},4)+1,"00000"))'
            sheet[f"C{current_row}"] = "Chưa thực hiện"
            sheet[f"D{current_row}"] = cancle_date
            sheet[f"L{current_row}"] = f"{vendor} PO{po_number}"
            sheet[f"G{current_row}"] = makhachhang
            sheet[f"Q{current_row}"] = item["Article"]
            sheet[f"V{current_row}"] = kho
            sheet[f"AE{current_row}"] = 10
            sheet[f"AJ{current_row}"] = khuvuc
            sheet[f"AM{current_row}"] = mien
            sheet[f"U{current_row}"] = "Không"
            sheet[f"AN{current_row}"] = congtrinh #viết mã công trình
            sheet[f"T{current_row}"] = "Không"
            sheet[f"E{current_row}"] = diachigiao
            sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(item["Article"])



            sku_ou = item["SKU/OU"]
            if isinstance(sku_ou, str):  # Nếu là chuỗi, loại bỏ dấu phẩy
                sku_ou = float(sku_ou.replace(",", ""))
            sku_ou = int(sku_ou) if sku_ou.is_integer() else sku_ou

            ou_qty = item["OU Qty"]
            if isinstance(ou_qty, str):  # Nếu là chuỗi, loại bỏ dấu phẩy
                ou_qty = float(ou_qty.replace(",", ""))
            ou_qty = int(ou_qty) if ou_qty.is_integer() else ou_qty

            qty_ord_pcs = ou_qty * sku_ou

            sheet[f"X{current_row}"] = qty_ord_pcs #ghi số lượng

            giathucte = ProcessHandler.find_price_by_sku(item["Article"],"GO")
            
            # 🔍 **Tìm giá trị trong CTKM.xlsx theo SKU & thời gian**
            value = ProcessHandler.find_value_by_sku_and_time(item["Article"], entry_date,"GO")
            if value:
                print(f"🔹 Đã lấy được giá trị: {value}, tiếp tục xử lý...")
                self.log_signal.emit(
    f'🎉 Mã hàng <b><span style="color: blue;">{item["Article"]}</span></b> '
    f'có chương trình khuyến mãi <b><span style="color: green;">{value}</span></b> 🛒'
)


                # ✅ Kiểm tra nếu value chứa "%", trích xuất và áp dụng chiết khấu
                discount = ProcessHandler.extract_discount(value)
                # Đảm bảo giathucte là float
                if isinstance(giathucte, str):
                    giathucte = giathucte.replace(",", "").strip()  # Loại bỏ dấu phẩy và khoảng trắng
                    try:
                        giathucte = float(giathucte) if giathucte else 0  # Chuyển về float, nếu rỗng thì gán 0
                    except ValueError:
                        print(f"❌ Lỗi chuyển đổi giathucte: {repr(giathucte)}")
                        giathucte = 0

                # Đảm bảo discount là float
                if isinstance(discount, str):
                    discount = discount.replace("%", "").strip()  # Loại bỏ dấu %
                    try:
                        discount = float(discount) if discount else 0
                    except ValueError:
                        print(f"❌ Lỗi chuyển đổi discount: {repr(discount)}")
                        discount = 0

                # Kiểm tra lần nữa nếu giathucte vẫn không phải là số
                if not isinstance(giathucte, (int, float)):
                    print(f"❌ giathucte không phải số: {repr(giathucte)}")
                    giathucte = 0
                

                giathucte = float(giathucte - (giathucte * discount / 100))

                # Ép kiểu giahoadon
                try:
                    giahoadon = float(item["Net Purchase Price"])
                except ValueError:
                    print(f"❌ Lỗi chuyển đổi giahoadon: {repr(item['Net Purchase Price'])}")
                    giahoadon = 0

                # Debug lại trước khi so sánh
                print(f"🔍 Debug: type(giathucte) = {type(giathucte)}, value = {giathucte}")
                print(f"🔍 Debug: type(giahoadon) = {type(giahoadon)}, value = {giahoadon}")
            

                if giathucte is not None and math.isclose(giahoadon, giathucte, rel_tol=1e-3):
                    sheet[f"Y{current_row}"] = giathucte
                    self.log_signal.emit(
    f'✅ Mã hàng <b><span style="color: blue;">{item["Article"]}</span></b> '
    f'có giá chính xác! 🎯 '
    f'Giá hóa đơn: <span style="color: green; font-weight: bold;">{giahoadon}</span>'
)

                else:
                    sheet[f"Y{current_row}"] = giathucte
                    self.log_signal.emit(
    f"⚠️ Mã hàng <b>{item['Article']}</b> sai giá! 🛑 Giá hóa đơn: <span style='color: red;'><b>{giahoadon}</b></span> - "
    f"Giá đúng: <span style='color: green;'><b>{giathucte}</b></span>"
)
                    saigia += 1
                    
                    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    sheet[f"Y{current_row}"].fill = red_fill
                    comment_text = f"Kiểm tra lại giá trị này! - Giá hóa đơn: {giahoadon} - Chênh lệch: {giahoadon - giathucte}"
                    comment = Comment(comment_text, "System")
                    sheet[f"Y{current_row}"].comment = comment
                
                sheet[f"Z{current_row}"] = f"=Y{current_row}*X{current_row}"

                
                kiemtra = ProcessHandler.check_value_in_sanpham(value)
                match = re.search(r"(\d+)\s*\+\s*1", value)  # Tìm số đầu tiên trong biểu thức X+1
                if match:
                    x = int(match.group(1))  # Chuyển đổi thành số nguyên
                    kiemtra = kiemtra or item["Article"]
                    
                    if x >= 2:  # Chỉ chia khi x >= 2, còn 1+1 thì giữ nguyên
                        qty_ord_pcs = math.floor(qty_ord_pcs / x)  # Chia rồi làm tròn xuống
                
                if kiemtra:
                    if isinstance(kiemtra, list):  # Nếu có nhiều mã SKU, nối thành chuỗi
                        kiemtra = ", ".join(kiemtra)
                    current_row += 1
                    self.log_signal.emit(
    f'🎁 Mã hàng <b><span style="color: blue;">{item["Article"]}</span></b> '
    f'đã thêm hàng khuyến mãi <b><span style="color: green;">{kiemtra}</span></b> tặng kèm! 🎉'
)
                    sheet[f"A{current_row}"] = entry_date
                    if current_row == 9:
                        sheet[f"B{current_row}"] = f'="ĐĐH{vendor}"&TEXT(G1,"00000")'
                    elif current_row > 9:
                        sheet[f"B{current_row}"] = f'=IF(L{current_row}=L{current_row - 1},B{current_row - 1},"ĐĐH{vendor}"&TEXT(RIGHT(B{current_row - 1},4)+1,"00000"))'
                    sheet[f"C{current_row}"] = "Chưa thực hiện"
                    sheet[f"D{current_row}"] = cancle_date
                    sheet[f"G{current_row}"] = makhachhang
                    sheet[f"L{current_row}"] = f"{vendor} PO{po_number}"
                    sheet[f"Q{current_row}"] = kiemtra
                    sheet[f"X{current_row}"] = qty_ord_pcs
                    sheet[f"V{current_row}"] = kho
                    sheet[f"Y{current_row}"] = 0
                    sheet[f"Z{current_row}"] = 0
                    sheet[f"U{current_row}"] = "Có"
                    sheet[f"AJ{current_row}"] = khuvuc
                    sheet[f"AE{current_row}"] = 10
                    sheet[f"AM{current_row}"] = mien
                    sheet[f"AN{current_row}"] = congtrinh
                    sheet[f"T{current_row}"] = "Không"
                    sheet[f"E{current_row}"] = diachigiao
                    sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(kiemtra)

            current_row += 1
        sheet["G1"] = STT_donhang
        # Lưu lại file Excel
        wb.save(file_path)
        print(f"✅ Đã ghi {len(products)} dòng vào '{file_path}', sheet '{sheet_name}'!")
        return saigia

    def write_to_dondathang_emart(self,items, po_number, entry_date, cancle_date, congtrinh,STT_donhang ,makhachhang="MN_MT_KH0032", vendor="Emart"):


        vendor = vendor.upper()
        """Ghi danh sách sản phẩm vào file dondathang.xlsx"""
        file_path = "dondathang.xlsx"
        sheet_name = "Don dat hang"
        
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]
        diachigiaohang = congtrinh

        start_row = sheet.max_row + 1
        current_row = start_row
        mapping = {
    "EMART GO VAP": "PVT",
    "EMART PHI": "PHI",
    "EMART SALA": "SALA"
}
        congtrinh = mapping.get(congtrinh, congtrinh)


        STT_donhang_str = f"{STT_donhang:05d}"    

        sku_mapping = ProcessHandler.load_sku_mapping()  # Load mapping SKU
        items = ProcessHandler.replace_sku_numbers(items, sku_mapping)  # Thay SKU
        saigia = 0
        if makhachhang[:2] == "MB":
            khuvuc = "MT_MB"
            kho = "TP_HN_5"
            mien = "HN"
        else:
            kho = "LA_KHO1"
            khuvuc = "MT_MN"
            mien = "LA"


        sheet[f"A{current_row}"] = entry_date
        if current_row == 9:
            sheet[f"B{current_row}"] = f'="ĐĐH{vendor}"&TEXT(G1,"00000")'
        elif current_row > 9:
            sheet[f"B{current_row}"] = f'=IF(L{current_row}=L{current_row - 1},B{current_row - 1},"ĐĐH{vendor}"&TEXT(RIGHT(B{current_row - 1},4)+1,"00000"))'
        sheet[f"C{current_row}"] = "Chưa thực hiện"
        sheet[f"D{current_row}"] = cancle_date
        sheet[f"G{current_row}"] = makhachhang
        sheet[f"L{current_row}"] = f"{vendor} PO{po_number}"
        sheet[f"V{current_row}"] = kho
        sheet[f"AE{current_row}"] = 10
        sheet[f"AJ{current_row}"] = khuvuc
        sheet[f"AM{current_row}"] = mien
        sheet[f"U{current_row}"] = "Không"
        sheet[f"Z{current_row}"] = 0
        sheet[f"S{current_row}"] = f"{vendor} PO{po_number}"
        sheet[f"T{current_row}"] = "Có"
        sheet[f"X{current_row}"] = 0
        sheet[f"Y{current_row}"] = 0
        sheet[f"E{current_row}"] = diachigiaohang

        
        sheet[f"AN{current_row}"] = 'PVT' #viết mã công trình
        current_row += 1

        for item in items:
            sheet[f"A{current_row}"] = entry_date
            if current_row == 9:
                sheet[f"B{current_row}"] = f'="ĐĐH{vendor}"&TEXT(G1,"00000")'
            elif current_row > 9:
                sheet[f"B{current_row}"] = f'=IF(L{current_row}=L{current_row - 1},B{current_row - 1},"ĐĐH{vendor}"&TEXT(RIGHT(B{current_row - 1},4)+1,"00000"))'
            sheet[f"C{current_row}"] = "Chưa thực hiện"
            sheet[f"D{current_row}"] = cancle_date
            sheet[f"L{current_row}"] = f"{vendor} PO{po_number}"
            sheet[f"G{current_row}"] = makhachhang
            sheet[f"Q{current_row}"] = item["Article"]
            sheet[f"V{current_row}"] = kho
            sheet[f"AE{current_row}"] = 10
            sheet[f"AJ{current_row}"] = khuvuc
            sheet[f"AM{current_row}"] = mien
            sheet[f"U{current_row}"] = "Không"
            sheet[f"E{current_row}"] = diachigiaohang

            sheet[f"AN{current_row}"] = congtrinh #viết mã công trình
            sheet[f"T{current_row}"] = "Không"
            sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(item["Article"])

            

            ou_qty = item["OU Qty"]
            if isinstance(ou_qty, str):  # Nếu là chuỗi, loại bỏ dấu phẩy
                ou_qty = float(ou_qty.replace(",", ""))
            ou_qty = int(ou_qty) if ou_qty.is_integer() else ou_qty

            qty_ord_pcs = ou_qty 

            sheet[f"X{current_row}"] = qty_ord_pcs #ghi số lượng

            giathucte = ProcessHandler.find_price_by_sku(item["Article"],"Emart")
            
            # 🔍 **Tìm giá trị trong CTKM.xlsx theo SKU & thời gian**
            value = ProcessHandler.find_value_by_sku_and_time(item["Article"], entry_date,"Emart")
            if value:
                print(f"🔹 Đã lấy được giá trị: {value}, tiếp tục xử lý...")
                self.log_signal.emit(
    f'🎉 Mã hàng <b><span style="color: blue;">{item["Article"]}</span></b> '
    f'có chương trình khuyến mãi <b><span style="color: green;">{value}</span></b> 🛒'
)


                # ✅ Kiểm tra nếu value chứa "%", trích xuất và áp dụng chiết khấu
                discount = ProcessHandler.extract_discount(value)
                if discount:
                    if isinstance(giathucte, str):
                        giathucte = giathucte.replace(",", "").strip()  # Loại bỏ dấu phẩy (nếu có)
                        giathucte = float(giathucte) if giathucte else 0  # Chuyển về float, nếu rỗng thì là 0
                    if isinstance(discount, str):  # Nếu discount là chuỗi
                        discount = discount.replace("%", "").strip()  # Loại bỏ dấu %
                        discount = float(discount) if discount else 0  # Chuyển về float, nếu rỗng thì là 0
                    
                    
                    
                    
                    giathucte = giathucte - (giathucte * discount / 100)

                
                
            
                giahoadon = float(item["Net Purchase Price"])
                
            
            
                giathucte = float(giathucte) if isinstance(giathucte, str) else giathucte
                giahoadon = float(giahoadon) if isinstance(giahoadon, str) else giahoadon

                if giathucte is not None and math.isclose(giahoadon, giathucte, rel_tol=1e-3):
                    sheet[f"Y{current_row}"] = giathucte
                    self.log_signal.emit(
    f'✅ Mã hàng <b><span style="color: blue;">{item["Article"]}</span></b> '
    f'có giá chính xác! 🎯 '
    f'Giá hóa đơn: <span style="color: green; font-weight: bold;">{giahoadon}</span>'
)

                else:
                    sheet[f"Y{current_row}"] = giathucte
                    self.log_signal.emit(
    f"⚠️ Mã hàng <b>{item['Article']}</b> sai giá! 🛑 Giá hóa đơn: <span style='color: red;'><b>{giahoadon}</b></span> - "
    f"Giá đúng: <span style='color: green;'><b>{giathucte}</b></span>"
)
                    saigia += 1
                    

                    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    sheet[f"Y{current_row}"].fill = red_fill
                    comment_text = f"Kiểm tra lại giá trị này! - Giá hóa đơn: {giahoadon} - Chênh lệch: {giahoadon - giathucte}"
                    comment = Comment(comment_text, "System")
                    sheet[f"Y{current_row}"].comment = comment
                
                sheet[f"Z{current_row}"] = f"=Y{current_row}*X{current_row}"

                
                kiemtra = ProcessHandler.check_value_in_sanpham(value)

                match = re.search(r"(\d+)\s*\+\s*1", value)  # Tìm số đầu tiên trong biểu thức X+1
                if match:
                    x = int(match.group(1))  # Chuyển đổi thành số nguyên
                    kiemtra = kiemtra or item["Article"]
                    
                    if x >= 2:  # Chỉ chia khi x >= 2, còn 1+1 thì giữ nguyên
                        qty_ord_pcs = math.floor(qty_ord_pcs / x)  # Chia rồi làm tròn xuống
                
                if kiemtra:
                    if isinstance(kiemtra, list):  # Nếu có nhiều mã SKU, nối thành chuỗi
                        kiemtra = ", ".join(kiemtra)
                    current_row += 1
                    self.log_signal.emit(
    f'🎁 Mã hàng <b><span style="color: blue;">{item["Article"]}</span></b> '
    f'đã thêm hàng khuyến mãi <b><span style="color: green;">{kiemtra}</span></b> tặng kèm! 🎉'
)
                    sheet[f"A{current_row}"] = entry_date
                    if current_row == 9:
                        sheet[f"B{current_row}"] = f'="ĐĐH{vendor}"&TEXT(G1,"00000")'
                    elif current_row > 9:
                        sheet[f"B{current_row}"] = f'=IF(L{current_row}=L{current_row - 1},B{current_row - 1},"ĐĐH{vendor}"&TEXT(RIGHT(B{current_row - 1},4)+1,"00000"))'
                    sheet[f"C{current_row}"] = "Chưa thực hiện"
                    sheet[f"D{current_row}"] = cancle_date
                    sheet[f"G{current_row}"] = makhachhang
                    sheet[f"L{current_row}"] = f"{vendor} PO{po_number}"
                    sheet[f"Q{current_row}"] = kiemtra
                    sheet[f"X{current_row}"] = qty_ord_pcs
                    sheet[f"V{current_row}"] = kho
                    sheet[f"Y{current_row}"] = 0
                    sheet[f"Z{current_row}"] = 0
                    sheet[f"U{current_row}"] = "Có"
                    sheet[f"AJ{current_row}"] = khuvuc
                    sheet[f"AE{current_row}"] = 10
                    sheet[f"AM{current_row}"] = mien
                    sheet[f"AN{current_row}"] = congtrinh
                    sheet[f"T{current_row}"] = "Không"
                    sheet[f"E{current_row}"] = diachigiaohang
                    sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(kiemtra)

            current_row += 1
        sheet["G1"] = STT_donhang
        # Lưu lại file Excel
        wb.save(file_path)
        print(f"✅ Đã ghi {len(items)} dòng vào '{file_path}', sheet '{sheet_name}'!")
        return saigia


    def xoa_du_lieu_don_dat_hang():
        """
        Xóa toàn bộ dòng từ dòng 9 đến hết trong file 'dondathang.xlsx', sheet 'Don dat hang'.
        """
        file_path = "dondathang.xlsx"
        sheet_name = "Don dat hang"

        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]

        max_row = sheet.max_row

        if max_row >= 9:
            sheet.delete_rows(9, max_row - 8)  # Xóa từ dòng 9 đến hết

        wb.save(file_path)
        wb.close()
        print(f"Đã xóa toàn bộ dòng từ 9 đến {max_row} trong sheet '{sheet_name}'.")

    @staticmethod
    def clean_sku_number(sku):
        """Tìm định dạng 'xxxxxx-x', loại bỏ phần '-x' và bỏ các ký tự thừa."""
        match = re.search(r"(\d{7})-\d", sku)  # Tìm chuỗi có dạng "xxxxxx-x"
        return match.group(1) if match else sku  # Giữ lại phần 'xxxxxx' nếu tìm thấy


    def load_sku_mapping():
        """Tạo mapping SKU từ sheet 'SanPham' (tìm từ cột C đến cuối, lấy cột A nếu trùng)"""
        file_path = "data.xlsx"
        sheet_name = "SanPham"

        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb[sheet_name]

        sku_mapping = {}
        
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Bỏ qua tiêu đề
            sku_code = str(row[0]).strip() if row[0] else None  # Cột A (SKU mới)
            possible_values = [str(cell).strip() for cell in row[2:] if cell]  # Cột C -> cuối cùng

            for value in possible_values:
                if value:  # Nếu có giá trị
                    sku_mapping[value] = sku_code  # Gán vào mapping
        return sku_mapping



    def replace_sku_numbers(products, sku_mapping):
        print(products)
        """Thay thế SKU Number hoặc Article nếu tìm thấy trong mapping"""
        for product in products:
            for key in ["SKU Number", "Article", "Barcode"]:  # Kiểm tra cả 2 trường
                if key in product:
                    clean_sku = ProcessHandler.clean_sku_number(product[key])  # Làm sạch mã
                    product[key] = ProcessHandler.clean_sku_number(product[key])
                    if clean_sku in sku_mapping:
                        product[key] = sku_mapping[clean_sku]  # Thay bằng mã mới

        return products


    def convert_entry_date(entry_date):
        """Chuyển đổi Entry Date từ dd/mm/yy sang dd/mm/yyyy với yy thuộc 2000+"""
        match = re.match(r"(\d{2})/(\d{2})/(\d{2})", entry_date)
        if match:
            day, month, year = match.groups()
            year = f"20{year}"  # Luôn giả định năm là từ 2000 trở đi
            return f"{day}/{month}/{year}"
        return entry_date  # Trả về nguyên nếu không đúng định dạng
    

    @staticmethod
    def demsodonhang1trang_coop(text):
        # Chuẩn hóa: thay \xa0 bằng khoảng trắng thông thường
        text = text.replace("\xa0", " ").replace("pom343", "POM343").replace("pom346", "POM346")
        
        pom34_pattern = r'POM34[36]\b'  # Tìm cả POM343 và POM346
        sub_total_pattern = r'\bSub\s*(?:Total|Tot\s*al)\b'

        
        pom34_count = len(re.findall(pom34_pattern, text))
        sub_total_count = len(re.findall(sub_total_pattern, text))
        
        return {"POM343": pom34_count, "Sub Total": sub_total_count}



    def process_coop_invoice(self,text,stt,path,page_label):
        file_name = os.path.basename(path)

        """Xử lý nếu Vendor là Coop"""
        print("\n🔹 Vendor xác nhận là Coop. Tiến hành xử lý...")

        
        # 📍 Lấy thông tin từ extract_info()
        info = ProcessHandler.extract_info(text)

        # 🟢 3. Xóa cụm từ không mong muốn
        # Danh sách các cụm từ cần loại bỏ (không phân biệt hoa thường, khoảng trắng tùy ý)
        patterns = [
            r"X\s*i\s*n\s*v\s*u\s*i\s*l\s*o\s*n\s*g\s*k\s*e\s*m\s*D\s*D\s*H\s*k\s*h\s*i\s*g\s*i\s*a\s*o\s*h\s*a\s*n\s*g",
            r"\*\s*=\s*T\s*h\s*i\s*s\s*S\s*K\s*U\s*D\s*i\s*s\s*c\s*o\s*u\s*n\s*t\s*e\s*d",
            r"M\s*o\s*t\s*H\s*o\s*a\s*D\s*o\s*n\s*c\s*h\s*i\s*x\s*u\s*a\s*t\s*c\s*h\s*o\s*m\s*o\s*t\s*P\s*O",
            r"m\s*u\s*a\s*1\s*T\s*A\s*N\s*G\s*1\s*C\s*U\s*N\s*G\s*L\s*O\s*A\s*I",
            r"1\s*T\s*A\s*N\s*G\s*1\s*C\s*U\s*N\s*G\s*L\s*O\s*A\s*I"
        ]

        # Tìm nội dung giữa "Notes" (hoặc "No tes") và "FOB"
        match = re.search(r"N\s*o\s*t\s*e\s*s\s*-\s*(.*?)\s*FOB", text, re.DOTALL | re.IGNORECASE)
        ghichu = match.group(1).strip() if match else ""

        # Loại bỏ các cụm từ không mong muốn
        for pattern in patterns:
            ghichu = re.sub(pattern, "", ghichu, flags=re.IGNORECASE)

        # Xóa khoảng trắng dư thừa
        ghichu = re.sub(r"\s+", " ", ghichu).strip()

        # Loại bỏ trùng lặp từ nhưng vẫn giữ thứ tự gốc
        words = ghichu.split()
        ghichu = " ".join(dict.fromkeys(words))

        print(f"Ghi chú: {ghichu}")
        po_number = info.get("P/O Number", "Không tìm thấy")
        po_location = info.get("P/O Location", "Không tìm thấy")


        shipto = text
        shipto = re.sub(r"\s+", " ", shipto)  # Thay thế nhiều khoảng trắng bằng một space


        # Regex để tìm "Ship To"
        pattern = r"S\s*h\s*i\s*p\s*T\s*o\s*:\s*S\s*t\s*a\s*t\s*u\s*s\s*-\s*\d+\s*R\s*E\s*L\s*E\s*A\s*S\s*E\s*D\s*(.*?)\s*C\s*o\s*n\s*t\s*a\s*c\s*t\s*-"

        match = re.search(pattern, shipto, re.DOTALL | re.IGNORECASE)
        
        if match:
            shipto = match.group(1).strip()  # Lấy nội dung và loại bỏ khoảng trắng đầu/cuối
            print("Ship To:", shipto)
        else:
            pattern = r"S\s*t\s*o\s*r\s*e\s*-\s*(.*?)\s*V\s*e\s*n\s*d\s*o\s*r"
            match = re.search(pattern, shipto, re.DOTALL | re.IGNORECASE)
            
            if match:
                shipto = match.group(1).strip()  # Lấy nội dung và loại bỏ khoảng trắng đầu/cuối
                print("Ship To:", shipto)
            else:
                print("Không tìm thấy thông tin Ship To.")
                shipto = ""

        

        entry_date_str = info.get("Entry Date", "Không tìm thấy")
        cancle_date_str = info.get("Cancel Date", "Không tìm thấy")

        def convert_date_format(date_str):
            """ Chuyển ngày từ 'dd/MM/yy' thành 'dd/MM/yyyy' """
            if date_str and date_str != "Không tìm thấy":
                try:
                    # Đọc theo định dạng "dd/MM/yy"
                    date_obj = datetime.strptime(date_str, "%d/%m/%y")
                    # Chuyển sang định dạng "dd/MM/yyyy"
                    return date_obj.strftime("%d/%m/%Y")
                except ValueError:
                    return "Không hợp lệ"  # Tránh lỗi khi format sai
            return "Không tìm thấy"

        # Chuyển đổi ngày
        entry_date = convert_date_format(entry_date_str)
        cancle_date = convert_date_format(cancle_date_str)

        # Nếu cancle_date không có, cộng 65 ngày vào entry_date
        if cancle_date == "Không tìm thấy" and entry_date != "Không tìm thấy":
            entry_date_obj = datetime.strptime(entry_date, "%d/%m/%Y")
            cancle_date = (entry_date_obj + timedelta(days=65)).strftime("%d/%m/%Y")


        



        print(f"📍 PO Number: {po_number}")
        print(f"📍 P/O Location: {po_location}")
        print(f"📅 Entry Date: {entry_date}")
        print(f"⏳ Cancel Date: {cancle_date}")
    
        self.log_signal.emit(f"<b>📦 Số PO:</b> <span style='color:blue;'>{po_number}</span>")
        self.log_signal.emit(f"<b>🗓️ Ngày đặt hàng:</b> <span style='color:green;'>{entry_date}</span>")
        self.log_signal.emit(f"<b>⏳ Hạn đơn hàng:</b> <span style='color:red;'>{cancle_date}</span>")

        print(f"📅 Cancle Date: {cancle_date}")
        # 🏢 Lấy Mã Khách Hàng
        makhachhang = ProcessHandler.get_makhachhang(po_location) if po_location else "Không tìm thấy"
        print(f"🏢 Mã Khách Hàng: {makhachhang}")

        # 📦 Trích xuất danh sách sản phẩm
        products = ProcessHandler.extract_products(text)
        soluongsanpham = len(products)
        self.log_signal.emit(f"<b>📊 Tổng số lượng sản phẩm:</b> <span style='color:purple;'>{soluongsanpham}</span>")

        # 🔽 Thay thế SKU Number từ Excel
        if products:
            sku_mapping = ProcessHandler.load_sku_mapping()  # Load mapping SKU
            products = ProcessHandler.replace_sku_numbers(products, sku_mapping)  # Thay SKU
            print(products)
            # Ghi vào dondathang.xls (truyền thêm `entry_date`)
            hethong = ProcessHandler.layhethong_COOP(makhachhang)
            saigia = ProcessHandler.write_to_dondathang(self,products, makhachhang, po_number, entry_date,cancle_date,stt,"COOP",ghichu,shipto)
            if not makhachhang: 
                makhachhang = "Không xác định"
                
            if saigia > 0 or makhachhang == "Không xác định":
                if saigia == 0:
                    saigia = "Đúng giá"
                else:
                    saigia = f'Có {saigia} mã sai giá'
                hoanthanh = "⚠️Hoàn Thành"
            else:
                saigia = "Đúng giá"
                hoanthanh = "✅Hoàn Thành"
            

            
            self.table_signal.emit(file_name, page_label,hethong, makhachhang ,po_number, saigia, hoanthanh)  # Phát tín hiệu cập nhật bảng
            
        stt = stt + 1

        print("\n📦 Danh sách sản phẩm:")
        for product in products:
            print(product)
        
    def parse_date(date_str):
        """Chuyển đổi chuỗi ngày tháng về định dạng chuẩn dd/mm/yyyy"""
        if not date_str or date_str in ["Không tìm thấy", ""]:  
            return None  # Trả về None nếu dữ liệu không hợp lệ
        
        if isinstance(date_str, datetime):
            return date_str  # Nếu đã là datetime, không cần xử lý

        if isinstance(date_str, str):
            parts = date_str.split('/')
            if len(parts) == 3:
                day, month, year = parts
                if not (day.isdigit() and month.isdigit() and year.isdigit()):
                    return None  # Nếu có ký tự không phải số -> bỏ qua
                
                if len(year) == 2:  # Nếu năm chỉ có 2 chữ số, thêm '20' phía trước
                    year = "20" + year

                date_str = f"{day.zfill(2)}/{month.zfill(2)}/{year}"

                try:
                    return datetime.strptime(date_str, "%d/%m/%Y")
                except ValueError:
                    return None  # Ngăn lỗi nếu dữ liệu không hợp lệ

        return None  # Nếu không thuộc kiểu string hoặc datetime
    
    def get_gid(sheet_name):
        """ Truy xuất GID từ tên sheet trong thẻ <gid>...</gid> của setting.ini """
        with open(CONFIG_FILE, "r", encoding="utf-8") as file:
            content = file.read()
        
        # Tìm nội dung bên trong thẻ <gid>...</gid>
        match = re.search(r"<gid>(.*?)</gid>", content, re.DOTALL)
        if not match:
            return None  # Không tìm thấy thẻ <gid>

        gid_section = match.group(1)  # Lấy nội dung bên trong <gid>...</gid>
        
        # Chuyển đổi thành dictionary
        gid_dict = {}
        for line in gid_section.strip().split("\n"):
            if "=" in line:
                key, value = line.strip().split("=")
                gid_dict[key.strip()] = value.strip()
        
        return gid_dict.get(sheet_name, None)  # Trả về GID theo tên sheet (hoặc None nếu không có)
    

    def find_price_by_sku(sku_number, sheet_name="Coop"):
        sheet_id = '1yvxE_SPYXKhofcZdhv1CSKAyiwdY1Mf4pFlsiMbtOr4'
        gid = ProcessHandler.get_gid(sheet_name)
        print(f"Mã hàng: {sku_number}")

        # Tạo URL tải dữ liệu từ Google Sheets
        sheet_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?tqx=out:csv&gid={gid}"

        # Tải dữ liệu từ Google Sheets, chỉ giữ 4 cột đầu
        df = pd.read_csv(sheet_url, dtype=str, usecols=[0, 1, 2, 3])  
        df.fillna("", inplace=True)  # Thay NaN bằng chuỗi rỗng

        # Tìm hàng có SKU trong cột thứ 2 (index = 1)
        matched_row = df[df.iloc[:, 1] == sku_number]

        if not matched_row.empty:
            price = matched_row.iloc[0, 3]  # Cột thứ 4 (index = 3)
            price = price.replace(".", "")  # Xóa dấu "." và ","
            
            print(price if price.strip() else "Không có giá trị")
            return price if price.strip() else "Không có giá trị"
        else:
            return "Không tìm thấy SKU"




    def find_value_by_sku_and_time(sku_code, time_to_check, sheet_name="Coop"):
        sheet_id = '1yvxE_SPYXKhofcZdhv1CSKAyiwdY1Mf4pFlsiMbtOr4'
        gid = ProcessHandler.get_gid(sheet_name)

        # 1️⃣ Tải dữ liệu từ Google Sheets
        sheet_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?tqx=out:csv&gid={gid}"
        df = pd.read_csv(sheet_url, dtype=str)  # Đọc dữ liệu, giữ nguyên định dạng chuỗi
        df.fillna("", inplace=True)  # Thay NaN bằng chuỗi rỗng để tránh lỗi

        # Chuẩn hóa tiêu đề cột
        df.columns = [col.strip().replace("\n", " ") for col in df.columns]

        # Tìm cột "Mã hàng"
        sku_col = next((col for col in df.columns if "Mã hàng" in col), None)
        if not sku_col:
            return "Không có cột Mã hàng"

        # Kiểm tra SKU có tồn tại không
        sku_rows = df[df[sku_col] == sku_code]
        if sku_rows.empty:
            return "Không có giá trị"

        sku_row_idx = sku_rows.index[0]  # Lấy chỉ mục dòng chứa SKU

        def normalize_date(date_str):
            """Chuyển ngày về dạng dd/mm"""
            day, month = date_str.split("/")
            return f"{int(day):02}/{int(month):02}"  # Đảm bảo 2 chữ số

        def is_within_date_range(time_to_check, column_name):
            match = re.search(r"(\d{1,2}/\d{1,2})-(\d{1,2}/\d{1,2})", column_name)
            if match:
                start_date, end_date = match.groups()
                time_to_check = re.sub(r"/\d{4}$", "", time_to_check)  # Xóa /2025 nếu có

                # Chuẩn hóa ngày
                start_date, end_date = normalize_date(start_date), normalize_date(end_date)
                current_year = datetime.now().year  # Lấy năm hiện tại

                time_to_check_dt = datetime.strptime(f"{time_to_check}/{current_year}", "%d/%m/%Y")
                start_date_dt = datetime.strptime(f"{start_date}/{current_year}", "%d/%m/%Y")
                end_date_dt = datetime.strptime(f"{end_date}/{current_year}", "%d/%m/%Y")

                return start_date_dt <= time_to_check_dt <= end_date_dt
            return False

        # Lưu danh sách các cột khớp thời gian
        matching_cols = [col for col in df.columns if is_within_date_range(time_to_check, col)]
        
        # Kiểm tra và lấy giá trị từ cột đầu tiên có dữ liệu
        for col in matching_cols:
            value = df.at[sku_row_idx, col]
            if value.strip():  # Nếu có dữ liệu, trả về ngay
                print(f"✅ Tìm thấy giá trị '{value}' tại SKU {sku_code}, cột {col}")
                return value
        
        # Nếu tất cả cột khớp thời gian đều rỗng
        return "Không có giá trị"
    




    def kiemtrakhuyenmaihoadon( time_to_check, sheet_name="Coop"):
        sheet_id = '1yvxE_SPYXKhofcZdhv1CSKAyiwdY1Mf4pFlsiMbtOr4'
        gid = ProcessHandler.get_gid(sheet_name)

        # 1️⃣ Tải dữ liệu từ Google Sheets
        sheet_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?tqx=out:csv&gid={gid}"
        df = pd.read_csv(sheet_url, dtype=str)  # Đọc dữ liệu, giữ nguyên định dạng chuỗi
        df.fillna("", inplace=True)  # Thay NaN bằng chuỗi rỗng để tránh lỗi

        # Chuẩn hóa tiêu đề cột
        df.columns = [col.strip().replace("\n", " ") for col in df.columns]

        # Tìm cột "Mã hàng"
        sku_col = next((col for col in df.columns if "TÊN SẢN PHẨM" in col), None)
        if not sku_col:
            return "Không có cột TÊN SẢN PHẨM"

        # Kiểm tra SKU có tồn tại không
        sku_rows = df[df[sku_col] == "Khuyến mãi trên 1 Hoá đơn"]
        if sku_rows.empty:
            return "Không có giá trị"

        sku_row_idx = sku_rows.index[0]  # Lấy chỉ mục dòng chứa SKU

        def normalize_date(date_str):
            """Chuyển ngày về dạng dd/mm"""
            day, month = date_str.split("/")
            return f"{int(day):02}/{int(month):02}"  # Đảm bảo 2 chữ số

        def is_within_date_range(time_to_check, column_name):
            match = re.search(r"(\d{1,2}/\d{1,2})-(\d{1,2}/\d{1,2})", column_name)
            if match:
                start_date, end_date = match.groups()
                time_to_check = re.sub(r"/\d{4}$", "", time_to_check)  # Xóa /2025 nếu có

                # Chuẩn hóa ngày
                start_date, end_date = normalize_date(start_date), normalize_date(end_date)
                current_year = datetime.now().year  # Lấy năm hiện tại

                time_to_check_dt = datetime.strptime(f"{time_to_check}/{current_year}", "%d/%m/%Y")
                start_date_dt = datetime.strptime(f"{start_date}/{current_year}", "%d/%m/%Y")
                end_date_dt = datetime.strptime(f"{end_date}/{current_year}", "%d/%m/%Y")

                return start_date_dt <= time_to_check_dt <= end_date_dt
            return False

        # Lưu danh sách các cột khớp thời gian
        matching_cols = [col for col in df.columns if is_within_date_range(time_to_check, col)]
        
        # Kiểm tra và lấy giá trị từ cột đầu tiên có dữ liệu
        for col in matching_cols:
            value = df.at[sku_row_idx, col]
            if value.strip():  # Nếu có dữ liệu, trả về ngay
                print(f"✅ Tìm thấy giá trị '{value}' tại Khuyến mãi trên 1 Hoá đơn, cột {col}")
                return value
        
        # Nếu tất cả cột khớp thời gian đều rỗng
        return "Không có giá trị"



        
        


    def laydanhsachsanpham_bigc(text):
        """
        Trích xuất danh sách sản phẩm từ nội dung văn bản BigC (bỏ qua Article Description).
        """


            # 1️⃣ Xác định vị trí danh sách sản phẩm (từ dòng chứa "Article")
        match_start = re.search(r"\bArticle\b", text)
    
        if not match_start:
            print("Không tìm thấy tiêu đề danh sách sản phẩm!")
            return []
        text = text[match_start.start():].strip()
            

        pattern = re.compile(
            r"(\d{13})\s+.+?\s+Pack\s+\d+\s+(\d+)\s+(\d+)\s+\d+\s+([\d,]+)\s+\w+\s+([\d,]+)",
            re.DOTALL
        )

        matches = pattern.findall(text)

        product_list = []
        for match in matches:
            if len(match) != 5:
                print(f"LỖI: Không đủ dữ liệu trong dòng: {match}")
                continue  # Bỏ qua dòng bị thiếu dữ liệu

            article, sku_ou, ou_qty, net_price, total_price = match

            # Chuẩn hóa dữ liệu (bỏ dấu phẩy và chuyển thành số)
            net_price = int(net_price.replace(",", ""))
            total_price = int(total_price.replace(",", ""))

            product_list.append({
                "Article": article.strip(),
                "SKU/OU": sku_ou.strip(),
                "OU Qty": ou_qty.strip(),
                "Net Purchase Price": net_price,
                "Total Net Purchase Price": total_price
            })

        return product_list




    def lay_ten_store(text):
        # Tìm vị trí của kho "FM LOGISTIC VSIP 2"
        match = re.search(r"(FM LOGISTIC VSIP 2|LINFOX WAREHOUSE \(802\)).*?Vietnam\s*\n(.*?)\n", text, re.DOTALL)
        
        if match:
            return match.group(2).strip()  # Trả về dòng ngay sau đó (tên cửa hàng)
        return None



    def ghepgia_donhangbigc(items, products):
        # Chuyển đổi products thành dict để tra cứu nhanh hơn
        product_dict = {p["Article"]: p["Net Purchase Price"] for p in products}

        # Thêm giá vào mỗi item
        for item in items:
            article = item["Article"]
            item["Net Purchase Price"] = product_dict.get(article, 0)  # Nếu không có thì báo lỗi

        return items


    def trichxuatdanhsachforstore_bigc(text):
        # Regex để tìm các dòng chứa thông tin sản phẩm
        pattern = r"(?<=\n)(\d{13})\s*\n(.*?)\s*\nPack\s*\n\d+\s*\n(\d+)\s*\n(\d+)"
        matches = re.findall(pattern, text, re.DOTALL)

        # Chuyển kết quả thành danh sách từ điển
        data = [{"Article": m[0], "SKU/OU": m[2], "OU Qty": m[3]} for m in matches]
        return data
    # --------- Chạy chương trình ---------
    def read_text_file(file_path):
        """Đọc file TXT với khả năng xử lý nhiều kiểu mã hóa"""
        encodings = ["utf-8", "utf-16", "utf-16-le", "utf-16-be", "latin-1"]

        for encoding in encodings:
            try:
                with open(file_path, "r", encoding=encoding) as f:
                    return f.read()
            except UnicodeDecodeError:
                continue  # Thử mã hóa tiếp theo

        raise UnicodeDecodeError("Không thể đọc file TXT với các bộ mã hóa phổ biến!")




    def ghi_gia_tri_G1(stt):
        """
        Ghi giá trị 10 vào ô G1 trong file 'dondathang.xlsx', sheet 'Don dat hang'.
        """
        file_path = "dondathang.xlsx"
        sheet_name = "Don dat hang"

        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]

        sheet["G1"].value = stt  # Ghi giá trị 10 vào ô G1

        wb.save(file_path)
        wb.close()


    def trichxuatinfo_donbigc(text):
        """Trích xuất PO Number, Entry Date và Cancel Date từ văn bản"""

        # Chuẩn hóa văn bản (xoá khoảng trắng thừa)
        cleaned_text = re.sub(r"\s+", " ", text).strip()

        # 1️⃣ Tìm PO Number & Entry Date
        match_po = re.search(r"(\d{13,})\s+(\d{2}/\d{2}/\d{2})", cleaned_text)
        po_number = match_po.group(1) if match_po else None
        entry_date = match_po.group(2) if match_po else None

        # 2️⃣ Tìm tất cả vị trí "Total Net Purchase Price"
        matches = list(re.finditer(r"Total Net Purchase Price", cleaned_text))

        # 3️⃣ Nếu tìm thấy, lấy ngày sau lần xuất hiện CUỐI CÙNG
        if matches:
            last_match = matches[-1]  # Lấy lần xuất hiện cuối
            search_region = cleaned_text[last_match.end():]  # Lấy phần sau đó
            match_cancel = re.search(r"(\d{2}/\d{2}/\d{2})", search_region)
            cancel_date = match_cancel.group(1) if match_cancel else None
        else:
            cancel_date = None

        # Nếu không tìm thấy cancel_date, đặt cancel_date = entry_date + 5 ngày
        if cancel_date is None and entry_date:
            try:
                entry_dt = datetime.strptime(entry_date, "%d/%m/%y")
                cancel_dt = entry_dt + timedelta(days=5)
                cancel_date = cancel_dt.strftime("%d/%m/%y")  # Chuyển lại thành chuỗi
            except ValueError:
                cancel_date = None  # Nếu entry_date không hợp lệ, giữ cancel_date là None

        return po_number, ProcessHandler.convert_entry_date(entry_date), ProcessHandler.convert_entry_date(cancel_date)




    def catdonra_nhieutrang(text):
        segments = []
        
        # Tách đoạn đầu tiên từ đầu đến "Sub Total" đầu tiên
        parts = text.split("Sub Total", 1)
        if len(parts) > 1:
            segments.append(parts[0])  # Đoạn trước "Sub Total"
            text = parts[-1]  # Giữ lại phần còn lại để tiếp tục xử lý
        
        # Tách các đoạn bắt đầu từ "POM343" hoặc "POM346"
        while any(keyword in text.lower() for keyword in ["pom343", "pom346"]) and "Sub Total" in text:
            for keyword in ["pom343", "pom346"]:
                if keyword in text.lower():
                    text = text.split(keyword, 1)[-1]  # Cắt bỏ phần trước từ khóa
                    parts = text.split("Sub Total", 1)  # Cắt đến "Sub Total" tiếp theo
                    if len(parts) > 1:
                        segments.append(keyword.upper() + parts[0])  # Giữ lại phần có từ khóa
                        text = parts[-1]  # Tiếp tục với phần còn lại
                    else:
                        segments.append(keyword.upper() + text)  # Nếu không còn "Sub Total", lấy hết và kết thúc
                        return segments
        
        return segments



    def get_grouped_data_xlsx(file_path: str, sheet_name: str = "Sheet1"):
        wb = openpyxl.load_workbook(filename=file_path, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active  # Chọn sheet

        products = []
        po_location = po_number = entry_date = cancel_date = None

        first_po_number = None  # Biến lưu P/O Number đầu tiên để so sánh
        rows_to_delete = []  # Danh sách lưu index các dòng cần xóa

        for idx, row in enumerate(ws.iter_rows(min_row=2, max_col=8, values_only=True), start=2):
            if row[1] is None:  # Bỏ qua dòng không có P/O Number
                continue

            if first_po_number is None:  # Lấy P/O Number đầu tiên
                first_po_number = row[1]
                po_location = row[0]  
                po_number = row[1]  
                entry_date = row[2]  
                cancel_date = row[3]  

            if row[1] != first_po_number:  # Nếu gặp P/O Number mới, dừng lấy dữ liệu
                break

            sku = row[4]  
            net_buy_cost = row[5]  
            qty_ord_pcs = row[6]  
            extended_cost = row[7]  

            products.append({
                "SKU Number": sku,
                "Net Buy Cost": net_buy_cost,
                "Qty Ord/Pcs": qty_ord_pcs,
                "Extended Cost": extended_cost
            })

            rows_to_delete.append(idx)  # Đánh dấu dòng cần xóa

        # Xóa các dòng đã lấy từ dưới lên để tránh lỗi chỉ mục
        for row_idx in reversed(rows_to_delete):
            ws.delete_rows(row_idx)

        # Lưu lại file sau khi xóa dữ liệu
        wb.save(file_path)

        return po_location, po_number, entry_date, cancel_date, products
    
    def tachcancledate_lotte(text, podonhang):
        start_pattern = re.escape(podonhang)  # Tìm dòng bắt đầu
        end_marker = "00:00"  # Dòng kết thúc
        date_pattern = r"\d{1,2}/\d{1,2}/\d{4}"  # Định dạng ngày (dd/mm/yyyy)

        lines = text.split("\n")
        start_index = None
        end_index = None

        for i, line in enumerate(lines):
            if start_index is None and re.match(start_pattern, line):  # Tìm dòng đầu tiên
                start_index = i
            if line.strip() == end_marker:  # Tìm dòng cuối cùng
                end_index = i
                break  # Dừng ngay khi tìm thấy dòng cuối

        if start_index is not None and end_index is not None and start_index < end_index:
            filtered_dates = [line.strip() for line in lines[start_index + 1:end_index] if re.search(date_pattern, line)]
            return "\n".join(filtered_dates)  # Trả về chỉ các dòng có ngày tháng

        return None  # Không tìm thấy hai dòng
    

    def tachsanpham_lotte(text):
        text = ProcessHandler.lamsachdonhang_lotte(text)
        pattern = r'(\d{1,2}-\d{6}-\d{3})\s+(\d{12,13})[\s\S]*?(\d+)\s+BOX\s+(\d+)\s+(\d+)\s+[\d,]+\s+([\d,]+)'
        matches = re.findall(pattern, text)
        
        product_data = []
        for match in matches:
            product_code, barcode, unit_count, box_qty, loose_qty, total_price = match
            product_data.append({
                "Product Code": product_code,
                "Barcode": barcode,
                "Qty-Box": int(unit_count),
                "Box Quantity": int(box_qty),
                "Loose Quantity": int(loose_qty),
                "Total Price": float(total_price.replace(',', ''))
            })
        
        return product_data

    def lamsachdonhang_lotte(text):
        start_pattern = "Sply qty"  # Tìm dòng bắt đầu
        end_marker = "Tot add tax"  # Dòng kết thúc

        lines = text.split("\n")
        start_index = None
        end_index = None

        for i, line in enumerate(lines):
            if start_index is None and re.match(start_pattern, line):  # Tìm dòng đầu tiên
                start_index = i
            if line.strip() == end_marker:  # Tìm dòng cuối cùng
                end_index = i
                break  # Dừng ngay khi tìm thấy dòng cuối

        if start_index is not None and end_index is not None and start_index < end_index:
            return "\n".join(lines[start_index + 1:end_index])  # Lấy nội dung giữa hai dòng

        return None  # Không tìm thấy hai dòng

    def trichxuatsanpham_satra(text):
        text = re.split(r"\bTổng cộng\b", text, maxsplit=1)[0].strip()
        pattern = re.compile(r"^\s*(\d+)\s+\d+\s*\n\s*(\d{13})", re.VERBOSE | re.MULTILINE)
        so_luong_pattern = re.compile(r"\b(\d{1,3}),000\b")  # Tìm số có dạng "x,000", "xx,000", "xxx,000"

        matches = list(pattern.finditer(text))
        if not matches:
            print("⚠️ Không tìm thấy dữ liệu phù hợp với regex!")
            return []

        positions = [(m.start(), m.group(2)) for m in matches]
        positions.append((len(text), None))  

        nhom_sanpham = []
        for i in range(len(positions) - 1):
            start, barcode = positions[i]
            end, _ = positions[i + 1]
            lines = [line.replace(" ", "") for line in text[start:end].strip().split("\n") if line.strip()]

            so_luong_index = next((i for i, line in enumerate(lines) if so_luong_pattern.search(line)), None)
            if so_luong_index is not None:
                so_luong = so_luong_pattern.sub(r"\1", lines[so_luong_index])
                thanh_tien = lines[so_luong_index + 1] if so_luong_index + 1 < len(lines) else ""  
                thanh_tien = re.sub(r",00$", "", thanh_tien)  # Bỏ phần ",00" nếu có
                nhom_sanpham.append({
                "Article": barcode,
                "OU Qty": so_luong,
                "Net Purchase Price": thanh_tien
            })

        return nhom_sanpham 


    def laydanhsanpham_emart(text):
        # Regex sửa đổi để khớp chính xác hơn
        product_pattern = re.finditer(r"""
            (?P<article_code>\d{7})\s*        # Article Code (7 chữ số)
            (?P<barcode>\d{12,13})\s*        # Barcode (12-13 chữ số)
            \s*(?P<description>.+?)\s+        # Description (kết thúc trước unit, ít nhất 1 khoảng trắng)
            (?P<unit>[A-Z]{2,})\s+           # Unit (ít nhất 2 chữ cái, như EA, BOX)
            \s*(?P<qty_in_box>\d+)\s+        # Qty in Box (yêu cầu khoảng trắng sau)
            \s*(?P<quantity>\d+)\s+          # PO Quantity (yêu cầu khoảng trắng sau)
            \s*(?P<purchase_price>[\d\.,]+)  # Purchase Price
        """, text.strip(), re.VERBOSE | re.DOTALL)

        results = []
        for match in product_pattern:
            purchase_price = match.group("purchase_price").replace(".", "")  # Xóa dấu chấm
            purchase_price_value = float(purchase_price.replace(",", "."))  # Chuyển đổi thành số
            
            if purchase_price_value == 0:
                continue  # Bỏ qua nếu giá trị "Net Purchase Price" bằng 0

            extracted_data = {
                "Article": match.group("barcode"),
                "OU Qty": int(match.group("quantity")),  # Ép kiểu thành int
                "Net Purchase Price": purchase_price
            }
            results.append(extracted_data)

        if not results:
            print("No valid matches found. Please check input format.")

        return results


    def laytenstore_lotte(text, podonhang):
        start_pattern = "DOAN TUAN ANH"  # Dòng bắt đầu
        end_marker = podonhang  # Dòng kết thúc

        lines = text.split("\n")
        start_index = None
        end_index = None

        for i, line in enumerate(lines):
            if start_index is None and re.match(start_pattern, line):  # Tìm dòng đầu tiên
                start_index = i
            if line.strip() == end_marker:  # Tìm dòng cuối cùng
                end_index = i
                break  # Dừng ngay khi tìm thấy dòng cuối

        if start_index is not None and end_index is not None and start_index < end_index:
            last_line = lines[end_index - 1].strip()  # Lấy dòng cuối cùng trước `end_marker`
            return last_line if last_line else None  # Trả về dòng cuối cùng (không rỗng)

        return None  # Không tìm thấy hai dòng
    




    def laydanhsachsanpham_kingfood(text):
        pattern = re.compile(r"""
        (?P<barcode>\d{13})\n  # Barcode (13 digits)
        (?P<name>.+?(?:\n.+?)?)\n   # Product name (1 or 2 lines)
        (?P<unit>.+?)\n        # Unit
        (?P<quantity>\d+)\n    # Quantity
        .+?\n                  # Skip 1st extra line
        .+?\n                  # Skip 2nd extra line
        .+?\n                  # Skip 3rd extra line
        .+?\n
        (?P<price>[-\d.,]+)    # Final price (-VAT), 4 lines after quantity
    """, re.VERBOSE)
        
        products = []
        for match in pattern.finditer(text):
            barcode = match.group("barcode")
            quantity = int(match.group("quantity"))
            price = match.group("price").replace('.', '').replace(',', '.')  # Remove commas for numerical processing
            
            if quantity > 0 and float(price) > 0:  # Exclude promotional items
                products.append({
                    "Article": barcode,
                    "OU Qty": quantity,
                    "Net Purchase Price": float(price)
                })
        
        return products
    




    def trichxuatsanpham_winmart(text):
        # Nếu đầu vào là danh sách, chuyển thành chuỗi với dấu xuống dòng
        if isinstance(text, list):
            text = "\n".join(text)
        
        # Tách văn bản thành các dòng
        lines = text.split("\n")
        start_idx, end_idx = None, None

        # Tìm chỉ số bắt đầu và kết thúc
        for i, line in enumerate(lines):
            if "(No.)" in line or "Stt" in line:  # Nhận diện dòng tiêu đề
                start_idx = i
            if "Tổng giá trị trước thuế (Total excl VAT amount)" in line and start_idx is not None:
                end_idx = i
                break

        if start_idx is not None and end_idx is not None:
            # Lấy các dòng giữa tiêu đề và kết thúc
            result = lines[start_idx + 1:end_idx]
            text_block = "\n".join(result)  # Ghép lại thành một đoạn văn bản
            print("Dòng được trích xuất:", result)

            # Mẫu regex để khớp với dữ liệu
            pattern = re.compile(r"""
                (?P<no>\d+)\n                # Số thứ tự (No.)
                (?P<article>\d+)\n          # Mã hàng
                (?P<barcode>\d+)\n          # Mã vạch (Barcode)
                (?P<quantity>[\d,]+)\n      # Số lượng (Quantity) - có thể chứa dấu phẩy
                \S+\n                       # ĐVT (Unit) - bỏ qua
                (?P<unit_price>[\d,]+)\n    # Đơn giá (Unit Price)
                (?P<amount>[\d,]+)          # Thành tiền (Amount)
            """, re.VERBOSE)
            
            ket_qua = []
            for match in pattern.finditer(text_block):
                thanh_tien = match.group("amount").replace(",", "")  # Xóa dấu phẩy
                if int(thanh_tien) > 0:  # Chỉ thêm nếu thành tiền > 0
                    ket_qua.append({
                        "Article": match.group("barcode"),  # Dùng mã vạch làm mã hàng
                        "Net Purchase Price": match.group("amount").replace(",", ""),
                        "OU Qty": match.group("quantity").replace(",", "")
                    })
            return ket_qua

        return []
    





    def tach_san_pham_BHX(text: str) -> list:
        lines = [line.strip() for line in text.replace('\r', '\n').split('\n') if line.strip()]

        # Xác định block bắt đầu bằng số thứ tự và tiếp theo là "C"
        product_indices = [
            i for i in range(len(lines) - 1)
            if re.fullmatch(r"\d{1,3}", lines[i]) and lines[i + 1] == "C"
        ]

        ket_qua = []

        for idx, start in enumerate(product_indices):
            end = product_indices[idx + 1] if idx + 1 < len(product_indices) else len(lines)

            block = lines[start:end]

            if len(block) < 12:
                print("❌ Block quá ngắn:", block)
                continue

            try:
                article = block[3]         # barcode
                quantity = block[7]        # số lượng thực tế (OU Qty)
                amount = block[11]         # tổng tiền

                ket_qua.append({
                    "Article": article,
                    "Net Purchase Price": amount.replace(",", ""),
                    "OU Qty": quantity.replace(",", "")
                })
            except Exception as e:
                print("❌ Lỗi xử lý block:", block, "→", e)

        return ket_qua
    



    def tach_san_pham_Fujimart(text: str) -> list:
        lines = [line.strip() for line in text.strip().splitlines() if line.strip()]

        # Tìm vị trí dòng đầu tiên là '1'
        start_index = next((i for i, line in enumerate(lines) if line == "1"), None)
        if start_index is None:
            print("❌ Không tìm thấy dòng bắt đầu là '1'")
            return []

        # Cắt bỏ những dòng rác trước dòng "1"
        lines = lines[start_index:]

        ket_qua = []

        # Chia theo block 8 dòng
        for i in range(0, len(lines), 8):
            block = lines[i:i+8]
            if len(block) < 8:
                continue  # bỏ qua block không đủ 8 dòng (ví dụ dòng tổng tiền cuối)

            try:
                quantity = block[1]
                amount = block[2]
                barcode = block[6]
                ket_qua.append({
                    "Article": barcode,
                    "Net Purchase Price": amount.replace(",", ""),
                    "OU Qty": quantity.replace(",", "")
                })
            except Exception as e:
                print("❌ Lỗi xử lý block:", block, "→", e)

        return ket_qua
    
    def tach_san_pham_Farmer(text: str) -> list:
        def parse_number(value: str) -> float:
            """Chuyển '765.419,00' → 765419.00"""
            value = value.strip()
            if ',' in value:
                value = value.replace('.', '').replace(',', '.')
            return float(value)

        lines = [line.strip() for line in text.strip().splitlines() if line.strip()]
        start_index = next((i for i, line in enumerate(lines) if line.strip() == "1"), None)
        if start_index is None:
            print("❌ Không tìm thấy dòng bắt đầu là '1'")
            return []

        lines = lines[start_index:]
        ket_qua = []
        i = 0

        while i < len(lines):
            line = lines[i]
            barcode_match = re.match(r"^\d{13}\b", line)
            if barcode_match:
                barcode = barcode_match.group()
                block = [line]
                j = i + 1
                while j < len(lines) and len(block) < 9:
                    block.append(lines[j])
                    j += 1

                try:
                    # Tìm số lượng: số đầu tiên có dạng xxx,00 sau barcode
                    quantity = None
                    for line in block:
                        numbers = re.findall(r"\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{2})", line)
                        for num in numbers:
                            val = parse_number(num)
                            if val.is_integer() and val < 1000:  # Số lượng thường nhỏ hơn 1000
                                quantity = int(val)
                                break
                        if quantity:
                            break

                    # Tìm amount: số tiền cuối cùng
                    all_numbers = re.findall(r"\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{2})", " ".join(block))
                    amount = parse_number(all_numbers[-1]) if all_numbers else None

                    if not quantity or not amount:
                        raise ValueError("Không tìm được đầy đủ dữ liệu.")

                    ket_qua.append({
                        "Article": barcode,
                        "Net Purchase Price": f"{amount:.2f}",
                        "OU Qty": f"{quantity}"
                    })

                except Exception as e:
                    print(f"⚠️ Lỗi block: {block} → {e}")

                i = j
            else:
                i += 1

        return ket_qua

        

    def fix_vni_to_unicode(text):
        vni_to_unicode = {
        "¸": "á", "µ": "à", "¶": "ả", "·": "ã", "¹": "ạ",
        "¨": "â", "Ç": "ầ", "©": "ấ", "Ê": "ẩ", "È": "ẫ", "É": "ậ",
        "¢": "ă", "¤": "ằ", "¡": "ắ", "£": "ẳ", "¥": "ẵ", "¦": "ặ",

        "è": "è", "é": "é", "ê": "ẻ", "ë": "ẽ", "ç": "ẹ",
        "ª": "ê", "Ì": "ề", "Ñ": "ế", "Í": "ể", "Î": "ễ", "Ï": "ệ",

        "ï": "ì", "Ý": "í", "ó": "ỉ", "ò": "ĩ", "î": "ị",

        "ó": "ó", "ò": "ò", "ô": "ỏ", "õ": "õ", "ö": "ọ",
        "¯": "ô", "×": "ố", "Ø": "ồ", "Ù": "ổ", "Ú": "ỗ", "Ü": "ộ",
        "­": "ơ", "¬": "ờ", "½": "ớ", "¾": "ở", "¿": "ỡ", "À": "ợ",

        "ú": "ú", "ù": "ù", "û": "ủ", "ü": "ũ", "þ": "ụ",
        "®": "ư", "Å": "ứ", "Ä": "ừ", "Æ": "ử", "Ç": "ữ", "È": "ự",

        "ý": "ý", "ÿ": "ỷ", "×": "ỹ", "Ø": "ỳ", "Þ": "ỵ",

        "§": "đ",
    }

        for vni_char, unicode_char in vni_to_unicode.items():
            text = text.replace(vni_char, unicode_char)
        return text


    

    def process_file(self, file_path,stt = 1):
        """Xử lý PDF, TXT hoặc XLSX"""
        self.log_signal.emit(f"🔄 Đang xử lý file: {file_path}")
        file_name = os.path.basename(file_path)
        
        


        if file_path.lower().endswith(".pdf"):
            doc = fitz.open(file_path)

            for page_num in range(len(doc)):
                text = doc[page_num].get_text("text")
                
                vendor = ProcessHandler.identify_vendor(text)
                page_label = f'{page_num + 1}/{len(doc)}'  # đổi tên biến này
                
                

                
                if text == "":
                    # Chuyển trang PDF thành ảnh
                    page = doc.load_page(page_num)  # thêm dòng này để lấy đúng trang PDF
                    # Dùng dpi cao để ảnh rõ
                    pix = page.get_pixmap(dpi=500)
                    img = Image.open(io.BytesIO(pix.tobytes("png")))

                    # Tiền xử lý ảnh: grayscale + tăng tương phản
                    img = img.convert("L")  # grayscale
                    img = ImageEnhance.Contrast(img).enhance(2.5)

                    # OCR tiếng Việt với cấu hình tốt
                    text = pytesseract.image_to_string(img, lang='vie', config='--psm 11')

                    print(f"\n🧾 Trang {page_num + 1}/{len(doc)}:\n{text}\n{'='*80}")

                    

                
                
                
                
              
                if page_num == 0:
                    self.log_signal.emit(f"<b style='color:blue;'>📌 Phát hiện Đơn Hàng:</b> <span style='color:green;'>{vendor}</span>")
                    self.log_signal.emit(f"<b style='color:purple;'>📄 Tổng số trang:</b> <span style='color:red;'>{len(doc)}</span>")
                    self.log_signal.emit(f"<b style='color:orange;'>📝 Đang xử lý trang:</b> <span style='color:brown;'>{page_num + 1}</span>")

                

                if vendor == "Coop":
                    sodon = process_handler.demsodonhang1trang_coop(text)
                    
                    pom343 = sodon.get("POM343")
                    sub_total = sodon.get("Sub Total")
                    
                   
                
                    print(f"Trang {page_num + 1}")
                    print(f"POM343: {pom343}, Sub Total: {sub_total}")
                    
                    if pom343 == 1 and sub_total == 1:
                        ProcessHandler.process_coop_invoice(self,text,stt,file_path,page_label)
                        stt = stt +1
                    elif pom343 > 1 and sub_total > 1:
                        if pom343 == sub_total:
                            
                            catdon = ProcessHandler.catdonra_nhieutrang(text)
                            for idx, segment in enumerate(catdon):
                                ProcessHandler.process_coop_invoice(self,segment,stt,file_path,page_label)
                                stt = stt + 1
                        else:
                            print("khác nhau")


                elif vendor == "BHX":
                    print(text)
                    lines = text.split("\n")
                    idx = next((i for i, line in enumerate(lines) if "Mã đơn hàng:" in line), -1)
                    POdonhang = lines[idx + 1].strip() if idx != -1 and idx + 1 < len(lines) else None
                    POdonhang = POdonhang.replace('.','/')
                    print(POdonhang)


                    # Tìm địa chỉ giao hàng
                    delivery_idx = next((i for i, line in enumerate(lines) if "Kho NCC giao" in line), -1)
                    delivery_address = lines[delivery_idx + 2].strip() if delivery_idx != -1 and delivery_idx + 2 < len(lines) else None
                    print("Địa chỉ giao hàng:", delivery_address)


                    # Tìm dòng chứa "Ngày đặt hàng:"
                    entry_idx = next((i for i, line in enumerate(lines) if "Ngày đặt hàng:" in line), -1)
                    entry_date = None

                    if entry_idx != -1:
                        parts = lines[entry_idx].split(":")
                        if len(parts) > 1:
                            date_time_str = parts[1].strip()
                            entry_date = date_time_str.split(" ")[0]  # Lấy phần ngày

                    print("entry_date:", entry_date)


                    # Tìm dòng chứa "Ngày NCC giao hàng:"
                    cancel_idx = next((i for i, line in enumerate(lines) if "Ngày NCC giao hàng" in line), -1)
                    cancel_date = None

                    if cancel_idx != -1 and cancel_idx + 1 < len(lines):
                        date_range = lines[cancel_idx + 1].strip()
                        if " - " in date_range:
                            cancel_date = date_range.split(" - ")[1].strip()

                    print("cancel_date:", cancel_date)



                    

                    start = text.rfind("Thành tiền")  # tìm vị trí xuất hiện cuối cùng
                    end = text.find("Tổng")

                    product_block = text[start + len("Thành tiền"):end].strip()
                    lines = product_block.splitlines()
                    result = "\n".join(lines[1:])  # bỏ dòng đầu
                    print(result)
                    products = ProcessHandler.tach_san_pham_BHX(result)
                    print(products)
                    if products:
                        sku_mapping = ProcessHandler.load_sku_mapping()  # Load mapping SKU
                        products = ProcessHandler.replace_sku_numbers(products, sku_mapping)  # Thay SKU
                        soluongsanphamstore = len(products)
                        
                        socuahang = delivery_address.split()[0]
                        makhachhang = ProcessHandler.laymakhachhang_fujimart(f'BHX{socuahang}')


                        self.log_signal.emit(f"<b>🏬 Đơn hàng Store:</b> <span style='color:orange;'>{delivery_address}</span>")
                        self.log_signal.emit(f"<b>📦 Số PO:</b> <span style='color:blue;'>{POdonhang}</span>")
                        self.log_signal.emit(f"<b>🗓️ Ngày đặt hàng:</b> <span style='color:green;'>{entry_date}</span>")
                        self.log_signal.emit(f"<b>⏳ Hạn đơn hàng:</b> <span style='color:red;'>{cancel_date}</span>")
                        self.log_signal.emit(f"<b>🛒 Số lượng sản phẩm đặt của Store:</b> <span style='color:brown;'>{soluongsanphamstore}</span>")

                        saigia = ProcessHandler.write_to_dondathang_bhx(self,products,makhachhang,POdonhang,entry_date,cancel_date,stt,vendor,delivery_address)
                        
                        if not makhachhang: 
                            makhachhang = "Không xác định"
                            
                        if saigia > 0 or makhachhang == "Không xác định":
                            
                            if saigia == 0:
                                saigia = "Đúng giá"
                            else:
                                saigia = f'Có {saigia} mã sai giá'
                            hoanthanh = "⚠️Hoàn Thành"
                        else:
                            saigia = "Đúng giá"
                            hoanthanh = "✅Hoàn Thành"
                        self.table_signal.emit(file_name, page_label,vendor,makhachhang, POdonhang,saigia, hoanthanh)  # Phát tín hiệu cập nhật bảng




                elif vendor == "Farmer":
                    print(text)
                    
                    POdonhang = None  # Khởi tạo mặc định
                    match = re.search(r"Số hiệu chứng từ\s*:\s*(PO-\w+)", text)
                    if match:
                        POdonhang = match.group(1)

                    print(POdonhang)
                    entry_date = None
                    match = re.search(r"Ngày chứng từ\s*:\s*(\d{1,2}/\d{1,2}/\d{4})", text)
                    if match:
                        entry_date = match.group(1)
                    print(entry_date)

                    tenstore = None
                    match = re.search(r"Chi nhánh yêu cầu\s*:\s*(.+)", text)
                    if match:
                        tenstore = match.group(1).strip()

                    


                    cancel_date = None
                    match = re.search(r"Ngày giao hàng\s*:\s*(\d{1,2}/\d{1,2}/\d{4})", text)
                    if match:
                        cancel_date = match.group(1)
                    print(cancel_date)

                    start = text.rfind("(VND)")  # tìm vị trí xuất hiện cuối cùng
                    end = text.find("TỔNG CỘNG TRƯỚC THUẾ VAT")

                    product_block = text[start + len("(VND)"):end].strip()
                    result = "\n".join(product_block.splitlines())
                    print(result)

                    products = ProcessHandler.tach_san_pham_Farmer(result)

                    print(products)

                    if products:
                        sku_mapping = ProcessHandler.load_sku_mapping()  # Load mapping SKU
                        products = ProcessHandler.replace_sku_numbers(products, sku_mapping)  # Thay SKU
                        soluongsanphamstore = len(products)
                        
                        
                        makhachhang = "MN_MT_FMDC01"
                        

                        self.log_signal.emit(f"<b>🏬 Đơn hàng Store:</b> <span style='color:orange;'>{tenstore}</span>")
                        self.log_signal.emit(f"<b>📦 Số PO:</b> <span style='color:blue;'>{POdonhang}</span>")
                        self.log_signal.emit(f"<b>🗓️ Ngày đặt hàng:</b> <span style='color:green;'>{entry_date}</span>")
                        self.log_signal.emit(f"<b>⏳ Hạn đơn hàng:</b> <span style='color:red;'>{cancel_date}</span>")
                        self.log_signal.emit(f"<b>🛒 Số lượng sản phẩm đặt của Store:</b> <span style='color:brown;'>{soluongsanphamstore}</span>")
                        
                        


                        saigia = ProcessHandler.write_to_dondathang_farmer(self,products,makhachhang,POdonhang,entry_date,cancel_date,stt,vendor,tenstore)
                        
                        if not makhachhang: 
                            makhachhang = "Không xác định"
                            
                        if saigia > 0 or makhachhang == "Không xác định":
                            
                            if saigia == 0:
                                saigia = "Đúng giá"
                            else:
                                saigia = f'Có {saigia} mã sai giá'
                            hoanthanh = "⚠️Hoàn Thành"
                        else:
                            saigia = "Đúng giá"
                            hoanthanh = "✅Hoàn Thành"
                        self.table_signal.emit(file_name, page_label,vendor,makhachhang, POdonhang,saigia, hoanthanh)  # Phát tín hiệu cập nhật bảng










                elif vendor == "FujiMart":

                    
                    
                    # Chuyển trang PDF thành ảnh
                    page = doc.load_page(page_num)  # thêm dòng này để lấy đúng trang PDF
                    # Dùng dpi cao để ảnh rõ
                    pix = page.get_pixmap(dpi=500)
                    img = Image.open(io.BytesIO(pix.tobytes("png")))

                    # Tiền xử lý ảnh: grayscale + tăng tương phản
                    img = img.convert("L")  # grayscale
                    img = ImageEnhance.Contrast(img).enhance(2.5)

                    # OCR tiếng Việt với cấu hình tốt
                    info = pytesseract.image_to_string(img, lang='vie')
                    

                    
                    POdonhang = info.splitlines()[0]
                    POdonhang = re.sub(r'\D', '', POdonhang)
                    
                    lines = text.splitlines()
                    # Tìm dòng trước PO
                    entry_date = next(lines[i - 1] for i in range(1, len(lines)) if POdonhang in lines[i])

                    # Tìm ngày giao
                    cancel_date = next((line.split("Ngµy giao:")[1].strip() for line in text.splitlines() if "Ngµy giao:" in line), "")

                    # Kiểm tra ngày giao hợp lệ
                    if not re.match(r"\d{2}/\d{2}/\d{4}$", cancel_date):
                        
                        cancel_date = "Không tìm thấy"

                        # Chỉ cộng thêm nếu entry_date hợp lệ
                        if cancel_date == "Không tìm thấy" and re.match(r"\d{2}/\d{2}/\d{4}$", entry_date):
                            entry_date_obj = datetime.strptime(entry_date, "%d/%m/%Y")
                            cancel_date = (entry_date_obj + timedelta(days=2)).strftime("%d/%m/%Y")
                    

                    print(info)

                    tenstore = ""
                    match = re.search(r"N\s*ơ\s*i\s*[\s]*n\s*h\s*ậ\s*n\s*:\s*(.+?)(?=\n|$)", info, re.IGNORECASE)

                    if match:
                        tenstore = match.group(1)


                    

                    start = text.rfind("§Þa chØ:")  # tìm vị trí xuất hiện cuối cùng
                    end = text.find("ng­êi ®Æt ®¬n")

                    product_block = text[start + len("§Þa chØ:"):end].strip()
                    result = "\n".join(product_block.splitlines())
                    print(result)
                    products = ProcessHandler.tach_san_pham_Fujimart(result)
                    print(products)
                    if products:
                        sku_mapping = ProcessHandler.load_sku_mapping()  # Load mapping SKU
                        products = ProcessHandler.replace_sku_numbers(products, sku_mapping)  # Thay SKU
                        soluongsanphamstore = len(products)

                        
                        socuahang = tenstore.split()[0]
                        
                        makhachhang = ProcessHandler.laymakhachhang_fujimart(f'FJ{socuahang}')
                        print(f'SỐ cửa hàng: FJ{socuahang}')
                        

                        self.log_signal.emit(f"<b>🏬 Đơn hàng Store:</b> <span style='color:orange;'>{tenstore}</span>")
                        self.log_signal.emit(f"<b>📦 Số PO:</b> <span style='color:blue;'>{POdonhang}</span>")
                        self.log_signal.emit(f"<b>🗓️ Ngày đặt hàng:</b> <span style='color:green;'>{entry_date}</span>")
                        self.log_signal.emit(f"<b>⏳ Hạn đơn hàng:</b> <span style='color:red;'>{cancel_date}</span>")
                        self.log_signal.emit(f"<b>🛒 Số lượng sản phẩm đặt của Store:</b> <span style='color:brown;'>{soluongsanphamstore}</span>")
                        
                        


                        saigia = ProcessHandler.write_to_dondathang_fujimart(self,products,makhachhang,POdonhang,entry_date,cancel_date,stt,vendor,tenstore)
                        
                        if not makhachhang: 
                            makhachhang = "Không xác định"
                            
                        if saigia > 0 or makhachhang == "Không xác định":
                            
                            if saigia == 0:
                                saigia = "Đúng giá"
                            else:
                                saigia = f'Có {saigia} mã sai giá'
                            hoanthanh = "⚠️Hoàn Thành"
                        else:
                            saigia = "Đúng giá"
                            hoanthanh = "✅Hoàn Thành"
                        self.table_signal.emit(file_name, page_label,vendor,makhachhang, POdonhang,saigia, hoanthanh)  # Phát tín hiệu cập nhật bảng

                     




                    
                    
                    
                    





                

                elif vendor == "Winmart":
                    
                    lines = text.split("\n")
                    idx = next((i for i, line in enumerate(lines) if "Ngày đặt hàng (PO date)" in line), -1)
                    entry_date = lines[idx + 1].strip() if idx != -1 and idx + 1 < len(lines) else None
                    entry_date = entry_date.replace('.','/')
                    

                    ghichu = "\n".join(text.split("Ghi chú")[1].split("Nhà cung cấp (Supplier): 0002011398")[0].strip().splitlines()[:-1])
                    ghichu = ghichu.replace('\n',' ')
                    
                    
                    
                    

                    idx = next((i for i, line in enumerate(lines) if "Số đơn hàng (PO No.)" in line), -1)
                    POdonhang = lines[idx + 1].strip() if idx != -1 and idx + 1 < len(lines) else None

                    idx = next((i for i, line in enumerate(lines) if "Ngày giao (Delivery Date)" in line), -1)
                    cancel_date = lines[idx + 1].strip() if idx != -1 and idx + 1 < len(lines) else None
                    cancel_date = cancel_date.replace('.','/')

                    idx = next((i for i, line in enumerate(lines) if "Địa chỉ giao hàng (Delivery Address)" in line), -1)
                    diachigiaohang = lines[idx + 1].strip() if idx != -1 and idx + 1 < len(lines) else None
                    diachigiaohang = diachigiaohang.replace('.','/')
                    
                   
                    
                    self.log_signal.emit(f"<b>📦 Số PO:</b> <span style='color:blue;'>{POdonhang}</span>")
                    self.log_signal.emit(f"<b>🗓️ Ngày đặt hàng:</b> <span style='color:green;'>{entry_date}</span>")

                    self.log_signal.emit(f"<b>⏳ Hạn đơn hàng:</b> <span style='color:red;'>{cancel_date}</span>")
                    

                    # Ghép các dòng lại để tránh trường hợp bị xuống dòng mất chữ

                    
                    

                    idx = -1
                    for i in range(len(lines) - 1):  # -1 để tránh vượt chỉ số khi kiểm tra dòng tiếp theo
                        line_lower = lines[i].lower()  # Chuyển dòng hiện tại về chữ thường
                        next_line_lower = lines[i + 1].lower()  # Chuyển dòng tiếp theo về chữ thường
                        
                        # Kiểm tra "TỔNG HỢP" và "WINCOMMERCE" ở hai dòng liên tiếp
                        if "tổng hợp" in line_lower and "wincommerce" in next_line_lower:
                            idx = i  # Lấy dòng đầu tiên làm mốc
                            break
                        # Kiểm tra "WINCOMMERCE" hoặc "TỔNG HỢP WINCOMMERCE" trong cùng một dòng
                        elif "wincommerce" in line_lower:
                            idx = i
                            break

                    if idx != -1:
                        diachi = []
                        for i in range(idx + 1, len(lines)):
                            # Dừng lại khi gặp "MST" hoặc "Địa chỉ giao hàng", không phân biệt hoa/thường
                            line_lower = lines[i].lower()
                            if "mst" in line_lower or "địa chỉ giao hàng" in line_lower:
                                break
                            diachi.append(lines[i].strip())
                        
                        diachi = " ".join(diachi)
                    else:
                        diachi = None
                    
                    

                    


                    products = ProcessHandler.trichxuatsanpham_winmart(text)
                    if products:
                        sku_mapping = ProcessHandler.load_sku_mapping()  # Load mapping SKU
                        products = ProcessHandler.replace_sku_numbers(products, sku_mapping)  # Thay SKU
                        soluongsanphamstore = len(products)
                        self.log_signal.emit(f"<b>🛒 Số lượng sản phẩm đặt của Store:</b> <span style='color:brown;'>{soluongsanphamstore}</span>")
                        

                        makhachhang = ProcessHandler.laymakhachhang_satra(diachi,"WINMART")
                        
                        saigia = ProcessHandler.write_to_dondathang_winmart(self,products,makhachhang,POdonhang,entry_date,cancel_date,stt,vendor,diachigiaohang,ghichu)

                        if not makhachhang: 
                            makhachhang = "Không xác định"
                        if saigia > 0 or makhachhang == "Không xác định":
                            if saigia == 0:
                                saigia = "Đúng giá"
                            else:
                                saigia = f'Có {saigia} mã sai giá'
                            hoanthanh = "⚠️Hoàn Thành"
                        else:
                            saigia = "Đúng giá"
                            hoanthanh = "✅Hoàn Thànhh"
                        self.table_signal.emit(file_name, page_label,vendor,makhachhang, POdonhang, saigia, hoanthanh)  # Phát tín hiệu cập nhật bảng










                elif vendor == "Lotte":
                    
                    lines = text.splitlines()
                    POdonhang = lines[1] if len(lines) > 1 else ""
                    if len(POdonhang) >= 7:
                        POdonhang = POdonhang[:6] + "-" + POdonhang[6:]
                    if len(POdonhang) >= 12:  # 7 + 1 (dấu -) + 4 = 12
                        POdonhang = POdonhang[:12] + "-" + POdonhang[12:]

                    
                    time_part, store_code, order_number = POdonhang.split("-")

                    # Chuyển đổi thời gian từ yyMMdd sang dd/MM/yyyy
                    entry_date = datetime.strptime(time_part, "%y%m%d").strftime("%d/%m/%Y")

                    cancel_date = ProcessHandler.tachcancledate_lotte(text,POdonhang)
                    tenstore = ProcessHandler.laytenstore_lotte(text,POdonhang)
                    self.log_signal.emit(f"<b>🏬 Đơn hàng Store:</b> <span style='color:orange;'>{tenstore}</span>")
                    self.log_signal.emit(f"<b>📦 Số PO:</b> <span style='color:blue;'>{POdonhang}</span>")
                    self.log_signal.emit(f"<b>🗓️ Ngày đặt hàng:</b> <span style='color:green;'>{entry_date}</span>")
                    self.log_signal.emit(f"<b>⏳ Hạn đơn hàng:</b> <span style='color:red;'>{cancel_date}</span>")
                   
                    diachigiaohang = "Lotte " + tenstore

                    
                    product_details = ProcessHandler.tachsanpham_lotte(text)
                    soluongsanpham = len(product_details)
                    self.log_signal.emit(f"<b>📊 Tổng số lượng sản phẩm:</b> <span style='color:purple;'>{soluongsanpham}</span>")
                    store_code = ProcessHandler.get_makhachhang_lotte(store_code[1:])
                    saigia = ProcessHandler.write_to_dondathang_lotte(self,product_details,store_code,POdonhang,entry_date,cancel_date,stt,vendor,diachigiaohang)
                    
                    if not store_code: 
                        store_code = "Không xác định"
                    if saigia > 0 or store_code == "Không xác định":
                        if saigia == 0:
                            saigia = "Đúng giá"
                        else:
                            saigia = f'Có {saigia} mã sai giá'
                        hoanthanh = "⚠️Hoàn Thành"
                    else:
                        saigia = "Đúng giá"
                        hoanthanh = "✅Hoàn Thành"
                    self.table_signal.emit(file_name, page_label,vendor,store_code, POdonhang,saigia, hoanthanh)  # Phát tín hiệu cập nhật bảng
                


                elif vendor == "Kingfood":
                    text = doc[page_num].get_text("text")
                    


                    if text:
                        tranggoc = doc[page_num].get_text("text")
                        

                        po_number = re.search(r"PO Number:\s*\n([^\n]*\n)?([^\n]*)", tranggoc)
                        po_number = po_number.group(1) if po_number else "Không tìm thấy PO Number"
                        self.log_signal.emit(f"<b>📦 Số PO:</b> <span style='color:blue;'>{po_number}</span>")

                        entry_date = re.search(r"Ngày Đặt Hàng:\s*\n([^\n]*\n)?([^\n]*)", tranggoc)
                        entry_date = entry_date.group(1).replace("-","/").strip() if entry_date else "Không tìm thấy ngày đặt hàng"
                        entry_date = datetime.strptime(entry_date, "%d/%m/%Y")  # Chuyển sang datetime
                        entry_date = entry_date.strftime("%d/%m/%Y")
                        self.log_signal.emit(f"<b>🗓️ Ngày đặt hàng:</b> <span style='color:green;'>{entry_date}</span>")

                        cancel_date = re.search(r"Ngày Giao Hàng NCC Xác Nhận:\s*\n([^\n]*\n)?([^\n]*)", tranggoc)
                        cancel_date = cancel_date.group(1).replace("-","/").strip() if cancel_date else "Không tìm thấy ngày giao hàng"
                        cancel_date = datetime.strptime(cancel_date, "%d/%m/%Y")  # Chuyển sang datetime
                        cancel_date = cancel_date.strftime("%d/%m/%Y")
                        self.log_signal.emit(f"<b>⏳ Hạn đơn hàng:</b> <span style='color:red;'>{cancel_date}</span>")

                        products = ProcessHandler.laydanhsachsanpham_kingfood(text)
                        
                        if products:
                            sku_mapping = ProcessHandler.load_sku_mapping()  # Load mapping SKU
                            products = ProcessHandler.replace_sku_numbers(products, sku_mapping)  # Thay SKU
                            soluongsanphamstore = len(products)
                            self.log_signal.emit(f"<b>🛒 Số lượng sản phẩm đặt của Store:</b> <span style='color:brown;'>{soluongsanphamstore}</span>")
                            store_code = "MN_MT_KFMSL"
                        
                        
                            saigia = ProcessHandler.write_to_dondathang_kingfood(self,products,store_code,po_number,entry_date,cancel_date,stt,vendor)
                            if not store_code: 
                                store_code = "Không xác định"
                            if saigia > 0 or store_code == "Không xác định":
                                if saigia == 0:
                                    saigia = "Đúng giá"
                                else:
                                    saigia = f'Có {saigia} mã sai giá'
                                    hoanthanh = "⚠️Hoàn Thành"
                            else:
                                saigia = "Đúng giá"
                                hoanthanh = "✅Hoàn Thành"
                            self.table_signal.emit(file_name, page_label,vendor,store_code, po_number,saigia, hoanthanh)  # Phát tín hiệu cập nhật bảng





                elif vendor == "Emart":
                    
                    po_number = re.search(r"PO No\.\s*\n\s*:? ?([^\n]+)", text)

                    if po_number:
                        po_number = po_number.group(1).strip()  # Loại bỏ dấu ":" và trim khoảng trắng
                        self.log_signal.emit(f"<b>📦 Số PO:</b> <span style='color:blue;'>{po_number}</span>")
                        
                    entry_date = re.search(r"Order By / Date\s*\n\s*:? ?([^\n]+)", text)
                    if entry_date:
                        entry_date = entry_date.group(1).strip()  # Loại bỏ dấu ":" và trim khoảng trắng
                        entry_date = entry_date[:10].replace(".", "/")
                        self.log_signal.emit(f"<b>🗓️ Ngày đặt hàng:</b> <span style='color:green;'>{entry_date}</span>")
                    cancel_date = re.search(r"Delivery Date\s*\n\s*:? ?([^\n]+)", text)
                    if cancel_date:
                        cancel_date = cancel_date.group(1).strip()  # Loại bỏ dấu ":" và trim khoảng trắng
                        cancel_date = cancel_date[:10].replace(".", "/")
                        self.log_signal.emit(f"<b>⏳ Hạn đơn hàng:</b> <span style='color:red;'>{cancel_date}</span>")

                    tenstore = re.search(r"^Delivery to :\s*(.+)", text, re.MULTILINE)
                    if tenstore:
                        tenstore = tenstore.group(1).split("   ")[0]
                    else:
                        print("Không tìm thấy thông tin giao hàng.")
                    self.log_signal.emit(f"<b>🏬 Đơn hàng Store:</b> <span style='color:orange;'>{tenstore}</span>")
                    text =  re.search(r"Article Code\s*(.*?)\s*Total Amount\(without VAT\) :", text, re.DOTALL)
                    text = text.group(1).strip()
                    products = ProcessHandler.laydanhsanpham_emart(text)
                    if products:
                        sku_mapping = ProcessHandler.load_sku_mapping()  # Load mapping SKU
                        products = ProcessHandler.replace_sku_numbers(products, sku_mapping)  # Thay SKU
                    soluongsanphamstore = len(products)
                    self.log_signal.emit(f"<b>🛒 Số lượng sản phẩm đặt của Store:</b> <span style='color:brown;'>{soluongsanphamstore}</span>")
                    saigia = ProcessHandler.write_to_dondathang_emart(self,products,po_number,entry_date,cancel_date,tenstore,stt)
                    if not tenstore: 
                        tenstore = "Không xác định"
                        
                    if saigia > 0 or tenstore == "Không xác định":
                        if saigia == 0:
                            saigia = "Đúng giá"
                        else:
                            saigia = f'Có {saigia} mã sai giá'
                        hoanthanh = "⚠️Hoàn Thành"

                    else:
            
                        saigia = "Đúng giá"
                        hoanthanh ="✅Hoàn Thành"
                    self.table_signal.emit(file_name, page_label,vendor,tenstore, po_number,saigia,hoanthanh )  # Phát tín hiệu cập nhật bảng


                elif vendor == "Satra":
                    
                    po_number =  re.search(r"\*P-[^*]+\*", text)
                    po_number =  po_number.group(0)[1:-1]
                    self.log_signal.emit(f"<b>📦 Số PO:</b> <span style='color:blue;'>{po_number}</span>")
                    diachi = re.search(r"Địa chỉ giao hàng:\s*\n(.*)", text)
                    if diachi:
                        diachi = diachi.group(1)
                        makhachhang = ProcessHandler.laymakhachhang_satra(diachi,"SATRA")
                    
                    entry_date = re.search(r"(.*?)\nNgày đặt hàng:", text, re.DOTALL)
                    if entry_date:
                        entry_date = entry_date.group(1).split("\n")[-1]  # Lấy dòng cuối cùng trước "Ngày đặt hàng:"
                        entry_date = datetime.strptime(entry_date, "%m/%d/%Y")
                        entry_date = entry_date.strftime("%d/%m/%Y")
                        if entry_date == '01/01/0001':
                            entry_date = re.search(r"(.*?)\nNgày in:", text, re.DOTALL)
                            if entry_date:
                                entry_date = entry_date.group(1).split("\n")[-1]  # Lấy dòng cuối cùng trước "Ngày đặt hàng:"
                                entry_date = datetime.strptime(entry_date, "%m/%d/%Y")
                                entry_date = entry_date.strftime("%d/%m/%Y")
                    self.log_signal.emit(f"<b>🗓️ Ngày đặt hàng:</b> <span style='color:green;'>{entry_date}</span>")
                    
                    cancel_date = re.search(r"Ngày giao hàng:\s*(.*?)\s*Địa chỉ giao hàng:", text, re.DOTALL)
                    if cancel_date:
                        cancel_date = cancel_date.group(1).strip()
                        pattern = r"(\d{1,2}/\d{1,2}/\d{4})"
                        for line in cancel_date.split("\n"):
                            if re.search(pattern, line):
                                cancel_date = line
                                cancel_date = datetime.strptime(cancel_date, "%m/%d/%Y")
                                cancel_date = cancel_date.strftime("%d/%m/%Y")
                    self.log_signal.emit(f"<b>⏳ Hạn đơn hàng:</b> <span style='color:red;'>{cancel_date}</span>")
                    text =  re.search(r"STT\s*(.*?)\s*Hàng phục vụ cho:", text, re.DOTALL)
                    text = text.group(1).strip()
                   
                    products = ProcessHandler.trichxuatsanpham_satra(text)
                    if products:
                        sku_mapping = ProcessHandler.load_sku_mapping()  # Load mapping SKU
                        products = ProcessHandler.replace_sku_numbers(products, sku_mapping)  # Thay SKU
                    soluongsanphamstore = len(products)
                    self.log_signal.emit(f"<b>🛒 Số lượng sản phẩm đặt của Store:</b> <span style='color:brown;'>{soluongsanphamstore}</span>")
                    saigia = ProcessHandler.write_to_dondathang_satra(self,products,makhachhang,po_number,entry_date,cancel_date,stt,vendor,diachi)
                    
                    if not makhachhang: 
                        makhachhang = "Không xác định"
                        
                    if saigia > 0 or makhachhang == "Không xác định":
                        
                        if saigia == 0:
                            saigia = "Đúng giá"
                        else:
                            saigia = f'Có {saigia} mã sai giá'
                        hoanthanh = "⚠️Hoàn Thành"
                    else:
                        saigia = "Đúng giá"
                        hoanthanh = "✅Hoàn Thành"
                    self.table_signal.emit(file_name, page_label,vendor,makhachhang, po_number,saigia, hoanthanh)  # Phát tín hiệu cập nhật bảng
                    
                    
                elif vendor == "BigC":
                    
                    if page_num == 0:
                        trangdaubigc = text
                        if "LINFOX WAREHOUSE (802)" in trangdaubigc:
                            makhachhang = "MB_MT_BIGC"
                            diachigiao = "LINFOX WAREHOUSE (802)"
                        else:
                            makhachhang = "MN_MT_BIGCAC"
                            diachigiao = "FM LOGISTIC VSIP 2 (806)"
                            
                    
                        po_number, entry_date, cancel_date = ProcessHandler.trichxuatinfo_donbigc(trangdaubigc)
                        self.log_signal.emit(f"<b>📦 Số PO:</b> <span style='color:blue;'>{po_number}</span>")
                        self.log_signal.emit(f"<b>🗓️ Ngày đặt hàng:</b> <span style='color:green;'>{entry_date}</span>")
                        self.log_signal.emit(f"<b>⏳ Hạn đơn hàng:</b> <span style='color:red;'>{cancel_date}</span>")
                        products = ProcessHandler.laydanhsachsanpham_bigc(trangdaubigc)
                        soluongsanpham = len(products)
                        self.log_signal.emit(f"<b>📊 Tổng số lượng sản phẩm:</b> <span style='color:purple;'>{soluongsanpham}</span>")

                    elif page_num > 0:
                        tenstore = ProcessHandler.lay_ten_store(text)
                    
                        items = ProcessHandler.trichxuatdanhsachforstore_bigc(text)
                        self.log_signal.emit(f"<b>🏬 Đơn hàng Store:</b> <span style='color:orange;'>{tenstore}</span>")
                        soluongsanphamstore = len(items)
                        self.log_signal.emit(f"<b>🛒 Số lượng sản phẩm đặt của Store:</b> <span style='color:brown;'>{soluongsanphamstore}</span>")
                        

                        saigia = ProcessHandler.write_to_dondathang_bigc(self,products,items,po_number,entry_date,cancel_date,tenstore,stt,makhachhang,vendor,page_num,diachigiao)
                        
                        if not tenstore: 
                            tenstore = "Không xác định"
                            
                        if saigia > 0 or tenstore == "Không xác định":
                            if saigia == 0:
                                saigia = "Đúng giá"
                            else:
                                saigia = f'Có {saigia} mã sai giá'
                            hoanthanh = "⚠️Hoàn Thành"
                        else:
                            saigia = "Đúng giá"
                            hoanthanh = "✅Hoàn Thành"
                        self.table_signal.emit(file_name, page_label,vendor,tenstore, po_number, saigia,hoanthanh)  # Phát tín hiệu cập nhật bảng


                        # (Xử lý BigC ở đây)

                else:
                    self.log_signal.emit("❌ Vendor không hỗ trợ hoặc không xác định!")
                    self.table_signal.emit(file_name, page_label,"Không xác định", "", "","", "⛔Thất bại")  # Phát tín hiệu cập nhật bảng


            self.log_signal.emit("✅ Xử lý PDF hoàn tất!")

        elif file_path.lower().endswith(".txt"):
            try:
                text = ProcessHandler.read_text_file(file_path)
                
                vendor = ProcessHandler.identify_vendor(text)

                if vendor == "Coop":
                    ProcessHandler.process_coop_invoice(self,text,stt,file_path,"1/1")
                    self.log_signal.emit("✅ Xử lý TXT hoàn tất!")
                else:
                    self.log_signal.emit("❌ Vendor không hỗ trợ hoặc không xác định!")
            except UnicodeDecodeError as e:
                self.log_signal.emit(f"⚠ Lỗi đọc file TXT: {e}")

        elif file_path.lower().endswith(".xlsx"):
            self.log_signal.emit("📊 Đang xử lý file Excel...")

            while True:
                    # Lấy dữ liệu từ file
                po_location, po_number, entry_date, cancel_date, products = ProcessHandler.get_grouped_data_xlsx(file_path)

                    # Kiểm tra nếu không còn dữ liệu để xử lý, thì dừng vòng lặp
                wb = openpyxl.load_workbook(filename=file_path, data_only=True)
                ws = wb["Sheet1"]
        
                makhachhang = ProcessHandler.get_makhachhang(po_location) if po_location else "Không tìm thấy"

                if products:
                    sku_mapping = ProcessHandler.load_sku_mapping()  # Load mapping SKU
                    products = ProcessHandler.replace_sku_numbers(products, sku_mapping)  # Thay SKU

                    # Ghi vào file "dondathang.xlsx"
                    saigia = ProcessHandler.write_to_dondathang(self,products, makhachhang, po_number, entry_date, cancel_date,stt,"COOP","")
                    if not makhachhang: 
                        makhachhang = "Không xác định"
                        
                    if saigia > 0 or makhachhang == "Không xác định":
                        if saigia == 0:
                            saigia = "Đúng giá"
                        else:
                            saigia = f'Có {saigia} mã sai giá'
                        hoanthanh = "⚠️Hoàn Thành"
                    else:
                        saigia = "Đúng giá"
                        hoanthanh = "✅Hoàn Thành"

                    hethong = ProcessHandler.layhethong_COOP(makhachhang)
                    self.table_signal.emit(file_name, "1/1",hethong,makhachhang, po_number, saigia, hoanthanh)  # Phát tín hiệu cập nhật bảng
                if ws.max_row == 1:  # Chỉ còn lại dòng tiêu đề
                    print("✅ Đã xử lý xong toàn bộ dữ liệu! Dừng lại.")
                    return
                print("✅ Đã xử lý xong một đơn hàng. Tiếp tục...")


                # Xử lý file Excel
                self.log_signal.emit("✅ Xử lý Excel hoàn tất!")

        else:
            self.log_signal.emit("⚠ Định dạng file không được hỗ trợ!")

        self.log_signal.emit("✅ Xử lý hoàn tất!")







        
    


    




process_handler = ProcessHandler()
def lay_gia_tri_G1():
    """
    Lấy giá trị ô G1 từ file 'dondathang.xlsx', sheet 'Don dat hang'.
    Trả về giá trị của ô G1 hoặc None nếu ô trống.
    """
    file_path = "dondathang.xlsx"
    sheet_name = "Don dat hang"

    wb = openpyxl.load_workbook(file_path, data_only=True)  # Đọc giá trị thực (không phải công thức)
    sheet = wb[sheet_name]

    value_G1 = sheet["G1"].value  # Lấy giá trị ô G1

    wb.close()
    return value_G1