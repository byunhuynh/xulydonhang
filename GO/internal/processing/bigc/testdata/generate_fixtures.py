"""
Throwaway dev tool — NOT part of the shipped Go or Python app.

Runs the CURRENT xulydonhang.py BigC pipeline against every real PDF in
đơn hàng/08-2026/ that identify_vendor recognizes as BigC, capturing the
resulting dondathang.xlsx rows (and the live-fetched Google Sheets
price/promotion data for the BIGC sheet) into JSON fixtures under
GO/internal/processing/bigc/testdata/fixtures/. The Go golden test
(Task 8) diffs RealProcessor's output against these fixtures instead of
against a live Google Sheets fetch, so it's deterministic and offline.

Unlike Coop/Lotte/Satra's harnesses (one page == one write call, snapshot
immediately after), BigC's write_to_dondathang_bigc is called once PER
STORE PAGE within a single file (page 0 sets up shared state; pages
1..N-1 each write one store's rows), so this harness mirrors process_file's
BigC branch (xulydonhang.py:9404-9536) across a WHOLE file before taking
one snapshot per PDF, spanning every store's rows.

Run from the repo root:
    .venv/Scripts/python.exe GO/internal/processing/bigc/testdata/generate_fixtures.py
"""
import glob
import json
import os
import shutil
import sys
import time

# Same depth as Satra's harness: this script sits 5 directory levels below
# repo root (GO/internal/processing/bigc/testdata/generate_fixtures.py),
# so reaching repo root from os.path.abspath(__file__) requires 6
# dirname() calls (one to strip the filename, five more to strip
# GO/internal/processing/bigc/testdata).
REPO_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))))
sys.path.insert(0, REPO_ROOT)
os.chdir(REPO_ROOT)  # xulydonhang.py's functions use relative paths ("data.xlsx", "settings.ini")

# See Coop's harness for why this is needed: process_file's debug print()
# calls contain emoji that the legacy cp1252 console codepage can't
# encode, aborting processing partway through if not fixed here first.
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="backslashreplace")

import openpyxl  # noqa: E402
import xulydonhang  # noqa: E402

FIXTURES_DIR = os.path.join(
    REPO_ROOT, "GO", "internal", "processing", "bigc", "testdata", "fixtures"
)
TEMPLATE_XLSX = os.path.join(REPO_ROOT, "dondathang.xlsx")
SCRATCH_XLSX = os.path.join(REPO_ROOT, "dondathang_fixture_scratch.xlsx")

# --- Monkey-patch network/upload side effects out (identical shape to
# Coop/Satra's harnesses; find_price_by_sku/find_all_promotions_by_sku_and_time
# are already generic over sheet_name, so this works for "BIGC" too
# with no changes to the caching logic itself) ---

_price_cache = {}
_promo_cache = {}
_promo_raw_rows = None


def _cached_find_price_by_sku(sku_number, sheet_name="COOP"):
    key = (sku_number, sheet_name)
    if key not in _price_cache:
        _price_cache[key] = _real_find_price_by_sku(sku_number, sheet_name)
    return _price_cache[key]


def _cached_find_all_promotions(sku_code, time_to_check, sheet_name="Coop"):
    global _promo_raw_rows
    if _promo_raw_rows is None:
        _capture_promo_raw_rows(sheet_name)
    key = (sku_code, time_to_check, sheet_name)
    if key not in _promo_cache:
        _promo_cache[key] = _real_find_all_promotions(sku_code, time_to_check, sheet_name)
    return _promo_cache[key]


def _capture_promo_raw_rows(sheet_name):
    global _promo_raw_rows
    import pandas as pd

    gid = xulydonhang.ProcessHandler.get_gid(sheet_name)
    if not gid:
        _promo_raw_rows = []
        return
    sheet_id = "1yvxE_SPYXKhofcZdhv1CSKAyiwdY1Mf4pFlsiMbtOr4"
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?tqx=out:csv&gid={gid}"
    df = pd.read_csv(url, dtype=str, header=None)
    df.fillna("", inplace=True)
    _promo_raw_rows = df.values.tolist()


def _noop_upload_file_to_drive(path, output_name=None, vendor=None, entry_date=None, makhachhang=None, cancle_date=None):
    return {"url": "https://example.invalid/skipped-during-fixture-generation"}


_real_find_price_by_sku = xulydonhang.ProcessHandler.find_price_by_sku
_real_find_all_promotions = xulydonhang.ProcessHandler.find_all_promotions_by_sku_and_time
xulydonhang.ProcessHandler.find_price_by_sku = staticmethod(_cached_find_price_by_sku)
xulydonhang.ProcessHandler.find_all_promotions_by_sku_and_time = staticmethod(_cached_find_all_promotions)
xulydonhang.ProcessHandler.upload_file_to_drive = staticmethod(_noop_upload_file_to_drive)


# --- Excel row capture (same columns as Coop/Satra's harnesses — same sheet, same layout) ---

COLUMNS = [
    "A", "B", "C", "D", "E", "G", "L", "Q", "S", "T", "U", "V", "X", "Y", "Z",
    "AE", "AJ", "AM", "AO", "AP", "AQ", "AT", "AU", "AV",
]


def snapshot_rows(path, start_row):
    wb = openpyxl.load_workbook(path)
    sheet = wb["Don dat hang"]
    rows = []
    for r in range(start_row, sheet.max_row + 1):
        row = {"row_number_offset": r - start_row}
        for col in COLUMNS:
            cell = sheet[f"{col}{r}"]
            value = cell.value
            row[col] = value
            if col == "Z":
                row["Z_is_formula"] = isinstance(value, str) and value.startswith("=")
        comment = sheet[f"Y{r}"].comment
        row["Y_has_comment"] = comment is not None
        row["Y_fill"] = sheet[f"Y{r}"].fill.fgColor.rgb if sheet[f"Y{r}"].fill else None
        rows.append(row)
    wb.close()
    return rows


def is_bigc_pdf(path):
    doc = xulydonhang.fitz.open(path)
    try:
        text = ""
        for page in doc:
            text += page.get_text("text")
        return xulydonhang.ProcessHandler.identify_vendor(text) == "BigC"
    finally:
        doc.close()


def compute_start_row(path):
    """Mirrors process_file's BigC page-0 setup (xulydonhang.py:9447-9455)
    EXACTLY: open with data_only=True, then walk max_row backwards over
    fully-blank rows (all cells None/""/" ") before settling on
    last_row + 1. NOT a naive `sheet.max_row + 1` — the real code skips
    trailing blank rows first, and an off-by-one here would corrupt the
    row range every fixture captures."""
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb["Don dat hang"]
    last_row = sheet.max_row
    while last_row > 0 and all(c.value in (None, "", " ") for c in sheet[last_row]):
        last_row -= 1
    wb.close()
    return last_row + 1


def process_one_pdf(path, start_row):
    """Mirrors the BigC branch of process_file (xulydonhang.py:9404-9536)
    for a whole file: page 0 sets up shared state (po_number, entry_date,
    cancel_date, products, makhachhang, diachigiao), then every
    subsequent page calls write_to_dondathang_bigc once as a store page —
    skipping the Google Drive upload side effect (already no-op'd by the
    monkeypatch above). `start_row` is the value process_file itself
    would have captured at page 0 (see compute_start_row above) and is
    passed through unchanged as the `bat_dau` argument on the LAST page's
    call only, exactly mirroring xulydonhang.py:9455/9508."""
    handler = xulydonhang.process_handler
    doc = xulydonhang.fitz.open(path)
    try:
        page0_text = doc[0].get_text("text")
        if xulydonhang.ProcessHandler.identify_vendor(page0_text) != "BigC":
            return  # first page isn't BigC after all; skip whole file

        po_number, entry_date, cancel_date = xulydonhang.ProcessHandler.trichxuatinfo_donbigc(page0_text)
        products = xulydonhang.ProcessHandler.laydanhsachsanpham_bigc(page0_text)

        trangdaubigc = page0_text
        if "3006900" in trangdaubigc and "LINFOX WAREHOUSE (802)" in trangdaubigc:
            makhachhang = "MB_GC_BIGC"; diachigiao = "LINFOX WAREHOUSE (802)"
        elif "3005382" in trangdaubigc and "LINFOX WAREHOUSE (802)" in trangdaubigc:
            makhachhang = "MB_MT_BIGC"; diachigiao = "LINFOX WAREHOUSE (802)"
        elif "3005382" in trangdaubigc and "FM LOGISTIC VSIP 2 (806)" in trangdaubigc:
            makhachhang = "MN_MT_BIGCAC"; diachigiao = "FM LOGISTIC VSIP 2 (806)"
        elif "3006900" in trangdaubigc and "FM LOGISTIC VSIP 2 (806)" in trangdaubigc:
            makhachhang = "MN_GC_BIGCAC"; diachigiao = "FM LOGISTIC VSIP 2 (806)"
        else:
            makhachhang = "MN_MT_BIGCAC"; diachigiao = "FM LOGISTIC VSIP 2 (806)"

        vendor = "BigC"

        for page_num in range(1, len(doc)):
            text = doc[page_num].get_text("text")
            tenstore = xulydonhang.ProcessHandler.lay_ten_store(text)
            items = xulydonhang.ProcessHandler.trichxuatdanhsachforstore_bigc(text)

            if page_num < len(doc) - 1:
                # Middle store page: matches xulydonhang.py:9478 exactly
                # (bat_dau=False, url=None).
                bat_dau = False
                url = None
            else:
                # Last page: matches xulydonhang.py:9508 exactly — bat_dau
                # is the row process_file captured once at page 0, url is
                # whatever upload_file_to_drive returned (no-op'd above to
                # a fixed placeholder, so we inline that same value here
                # instead of calling the no-op indirectly).
                bat_dau = start_row
                url = "https://example.invalid/skipped-during-fixture-generation"

            xulydonhang.ProcessHandler.write_to_dondathang_bigc(
                handler, products, items, po_number, entry_date, cancel_date,
                tenstore, 1, makhachhang, vendor, page_num, diachigiao, bat_dau, url,
            )
    finally:
        doc.close()


def _remove_with_retry(path, attempts=5, delay=0.5):
    """os.remove wrapped with retry-with-backoff. Windows Defender's
    real-time scanner can transiently hold a lock on a freshly-saved
    .xlsx right after openpyxl's wb.save() closes it, which surfaces here
    as a PermissionError ([WinError 5] Access is denied). Retrying a few
    times with a short delay lets the scan finish and the lock clear
    before we give up and let the exception propagate for real."""
    for i in range(attempts):
        try:
            os.remove(path)
            return
        except PermissionError:
            if i == attempts - 1:
                raise
            time.sleep(delay)


def _move_with_retry(src, dst, attempts=5, delay=0.5):
    """shutil.move wrapped with the same retry-with-backoff as
    _remove_with_retry, and for the same reason (transient AV lock)."""
    for i in range(attempts):
        try:
            shutil.move(src, dst)
            return
        except PermissionError:
            if i == attempts - 1:
                raise
            time.sleep(delay)


def main():
    os.makedirs(FIXTURES_DIR, exist_ok=True)

    pdf_paths = sorted(glob.glob(os.path.join(REPO_ROOT, "đơn hàng", "08-2026", "*.pdf")))
    print(f"Found {len(pdf_paths)} candidate PDFs")

    generated = 0
    skipped = 0
    for path in pdf_paths:
        try:
            if not is_bigc_pdf(path):
                continue
        except Exception as e:
            print(f"SKIP (vendor check failed) {os.path.basename(path)}: {e}")
            skipped += 1
            continue

        shutil.copyfile(TEMPLATE_XLSX, SCRATCH_XLSX)
        os.chdir(REPO_ROOT)
        real_target = os.path.join(REPO_ROOT, "dondathang.xlsx")
        backup = real_target + ".fixture_backup"
        shutil.move(real_target, backup)
        shutil.copyfile(SCRATCH_XLSX, real_target)
        try:
            start_row = compute_start_row(real_target)

            process_one_pdf(path, start_row)

            rows = snapshot_rows(real_target, start_row)
        except Exception as e:
            print(f"SKIP (processing failed) {os.path.basename(path)}: {e}")
            skipped += 1
            rows = None
        finally:
            _remove_with_retry(real_target)
            _move_with_retry(backup, real_target)
            if os.path.exists(SCRATCH_XLSX):
                os.remove(SCRATCH_XLSX)

        if rows is None:
            continue

        fixture = {"source_pdf": os.path.basename(path), "rows": rows}
        fixture_name = os.path.splitext(os.path.basename(path))[0] + ".json"
        with open(os.path.join(FIXTURES_DIR, fixture_name), "w", encoding="utf-8") as f:
            json.dump(fixture, f, ensure_ascii=False, indent=2, default=str)
        generated += 1
        print(f"OK {os.path.basename(path)} -> {len(rows)} rows")

    if _promo_raw_rows is None:
        _capture_promo_raw_rows("BIGC")
    frozen_pricing = {"raw_rows": _promo_raw_rows or []}
    with open(os.path.join(FIXTURES_DIR, "_frozen_pricing.json"), "w", encoding="utf-8") as f:
        json.dump(frozen_pricing, f, ensure_ascii=False, indent=2, default=str)

    print(f"\nDone: {generated} fixtures generated, {skipped} PDFs skipped.")


if __name__ == "__main__":
    main()
