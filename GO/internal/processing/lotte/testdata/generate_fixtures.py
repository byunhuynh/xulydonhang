"""
Throwaway dev tool — NOT part of the shipped Go or Python app.

Runs the CURRENT xulydonhang.py Lotte pipeline against every real PDF in
lotte/testdata/realpdfs/ (git-tracked, stable copies located from the live
đơn hàng/08-2026/ folder and the đơn hàng/mẫu đơn hàng/ archive — see the
Lotte testdata migration report) that identify_vendor recognizes as Lotte,
capturing the resulting dondathang.xlsx rows (and the live-fetched Google
Sheets price/promotion data for the LOTTE sheet) into JSON fixtures under
GO/internal/processing/lotte/testdata/fixtures/. The Go golden test
(Task 9) diffs RealProcessor's output against these fixtures instead of
against a live Google Sheets fetch, so it's deterministic and offline.

Run from the repo root:
    .venv/Scripts/python.exe GO/internal/processing/lotte/testdata/generate_fixtures.py
"""
import glob
import json
import os
import shutil
import sys
import time

# Same depth as Coop's harness: this script sits 5 directory levels below
# repo root (GO/internal/processing/lotte/testdata/generate_fixtures.py),
# so reaching repo root from os.path.abspath(__file__) requires 6
# dirname() calls (one to strip the filename, five more to strip
# GO/internal/processing/lotte/testdata).
REPO_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))))
sys.path.insert(0, REPO_ROOT)
os.chdir(REPO_ROOT)  # xulydonhang.py's functions use relative paths ("data.xlsx", "settings.ini")

# See Coop's harness for why this is needed: process_file's debug print()
# calls contain emoji that the legacy cp1252 console codepage can't
# encode, aborting processing partway through if not fixed here first.
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="backslashreplace")

import openpyxl  # noqa: E402
import xulydonhang  # noqa: E402

FIXTURES_DIR = os.path.join(
    REPO_ROOT, "GO", "internal", "processing", "lotte", "testdata", "fixtures"
)
REALPDFS_DIR = os.path.join(
    REPO_ROOT, "GO", "internal", "processing", "lotte", "testdata", "realpdfs"
)
TEMPLATE_XLSX = os.path.join(REPO_ROOT, "dondathang.xlsx")
SCRATCH_XLSX = os.path.join(REPO_ROOT, "dondathang_fixture_scratch.xlsx")


def _restore_with_retry(backup, real_target, attempts=20, delay=1.0):
    """Atomically restore real_target from backup via os.replace() (a single
    atomic overwrite, no separate remove step), with a generous retry
    budget for transient Windows file-lock flakiness (e.g. a concurrently
    running instance of the production app briefly holding the file open).

    This replaces the original two-step os.remove() + shutil.move()
    sequence, which had a real bug found during Coop's fixture
    regeneration: os.remove() could exhaust its own retry budget and raise
    *before* shutil.move() ever ran, abandoning the production-file
    restore mid-run inside the `finally:` block. os.replace() is atomic on
    Windows (backed by MoveFileExW with MOVEFILE_REPLACE_EXISTING), so
    there is no window where real_target is missing or where a failure can
    leave the restore half-done.
    """
    last_err = None
    for attempt in range(1, attempts + 1):
        try:
            os.replace(backup, real_target)
            return
        except OSError as e:
            last_err = e
            time.sleep(delay)
    raise RuntimeError(
        f"Failed to restore {real_target} from {backup} after {attempts} attempts: {last_err}"
    )

# --- Monkey-patch network/upload side effects out (identical shape to
# Coop's harness; find_price_by_sku/find_all_promotions_by_sku_and_time
# are already generic over sheet_name, so this works for "LOTTE" too
# with no changes to the caching logic itself) ---

_price_cache = {}
_promo_cache = {}
_promo_raw_rows = None


def _cached_find_price_by_sku(sku_number, sheet_name="COOP"):
    key = (sku_number, sheet_name)
    if key not in _price_cache:
        _price_cache[key] = _real_find_price_by_sku(sku_number, sheet_name)
    return _price_cache[key]


def _cached_find_all_promotions(sku_code, time_to_check, sheet_name="Coop"):
    global _promo_raw_rows
    if _promo_raw_rows is None:
        _capture_promo_raw_rows(sheet_name)
    key = (sku_code, time_to_check, sheet_name)
    if key not in _promo_cache:
        _promo_cache[key] = _real_find_all_promotions(sku_code, time_to_check, sheet_name)
    return _promo_cache[key]


def _capture_promo_raw_rows(sheet_name):
    global _promo_raw_rows
    import pandas as pd

    gid = xulydonhang.ProcessHandler.get_gid(sheet_name)
    if not gid:
        _promo_raw_rows = []
        return
    sheet_id = "1yvxE_SPYXKhofcZdhv1CSKAyiwdY1Mf4pFlsiMbtOr4"
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?tqx=out:csv&gid={gid}"
    df = pd.read_csv(url, dtype=str, header=None)
    df.fillna("", inplace=True)
    _promo_raw_rows = df.values.tolist()


def _noop_upload_file_to_drive(path, output_name=None, vendor=None, entry_date=None, makhachhang=None, cancle_date=None):
    return {"url": "https://example.invalid/skipped-during-fixture-generation"}


_real_find_price_by_sku = xulydonhang.ProcessHandler.find_price_by_sku
_real_find_all_promotions = xulydonhang.ProcessHandler.find_all_promotions_by_sku_and_time
xulydonhang.ProcessHandler.find_price_by_sku = staticmethod(_cached_find_price_by_sku)
xulydonhang.ProcessHandler.find_all_promotions_by_sku_and_time = staticmethod(_cached_find_all_promotions)
xulydonhang.ProcessHandler.upload_file_to_drive = staticmethod(_noop_upload_file_to_drive)


# --- Excel row capture (same columns as Coop's harness — same sheet, same layout) ---

COLUMNS = [
    "A", "B", "C", "D", "E", "G", "L", "Q", "S", "T", "U", "V", "X", "Y", "Z",
    "AE", "AJ", "AM", "AO", "AP", "AQ", "AT", "AU", "AV",
]


def snapshot_rows(path, start_row):
    wb = openpyxl.load_workbook(path)
    sheet = wb["Don dat hang"]
    rows = []
    for r in range(start_row, sheet.max_row + 1):
        row = {"row_number_offset": r - start_row}
        for col in COLUMNS:
            cell = sheet[f"{col}{r}"]
            value = cell.value
            row[col] = value
            if col == "Z":
                row["Z_is_formula"] = isinstance(value, str) and value.startswith("=")
        comment = sheet[f"Y{r}"].comment
        row["Y_has_comment"] = comment is not None
        row["Y_fill"] = sheet[f"Y{r}"].fill.fgColor.rgb if sheet[f"Y{r}"].fill else None
        rows.append(row)
    wb.close()
    return rows


def is_lotte_pdf(path):
    doc = xulydonhang.fitz.open(path)
    try:
        text = ""
        for page in doc:
            text += page.get_text("text")
        return xulydonhang.ProcessHandler.identify_vendor(text) == "Lotte"
    finally:
        doc.close()


def process_one_pdf(path):
    """Mirrors the Lotte branch of process_file (xulydonhang.py:9079-9139)
    for every page identify_vendor recognizes as Lotte, skipping the
    Google Drive upload / current-page-extraction side effects (not
    needed to capture the Excel row output this harness cares about)."""
    handler = xulydonhang.process_handler
    doc = xulydonhang.fitz.open(path)
    try:
        for page_num in range(len(doc)):
            text = doc[page_num].get_text("text")
            if xulydonhang.ProcessHandler.identify_vendor(text) != "Lotte":
                continue

            lines = text.splitlines()
            po_number = lines[1] if len(lines) > 1 else ""
            if len(po_number) >= 7:
                po_number = po_number[:6] + "-" + po_number[6:]
            if len(po_number) >= 12:
                po_number = po_number[:12] + "-" + po_number[12:]

            time_part, store_code, order_number = po_number.split("-")
            entry_date = xulydonhang.datetime.strptime(time_part, "%y%m%d").strftime("%d/%m/%Y")

            cancel_date = xulydonhang.ProcessHandler.tachcancledate_lotte(text, po_number)
            tenstore = xulydonhang.ProcessHandler.laytenstore_lotte(text, po_number)
            diachigiaohang = "Lotte " + (tenstore or "")

            product_details = xulydonhang.ProcessHandler.tachsanpham_lotte(text)
            store_code_resolved = xulydonhang.ProcessHandler.get_makhachhang_lotte(store_code[1:])

            xulydonhang.ProcessHandler.write_to_dondathang_lotte(
                handler, product_details, store_code_resolved, po_number,
                entry_date, cancel_date, 1, "Lotte", diachigiaohang, None,
            )
    finally:
        doc.close()


def main():
    os.makedirs(FIXTURES_DIR, exist_ok=True)

    pdf_paths = sorted(glob.glob(os.path.join(REALPDFS_DIR, "*.pdf")))
    print(f"Found {len(pdf_paths)} candidate PDFs")

    generated = 0
    skipped = 0
    for path in pdf_paths:
        try:
            if not is_lotte_pdf(path):
                continue
        except Exception as e:
            print(f"SKIP (vendor check failed) {os.path.basename(path)}: {e}")
            skipped += 1
            continue

        shutil.copyfile(TEMPLATE_XLSX, SCRATCH_XLSX)
        os.chdir(REPO_ROOT)
        real_target = os.path.join(REPO_ROOT, "dondathang.xlsx")
        backup = real_target + ".fixture_backup"
        shutil.move(real_target, backup)
        shutil.copyfile(SCRATCH_XLSX, real_target)
        try:
            wb = openpyxl.load_workbook(real_target)
            start_row = wb["Don dat hang"].max_row + 1
            wb.close()

            process_one_pdf(path)

            rows = snapshot_rows(real_target, start_row)
        except Exception as e:
            print(f"SKIP (processing failed) {os.path.basename(path)}: {e}")
            skipped += 1
            rows = None
        finally:
            _restore_with_retry(backup, real_target)
            if os.path.exists(SCRATCH_XLSX):
                os.remove(SCRATCH_XLSX)

        if rows is None:
            continue

        fixture = {"source_pdf": os.path.basename(path), "rows": rows}
        fixture_name = os.path.splitext(os.path.basename(path))[0] + ".json"
        with open(os.path.join(FIXTURES_DIR, fixture_name), "w", encoding="utf-8") as f:
            json.dump(fixture, f, ensure_ascii=False, indent=2, default=str)
        generated += 1
        print(f"OK {os.path.basename(path)} -> {len(rows)} rows")

    if _promo_raw_rows is None:
        _capture_promo_raw_rows("LOTTE")
    frozen_pricing = {"raw_rows": _promo_raw_rows or []}
    with open(os.path.join(FIXTURES_DIR, "_frozen_pricing.json"), "w", encoding="utf-8") as f:
        json.dump(frozen_pricing, f, ensure_ascii=False, indent=2, default=str)

    print(f"\nDone: {generated} fixtures generated, {skipped} PDFs skipped.")


if __name__ == "__main__":
    main()
