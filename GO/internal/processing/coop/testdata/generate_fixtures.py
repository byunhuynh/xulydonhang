"""
Throwaway dev tool — NOT part of the shipped Go or Python app.

Runs the CURRENT xulydonhang.py Coop pipeline against every real PDF
committed into GO/internal/processing/coop/testdata/realpdfs/ that
identify_vendor recognizes as Coop, capturing the resulting
dondathang.xlsx rows (and the live-fetched Google Sheets price/promotion
data) into JSON fixtures under GO/internal/processing/coop/testdata/fixtures/.
The Go golden test (Task 13) diffs RealProcessor's output against these
fixtures instead of against a live Google Sheets fetch, so it's
deterministic and offline.

Originally this harness read source PDFs from the live đơn hàng/08-2026/
folder, but that tree is continuously reorganized by a concurrently-running
production instance of this application (files get moved into a dated
archive and renamed), which broke every fixture generated that way. Fixed
by pointing at the git-tracked, stable coop/testdata/realpdfs/ snapshot
instead (151 of the original 155 fixtures' source PDFs were recoverable —
see commit 213157e), matching the pattern already established by
kingfood/testdata/generate_fixtures.py's own REALPDFS_DIR constant.

Run from the repo root:
    .venv/Scripts/python.exe GO/internal/processing/coop/testdata/generate_fixtures.py
"""
import glob
import json
import os
import re
import shutil
import sys
import time

# NOTE: the brief's literal code used 5 nested os.path.dirname() calls here,
# which resolves only to the "GO" directory, not the actual repo root — this
# script sits 5 directory levels below repo root
# (GO/internal/processing/coop/testdata/generate_fixtures.py), so reaching
# repo root from os.path.abspath(__file__) requires 6 dirname() calls (one to
# strip the filename, five more to strip GO/internal/processing/coop/testdata).
# Verified empirically before running against production data. Fixed by
# adding one more dirname() wrap below.
REPO_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))))
sys.path.insert(0, REPO_ROOT)
os.chdir(REPO_ROOT)  # xulydonhang.py's functions use relative paths ("data.xlsx", "settings.ini")

# xulydonhang.py's Coop pipeline (process_coop_invoice and callees) has many
# debug print() calls containing emoji (e.g. "🔹 Khuyen mai goc: ..."). When
# this harness runs under the repo's Windows Python via Git Bash, sys.stdout
# defaults to the legacy cp1252 console codepage, which cannot encode those
# characters — every one of those print() calls raised UnicodeEncodeError,
# which aborted process_coop_invoice partway through and was caught by this
# harness's own try/except as "processing failed", so ALL 155 candidate PDFs
# were skipped on first run. Confirmed by reproducing the exact same
# UnicodeEncodeError diagnosing a plain print of the offending character.
# Fixing it here (not in xulydonhang.py, which stays untouched) by forcing
# UTF-8 stdout/stderr with a non-fatal error handler before importing it.
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="backslashreplace")

import openpyxl  # noqa: E402
import xulydonhang  # noqa: E402

REALPDFS_DIR = os.path.join(
    REPO_ROOT, "GO", "internal", "processing", "coop", "testdata", "realpdfs"
)
FIXTURES_DIR = os.path.join(
    REPO_ROOT, "GO", "internal", "processing", "coop", "testdata", "fixtures"
)
TEMPLATE_XLSX = os.path.join(REPO_ROOT, "dondathang.xlsx")
SCRATCH_XLSX = os.path.join(REPO_ROOT, "dondathang_fixture_scratch.xlsx")

# --- Monkey-patch network/upload side effects out ---

_price_cache = {}
_promo_cache = {}
_promo_raw_rows = None  # captured once, shared across all SKUs (same sheet)


def _cached_find_price_by_sku(sku_number, sheet_name="COOP"):
    key = (sku_number, sheet_name)
    if key not in _price_cache:
        _price_cache[key] = _real_find_price_by_sku(sku_number, sheet_name)
    return _price_cache[key]


def _cached_find_all_promotions(sku_code, time_to_check, sheet_name="Coop"):
    global _promo_raw_rows
    if _promo_raw_rows is None:
        _capture_promo_raw_rows(sheet_name)
    key = (sku_code, time_to_check, sheet_name)
    if key not in _promo_cache:
        _promo_cache[key] = _real_find_all_promotions(sku_code, time_to_check, sheet_name)
    return _promo_cache[key]


def _capture_promo_raw_rows(sheet_name):
    """Fetch the raw CSV once and stash it so we can freeze it verbatim,
    in addition to letting the real function do its own fetch+parse for
    the actual lookup (keeps this harness simple — one extra fetch is
    fine for a one-time fixture-generation run)."""
    global _promo_raw_rows
    import pandas as pd

    gid = xulydonhang.ProcessHandler.get_gid(sheet_name)
    if not gid:
        _promo_raw_rows = []
        return
    sheet_id = "1yvxE_SPYXKhofcZdhv1CSKAyiwdY1Mf4pFlsiMbtOr4"
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?tqx=out:csv&gid={gid}"
    df = pd.read_csv(url, dtype=str, header=None)
    df.fillna("", inplace=True)
    _promo_raw_rows = df.values.tolist()


def _noop_upload_file_to_drive(path, output_name=None, vendor=None, entry_date=None, makhachhang=None, cancle_date=None):
    return {"url": "https://example.invalid/skipped-during-fixture-generation"}


_real_find_price_by_sku = xulydonhang.ProcessHandler.find_price_by_sku
_real_find_all_promotions = xulydonhang.ProcessHandler.find_all_promotions_by_sku_and_time
xulydonhang.ProcessHandler.find_price_by_sku = staticmethod(_cached_find_price_by_sku)
xulydonhang.ProcessHandler.find_all_promotions_by_sku_and_time = staticmethod(_cached_find_all_promotions)
xulydonhang.ProcessHandler.upload_file_to_drive = staticmethod(_noop_upload_file_to_drive)


# --- Excel row capture ---

COLUMNS = [
    "A", "B", "C", "D", "E", "G", "L", "Q", "S", "T", "U", "V", "X", "Y", "Z",
    "AE", "AJ", "AM", "AO", "AP", "AQ", "AT", "AU", "AV",
]


def snapshot_rows(path, start_row):
    wb = openpyxl.load_workbook(path)
    sheet = wb["Don dat hang"]
    rows = []
    for r in range(start_row, sheet.max_row + 1):
        row = {"row_number_offset": r - start_row}
        for col in COLUMNS:
            cell = sheet[f"{col}{r}"]
            value = cell.value
            row[col] = value
            if col == "Z":
                row["Z_is_formula"] = isinstance(value, str) and value.startswith("=")
        comment = sheet[f"Y{r}"].comment
        row["Y_has_comment"] = comment is not None
        row["Y_fill"] = sheet[f"Y{r}"].fill.fgColor.rgb if sheet[f"Y{r}"].fill else None
        rows.append(row)
    wb.close()
    return rows


def is_coop_pdf(path):
    doc = xulydonhang.fitz.open(path)
    try:
        text = ""
        for page in doc:
            text += page.get_text("text")
            if len(text) > 3000:
                break
        return xulydonhang.ProcessHandler.identify_vendor(text) == "Coop"
    finally:
        doc.close()


def process_one_pdf(path):
    """Runs the Coop dispatch branch of process_file for a single-page
    (or multi-page) PDF, mirroring the relevant slice of process_file's
    logic without needing the full GUI-oriented method."""
    handler = xulydonhang.process_handler
    doc = xulydonhang.fitz.open(path)
    try:
        for page_num in range(len(doc)):
            text = doc[page_num].get_text("text")
            if xulydonhang.ProcessHandler.identify_vendor(text) != "Coop":
                continue

            counts = xulydonhang.ProcessHandler.demsodonhang1trang_coop(text)
            pom_count, sub_count = counts["POM343"], counts["Sub Total"]
            if pom_count == 0 or sub_count == 0 or pom_count != sub_count:
                continue

            if pom_count == 1:
                segments = [text]
            else:
                segments = xulydonhang.ProcessHandler.catdonra_nhieutrang(text)

            for i, segment in enumerate(segments):
                page_label = "1/1" if len(segments) == 1 else f"{i + 1}/{len(segments)}"
                handler.process_coop_invoice(segment, 1, path, page_label, doc)
    finally:
        doc.close()


def _restore_with_retry(backup, real_target, attempts=20, delay=1.0):
    """Restore the production dondathang.xlsx from its per-iteration backup.

    Uses a single os.replace() (atomic overwrite-in-place on Windows, no
    separate remove-then-move) instead of the previous remove()+move()
    pair, and a much larger retry budget than the original 5x0.5s=2.5s.
    Found genuinely necessary, not just defensive: during the first live
    151-PDF run, the 3rd PDF hit `PermissionError: [WinError 5] Access is
    denied` on os.remove(real_target) after exhausting 5 retries (2.5s) —
    almost certainly a transient exclusive lock from Windows Defender/
    OneDrive scanning the just-copied file, since a manual os.open(path,
    O_RDWR) succeeded moments later with no process holding it open. That
    exception escaped the old `finally:` block BEFORE the restore
    (_move_with_retry) ran, silently skipping the production-file restore
    entirely and only avoiding real damage because the file it failed to
    remove already happened to hold pristine (unwritten) content that
    iteration. This is a real safety gap, not a hypothetical one: if
    real_target had already contained newly-written rows at the moment of
    the lock, those rows would have been left in place instead of being
    reverted to the backup. Fixed by (a) collapsing to one atomic
    operation so there's no window between "old content gone" and "backup
    restored", and (b) raising the retry budget to something that
    comfortably outlasts a multi-second AV/sync scan. If this still fails
    after all attempts, we deliberately let the exception propagate and
    crash the run — per this project's file-safety protocol, a still-
    unrestored production file is a STOP-and-report-BLOCKED condition, not
    something to paper over. The backup file at `backup` is never deleted
    until this call succeeds, so nothing is lost even in that case.
    """
    last_err = None
    for i in range(attempts):
        try:
            os.replace(backup, real_target)
            return
        except OSError as e:
            last_err = e
            if i == attempts - 1:
                raise
            time.sleep(delay)


def main():
    os.makedirs(FIXTURES_DIR, exist_ok=True)

    pdf_paths = sorted(set(
        glob.glob(os.path.join(REALPDFS_DIR, "*.pdf")) +
        glob.glob(os.path.join(REALPDFS_DIR, "*.PDF"))
    ))
    print(f"Found {len(pdf_paths)} candidate PDFs in {REALPDFS_DIR}")

    generated = 0
    skipped = 0
    for path in pdf_paths:
        try:
            if not is_coop_pdf(path):
                continue
        except Exception as e:
            print(f"SKIP (vendor check failed) {os.path.basename(path)}: {e}")
            skipped += 1
            continue

        shutil.copyfile(TEMPLATE_XLSX, SCRATCH_XLSX)
        os.chdir(REPO_ROOT)
        real_target = os.path.join(REPO_ROOT, "dondathang.xlsx")
        backup = real_target + ".fixture_backup"
        shutil.move(real_target, backup)
        shutil.copyfile(SCRATCH_XLSX, real_target)
        try:
            wb = openpyxl.load_workbook(real_target)
            start_row = wb["Don dat hang"].max_row + 1
            wb.close()

            process_one_pdf(path)

            rows = snapshot_rows(real_target, start_row)
        except Exception as e:
            print(f"SKIP (processing failed) {os.path.basename(path)}: {e}")
            skipped += 1
            rows = None
        finally:
            _restore_with_retry(backup, real_target)
            if os.path.exists(SCRATCH_XLSX):
                os.remove(SCRATCH_XLSX)

        if rows is None:
            continue

        fixture = {"source_pdf": os.path.basename(path), "rows": rows}
        fixture_name = os.path.splitext(os.path.basename(path))[0] + ".json"
        with open(os.path.join(FIXTURES_DIR, fixture_name), "w", encoding="utf-8") as f:
            json.dump(fixture, f, ensure_ascii=False, indent=2, default=str)
        generated += 1
        print(f"OK {os.path.basename(path)} -> {len(rows)} rows")

    # _promo_raw_rows is the ONE raw CSV snapshot Coop's Google Sheet
    # returns for gid=get_gid("COOP") — the same sheet find_price_by_sku
    # and find_all_promotions_by_sku_and_time both fetch (see Task 3's
    # design note). This single snapshot is what gets frozen; there is
    # no separate "prices" artifact, matching pricing.ParseIndex's
    # one-CSV-in, two-views-out design in the Go port.
    if _promo_raw_rows is None:
        _capture_promo_raw_rows("COOP")
    frozen_pricing = {"raw_rows": _promo_raw_rows or []}
    with open(os.path.join(FIXTURES_DIR, "_frozen_pricing.json"), "w", encoding="utf-8") as f:
        json.dump(frozen_pricing, f, ensure_ascii=False, indent=2, default=str)

    print(f"\nDone: {generated} fixtures generated, {skipped} PDFs skipped.")


if __name__ == "__main__":
    main()
