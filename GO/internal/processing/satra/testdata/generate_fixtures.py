"""
Throwaway dev tool — NOT part of the shipped Go or Python app.

Runs the CURRENT xulydonhang.py Satra pipeline against every real PDF
committed into GO/internal/processing/satra/testdata/realpdfs/ that
identify_vendor recognizes as Satra, capturing the resulting
dondathang.xlsx rows (and the live-fetched Google Sheets price/promotion
data for the SATRA sheet) into JSON fixtures under
GO/internal/processing/satra/testdata/fixtures/. The Go golden test
(Task 7) diffs RealProcessor's output against these fixtures instead of
against a live Google Sheets fetch, so it's deterministic and offline.

Originally this harness read source PDFs from the live đơn hàng/08-2026/
folder, but that tree is continuously reorganized by a concurrently-running
production instance of this application (files get moved into a dated
archive and renamed), which broke every fixture generated that way. Fixed
by pointing at the git-tracked, stable satra/testdata/realpdfs/ snapshot
instead (33 of the original 36 fixtures' source PDFs were recoverable —
see the migration commit preceding this one), matching the pattern already
established by Coop/Lotte's own REALPDFS_DIR constant.

Run from the repo root:
    .venv/Scripts/python.exe GO/internal/processing/satra/testdata/generate_fixtures.py
"""
import glob
import json
import os
import shutil
import sys
import time

# Same depth as Lotte's harness: this script sits 5 directory levels below
# repo root (GO/internal/processing/satra/testdata/generate_fixtures.py),
# so reaching repo root from os.path.abspath(__file__) requires 6
# dirname() calls (one to strip the filename, five more to strip
# GO/internal/processing/satra/testdata).
REPO_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))))
sys.path.insert(0, REPO_ROOT)
os.chdir(REPO_ROOT)  # xulydonhang.py's functions use relative paths ("data.xlsx", "settings.ini")

# See Coop's harness for why this is needed: process_file's debug print()
# calls contain emoji that the legacy cp1252 console codepage can't
# encode, aborting processing partway through if not fixed here first.
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="backslashreplace")

import openpyxl  # noqa: E402
import xulydonhang  # noqa: E402

REALPDFS_DIR = os.path.join(
    REPO_ROOT, "GO", "internal", "processing", "satra", "testdata", "realpdfs"
)
FIXTURES_DIR = os.path.join(
    REPO_ROOT, "GO", "internal", "processing", "satra", "testdata", "fixtures"
)
TEMPLATE_XLSX = os.path.join(REPO_ROOT, "dondathang.xlsx")
SCRATCH_XLSX = os.path.join(REPO_ROOT, "dondathang_fixture_scratch.xlsx")

# --- Monkey-patch network/upload side effects out (identical shape to
# Coop's harness; find_price_by_sku/find_all_promotions_by_sku_and_time
# are already generic over sheet_name, so this works for "SATRA" too
# with no changes to the caching logic itself) ---

_price_cache = {}
_promo_cache = {}
_promo_raw_rows = None


def _cached_find_price_by_sku(sku_number, sheet_name="COOP"):
    key = (sku_number, sheet_name)
    if key not in _price_cache:
        _price_cache[key] = _real_find_price_by_sku(sku_number, sheet_name)
    return _price_cache[key]


def _cached_find_all_promotions(sku_code, time_to_check, sheet_name="Coop"):
    global _promo_raw_rows
    if _promo_raw_rows is None:
        _capture_promo_raw_rows(sheet_name)
    key = (sku_code, time_to_check, sheet_name)
    if key not in _promo_cache:
        _promo_cache[key] = _real_find_all_promotions(sku_code, time_to_check, sheet_name)
    return _promo_cache[key]


def _capture_promo_raw_rows(sheet_name):
    global _promo_raw_rows
    import pandas as pd

    gid = xulydonhang.ProcessHandler.get_gid(sheet_name)
    if not gid:
        _promo_raw_rows = []
        return
    sheet_id = "1yvxE_SPYXKhofcZdhv1CSKAyiwdY1Mf4pFlsiMbtOr4"
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?tqx=out:csv&gid={gid}"
    df = pd.read_csv(url, dtype=str, header=None)
    df.fillna("", inplace=True)
    _promo_raw_rows = df.values.tolist()


def _noop_upload_file_to_drive(path, output_name=None, vendor=None, entry_date=None, makhachhang=None, cancle_date=None):
    return {"url": "https://example.invalid/skipped-during-fixture-generation"}


_real_find_price_by_sku = xulydonhang.ProcessHandler.find_price_by_sku
_real_find_all_promotions = xulydonhang.ProcessHandler.find_all_promotions_by_sku_and_time
xulydonhang.ProcessHandler.find_price_by_sku = staticmethod(_cached_find_price_by_sku)
xulydonhang.ProcessHandler.find_all_promotions_by_sku_and_time = staticmethod(_cached_find_all_promotions)
xulydonhang.ProcessHandler.upload_file_to_drive = staticmethod(_noop_upload_file_to_drive)


# --- Excel row capture (same columns as Coop's harness — same sheet, same layout) ---

COLUMNS = [
    "A", "B", "C", "D", "E", "G", "L", "Q", "S", "T", "U", "V", "X", "Y", "Z",
    "AE", "AJ", "AM", "AO", "AP", "AQ", "AT", "AU", "AV",
]


def snapshot_rows(path, start_row):
    wb = openpyxl.load_workbook(path)
    sheet = wb["Don dat hang"]
    rows = []
    for r in range(start_row, sheet.max_row + 1):
        row = {"row_number_offset": r - start_row}
        for col in COLUMNS:
            cell = sheet[f"{col}{r}"]
            value = cell.value
            row[col] = value
            if col == "Z":
                row["Z_is_formula"] = isinstance(value, str) and value.startswith("=")
        comment = sheet[f"Y{r}"].comment
        row["Y_has_comment"] = comment is not None
        row["Y_fill"] = sheet[f"Y{r}"].fill.fgColor.rgb if sheet[f"Y{r}"].fill else None
        rows.append(row)
    wb.close()
    return rows


def is_satra_pdf(path):
    doc = xulydonhang.fitz.open(path)
    try:
        text = ""
        for page in doc:
            text += page.get_text("text")
        return xulydonhang.ProcessHandler.identify_vendor(text) == "Satra"
    finally:
        doc.close()


def process_one_pdf(path):
    """Mirrors the Satra branch of process_file (xulydonhang.py:9303-9394)
    for every page identify_vendor recognizes as Satra, skipping the
    Google Drive upload / current-page-extraction side effects."""
    handler = xulydonhang.process_handler
    doc = xulydonhang.fitz.open(path)
    try:
        for page_num in range(len(doc)):
            text = doc[page_num].get_text("text")
            if xulydonhang.ProcessHandler.identify_vendor(text) != "Satra":
                continue

            po_number = xulydonhang.re.search(r"\*P-[^*]+\*", text)
            po_number = po_number.group(0)[1:-1]

            makhachhang = None
            match = xulydonhang.re.search(r"Địa chỉ giao hàng:\s*((?:.*\n)+?)Địa chỉ thanh toán:", text)
            diachi = ""
            if match:
                diachi = match.group(1).strip().replace("\n", " ").replace("  ", " ")
                makhachhang = xulydonhang.ProcessHandler.laymakhachhang_satra(diachi, "SATRA")

            entry_date = xulydonhang.re.search(r"(.*?)\nNgày đặt hàng:", text, xulydonhang.re.DOTALL)
            if entry_date:
                entry_date = entry_date.group(1).split("\n")[-1]
                entry_date = xulydonhang.datetime.strptime(entry_date, "%m/%d/%Y")
                entry_date = entry_date.strftime("%d/%m/%Y")
                if entry_date == "01/01/0001":
                    entry_date = xulydonhang.re.search(r"(.*?)\nNgày in:", text, xulydonhang.re.DOTALL)
                    if entry_date:
                        entry_date = entry_date.group(1).split("\n")[-1]
                        entry_date = xulydonhang.datetime.strptime(entry_date, "%m/%d/%Y")
                        entry_date = entry_date.strftime("%d/%m/%Y")

            cancel_date = xulydonhang.re.search(r"Ngày giao hàng:\s*(.*?)\s*Địa chỉ giao hàng:", text, xulydonhang.re.DOTALL)
            if cancel_date:
                cancel_date = cancel_date.group(1).strip()
                pattern = r"(\d{1,2}/\d{1,2}/\d{4})"
                for line in cancel_date.split("\n"):
                    if xulydonhang.re.search(pattern, line):
                        cancel_date = line
                        cancel_date = xulydonhang.datetime.strptime(cancel_date, "%m/%d/%Y")
                        cancel_date = cancel_date.strftime("%d/%m/%Y")
                        break

            product_text = xulydonhang.re.search(r"STT\s*(.*?)\s*Hàng phục vụ cho:", text, xulydonhang.re.DOTALL)
            product_text = product_text.group(1).strip()
            products = xulydonhang.ProcessHandler.trichxuatsanpham_satra(product_text)
            if products:
                sku_mapping = xulydonhang.ProcessHandler.load_sku_mapping()
                products = xulydonhang.ProcessHandler.replace_sku_numbers(products, sku_mapping)

            xulydonhang.ProcessHandler.write_to_dondathang_satra(
                handler, products, makhachhang, po_number, entry_date, cancel_date,
                1, "Satra", diachi, None,
            )
    finally:
        doc.close()


def _restore_with_retry(backup, real_target, attempts=20, delay=1.0):
    """Restore the production dondathang.xlsx from its per-iteration backup.

    Uses a single os.replace() (atomic overwrite-in-place on Windows, no
    separate remove-then-move) instead of a remove()-then-move() pair, and
    a large retry budget. This is the exact fix discovered during Coop's
    fixture regeneration (see coop/testdata/generate_fixtures.py and the
    Coop/Lotte migration reports) — applied here directly rather than
    rediscovered: the original remove()+move() pair could have os.remove()
    hit a transient Windows file lock (e.g. Defender/OneDrive scanning the
    just-copied file), exhaust its retry budget, and raise BEFORE the
    restore ever ran, silently abandoning the production-file restore
    mid-run. Collapsing to one atomic os.replace() closes that window; the
    backup file is never deleted until this call succeeds, so nothing is
    lost even if all attempts fail (in which case the exception is left to
    propagate and stop the run, per this project's file-safety protocol).
    """
    last_err = None
    for i in range(attempts):
        try:
            os.replace(backup, real_target)
            return
        except OSError as e:
            last_err = e
            if i == attempts - 1:
                raise
            time.sleep(delay)


def main():
    os.makedirs(FIXTURES_DIR, exist_ok=True)

    pdf_paths = sorted(set(
        glob.glob(os.path.join(REALPDFS_DIR, "*.pdf")) +
        glob.glob(os.path.join(REALPDFS_DIR, "*.PDF"))
    ))
    print(f"Found {len(pdf_paths)} candidate PDFs in {REALPDFS_DIR}")

    generated = 0
    skipped = 0
    for path in pdf_paths:
        try:
            if not is_satra_pdf(path):
                continue
        except Exception as e:
            print(f"SKIP (vendor check failed) {os.path.basename(path)}: {e}")
            skipped += 1
            continue

        shutil.copyfile(TEMPLATE_XLSX, SCRATCH_XLSX)
        os.chdir(REPO_ROOT)
        real_target = os.path.join(REPO_ROOT, "dondathang.xlsx")
        backup = real_target + ".fixture_backup"
        shutil.move(real_target, backup)
        shutil.copyfile(SCRATCH_XLSX, real_target)
        try:
            wb = openpyxl.load_workbook(real_target)
            start_row = wb["Don dat hang"].max_row + 1
            wb.close()

            process_one_pdf(path)

            rows = snapshot_rows(real_target, start_row)
        except Exception as e:
            print(f"SKIP (processing failed) {os.path.basename(path)}: {e}")
            skipped += 1
            rows = None
        finally:
            _restore_with_retry(backup, real_target)
            if os.path.exists(SCRATCH_XLSX):
                os.remove(SCRATCH_XLSX)

        if rows is None:
            continue

        fixture = {"source_pdf": os.path.basename(path), "rows": rows}
        fixture_name = os.path.splitext(os.path.basename(path))[0] + ".json"
        with open(os.path.join(FIXTURES_DIR, fixture_name), "w", encoding="utf-8") as f:
            json.dump(fixture, f, ensure_ascii=False, indent=2, default=str)
        generated += 1
        print(f"OK {os.path.basename(path)} -> {len(rows)} rows")

    if _promo_raw_rows is None:
        _capture_promo_raw_rows("SATRA")
    frozen_pricing = {"raw_rows": _promo_raw_rows or []}
    with open(os.path.join(FIXTURES_DIR, "_frozen_pricing.json"), "w", encoding="utf-8") as f:
        json.dump(frozen_pricing, f, ensure_ascii=False, indent=2, default=str)

    print(f"\nDone: {generated} fixtures generated, {skipped} PDFs skipped.")


if __name__ == "__main__":
    main()
