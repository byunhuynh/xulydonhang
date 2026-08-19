"""
Throwaway dev tool — NOT part of the shipped Go or Python app.

Runs the CURRENT xulydonhang.py Kingfood pipeline against the 9 real
Kingfood PDFs copied into
GO/internal/processing/kingfood/testdata/realpdfs/ (see this task's own
Step 1 — committed directly into this repo's testdata from the start,
per explicit project-owner confirmation for this vendor), capturing the
resulting dondathang.xlsx rows (and the live-fetched Google Sheets
price/promotion data for the KINGFOOD sheet) into JSON fixtures under
GO/internal/processing/kingfood/testdata/fixtures/. The Go golden test
(Task 6) diffs RealProcessor's output against these fixtures instead of
against a live Google Sheets fetch, so it's deterministic and offline.

Like Satra/Lotte/Winmart/Emart/FujiMart (one page == one order,
write_to_dondathang_kingfood appends immediately, no explicit start-row
argument needed), and UNLIKE BigC, this harness computes start_row once
up front and takes a single snapshot after process_one_pdf's per-page
loop has finished.

Run from the repo root:
    .venv/Scripts/python.exe GO/internal/processing/kingfood/testdata/generate_fixtures.py
"""
import glob
import json
import os
import shutil
import sys
import time

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))))
sys.path.insert(0, REPO_ROOT)
os.chdir(REPO_ROOT)  # xulydonhang.py's functions use relative paths ("data.xlsx", "settings.ini")

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="backslashreplace")

import openpyxl  # noqa: E402
import xulydonhang  # noqa: E402

REALPDFS_DIR = os.path.join(
    REPO_ROOT, "GO", "internal", "processing", "kingfood", "testdata", "realpdfs"
)
FIXTURES_DIR = os.path.join(
    REPO_ROOT, "GO", "internal", "processing", "kingfood", "testdata", "fixtures"
)
TEMPLATE_XLSX = os.path.join(REPO_ROOT, "dondathang.xlsx")
SCRATCH_XLSX = os.path.join(REPO_ROOT, "dondathang_fixture_scratch.xlsx")

_price_cache = {}
_promo_cache = {}
_promo_raw_rows = None


def _cached_find_price_by_sku(sku_number, sheet_name="COOP"):
    key = (sku_number, sheet_name)
    if key not in _price_cache:
        _price_cache[key] = _real_find_price_by_sku(sku_number, sheet_name)
    return _price_cache[key]


def _cached_find_all_promotions(sku_code, time_to_check, sheet_name="Coop"):
    global _promo_raw_rows
    if _promo_raw_rows is None:
        _capture_promo_raw_rows(sheet_name)
    key = (sku_code, time_to_check, sheet_name)
    if key not in _promo_cache:
        _promo_cache[key] = _real_find_all_promotions(sku_code, time_to_check, sheet_name)
    return _promo_cache[key]


def _capture_promo_raw_rows(sheet_name):
    global _promo_raw_rows
    import pandas as pd

    gid = xulydonhang.ProcessHandler.get_gid(sheet_name)
    if not gid:
        _promo_raw_rows = []
        return
    sheet_id = "1yvxE_SPYXKhofcZdhv1CSKAyiwdY1Mf4pFlsiMbtOr4"
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?tqx=out:csv&gid={gid}"
    df = pd.read_csv(url, dtype=str, header=None)
    df.fillna("", inplace=True)
    _promo_raw_rows = df.values.tolist()


def _noop_upload_file_to_drive(path, output_name=None, vendor=None, entry_date=None, makhachhang=None, cancle_date=None):
    return {"url": "https://example.invalid/skipped-during-fixture-generation"}


_real_find_price_by_sku = xulydonhang.ProcessHandler.find_price_by_sku
_real_find_all_promotions = xulydonhang.ProcessHandler.find_all_promotions_by_sku_and_time
xulydonhang.ProcessHandler.find_price_by_sku = staticmethod(_cached_find_price_by_sku)
xulydonhang.ProcessHandler.find_all_promotions_by_sku_and_time = staticmethod(_cached_find_all_promotions)
xulydonhang.ProcessHandler.upload_file_to_drive = staticmethod(_noop_upload_file_to_drive)


COLUMNS = [
    "A", "B", "C", "D", "E", "G", "L", "Q", "S", "T", "U", "V", "X", "Y", "Z",
    "AE", "AJ", "AM", "AO", "AP", "AQ", "AT", "AU", "AV",
]


def snapshot_rows(path, start_row):
    wb = openpyxl.load_workbook(path)
    sheet = wb["Don dat hang"]
    rows = []
    for r in range(start_row, sheet.max_row + 1):
        row = {"row_number_offset": r - start_row}
        for col in COLUMNS:
            cell = sheet[f"{col}{r}"]
            value = cell.value
            row[col] = value
            if col == "Z":
                row["Z_is_formula"] = isinstance(value, str) and value.startswith("=")
        comment = sheet[f"Y{r}"].comment
        row["Y_has_comment"] = comment is not None
        row["Y_fill"] = sheet[f"Y{r}"].fill.fgColor.rgb if sheet[f"Y{r}"].fill else None
        rows.append(row)
    wb.close()
    return rows


def process_one_pdf(path):
    """Mirrors the Kingfood branch of process_file (xulydonhang.py:9230-
    9310) for every page identify_vendor recognizes as Kingfood, skipping
    the Google Drive upload side effect (monkeypatched to a no-op above).
    No OCR needed for Kingfood (unlike FujiMart's harness)."""
    handler = xulydonhang.process_handler
    doc = xulydonhang.fitz.open(path)
    try:
        for page_num in range(len(doc)):
            text = doc[page_num].get_text("text")
            if xulydonhang.ProcessHandler.identify_vendor(text) != "Kingfood":
                continue

            import re
            from datetime import datetime

            tranggoc = doc[0].get_text("text")

            po_number = re.search(r"PO Number:\s*\n([^\n]*\n)?([^\n]*)", tranggoc)
            po_number = po_number.group(1).strip() if po_number else "Không tìm thấy PO Number"

            entry_date = re.search(r"Ngày Đặt Hàng:\s*\n([^\n]*\n)?([^\n]*)", tranggoc)
            entry_date = entry_date.group(1).replace("-", "/").strip() if entry_date else "Không tìm thấy ngày đặt hàng"
            entry_date = datetime.strptime(entry_date, "%d/%m/%Y")
            entry_date = entry_date.strftime("%d/%m/%Y")

            cancel_date = re.search(
                r"Ngày\s*Giao\s*Hàng\s*NCC\s*Xác\s*Nhận:\s*\n*([^\n]*\n)?([^\n]*)",
                tranggoc,
                re.IGNORECASE,
            )
            cancel_date = cancel_date.group(1).replace("-", "/").strip() if cancel_date else "Không tìm thấy ngày giao hàng"
            cancel_date = datetime.strptime(cancel_date, "%d/%m/%Y")
            cancel_date = cancel_date.strftime("%d/%m/%Y")

            products = xulydonhang.ProcessHandler.laydanhsachsanpham_kingfood(text)
            if not products:
                continue
            sku_mapping = xulydonhang.ProcessHandler.load_sku_mapping()
            products = xulydonhang.ProcessHandler.replace_sku_numbers(products, sku_mapping)

            store_code = "MN_MT_KFMSL"
            delivery = "Số 324, đường ĐT743A, Phường Đông Hoà, Thành phố Hồ Chí Minh"

            xulydonhang.ProcessHandler.write_to_dondathang_kingfood(
                handler, products, store_code, po_number, entry_date, cancel_date,
                1, "Kingfood", delivery, None,
            )
    finally:
        doc.close()


def _remove_with_retry(path, attempts=5, delay=0.5):
    for i in range(attempts):
        try:
            os.remove(path)
            return
        except PermissionError:
            if i == attempts - 1:
                raise
            time.sleep(delay)


def _move_with_retry(src, dst, attempts=5, delay=0.5):
    for i in range(attempts):
        try:
            shutil.move(src, dst)
            return
        except PermissionError:
            if i == attempts - 1:
                raise
            time.sleep(delay)


def main():
    os.makedirs(FIXTURES_DIR, exist_ok=True)

    pdf_paths = sorted(set(
        glob.glob(os.path.join(REALPDFS_DIR, "*.pdf")) +
        glob.glob(os.path.join(REALPDFS_DIR, "*.PDF"))
    ))
    print(f"Found {len(pdf_paths)} candidate PDFs in {REALPDFS_DIR}")

    generated = 0
    skipped = 0
    for path in pdf_paths:
        shutil.copyfile(TEMPLATE_XLSX, SCRATCH_XLSX)
        os.chdir(REPO_ROOT)
        real_target = os.path.join(REPO_ROOT, "dondathang.xlsx")
        backup = real_target + ".fixture_backup"
        shutil.move(real_target, backup)
        shutil.copyfile(SCRATCH_XLSX, real_target)
        try:
            wb = openpyxl.load_workbook(real_target)
            start_row = wb["Don dat hang"].max_row + 1
            wb.close()

            process_one_pdf(path)

            rows = snapshot_rows(real_target, start_row)
        except Exception as e:
            print(f"SKIP (processing failed) {os.path.basename(path)}: {e}")
            skipped += 1
            rows = None
        finally:
            _remove_with_retry(real_target)
            _move_with_retry(backup, real_target)
            if os.path.exists(SCRATCH_XLSX):
                os.remove(SCRATCH_XLSX)

        if rows is None or len(rows) == 0:
            if rows is not None:
                print(f"SKIP (no rows written, likely no products extracted) {os.path.basename(path)}")
                skipped += 1
            continue

        fixture = {"source_pdf": os.path.basename(path), "rows": rows}
        fixture_name = os.path.splitext(os.path.basename(path))[0] + ".json"
        with open(os.path.join(FIXTURES_DIR, fixture_name), "w", encoding="utf-8") as f:
            json.dump(fixture, f, ensure_ascii=False, indent=2, default=str)
        generated += 1
        print(f"OK {os.path.basename(path)} -> {len(rows)} rows")

    if _promo_raw_rows is None:
        _capture_promo_raw_rows("KINGFOOD")
    frozen_pricing = {"raw_rows": _promo_raw_rows or []}
    with open(os.path.join(FIXTURES_DIR, "_frozen_pricing.json"), "w", encoding="utf-8") as f:
        json.dump(frozen_pricing, f, ensure_ascii=False, indent=2, default=str)

    print(f"\nDone: {generated} fixtures generated, {skipped} PDFs skipped.")


if __name__ == "__main__":
    main()
