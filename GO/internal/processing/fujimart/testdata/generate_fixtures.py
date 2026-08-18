"""
Throwaway dev tool — NOT part of the shipped Go or Python app.

Runs the CURRENT xulydonhang.py FujiMart pipeline against the 15 real
FujiMart PDFs copied into
GO/internal/processing/fujimart/testdata/realpdfs/ (see this task's own
Step 1 — committed directly into this repo's testdata from the start,
per explicit project-owner confirmation for this vendor, unlike the
first 5 ported vendors' harnesses which still read from the live
đơn hàng/ tree at test-run time), capturing the resulting dondathang.xlsx
rows (and the live-fetched Google Sheets price/promotion data for the
FUJIMART sheet) into JSON fixtures under
GO/internal/processing/fujimart/testdata/fixtures/. The Go golden test
(Task 6) diffs RealProcessor's output against these fixtures instead of
against a live Google Sheets fetch, so it's deterministic and offline.

Like Satra/Lotte/Winmart/Emart (one page == one order,
write_to_dondathang_fujimart appends immediately, no explicit start-row
argument needed), and UNLIKE BigC, this harness computes start_row once
up front and takes a single snapshot after process_one_pdf's per-page
loop has finished.

Run from the repo root:
    .venv/Scripts/python.exe GO/internal/processing/fujimart/testdata/generate_fixtures.py
"""
import glob
import json
import os
import shutil
import sys
import time

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))))
sys.path.insert(0, REPO_ROOT)
os.chdir(REPO_ROOT)  # xulydonhang.py's functions use relative paths ("data.xlsx", "settings.ini")

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="backslashreplace")

import openpyxl  # noqa: E402
import xulydonhang  # noqa: E402

REALPDFS_DIR = os.path.join(
    REPO_ROOT, "GO", "internal", "processing", "fujimart", "testdata", "realpdfs"
)
FIXTURES_DIR = os.path.join(
    REPO_ROOT, "GO", "internal", "processing", "fujimart", "testdata", "fixtures"
)
TEMPLATE_XLSX = os.path.join(REPO_ROOT, "dondathang.xlsx")
SCRATCH_XLSX = os.path.join(REPO_ROOT, "dondathang_fixture_scratch.xlsx")

_price_cache = {}
_promo_cache = {}
_promo_raw_rows = None


def _cached_find_price_by_sku(sku_number, sheet_name="COOP"):
    key = (sku_number, sheet_name)
    if key not in _price_cache:
        _price_cache[key] = _real_find_price_by_sku(sku_number, sheet_name)
    return _price_cache[key]


def _cached_find_all_promotions(sku_code, time_to_check, sheet_name="Coop"):
    global _promo_raw_rows
    if _promo_raw_rows is None:
        _capture_promo_raw_rows(sheet_name)
    key = (sku_code, time_to_check, sheet_name)
    if key not in _promo_cache:
        _promo_cache[key] = _real_find_all_promotions(sku_code, time_to_check, sheet_name)
    return _promo_cache[key]


def _capture_promo_raw_rows(sheet_name):
    global _promo_raw_rows
    import pandas as pd

    gid = xulydonhang.ProcessHandler.get_gid(sheet_name)
    if not gid:
        _promo_raw_rows = []
        return
    sheet_id = "1yvxE_SPYXKhofcZdhv1CSKAyiwdY1Mf4pFlsiMbtOr4"
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?tqx=out:csv&gid={gid}"
    df = pd.read_csv(url, dtype=str, header=None)
    df.fillna("", inplace=True)
    _promo_raw_rows = df.values.tolist()


def _noop_upload_file_to_drive(path, output_name=None, vendor=None, entry_date=None, makhachhang=None, cancle_date=None):
    return {"url": "https://example.invalid/skipped-during-fixture-generation"}


_real_find_price_by_sku = xulydonhang.ProcessHandler.find_price_by_sku
_real_find_all_promotions = xulydonhang.ProcessHandler.find_all_promotions_by_sku_and_time
xulydonhang.ProcessHandler.find_price_by_sku = staticmethod(_cached_find_price_by_sku)
xulydonhang.ProcessHandler.find_all_promotions_by_sku_and_time = staticmethod(_cached_find_all_promotions)
xulydonhang.ProcessHandler.upload_file_to_drive = staticmethod(_noop_upload_file_to_drive)


COLUMNS = [
    "A", "B", "C", "D", "E", "G", "L", "Q", "S", "T", "U", "V", "X", "Y", "Z",
    "AE", "AJ", "AM", "AO", "AP", "AQ", "AT", "AU", "AV",
]


def snapshot_rows(path, start_row):
    wb = openpyxl.load_workbook(path)
    sheet = wb["Don dat hang"]
    rows = []
    for r in range(start_row, sheet.max_row + 1):
        row = {"row_number_offset": r - start_row}
        for col in COLUMNS:
            cell = sheet[f"{col}{r}"]
            value = cell.value
            row[col] = value
            if col == "Z":
                row["Z_is_formula"] = isinstance(value, str) and value.startswith("=")
        comment = sheet[f"Y{r}"].comment
        row["Y_has_comment"] = comment is not None
        row["Y_fill"] = sheet[f"Y{r}"].fill.fgColor.rgb if sheet[f"Y{r}"].fill else None
        rows.append(row)
    wb.close()
    return rows


def process_one_pdf(path):
    """Mirrors the FujiMart branch of process_file (xulydonhang.py:8831-
    8964) for every page identify_vendor recognizes as FujiMart, skipping
    the Google Drive upload / current-page-extraction side effects and
    the OCR step (this harness runs the REAL Python function, which still
    does OCR internally — that's fine, this harness's job is only to
    capture ground-truth OUTPUT, not to validate the Go port's own
    no-OCR design, which Task 2/6 validate separately)."""
    handler = xulydonhang.process_handler
    doc = xulydonhang.fitz.open(path)
    try:
        for page_num in range(len(doc)):
            text = doc[page_num].get_text("text")
            if xulydonhang.ProcessHandler.identify_vendor(text) != "FujiMart":
                continue

            import io
            import re
            from datetime import datetime, timedelta
            from PIL import Image, ImageEnhance
            import pytesseract

            page = doc.load_page(page_num)
            pix = page.get_pixmap(dpi=500)
            img = Image.open(io.BytesIO(pix.tobytes("png")))
            img = img.convert("L")
            img = ImageEnhance.Contrast(img).enhance(2.5)
            info = pytesseract.image_to_string(img, lang="vie")

            lines = [line.strip() for line in text.strip().splitlines() if line.strip()]
            entry_date = None
            for i, line in enumerate(lines):
                if "Sè §¬n:" in line and i >= 3:
                    entry_date = lines[i - 3]

            cancel_date = next((line.split("Ngµy giao:")[1].strip() for line in text.splitlines() if "Ngµy giao:" in line), "")

            if not re.match(r"\d{2}/\d{2}/\d{4}$", cancel_date):
                cancel_date = "Không tìm thấy"
                if cancel_date == "Không tìm thấy" and entry_date and re.match(r"\d{2}/\d{2}/\d{4}$", entry_date):
                    entry_date_obj = datetime.strptime(entry_date, "%d/%m/%Y")
                    cancel_date = (entry_date_obj + timedelta(days=2)).strftime("%d/%m/%Y")

            if not entry_date or not re.match(r"\d{2}/\d{2}/\d{4}$", entry_date):
                entry_date = "Không tìm thấy"
                if entry_date == "Không tìm thấy" and re.match(r"\d{2}/\d{2}/\d{4}$", cancel_date):
                    cancel_date_obj = datetime.strptime(cancel_date, "%d/%m/%Y")
                    entry_date = (cancel_date_obj - timedelta(days=2)).strftime("%d/%m/%Y")

            m = re.search(rf"^{entry_date}\s*\n(.+)", text, re.MULTILINE)
            po_number = m.group(1) if m else None

            tenstore = ""
            match = re.search(r"N\s*ơ\s*i\s*[\s]*n\s*h\s*ậ\s*n\s*:\s*(.+?)(?=\n|$)", info, re.IGNORECASE)
            if match:
                tenstore = match.group(1)

            start = text.rfind("§Þa chØ:")
            end = text.find("VAT")
            product_block = text[start + len("§Þa chØ:"):end].strip()
            result = "\n".join(product_block.splitlines())

            products = xulydonhang.ProcessHandler.tach_san_pham_Fujimart(result)
            if not products:
                continue
            sku_mapping = xulydonhang.ProcessHandler.load_sku_mapping()
            products = xulydonhang.ProcessHandler.replace_sku_numbers(products, sku_mapping)

            xulydonhang.ProcessHandler.write_to_dondathang_fujimart(
                handler, products, "MB_MT_FUJI", po_number, entry_date, cancel_date,
                1, "FujiMart", tenstore, None,
            )
    finally:
        doc.close()


def _remove_with_retry(path, attempts=5, delay=0.5):
    for i in range(attempts):
        try:
            os.remove(path)
            return
        except PermissionError:
            if i == attempts - 1:
                raise
            time.sleep(delay)


def _move_with_retry(src, dst, attempts=5, delay=0.5):
    for i in range(attempts):
        try:
            shutil.move(src, dst)
            return
        except PermissionError:
            if i == attempts - 1:
                raise
            time.sleep(delay)


def main():
    os.makedirs(FIXTURES_DIR, exist_ok=True)

    pdf_paths = sorted(set(
        glob.glob(os.path.join(REALPDFS_DIR, "*.pdf")) +
        glob.glob(os.path.join(REALPDFS_DIR, "*.PDF"))
    ))
    print(f"Found {len(pdf_paths)} candidate PDFs in {REALPDFS_DIR}")

    generated = 0
    skipped = 0
    for path in pdf_paths:
        shutil.copyfile(TEMPLATE_XLSX, SCRATCH_XLSX)
        os.chdir(REPO_ROOT)
        real_target = os.path.join(REPO_ROOT, "dondathang.xlsx")
        backup = real_target + ".fixture_backup"
        shutil.move(real_target, backup)
        shutil.copyfile(SCRATCH_XLSX, real_target)
        try:
            wb = openpyxl.load_workbook(real_target)
            start_row = wb["Don dat hang"].max_row + 1
            wb.close()

            process_one_pdf(path)

            rows = snapshot_rows(real_target, start_row)
        except Exception as e:
            print(f"SKIP (processing failed) {os.path.basename(path)}: {e}")
            skipped += 1
            rows = None
        finally:
            _remove_with_retry(real_target)
            _move_with_retry(backup, real_target)
            if os.path.exists(SCRATCH_XLSX):
                os.remove(SCRATCH_XLSX)

        if rows is None or len(rows) == 0:
            if rows is not None:
                print(f"SKIP (no rows written, likely no products extracted) {os.path.basename(path)}")
                skipped += 1
            continue

        fixture = {"source_pdf": os.path.basename(path), "rows": rows}
        fixture_name = os.path.splitext(os.path.basename(path))[0] + ".json"
        with open(os.path.join(FIXTURES_DIR, fixture_name), "w", encoding="utf-8") as f:
            json.dump(fixture, f, ensure_ascii=False, indent=2, default=str)
        generated += 1
        print(f"OK {os.path.basename(path)} -> {len(rows)} rows")

    if _promo_raw_rows is None:
        _capture_promo_raw_rows("FUJIMART")
    frozen_pricing = {"raw_rows": _promo_raw_rows or []}
    with open(os.path.join(FIXTURES_DIR, "_frozen_pricing.json"), "w", encoding="utf-8") as f:
        json.dump(frozen_pricing, f, ensure_ascii=False, indent=2, default=str)

    print(f"\nDone: {generated} fixtures generated, {skipped} PDFs skipped.")


if __name__ == "__main__":
    main()
