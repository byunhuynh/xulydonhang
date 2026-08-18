"""
Throwaway dev tool — NOT part of the shipped Go or Python app.

Runs the CURRENT xulydonhang.py Emart pipeline against the 9 real Emart
PDFs copied into GO/internal/processing/emart/testdata/realpdfs/ (see
this task's own Step 0 for why: the live đơn hàng/08-2026/ folder this
project's other vendor harnesses read from was reorganized mid-session by
a live, concurrently-running production instance of this same
application — real, external, out-of-repo data outside this project's
control. 8 of the original 17 real Emart PDFs are not yet available;
this generates fixtures for the 9 that are, and the 8 remaining can be
added later via the exact same process once located.), capturing the
resulting dondathang.xlsx rows (and the live-fetched Google Sheets
price/promotion data for the EMART sheet) into JSON fixtures under
GO/internal/processing/emart/testdata/fixtures/. The Go golden test
(Task 6) diffs RealProcessor's output against these fixtures instead of
against a live Google Sheets fetch, so it's deterministic and offline.

Like Satra/Lotte/Winmart (one page == one order, write_to_dondathang_emart
appends immediately, no explicit start-row argument needed), and UNLIKE
BigC, this harness computes start_row once up front and takes a single
snapshot after process_one_pdf's per-page loop has finished.

Run from the repo root:
    .venv/Scripts/python.exe GO/internal/processing/emart/testdata/generate_fixtures.py
"""
import glob
import json
import os
import re
import shutil
import sys
import time

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))))
sys.path.insert(0, REPO_ROOT)
os.chdir(REPO_ROOT)  # xulydonhang.py's functions use relative paths ("data.xlsx", "settings.ini")

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="backslashreplace")

import openpyxl  # noqa: E402
import xulydonhang  # noqa: E402

REALPDFS_DIR = os.path.join(
    REPO_ROOT, "GO", "internal", "processing", "emart", "testdata", "realpdfs"
)
FIXTURES_DIR = os.path.join(
    REPO_ROOT, "GO", "internal", "processing", "emart", "testdata", "fixtures"
)
TEMPLATE_XLSX = os.path.join(REPO_ROOT, "dondathang.xlsx")
SCRATCH_XLSX = os.path.join(REPO_ROOT, "dondathang_fixture_scratch.xlsx")

_price_cache = {}
_promo_cache = {}
_promo_raw_rows = None


def _cached_find_price_by_sku(sku_number, sheet_name="COOP"):
    key = (sku_number, sheet_name)
    if key not in _price_cache:
        _price_cache[key] = _real_find_price_by_sku(sku_number, sheet_name)
    return _price_cache[key]


def _cached_find_all_promotions(sku_code, time_to_check, sheet_name="Coop"):
    global _promo_raw_rows
    if _promo_raw_rows is None:
        _capture_promo_raw_rows(sheet_name)
    key = (sku_code, time_to_check, sheet_name)
    if key not in _promo_cache:
        _promo_cache[key] = _real_find_all_promotions(sku_code, time_to_check, sheet_name)
    return _promo_cache[key]


def _capture_promo_raw_rows(sheet_name):
    global _promo_raw_rows
    import pandas as pd

    gid = xulydonhang.ProcessHandler.get_gid(sheet_name)
    if not gid:
        _promo_raw_rows = []
        return
    sheet_id = "1yvxE_SPYXKhofcZdhv1CSKAyiwdY1Mf4pFlsiMbtOr4"
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?tqx=out:csv&gid={gid}"
    df = pd.read_csv(url, dtype=str, header=None)
    df.fillna("", inplace=True)
    _promo_raw_rows = df.values.tolist()


def _noop_upload_file_to_drive(path, output_name=None, vendor=None, entry_date=None, makhachhang=None, cancle_date=None):
    return {"url": "https://example.invalid/skipped-during-fixture-generation"}


_real_find_price_by_sku = xulydonhang.ProcessHandler.find_price_by_sku
_real_find_all_promotions = xulydonhang.ProcessHandler.find_all_promotions_by_sku_and_time
xulydonhang.ProcessHandler.find_price_by_sku = staticmethod(_cached_find_price_by_sku)
xulydonhang.ProcessHandler.find_all_promotions_by_sku_and_time = staticmethod(_cached_find_all_promotions)
xulydonhang.ProcessHandler.upload_file_to_drive = staticmethod(_noop_upload_file_to_drive)


COLUMNS = [
    "A", "B", "C", "D", "E", "G", "K", "L", "Q", "S", "T", "U", "V", "X", "Y", "Z",
    "AE", "AJ", "AM", "AN", "AO", "AP", "AQ", "AT", "AU", "AV",
]


def snapshot_rows(path, start_row):
    wb = openpyxl.load_workbook(path)
    sheet = wb["Don dat hang"]
    rows = []
    for r in range(start_row, sheet.max_row + 1):
        row = {"row_number_offset": r - start_row}
        for col in COLUMNS:
            cell = sheet[f"{col}{r}"]
            value = cell.value
            row[col] = value
            if col == "Z":
                row["Z_is_formula"] = isinstance(value, str) and value.startswith("=")
        comment = sheet[f"Y{r}"].comment
        row["Y_has_comment"] = comment is not None
        row["Y_fill"] = sheet[f"Y{r}"].fill.fgColor.rgb if sheet[f"Y{r}"].fill else None
        rows.append(row)
    wb.close()
    return rows


def process_one_pdf(path):
    """Mirrors the Emart branch of process_file (xulydonhang.py:9314-9384)
    for every page identify_vendor recognizes as Emart, skipping the
    Google Drive upload / current-page-extraction side effects."""
    handler = xulydonhang.process_handler
    doc = xulydonhang.fitz.open(path)
    try:
        for page_num in range(len(doc)):
            text = doc[page_num].get_text("text")
            if xulydonhang.ProcessHandler.identify_vendor(text) != "Emart":
                continue

            po_number = re.search(r"PO No\.\s*\n\s*:? ?([^\n]+)", text)
            if po_number:
                po_number = po_number.group(1).strip()

            entry_date = re.search(r"Order By / Date\s*\n\s*:? ?([^\n]+)", text)
            if entry_date:
                entry_date = entry_date.group(1).strip()
                entry_date = entry_date[:10].replace(".", "/")

            cancel_date = re.search(r"Delivery Date\s*\n\s*:? ?([^\n]+)", text)
            if cancel_date:
                cancel_date = cancel_date.group(1).strip()
                cancel_date = cancel_date[:10].replace(".", "/")

            tenstore = re.search(r"^Delivery to :\s*(.+)", text, re.MULTILINE)
            if tenstore:
                tenstore = tenstore.group(1).split("   ")[0]
            else:
                tenstore = None

            table_match = re.search(r"Article Code\s*(.*?)\s*Total Amount\(without VAT\) :", text, re.DOTALL)
            table_text = table_match.group(1).strip()
            products = xulydonhang.ProcessHandler.laydanhsanpham_emart(table_text)
            if products:
                sku_mapping = xulydonhang.ProcessHandler.load_sku_mapping()
                products = xulydonhang.ProcessHandler.replace_sku_numbers(products, sku_mapping)

            xulydonhang.ProcessHandler.write_to_dondathang_emart(
                handler, products, po_number, entry_date, cancel_date, tenstore,
                1, "MN_MT_KH0032", "Emart", None,
            )
    finally:
        doc.close()


def _remove_with_retry(path, attempts=5, delay=0.5):
    for i in range(attempts):
        try:
            os.remove(path)
            return
        except PermissionError:
            if i == attempts - 1:
                raise
            time.sleep(delay)


def _move_with_retry(src, dst, attempts=5, delay=0.5):
    for i in range(attempts):
        try:
            shutil.move(src, dst)
            return
        except PermissionError:
            if i == attempts - 1:
                raise
            time.sleep(delay)


def main():
    os.makedirs(FIXTURES_DIR, exist_ok=True)

    pdf_paths = sorted(glob.glob(os.path.join(REALPDFS_DIR, "*.pdf")))
    print(f"Found {len(pdf_paths)} candidate PDFs in {REALPDFS_DIR}")

    generated = 0
    skipped = 0
    for path in pdf_paths:
        shutil.copyfile(TEMPLATE_XLSX, SCRATCH_XLSX)
        os.chdir(REPO_ROOT)
        real_target = os.path.join(REPO_ROOT, "dondathang.xlsx")
        backup = real_target + ".fixture_backup"
        shutil.move(real_target, backup)
        shutil.copyfile(SCRATCH_XLSX, real_target)
        try:
            wb = openpyxl.load_workbook(real_target)
            start_row = wb["Don dat hang"].max_row + 1
            wb.close()

            process_one_pdf(path)

            rows = snapshot_rows(real_target, start_row)
        except Exception as e:
            print(f"SKIP (processing failed) {os.path.basename(path)}: {e}")
            skipped += 1
            rows = None
        finally:
            _remove_with_retry(real_target)
            _move_with_retry(backup, real_target)
            if os.path.exists(SCRATCH_XLSX):
                os.remove(SCRATCH_XLSX)

        if rows is None:
            continue

        fixture = {"source_pdf": os.path.basename(path), "rows": rows}
        fixture_name = os.path.splitext(os.path.basename(path))[0] + ".json"
        with open(os.path.join(FIXTURES_DIR, fixture_name), "w", encoding="utf-8") as f:
            json.dump(fixture, f, ensure_ascii=False, indent=2, default=str)
        generated += 1
        print(f"OK {os.path.basename(path)} -> {len(rows)} rows")

    if _promo_raw_rows is None:
        _capture_promo_raw_rows("EMART")
    frozen_pricing = {"raw_rows": _promo_raw_rows or []}
    with open(os.path.join(FIXTURES_DIR, "_frozen_pricing.json"), "w", encoding="utf-8") as f:
        json.dump(frozen_pricing, f, ensure_ascii=False, indent=2, default=str)

    print(f"\nDone: {generated} fixtures generated, {skipped} PDFs skipped.")


if __name__ == "__main__":
    main()
