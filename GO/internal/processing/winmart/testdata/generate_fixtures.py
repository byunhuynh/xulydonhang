"""
Throwaway dev tool — NOT part of the shipped Go or Python app.

Runs the CURRENT xulydonhang.py Winmart pipeline against every real PDF in
đơn hàng/08-2026/ that identify_vendor recognizes as Winmart, capturing the
resulting dondathang.xlsx rows (and the live-fetched Google Sheets
price/promotion data for the WINMART sheet) into JSON fixtures under
GO/internal/processing/winmart/testdata/fixtures/. The Go golden test
(Task 6) diffs RealProcessor's output against these fixtures instead of
against a live Google Sheets fetch, so it's deterministic and offline.

Like Satra/Lotte (one page == one order, write_to_dondathang_winmart
appends immediately, no explicit start-row argument needed), and UNLIKE
BigC (write_to_dondathang_bigc needs an explicit bat_dau/start-row
argument on its last call because it accumulates rows across an entire
multi-store-page document before finalizing), this harness computes
start_row once up front (max_row + 1, no compute_start_row helper needed)
and takes a single snapshot after process_one_pdf's per-page loop has
finished writing every Winmart page in the file — each write call commits
its rows to the sheet immediately, so the snapshot reflects the union of
every page's rows as soon as the loop completes.

4 of the 16 real Winmart PDFs found under đơn hàng/08-2026/ crash the
real, unmodified xulydonhang.py itself with "unsupported operand
type(s) for +=: 'float' and 'str'" inside write_to_dondathang_winmart —
a genuine, pre-existing bug in the legacy Python system, not something
this harness or the Go port is responsible for fixing. They are SKIPped
with that exact error printed, not silently dropped. As of this
writing the 4 are: 4194159303.pdf, 4194159910.pdf, 4194159918.pdf,
4901307989.pdf — all 4 have zero golden fixture coverage.

Run from the repo root:
    .venv/Scripts/python.exe GO/internal/processing/winmart/testdata/generate_fixtures.py
"""
import glob
import json
import os
import shutil
import sys
import time

# Same depth as BigC/Satra/Lotte's harnesses: this script sits 5 directory
# levels below repo root
# (GO/internal/processing/winmart/testdata/generate_fixtures.py), so
# reaching repo root from os.path.abspath(__file__) requires 6 dirname()
# calls (one to strip the filename, five more to strip
# GO/internal/processing/winmart/testdata).
REPO_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))))
sys.path.insert(0, REPO_ROOT)
os.chdir(REPO_ROOT)  # xulydonhang.py's functions use relative paths ("data.xlsx", "settings.ini")

# See Coop's harness for why this is needed: process_file's debug print()
# calls contain emoji that the legacy cp1252 console codepage can't
# encode, aborting processing partway through if not fixed here first.
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="backslashreplace")

import openpyxl  # noqa: E402
import xulydonhang  # noqa: E402

FIXTURES_DIR = os.path.join(
    REPO_ROOT, "GO", "internal", "processing", "winmart", "testdata", "fixtures"
)
TEMPLATE_XLSX = os.path.join(REPO_ROOT, "dondathang.xlsx")
SCRATCH_XLSX = os.path.join(REPO_ROOT, "dondathang_fixture_scratch.xlsx")

# --- Monkey-patch network/upload side effects out (identical shape to
# Coop/Satra/Lotte/BigC's harnesses; find_price_by_sku/
# find_all_promotions_by_sku_and_time are already generic over
# sheet_name, so this works for "WINMART" too with no changes to the
# caching logic itself) ---

_price_cache = {}
_promo_cache = {}
_promo_raw_rows = None


def _cached_find_price_by_sku(sku_number, sheet_name="COOP"):
    key = (sku_number, sheet_name)
    if key not in _price_cache:
        _price_cache[key] = _real_find_price_by_sku(sku_number, sheet_name)
    return _price_cache[key]


def _cached_find_all_promotions(sku_code, time_to_check, sheet_name="Coop"):
    global _promo_raw_rows
    if _promo_raw_rows is None:
        _capture_promo_raw_rows(sheet_name)
    key = (sku_code, time_to_check, sheet_name)
    if key not in _promo_cache:
        _promo_cache[key] = _real_find_all_promotions(sku_code, time_to_check, sheet_name)
    return _promo_cache[key]


def _capture_promo_raw_rows(sheet_name):
    global _promo_raw_rows
    import pandas as pd

    gid = xulydonhang.ProcessHandler.get_gid(sheet_name)
    if not gid:
        _promo_raw_rows = []
        return
    sheet_id = "1yvxE_SPYXKhofcZdhv1CSKAyiwdY1Mf4pFlsiMbtOr4"
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?tqx=out:csv&gid={gid}"
    df = pd.read_csv(url, dtype=str, header=None)
    df.fillna("", inplace=True)
    _promo_raw_rows = df.values.tolist()


def _noop_upload_file_to_drive(path, output_name=None, vendor=None, entry_date=None, makhachhang=None, cancle_date=None):
    return {"url": "https://example.invalid/skipped-during-fixture-generation"}


_real_find_price_by_sku = xulydonhang.ProcessHandler.find_price_by_sku
_real_find_all_promotions = xulydonhang.ProcessHandler.find_all_promotions_by_sku_and_time
xulydonhang.ProcessHandler.find_price_by_sku = staticmethod(_cached_find_price_by_sku)
xulydonhang.ProcessHandler.find_all_promotions_by_sku_and_time = staticmethod(_cached_find_all_promotions)
xulydonhang.ProcessHandler.upload_file_to_drive = staticmethod(_noop_upload_file_to_drive)


# --- Excel row capture (same columns as Coop/Satra/Lotte/BigC's harnesses — same sheet, same layout) ---

COLUMNS = [
    "A", "B", "C", "D", "E", "G", "L", "Q", "S", "T", "U", "V", "X", "Y", "Z",
    "AE", "AJ", "AM", "AO", "AP", "AQ", "AT", "AU", "AV",
]


def snapshot_rows(path, start_row):
    wb = openpyxl.load_workbook(path)
    sheet = wb["Don dat hang"]
    rows = []
    for r in range(start_row, sheet.max_row + 1):
        row = {"row_number_offset": r - start_row}
        for col in COLUMNS:
            cell = sheet[f"{col}{r}"]
            value = cell.value
            row[col] = value
            if col == "Z":
                row["Z_is_formula"] = isinstance(value, str) and value.startswith("=")
        comment = sheet[f"Y{r}"].comment
        row["Y_has_comment"] = comment is not None
        row["Y_fill"] = sheet[f"Y{r}"].fill.fgColor.rgb if sheet[f"Y{r}"].fill else None
        rows.append(row)
    wb.close()
    return rows


def is_winmart_pdf(path):
    doc = xulydonhang.fitz.open(path)
    try:
        text = ""
        for page in doc:
            text += page.get_text("text")
        return xulydonhang.ProcessHandler.identify_vendor(text) == "Winmart"
    finally:
        doc.close()


def process_one_pdf(path):
    """Mirrors the Winmart branch of process_file (xulydonhang.py:8984-9160)
    for every page identify_vendor recognizes as Winmart, skipping the
    Google Drive upload / current-page-extraction side effects."""
    handler = xulydonhang.process_handler
    doc = xulydonhang.fitz.open(path)
    try:
        for page_num in range(len(doc)):
            text = doc[page_num].get_text("text")
            if xulydonhang.ProcessHandler.identify_vendor(text) != "Winmart":
                continue

            lines = text.split("\n")

            idx = next((i for i, line in enumerate(lines) if "Ngày đặt hàng (PO date)" in line), -1)
            entry_date = lines[idx + 1].strip() if idx != -1 and idx + 1 < len(lines) else None
            entry_date = entry_date.replace('.', '/') if entry_date else None

            ghichu = "\n".join(
                text.split("Ghi chú")[1]
                .split("Nhà cung cấp (Supplier): 0002011398")[0]
                .strip()
                .splitlines()[:-1]
            )
            ghichu = ghichu.replace('\n', ' ')

            idx = next((i for i, line in enumerate(lines) if "Số đơn hàng (PO No.)" in line), -1)
            po_number = lines[idx + 1].strip() if idx != -1 and idx + 1 < len(lines) else None

            idx = next((i for i, line in enumerate(lines) if "Ngày giao (Delivery Date)" in line), -1)
            cancel_date = lines[idx + 1].strip() if idx != -1 and idx + 1 < len(lines) else None
            cancel_date = cancel_date.replace('.', '/') if cancel_date else None

            idx = next((i for i, line in enumerate(lines) if "Địa chỉ giao hàng (Delivery Address)" in line), -1)
            if idx != -1:
                ma_kho = lines[idx + 1].strip()
                address_lines = []
                for line in lines[idx + 2:]:
                    if "Thông tin đơn hàng (Information)" in line:
                        break
                    line = line.strip()
                    if "WM+" in line:
                        continue
                    if line:
                        address_lines.append(line)
                diachigiaohang = f"{ma_kho} - {' '.join(address_lines)}"
            else:
                diachigiaohang = None

            idx = -1
            for i in range(len(lines) - 1):
                line_lower = lines[i].lower()
                next_line_lower = lines[i + 1].lower()
                if "tổng hợp" in line_lower and "wincommerce" in next_line_lower:
                    idx = i
                    break
                elif "wincommerce" in line_lower:
                    idx = i
                    break
            if idx != -1:
                diachi_lines = []
                for i in range(idx + 1, len(lines)):
                    line_lower = lines[i].lower()
                    if "mst" in line_lower or "địa chỉ giao hàng" in line_lower:
                        break
                    diachi_lines.append(lines[i].strip())
                diachi = " ".join(diachi_lines)
            else:
                diachi = None

            products = xulydonhang.ProcessHandler.trichxuatsanpham_winmart(text)

            makhachhang = xulydonhang.ProcessHandler.laymakhachhang_satra(diachi, "WINMART")
            if not makhachhang:
                makhachhang = "Không xác định"

            xulydonhang.ProcessHandler.write_to_dondathang_winmart(
                handler, products, makhachhang, po_number, entry_date, cancel_date,
                1, "Winmart", diachigiaohang, ghichu, None,
            )
    finally:
        doc.close()


def _remove_with_retry(path, attempts=5, delay=0.5):
    """os.remove wrapped with retry-with-backoff. Windows Defender's
    real-time scanner can transiently hold a lock on a freshly-saved
    .xlsx right after openpyxl's wb.save() closes it, which surfaces here
    as a PermissionError ([WinError 5] Access is denied). Retrying a few
    times with a short delay lets the scan finish and the lock clear
    before we give up and let the exception propagate for real.

    (BigC's Task 7 added this hardening reactively, after a transient
    Windows file-lock crashed its harness mid-restore on the last PDF;
    this harness includes it from the start.)"""
    for i in range(attempts):
        try:
            os.remove(path)
            return
        except PermissionError:
            if i == attempts - 1:
                raise
            time.sleep(delay)


def _move_with_retry(src, dst, attempts=5, delay=0.5):
    """shutil.move wrapped with the same retry-with-backoff as
    _remove_with_retry, and for the same reason (transient AV lock)."""
    for i in range(attempts):
        try:
            shutil.move(src, dst)
            return
        except PermissionError:
            if i == attempts - 1:
                raise
            time.sleep(delay)


def main():
    os.makedirs(FIXTURES_DIR, exist_ok=True)

    pdf_paths = sorted(glob.glob(os.path.join(REPO_ROOT, "đơn hàng", "08-2026", "*.pdf")))
    print(f"Found {len(pdf_paths)} candidate PDFs")

    generated = 0
    skipped = 0
    for path in pdf_paths:
        try:
            if not is_winmart_pdf(path):
                continue
        except Exception as e:
            print(f"SKIP (vendor check failed) {os.path.basename(path)}: {e}")
            skipped += 1
            continue

        shutil.copyfile(TEMPLATE_XLSX, SCRATCH_XLSX)
        os.chdir(REPO_ROOT)
        real_target = os.path.join(REPO_ROOT, "dondathang.xlsx")
        backup = real_target + ".fixture_backup"
        shutil.move(real_target, backup)
        shutil.copyfile(SCRATCH_XLSX, real_target)
        try:
            wb = openpyxl.load_workbook(real_target)
            start_row = wb["Don dat hang"].max_row + 1
            wb.close()

            process_one_pdf(path)

            rows = snapshot_rows(real_target, start_row)
        except Exception as e:
            print(f"SKIP (processing failed) {os.path.basename(path)}: {e}")
            skipped += 1
            rows = None
        finally:
            _remove_with_retry(real_target)
            _move_with_retry(backup, real_target)
            if os.path.exists(SCRATCH_XLSX):
                os.remove(SCRATCH_XLSX)

        if rows is None:
            continue

        fixture = {"source_pdf": os.path.basename(path), "rows": rows}
        fixture_name = os.path.splitext(os.path.basename(path))[0] + ".json"
        with open(os.path.join(FIXTURES_DIR, fixture_name), "w", encoding="utf-8") as f:
            json.dump(fixture, f, ensure_ascii=False, indent=2, default=str)
        generated += 1
        print(f"OK {os.path.basename(path)} -> {len(rows)} rows")

    if _promo_raw_rows is None:
        _capture_promo_raw_rows("WINMART")
    frozen_pricing = {"raw_rows": _promo_raw_rows or []}
    with open(os.path.join(FIXTURES_DIR, "_frozen_pricing.json"), "w", encoding="utf-8") as f:
        json.dump(frozen_pricing, f, ensure_ascii=False, indent=2, default=str)

    print(f"\nDone: {generated} fixtures generated, {skipped} PDFs skipped.")


if __name__ == "__main__":
    main()
