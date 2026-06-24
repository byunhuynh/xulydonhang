import fitz  # PyMuPDF
import re
import openpyxl  # Đọc file Excel
from datetime import datetime,timedelta, date
import itertools
import math
import os
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment
from PySide6.QtCore import Signal, QObject
from fuzzywuzzy import fuzz
import configparser
import pandas as pd
import pytesseract
from PIL import Image, ImageEnhance, ImageOps
from statistics import mode
import random
import string
import io
import requests
import json
from collections import defaultdict
from zoneinfo import ZoneInfo  # Python 3.9+
import base64, mimetypes




# Đọc thông tin từ file setting.ini
CONFIG_FILE = "settings.ini"
config = configparser.ConfigParser()
songayno_TMDT = 15
songayno_MT = 60    

class ProcessHandler(QObject):
    table_signal = Signal(str, str, str,str, str, str,str)  # Tín hiệu mới để cập nhật bảng

    log_signal = Signal(str)  # Tín hiệu gửi log về giao diện
    

    


    def extract_sku(value):
        """Trích xuất mã SKU từ value (chuỗi có thể chứa ký tự chữ & số)."""
        match = re.search(r"\b[A-Z0-9]+\b", value)  # Tìm từ viết hoa + số
        return match.group(0) if match else None

    def check_value_in_sanpham(value):
        """Tìm tất cả mã SKU có trong value dựa trên danh sách SKU từ 'data.xlsx'."""
        file_path = "data.xlsx"
        sheet_name = "SanPham"

        if not value or not isinstance(value, str):
            return False

        try:
            # 📌 Tải danh sách SKU từ file Excel
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb[sheet_name]

            # Chuyển tất cả mã SKU thành chuỗi để tránh lỗi
            sku_list = {str(row[0]).strip() for row in sheet.iter_rows(min_col=1, max_col=1, values_only=True) if row[0]}

            # 🔍 Tìm tất cả mã SKU có trong chuỗi value
            pattern = r"\b(" + "|".join(map(re.escape, sku_list)) + r")\b"
            matches = re.findall(pattern, value)

            return matches if matches else False

        except Exception as e:
            print(f"⚠ Lỗi khi kiểm tra SKU: {e}")
            return False






    def extract_text_from_pdf(pdf_path):
        """Trích xuất nội dung từ PDF"""
        doc = fitz.open(pdf_path)
        text = "\n".join([page.get_text("text") for page in doc])
        return text


    def identify_vendor(text):
        """Xác định Vendor bất kể khoảng trắng hoặc xuống dòng"""
        cleaned_text = re.sub(r"\s+", " ", text).strip()  # Chuẩn hóa khoảng trắng
       
       

        if re.search(r"Vendor\s*[-:]\s*(21569|22856)", cleaned_text):
            return "Coop"
        
        if re.search(r"3005382", cleaned_text) or re.search(r"CTY TNHH DV EB", cleaned_text, re.IGNORECASE):
            return "BigC"
        
        if re.search(r"0107889783\s*009333", cleaned_text) or re.search(r"1102018142\s*010544", cleaned_text):
            return "Lotte"
        
        if re.search(r"VD-00002345", cleaned_text):
            return "Satra"
        
        if re.search(r"VD-00002547", cleaned_text):
            return "Satra"
        
        if re.search(r"CONG TY TNHH TMDV XNK HA THANH \(101017\)", cleaned_text) or re.search(r"THISO RETAIL COMPANY LIMITED", cleaned_text):
            return "Emart"
        
        if re.search(r"0313403198", cleaned_text):
            return "Kingfood"
        

        if re.search(r"CHI NHÁNH TP. HỒ CHÍ MINH - CÔNG TY TNHH TMDV XNK HÀ THÀNH", cleaned_text) or re.search(r"CÔNG TY TNHH SXTM P&D", cleaned_text):
            return "CN-HCM"

        if re.search(r"Nhà cung cấp \(Supplier\): 0002011398", cleaned_text):
            return "Winmart"
        

        if re.search(r"VN_CÔNG TY TNHH TMDV XNK HÀ THÀNH_CONSIGNMENT_C2M", cleaned_text):
            return "SHOPEE-CHOICE"
        
        if re.search(r"251000000161", cleaned_text):
            return "FujiMart"
        
        if re.search(r"TIKTOK PTE", cleaned_text):
            return "Tiktok"
        
        if re.search(r"Blue\s*VietNam", cleaned_text, re.IGNORECASE) \
        and not re.search(r"Mã\s*đơn\s*hàng", cleaned_text, re.IGNORECASE):
            return "KOC"
        

        if re.search(r"Clevy\s*VietNam", cleaned_text, re.IGNORECASE) \
        and not re.search(r"Mã\s*đơn\s*hàng", cleaned_text, re.IGNORECASE):
            return "KOC"



        if re.search(r"Đơn vị : HỆ THỐNG SIÊU THỊ JMART", cleaned_text):
            return "JMart"
        

        if re.search(r"0317734828", cleaned_text):
            return "MR.DIY"
        
        if re.search(r"Công Ty Cổ Phần Thương Mại Bách Hóa Xanh", cleaned_text):
            return "BHX"
        
        if re.search(r"CÔNG TY CỔ PHẦN ĐẦU TƯ BC", cleaned_text):
            return "BC Mart"
        

        if re.search(r"Mã NCC:V0001626", cleaned_text):
            return "Farmer"
        
        if re.search(r"JV Mart – Siêu Thị Đồ Vệ Sinh", cleaned_text):
            return "JV-Mart"
        
        if re.search(r"JV Mart - Chăm sóc Nhà cửa", cleaned_text):
            return "JV-Mart-2"
        
        if re.search(r"CLEVY VIỆT NAM", cleaned_text):
            return "Clevy"
        

        if re.search(r"Clevy VN – Care Plus", cleaned_text):
            return "Clevy-Care"


        if re.search(r"CÔNG TY CỔ PHẦN EDUCI GROUP", cleaned_text):
            return "Clevy-ViettelPost"

        return "Unknown"

    def extract_po_location(text):
        """Lấy P/O Location từ PDF"""
        match = re.search(r"P/O Location:\s*(.+)", text)
        return match.group(1).strip() if match else None

    def get_makhachhang(po_location):
        """Lấy Mã Khách Hàng từ file data.xlsx"""
        file_path = "data.xlsx"
        sheet_name = "MaKH"
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb[sheet_name]
        
        # Chuyển po_location thành chuỗi và đảm bảo chỉ chứa các chữ số
        po_location_str = str(po_location).strip()
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            col_A, col_B, col_C = row[0], row[2], row[2]
            
            if col_A and str(col_A).strip().upper() in ["COOP", "COOPFOOD"]:
                if col_B:
                    # Xử lý col_B để tách phần số cuối
                    col_B_str = str(col_B).strip()
                    
                    # Tìm phần số cuối bằng regex
                    import re
                    match = re.search(r'(\d+)$', col_B_str)
                    if match:
                        ending_number = match.group(1)
                        
                        # So sánh chính xác với po_location
                        if ending_number == po_location_str:
                            return str(col_C)
        
        return "Không tìm thấy"
    
    
    def normalize_text(text):
        """Chuẩn hóa chuỗi: viết thường, loại bỏ dấu câu, khoảng trắng thừa."""
        text = text.lower()
        text = re.sub(r'[^a-z0-9\s]', '', text)  # Xóa ký tự đặc biệt
        text = re.sub(r'\s+', ' ', text).strip()  # Chuẩn hóa khoảng trắng
        return text
    

    def kiemtravatach_nhieuKM(text):
        # Tìm vị trí các cụm khuyến mãi (vd: 1+1, 2+1, v.v.)
        pattern = r'\b\d+\s*\+\s*\d+\b'
        matches = list(re.finditer(pattern, text))

        # Nếu có ít nhất 2 cụm -> tách tại các điểm đó
        if len(matches) >= 2:
            parts = []
            for i in range(len(matches)):
                start = matches[i].start()
                end = matches[i+1].start() if i+1 < len(matches) else len(text)
                parts.append(text[start:end].strip())
            return parts
        else:
            return [text]




    def laymakhachhang_STF(po_location):
        """Tìm chính xác po_location trong cột B, trả về giá trị tương ứng ở cột C."""
        file_path = "data.xlsx"
        sheet_name = "MaKH"

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb[sheet_name]

            for row in sheet.iter_rows(min_row=2):
                ten_kh = row[1].value  # Cột B
                if ten_kh and str(ten_kh).strip().lower() == str(po_location).strip().lower():
                    return row[2].value  # Cột C
            return None

        except Exception as e:
            print("Lỗi:", e)
            return None
    
    def laymakhachhang_satra(po_location,hethong):
        """Lấy Mã Khách Hàng từ file data.xlsx với tìm kiếm gần đúng."""
        file_path = "data.xlsx"
        sheet_name = "MaKH"
        
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb[sheet_name]
        
        po_location_norm = ProcessHandler.normalize_text(po_location)  
        best_match = None
        best_score = 0
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            col_A, col_B, col_C, col_D = row[:4]

            if col_A and str(col_A).strip().upper() in hethong:
                if col_D:
                    col_D_norm = ProcessHandler.normalize_text(str(col_D))
                    
                    score = fuzz.partial_ratio(po_location_norm, col_D_norm)  
                    if score > best_score:
                        best_score = score
                        best_match = str(col_C)
        
        return best_match if best_score > 95 else None  # Giảm ngưỡng xuống 70%

    def layhethong_COOP(makhachhang):


        """Lấy Mã Khách Hàng từ file data.xlsx"""
        file_path = "data.xlsx"
        sheet_name = "MaKH"
        makhachhang = makhachhang.strip()
        
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb[sheet_name]

        for row in sheet.iter_rows(min_row=2, values_only=True):  # Bỏ qua tiêu đề
            col_A, col_B, col_C = row[0], row[1], row[2]  # Cột A, B, C trong Excel

            if col_C and str(col_C).strip() in [makhachhang]:
                return str(col_A)  # Trả về giá trị cột C dưới dạng chuỗi
        return None
    
    def get_makhachhang_lotte(storecode):
        """Lấy Mã Khách Hàng từ file data.xlsx"""
        file_path = "data.xlsx"
        sheet_name = "MaKH"
        
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb[sheet_name]

        for row in sheet.iter_rows(min_row=2, values_only=True):  # Bỏ qua tiêu đề
            col_A, col_B, col_C = row[0], row[1], row[2]  # Cột A, B, C trong Excel

            if col_A and str(col_A).strip().upper() in ["LOTTE"]:
                if col_C and str(col_C).strip().endswith(str(storecode)):  # Chuyển cả hai về chuỗi
                    return str(col_C)  # Trả về giá trị cột C dưới dạng chuỗi
        return None


    @staticmethod
    def clean_text(text):
        """Xóa dòng trống & gộp dòng liên quan để dễ xử lý"""
        lines = text.split("\n")
        cleaned_lines = [line.strip() for line in lines if line.strip()]  # Xóa dòng trống
        return " ".join(cleaned_lines)  # Gộp tất cả thành 1 đoạn

    @staticmethod
    def tachSP_text(text):
        """Trích xuất SKU, Net Buy Cost, Qty Ord/Pcs, Extended Cost từ văn bản"""
        text = process_handler.clean_text(text)  # Làm sạch và gộp nội dung


        # Regex tìm SKU
        sku_pattern = re.compile(r"(\d{6,}-\d+)")  # SKU có dạng 3573266-1 hoặc 3573267-6

        # Regex tìm giá trị số tiền (có thể có dấu phẩy và chấm)
        cost_pattern = re.compile(r"(\d{1,3}(?:,\d{3})*\.\d{2})")

        matches = sku_pattern.finditer(text)
        products = []

        for match in matches:
            sku = match.group(1)  # SKU
            start_idx = match.end()  # Vị trí kết thúc của SKU trong văn bản
            
            # Cắt phần văn bản sau SKU để tìm các số
            sub_text = text[start_idx:start_idx + 100]  # Lấy 100 ký tự sau SKU để tìm số liệu
            cost_values = cost_pattern.findall(sub_text)

            if len(cost_values) >= 3:
                net_buy_cost = cost_values[-3]  # Net Buy Cost
                qty_ord_pcs = cost_values[-2]  # Qty Ord/Pcs
                extended_cost = cost_values[-1]  # Extended Cost
                
                products.append({
                    "Barcode": sku,
                    "Net Buy Cost": net_buy_cost,
                    "Qty Ord/Pcs": qty_ord_pcs,
                    "Extended Cost": extended_cost
                })
            else:
                print(f"⚠️ Không tìm đủ dữ liệu cho SKU: {sku}")

        return products



    @staticmethod
    def extract_products(text):
        print("Bắt đầu trích xuất sản phẩm...\n")

        # Cắt phần dư thừa
        text = re.split(r"S\s*u\s*b\s*T\s*o\s*t\s*a\s*l", text, flags=re.IGNORECASE)[0]
        text = re.split(r"V\s*N\s*D\s*V\s*i\s*e\s*t\s*N\s*a\s*m\s*D\s*o\s*n\s*g", text, flags=re.IGNORECASE)[-1]

        # Chuẩn hóa dòng
        lines = [line.strip() for line in text.strip().split('\n') if line.strip()]
        print(f"DEBUG: Tổng số dòng sau chuẩn hóa -> {len(lines)}")

        # Tìm các vị trí có SKU
        sku_indices = [i for i, line in enumerate(lines) if re.search(r'\d{7}-\s*\d', line)]
        print(f"DEBUG: Chỉ số SKU được tìm thấy tại {sku_indices}\n")

        products = []

        for i in range(len(sku_indices)):
            start = sku_indices[i]
            end = sku_indices[i + 1] if i + 1 < len(sku_indices) else len(lines)
            block = lines[start:end]
            block = [l.replace(", ", ",").replace(". ", ".") for l in block]

            print(f"DEBUG: Xử lý Khối {i+1}:\n" + "\n".join(block) + "\n")

            item = {"Barcode": None, 'Qty Ord/Pcs': None, 'Extended Cost': None}

            # Trích SKU
            m = re.search(r'(\d{7}-\s*\d)', block[0])
            if m:
                item["Barcode"] = m.group(1).replace(" ", "")
            print(f"DEBUG: Số SKU -> {item["Barcode"]}")

            # Lấy tất cả số
            nums = re.findall(r'(?<![a-zA-Z])\d[\d,]*\.\d+(?![a-zA-Z])', " ".join(block))
            print(f"DEBUG: Tất cả các số -> {nums}")

            # Lọc số lớn chứa dấu phẩy
            large = [n for n in nums if ',' in n]
            print(f"DEBUG: Số tiền lớn trong block -> {large}")

            # Xử lý block cuối
            if i == len(sku_indices) - 1:
                if large:
                    print("DEBUG: 🔍 Áp dụng thông minh cho block cuối với việc thử 2 mức qty...")
                    # Sắp xếp các số large theo giá trị tăng dần
                    lf = [float(x.replace(",", "")) for x in large]
                    for idx in sorted(range(len(large)), key=lambda k: lf[k]):
                        cost = large[idx]
                        cost_idx = nums.index(cost)
                        # Thử qty tiên quyết (cost_idx - 1)
                        if cost_idx > 0:
                            qty1 = nums[cost_idx - 1]
                            up1 = float(cost.replace(",", "")) / float(qty1.replace(",", ""))
                            print(f"DEBUG: Thử 1 Qty={qty1}, Cost={cost}, UP={up1:.2f}")
                            if 1000 < up1 < 2000000:
                                item['Qty Ord/Pcs'] = qty1
                                item['Extended Cost'] = cost
                                print(f"DEBUG: ✅ Chọn 1 Qty={qty1}, Cost={cost}")
                                break
                        # Nếu thất bại, thử qty lùi thêm 1 dòng (cost_idx - 2)
                        if cost_idx > 1:
                            qty2 = nums[cost_idx - 2]
                            up2 = float(cost.replace(",", "")) / float(qty2.replace(",", ""))
                            print(f"DEBUG: Thử 2 Qty={qty2}, Cost={cost}, UP={up2:.2f}")
                            if 1000 < up2 < 2000000:
                                item['Qty Ord/Pcs'] = qty2
                                item['Extended Cost'] = cost
                                print(f"DEBUG: ✅ Chọn 2 Qty={qty2}, Cost={cost}")
                                break
                    if not item['Qty Ord/Pcs']:
                        print("DEBUG: ⚠️ Không tìm được cặp Qty/Cost hợp lý ở cuối, fallback...")
                        if len(nums) >= 2:
                            item['Qty Ord/Pcs'] = nums[-2]
                            item['Extended Cost'] = nums[-1]
                            print(f"DEBUG: ✅ Fallback cuối -> Qty={item['Qty Ord/Pcs']}, Cost={item['Extended Cost']}")
                        else:
                            print("DEBUG: ⚠️ Không đủ số để fallback ở cuối!")
                else:
                    # Không có số large -> fallback
                    if len(nums) >= 2:
                        item['Qty Ord/Pcs'] = nums[-2]
                        item['Extended Cost'] = nums[-1]
                        print(f"DEBUG: ✅ Fallback cuối -> Qty={item['Qty Ord/Pcs']}, Cost={item['Extended Cost']}")
                    else:
                        print("DEBUG: ⚠️ Không đủ số để trích xuất ở cuối!")
            else:
                # Block thường
                if len(large) >= 2:
                    cost = nums[-1]
                    first_large = large[0]
                    idx0 = nums.index(first_large)
                    qty = nums[idx0 - 1] if idx0 > 0 else None
                    item['Qty Ord/Pcs'] = qty
                    item['Extended Cost'] = cost
                    print(f"DEBUG: Block thường thông minh -> Qty={qty}, Cost={cost}")
                else:
                    # Fallback cũ
                    if len(nums) >= 2:
                        item['Qty Ord/Pcs'] = nums[-2]
                        item['Extended Cost'] = nums[-1]
                        print(f"DEBUG: ✅ Fallback thường -> Qty={item['Qty Ord/Pcs']}, Cost={item['Extended Cost']}")
                    else:
                        print("DEBUG: ⚠️ Không đủ số để trích xuất thường!")

            # Kiểm tra đơn giá
            if item['Qty Ord/Pcs'] and item['Extended Cost']:
                try:
                    q = float(item['Qty Ord/Pcs'].replace(",", ""))
                    c = float(item['Extended Cost'].replace(",", ""))
                    up = c / q if q else 0
                    print(f"DEBUG: Đơn giá tính được: {up:.2f}")
                    if not (1000 < up < 2000000):
                        print(f"DEBUG: ⚠️ Đơn giá không hợp lý: {up:.2f}")
                    else:
                        print("DEBUG: Đơn giá hợp lệ.")
                except Exception as e:
                    print(f"DEBUG: ⚠️ Lỗi tính đơn giá: {e}")

            if item["Barcode"]:
                products.append(item)

        print(f"Tổng sản phẩm đã trích xuất: {len(products)}\n")
        return products

        

    def is_special_keyword(line):
        keywords = {"Thùng", "Túi", "Can", "Lọ"}
        return any(keyword in line for keyword in keywords)
   

    # Phiên bản đơn giản hơn nếu cần
    def trichdon_CNHCM(text):
        """Phiên bản đơn giản - ít debug info"""
        start = "Số lượng"
        end = "TỔNG CỘNG"
        
        start_index = text.find(start)
        end_index = text.find(end)
        if start_index != -1 and end_index != -1:
            text = text[start_index + len(start):end_index].strip()


        print('bắt đầu')
        print(text)
        print('Kết thúc')
        
        lines = [line.strip() for line in text.strip().splitlines() if line.strip()]
        result = []
        i = 0
        common_units = {"Thùng", "Hộp", "Chai", "Lon", "Gói", "Cái", "kg", "Túi", "Can"}

        while i < len(lines):
            if lines[i].isdigit():
                # Quét động để tìm dòng Đơn vị tính (match chính xác)
                unit_idx = None
                for j in range(i + 2, min(i + 10, len(lines))):
                    if lines[j] in common_units:
                        unit_idx = j
                        break

                if unit_idx is not None and unit_idx + 1 < len(lines):
                    item = {
                        "Mã hàng": lines[i + 1],
                        "Tên sản phẩm": " ".join(lines[i + 2:unit_idx]),
                        "Đơn vị tính": lines[unit_idx],
                        "Số lượng": lines[unit_idx + 1],
                        "Đơn giá": 0
                    }
                    # Bỏ qua sub-unit và sub-qty nếu có
                    next_i = unit_idx + 2
                    if next_i < len(lines) and lines[next_i] in common_units:
                        next_i += 2
                    i = next_i
                    result.append(item)
                else:
                    i += 1
            else:
                i += 1

        print(result)
        
        return result





    def trichdon_CNHCM2(text):
        text = ProcessHandler.lamsachdon_CNHCM2(text)
        lines = [line.strip() for line in text.split('\n') if line.strip()]
        items = []
        
        i = 0
        while i < len(lines):
            item = {
                "Mã hàng": "",
                "Tên sản phẩm": "",
                "Đơn vị tính": "",
                "Số lượng": "",
                "Đơn giá": "",
                "Thành tiền": ""
            }
            
            # Mã hàng
            if i < len(lines):
                item["Mã hàng"] = lines[i]
                i += 1
            
            # Tên sản phẩm (có thể 1 hoặc 2 dòng)
            product_name = ""
            if i < len(lines):
                product_name = lines[i]
                i += 1
                
                # Kiểm tra xem dòng tiếp theo có phải là phần 2 của tên sản phẩm không
                if (i < len(lines) and 
                    lines[i] not in ['Can', 'Chai', 'Túi', 'Hộp', 'Gói', 'Thùng', 'Lọ', 'Bịch', 'kg', 'Lon'] and
                    not re.match(r'^\d+[.,]\d*$', lines[i])):  # không phải số lượng
                    
                    product_name += " " + lines[i]
                    i += 1
            
            item["Tên sản phẩm"] = product_name
            
            # Đơn vị tính
            if i < len(lines):
                item["Đơn vị tính"] = lines[i]
                i += 1
            
            # Số lượng
            if i < len(lines):
                qty = lines[i]
                # Kiểm tra nếu là số (có thể có dấu phay)
                if re.match(r'^\d+[.,]?\d*$', qty):
                    item["Số lượng"] = qty.replace(",00", "")
                    i += 1
            
            # Đơn giá
            if i < len(lines):
                price = lines[i]
                # Kiểm tra nếu là giá tiền (số có thể có dấu chấm phân cách nghìn)
                if re.match(r'^\d{1,3}(\.\d{3})*[.,]?\d*$', price):
                    item["Đơn giá"] = price.replace(",00", "")
                    i += 1
            
            # Thành tiền
            if i < len(lines):
                total = lines[i]
                # Kiểm tra nếu là số tiền
                if re.match(r'^\d{1,3}(\.\d{3})*[.,]?\d*$', total):
                    item["Thành tiền"] = total.replace(",00", "")
                    i += 1
            if item["Đơn giá"] or item["Thành tiền"]:
                items.append(item)
        
        return items
        


    

    def lamsachdon_CNHCM2(text):
        start_pattern = "Thành tiền"  # Tìm dòng bắt đầu
        end_marker = "Cộng tiền hàng:"  # Dòng kết thúc

        lines = text.split("\n")
        start_index = None
        end_index = None

        for i, line in enumerate(lines):
            if start_index is None and re.match(start_pattern, line):  # Tìm dòng đầu tiên
                start_index = i
            if line.strip() == end_marker:  # Tìm dòng cuối cùng
                end_index = i
                break  # Dừng ngay khi tìm thấy dòng cuối

        if start_index is not None and end_index is not None and start_index < end_index:
            return "\n".join(lines[start_index + 1:end_index])  # Lấy nội dung giữa hai dòng

        return None  # Không tìm thấy hai dòng

        
            
        





        

    def xoakhoangtrang(text):
        return re.sub(r'\s+', '', text)  # Xóa tất cả khoảng trắng trong chuỗi
    


    def remove_line_numbers(text):
        # Chia text thành các dòng
        lines = text.split('\n')
        # Danh sách để chứa các dòng đã xử lý
        cleaned_lines = []
        # Giới hạn số thứ tự cần xóa (1 đến 10)
        valid_numbers = set(str(i) for i in range(1, 11))  # Tập hợp "1" đến "10"
        
        for line in lines:
            # Kiểm tra nếu dòng bắt đầu bằng số
            if line.strip():
                first_part = line.split(maxsplit=1)[0]
                if first_part in valid_numbers:
                    # Nếu là số thứ tự từ 1-10, loại bỏ nó và giữ nội dung phía sau
                    if len(line.split(maxsplit=1)) > 1:  # Đảm bảo có nội dung phía sau
                        cleaned_line = line.split(maxsplit=1)[1]
                        cleaned_lines.append(cleaned_line)
                    # Nếu chỉ có số (không nội dung), bỏ qua dòng
                else:
                    # Giữ nguyên dòng nếu không phải số thứ tự từ 1-10
                    cleaned_lines.append(line)
            else:
                # Giữ nguyên dòng trống
                cleaned_lines.append(line)
        
        # Gộp lại các dòng thành text hoàn chỉnh
        return '\n'.join(cleaned_lines)

    def extract_info(text):
        
        """Trích xuất các thông tin quan trọng từ văn bản"""
        text = ProcessHandler.remove_line_numbers(text)
        text = re.sub(r"\s+", " ", text)
        text = ProcessHandler.xoakhoangtrang(text)
        text = text.split("Currency")[0] if "Currency" in text else text

        print(f'Bắt đầu trích xuất {text}')
        
        info = {
            "P/O Number": re.search(r"P\s*/\s*O\s*N\s*u\s*m\s*b\s*e\s*r\s*[:-]?\s*([\d-]+)", text),
            "P/O Location": re.search(r"P\s*/\s*O\s*L\s*o\s*c\s*a\s*t\s*i\s*o\s*n\s*:\s*(\d+)", text),
            "Entry Date": re.search(r"(?:Entry\s*Date|E\s*n\s*t\s*r\s*y\s*D\s*a\s*t\s*e)\s*-\s*([\d/]+)", text),
            "Cancel Date": re.search(r"Cancel\s*Date-\s*([\d/]+)", text)
        }

        if not info["P/O Location"]:
            info["P/O Location"] = re.search(r"Store-\s*(\d+)", text)

        extracted_info = {key: (match.group(1).strip() if match else "Không tìm thấy") for key, match in info.items()}

        return extracted_info
        
    def extract_discount(value):
        """Trích xuất số phần trăm giảm giá từ chuỗi."""
        if isinstance(value, list):  
            value = value[0] if value else ""

        if not isinstance(value, str):  
            value = str(value)  # Chuyển thành chuỗi nếu chưa phải chuỗi

        match = re.search(r"(\d+(\.\d+)?)%", value)  # Tìm số trước dấu "%"
        if match:
            return float(match.group(1))  # Chuyển thành số thực
        
        return 0  # Không có chiết khấu

    
        


    def extract_number(text):
        """Hàm trích xuất phần số từ chuỗi, nếu không có số thì trả về None"""
        match = re.search(r'\d+', str(text))
        return int(match.group()) if match else None
    

    def tachkhuyenmai_coop(text, hethong):
        text = text.strip()
        hethong = hethong.strip()

        cm_result = cf_result = ""

        # ✅ Tách CF trước (ưu tiên trong ngoặc)
        cf_paren_match = re.search(r'\(([^)]*cf[^)]*)\)', text, flags=re.IGNORECASE)
        if cf_paren_match:
            cf_result = cf_paren_match.group(1).strip()
        else:
            cf_match = re.search(r'(cf[^\n]*)', text, flags=re.IGNORECASE)
            if cf_match:
                cf_result = cf_match.group(1).strip()

        # ✅ Tách CM: từ CM đến trước CF (không cắt ở dấu ngoặc bình thường)
        if cf_result:
            # Nếu tìm được CF → cắt CM tới ngay trước CF
            cf_start = text.lower().find(cf_result.lower())
            cm_candidate = text[:cf_start]
            cm_match = re.search(r'(cm.*)', cm_candidate, flags=re.IGNORECASE)
            if cm_match:
                cm_result = cm_match.group(1).strip()
        else:
            # Nếu không có CF → lấy toàn bộ đoạn có CM
            cm_match = re.search(r'(cm[^\n]*)', text, flags=re.IGNORECASE)
            if cm_match:
                cm_result = cm_match.group(1).strip()

        if hethong.upper() == "COOPMART":
            return cm_result if cm_result else text
        elif hethong.upper() == "COOPFOOD":
            return cf_result if cf_result else text
        else:
            return text


    def timten_sanpham(search_value,file_path = "data.xlsx", sheet_name="SanPham"):
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name]

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2, values_only=True):
            key, value = row
            if key == search_value:
                return value  # Trả về giá trị cột B nếu tìm thấy

        return "Không thấy tên sản phẩm"  # Trả về None nếu không tìm thấy
    
    def timtrongluong_sanpham(search_value, file_path="data.xlsx", sheet_name="SanPham"):
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name]

        # Quét từ dòng 2 (bỏ tiêu đề), lấy cột A (key), cột B, cột C
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3, values_only=True):
            col_a, col_b, col_c = row
            if col_a == search_value:
                return col_c  # Trả về giá trị cột C nếu tìm thấy

        return "Không thấy trọng lượng sản phẩm"
    
    def timquycach_sanpham(search_value, file_path="data.xlsx", sheet_name="SanPham"):
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name]

        # Quét từ dòng 2 (bỏ tiêu đề), lấy cột A (key), cột B, cột C
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4, values_only=True):
            col_a, col_b, col_c, col_d = row
            if col_a == search_value:
                return col_d  # Trả về giá trị cột C nếu tìm thấy

        return "Không thấy Quy cách sản phẩm"


    def laydiachi_coopfood(makhachhang):
        file_path = "data.xlsx"
        sheet_name = "MaKH"

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb[sheet_name]

            for row in sheet.iter_rows(min_row=2, values_only=True):  # Bỏ qua dòng tiêu đề
                col_C = row[2]  # Cột C là index 2 (0-based)
                col_D = row[3] if len(row) > 3 else None  # Cột D là index 3
                if str(col_C).strip() == str(makhachhang).strip():
                    return col_D

            return "Không tìm thấy địa chỉ"

        except Exception as e:
            return f"Lỗi khi đọc file: {e}"
        

    def laycachbo_khuyenmai(value):
        match = re.search(r"\{(.*?)\}", value)
        if match:
            return match.group(1)  # Lấy nội dung bên trong dấu {}
        return None  # Trả về None nếu không tìm thấy
    

    def laytongtien(sheet, row, col="Z"):
        value = sheet[f"{col}{row}"].value
        if isinstance(value, str):
            if value.startswith("="):
                return 0
            try:
                return float(value.replace(",", ""))
            except ValueError:
                return 0
        return value if isinstance(value, (int, float)) else 0


    def tachtien_khuyenmai(text):
        """
        Trích xuất số tiền từ text:
        - Hỗ trợ dạng '199k', '199 K', '199K'
        - Hỗ trợ dạng đầy đủ như '199000'
        - Ưu tiên lấy số đầu tiên hợp lệ (nếu có nhiều)
        """
        print(f"[DEBUG] Đầu vào: {text}")

        # Tìm 'số + k' → nhân 1000
        match_k = re.search(r'\b(\d{1,3})(?:\s*)k\b', text, re.IGNORECASE)
        if match_k:
            value = int(match_k.group(1)) * 1000
            print(f"[DEBUG] Phát hiện dạng 'k': {match_k.group(0)} -> {value}")
            return value

        # Tìm số lớn nhất có 5 hoặc 6 chữ số (dạng 199000, 150000...)
        match_full = re.search(r'\b\d{5,6}\b', text)
        if match_full:
            value = int(match_full.group(0))
            print(f"[DEBUG] Phát hiện số đầy đủ: {match_full.group(0)} -> {value}")
            return value

        print("[DEBUG] Không tìm thấy số tiền hợp lệ.")
        return None  # Không tìm thấy

    def layduoi_mahang(text):
        base = text.split('_')[0]
        return base[-4:]

    def format_weight_kg(value: float) -> str:
        """
        Định dạng trọng lượng với đơn vị gốc là kg.
        < 1000 kg  -> hiển thị kg
        >= 1000 kg -> đổi sang tấn
        """
        if value >= 1000:
            return f"{round(value / 1000, 2)} tấn"
        else:
            return f"{round(value, 2)} kg"


    def ghi_message(text: str, file_path="message.txt"):
        """Ghi thêm 1 dòng vào file message.txt"""
        with open(file_path, "a", encoding="utf-8") as f:
            f.write(text.strip() + "\n")


    def format_timedelta(td):
        total_seconds = int(td.total_seconds())
        if total_seconds < 0:
            total_seconds = -total_seconds  # tránh số âm nếu đảo thứ tự

        minutes, seconds = divmod(total_seconds, 60)

        # Nếu có phút
        if minutes > 0:
            return f"{minutes}p{seconds}s"
        else:
            return f"{seconds}s"



    def write_to_dondathang(self,products, makhachhang, po_number, entry_date, cancle_date,STT_donhang,vendor,ghichu,shipto,url):

        now = datetime.now()

        #taongaunhien
        """Ghi danh sách sản phẩm vào file dondathang.xlsx"""
        file_path = "dondathang.xlsx"
        sheet_name = "Don dat hang"
        
        
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]

        start_row = sheet.max_row + 1
        current_row = start_row
        # Lấy giá trị G1, đảm bảo nó là số
        trongluong = 0
        sokienhang = 0
        
        
        print(f"📍 PO Number {po_number} ")
        print(f"📍 ")
        STT_donhang_str = f"-{po_number}"     
        hethong = ProcessHandler.layhethong_COOP(makhachhang)
        if hethong == 'COOPFOOD':
            diachimoi = ProcessHandler.laydiachi_coopfood(makhachhang)
            if diachimoi:
                shipto = shipto + " - " + diachimoi

        else:
            hethong = "COOPMART"


        now_str = now.strftime("%d/%m/%Y %H:%M:%S")


        
        ProcessHandler.ghi_message(f"[PO: {po_number}")
        ProcessHandler.ghi_message(f"store: {shipto}")
        ProcessHandler.ghi_message(f"vendor: {hethong}")
        ProcessHandler.ghi_message(f"Mã Khách hàng: {makhachhang}")
        ProcessHandler.ghi_message(f"start_time: {now_str}")
        if url:
            ProcessHandler.ghi_message(f"url: {url}")



        




        saigia = 0
        tongtien = current_row
        diengiai = f"{hethong} PO{po_number}" + (f" - {ghichu}" if ghichu else "")


        if makhachhang[:2] == "MB":
            khuvuc = "MT_MB"
            kho = "TP_HN_12"
            mien = "HN"
        else:
            kho = "LA_TP"
            khuvuc = "MT_MN"
            mien = "LA"




        
        sheet[f"A{current_row}"] = entry_date
        sheet[f"AV{current_row}"] = songayno_MT
        sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'

        sheet[f"C{current_row}"] = "Chưa thực hiện"
        sheet[f"D{current_row}"] = cancle_date
        sheet[f"E{current_row}"] = shipto
        sheet[f"G{current_row}"] = makhachhang
        sheet[f"L{current_row}"] = diengiai
        sheet[f"V{current_row}"] = kho
        sheet[f"AE{current_row}"] = 8
        sheet[f"AJ{current_row}"] = khuvuc
        sheet[f"AM{current_row}"] = mien
        sheet[f"U{current_row}"] = "Không"
        sheet[f"Z{current_row}"] = 0
        sheet[f"S{current_row}"] = f"{hethong} PO{po_number}"
        sheet[f"T{current_row}"] = "Có"
        sheet[f"X{current_row}"] = 0
        sheet[f"Y{current_row}"] = 0

        current_row += 1
            
        for product in products:
            sheet[f"A{current_row}"] = entry_date
            sheet[f"AV{current_row}"] = songayno_MT

            sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'

            sheet[f"C{current_row}"] = "Chưa thực hiện"
            sheet[f"D{current_row}"] = cancle_date
            sheet[f"E{current_row}"] = shipto
            sheet[f"G{current_row}"] = makhachhang
            sheet[f"L{current_row}"] = diengiai
            sheet[f"Q{current_row}"] = product["Barcode"]
            sheet[f"V{current_row}"] = kho
            sheet[f"AE{current_row}"] = 8
            sheet[f"AJ{current_row}"] = khuvuc
            sheet[f"AM{current_row}"] = mien
            sheet[f"U{current_row}"] = "Không"
            sheet[f"Y{current_row}"] = 0
            sheet[f"T{current_row}"] = "Không"
            sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(product["Barcode"])
            

            # ✅ Xử lý "Extended Cost"
            extended_cost = product["Extended Cost"]
            if isinstance(extended_cost, str):  # Nếu là chuỗi, loại bỏ dấu phẩy
                extended_cost = float(extended_cost.replace(",", ""))
            extended_cost = int(extended_cost) if extended_cost.is_integer() else extended_cost
            sheet[f"Z{current_row}"] = f"=Y{current_row}*X{current_row}"

            # ✅ Xử lý "Qty Ord/Pcs"
            qty_ord_pcs = product["Qty Ord/Pcs"]
            if isinstance(qty_ord_pcs, str):  # Nếu là chuỗi, loại bỏ dấu phẩy
                qty_ord_pcs = float(qty_ord_pcs.replace(",", ""))
            qty_ord_pcs = int(qty_ord_pcs) if qty_ord_pcs.is_integer() else qty_ord_pcs
            sheet[f"X{current_row}"] = qty_ord_pcs

            trongluong += (ProcessHandler.timtrongluong_sanpham(product["Barcode"]) * qty_ord_pcs)
            sokienhang += math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(product["Barcode"]))



            sheet[f"AU{current_row}"] = math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(product["Barcode"]))
            sheet[f"AT{current_row}"] = ProcessHandler.timtrongluong_sanpham(product["Barcode"]) * qty_ord_pcs

            # ✅ Tính giá hóa đơn
            giahoadon = extended_cost / qty_ord_pcs, 2



            giathuctegoc = ProcessHandler.find_price_by_sku(product["Barcode"])
            giathucte = giathuctegoc
            # 🔍 **Tìm giá trị trong CTKM.xlsx theo SKU & thời gian**
            
            results = ProcessHandler.find_all_promotions_by_sku_and_time(product["Barcode"], entry_date,vendor)
            slkhuyenmai = len(results)
            print(f'Giá thực tế: {giathucte}')
            print(f'🔎 Số lượng CTKM khớp: {slkhuyenmai}')
            khuyenmai = ''
            



            giakhop = False  # Cờ kiểm tra khớp giá

            for col, value in results:
                print(f'🧪 Thử CTKM: {col} → {value}')
                print(f"🔹 Khuyến mãi gốc: {value}")
                print(f"🔹 Hệ thống: {hethong}")
                value = ProcessHandler.tachkhuyenmai_coop(value, hethong)
                khuyenmai = value
                print(f"🔹 Đã tách nội dung: {value}")
                if value:
                    
                    discount = ProcessHandler.extract_discount(value)
                    print(f'🔻 Giảm giá: {discount} ({type(discount)})')

                    if discount != 0:
                        if isinstance(discount, str):
                            discount = discount.replace("%", "").strip()
                            discount = float(discount) if discount else 0

                        if isinstance(giathuctegoc, str):
                            giathucte = giathuctegoc.replace(",", "").strip()
                            giathucte = float(giathucte) if giathucte else 0

                        giathucte = giathucte - (giathucte * float(discount) / 100)
                    else:
                        # Nếu không có giảm giá, kiểm tra lại giá trị giathucte
                        if isinstance(giathuctegoc, str):
                            giathucte = giathuctegoc.replace(",", "").strip()
                            giathucte = float(giathucte) if giathucte else 0



                    giahoadon = float(giahoadon[0]) if isinstance(giahoadon, tuple) else float(giahoadon)
                    if isinstance(giathucte, str):
                        giathucte = giathucte.replace(",", "").strip()
                        giathucte = float(giathucte) if giathucte else 0

                    if math.isclose(giahoadon, giathucte, rel_tol=1e-4):
                        sheet[f"Y{current_row}"] = giathucte
                        self.log_signal.emit(
                        f'🎉 Mã hàng <b><span style="color: blue;">{product["Barcode"]}</span></b> '
                        f'có CTKM <b><span style="color: green;">{value}</span></b> từ <i>{col}</i> 🛒')
                        self.log_signal.emit(
                            f'✅ Mã hàng <b><span style="color: blue;">{product["Barcode"]}</span></b> '
                            f'có giá chính xác 🎯: <span style="color: green;"><b>{giahoadon}</b></span>'
                        )
                        
                        
                        giakhop = True
                        break  # ✅ THOÁT VÒNG LẶP khi đã khớp giá
                if giakhop == True:
                    break

            if not results:
                khuyenmai = ''
                giahoadon = float(giahoadon[0]) if isinstance(giahoadon, tuple) else float(giahoadon)
                if isinstance(giathucte, str):
                    giathucte = giathucte.replace(",", "").strip()
                    giathucte = float(giathucte) if giathucte else 0

                if math.isclose(giahoadon, giathucte, rel_tol=1e-4):
                    sheet[f"Y{current_row}"] = giathucte
                    self.log_signal.emit(
                    f'🎉 Mã hàng <b><span style="color: blue;">{product["Barcode"]}</span></b> '
                    f'có CTKM <b><span style="color: green;">{khuyenmai}</span></b> từ <i></i> 🛒')
                    self.log_signal.emit(
                        f'✅ Mã hàng <b><span style="color: blue;">{product["Barcode"]}</span></b> '
                        f'có giá chính xác 🎯: <span style="color: green;"><b>{giahoadon}</b></span>'
                    )
                    giakhop = True



            # ❌ Sau khi thử hết, không khớp giá nào:
            if not giakhop:
                
                sheet[f"Y{current_row}"] = giathucte
                red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                sheet[f"Y{current_row}"].fill = red_fill
                comment_text = f"Kiểm tra lại giá mã này! - Giá hóa đơn: {giahoadon} - Chênh lệch: {giahoadon - giathucte}"
                comment = Comment(comment_text, "System")
                sheet[f"Y{current_row}"].comment = comment

                self.log_signal.emit(
                    f"⚠️ Mã hàng <b>{product["Barcode"]}</b> sai giá! 🛑 "
                    f"Hóa đơn: <b><span style='color: red;'>{giahoadon}</span></b> – "
                    f"Giá hệ thống: <b><span style='color: green;'>{giathucte}</span></b>"
                )
                ProcessHandler.ghi_message(f"text: Mã hàng {product["Barcode"]} - {ProcessHandler.timten_sanpham(product["Barcode"])}, Giá trên PO {giahoadon:,}, Giá đúng: {giathucte:,}. Chênh lệch {(giahoadon - giathucte):,.0f}")
                saigia += 1
                sheet[f"AQ{current_row}"] = khuyenmai

            tongtien += giathucte * qty_ord_pcs

            

            nhieuCtkm = khuyenmai.split('|')

            for i, hangkm in enumerate(nhieuCtkm):
                print("Chương trình khuyến mãi tách được:")
                
                print(hangkm)

                kiemtra = ProcessHandler.check_value_in_sanpham(hangkm)
                match = re.search(r"(\d+)\s*\+\s*1", hangkm)  # Tìm số đầu tiên trong biểu thức X+1
                sheet[f"AQ{current_row}"] = khuyenmai
                
                
                    
                    
                if match:
                    x = int(match.group(1))  # Chuyển đổi thành số nguyên
                    kiemtra = kiemtra or product["Barcode"]
                    
                    if x >= 2:  # Chỉ chia khi x >= 2, còn 1+1 thì giữ nguyên
                        qty_ord_pcs = math.floor(qty_ord_pcs / x)  # Chia rồi làm tròn xuống
                
                if kiemtra:
                    if isinstance(kiemtra, list):  # Nếu có nhiều mã SKU, nối thành chuỗi
                        kiemtra = ", ".join(kiemtra)
                    cachbokem = ProcessHandler.laycachbo_khuyenmai(hangkm) or "KM Bó Kèm - Che Barcode"
                    if cachbokem:
                        if i == 0:
                            sheet[f"AO{current_row}"] = cachbokem
                            



                            cachbokem_lower = cachbokem.lower()
                            if "bó kèm" in cachbokem_lower or "quấn kèm" in cachbokem_lower:
                                sheet[f"AP{current_row}"] = f'{ProcessHandler.layduoi_mahang(product["Barcode"])}_{ProcessHandler.layduoi_mahang(kiemtra)}_1'
                                sheet[f"AP{current_row + 1}"] = f'{ProcessHandler.layduoi_mahang(product["Barcode"])}_{ProcessHandler.layduoi_mahang(kiemtra)}_1'
                        elif i > 0:
                            sheet[f"AO{current_row + 1}"] = cachbokem
                            

                            cachbokem_lower = cachbokem.lower()
                            if "bó kèm" in cachbokem_lower or "quấn kèm" in cachbokem_lower:
                                sheet[f"AP{current_row + 1}"] = f'{ProcessHandler.layduoi_mahang(product["Barcode"])}_{ProcessHandler.layduoi_mahang(kiemtra)}_1'
                             




                    self.log_signal.emit(
        f'🎁 Mã hàng <b><span style="color: blue;">{product["Barcode"]}</span></b> '
        f'đã thêm hàng khuyến mãi <b><span style="color: green;">{kiemtra}</span></b> tặng kèm! 🎉'
        )
                    current_row += 1
                    sheet[f"A{current_row}"] = entry_date
                    sheet[f"AV{current_row}"] = songayno_MT
                    sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'
                    sheet[f"C{current_row}"] = "Chưa thực hiện"
                    sheet[f"D{current_row}"] = cancle_date
                    sheet[f"G{current_row}"] = makhachhang
                    sheet[f"L{current_row}"] = diengiai
                    sheet[f"Q{current_row}"] = kiemtra
                    sheet[f"X{current_row}"] = qty_ord_pcs
                    sheet[f"V{current_row}"] = kho
                    sheet[f"Y{current_row}"] = 0
                    sheet[f"Z{current_row}"] = 0
                    sheet[f"U{current_row}"] = "Có"
                    sheet[f"AJ{current_row}"] = khuvuc
                    sheet[f"AE{current_row}"] = 8
                    sheet[f"AM{current_row}"] = mien
                    sheet[f"T{current_row}"] = "Không"
                    sheet[f"E{current_row}"] = shipto
                    sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(kiemtra)
                    trongluong += (ProcessHandler.timtrongluong_sanpham(kiemtra) * qty_ord_pcs)
                    sokienhang += math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(kiemtra))
                    sheet[f"AU{current_row}"] = math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(kiemtra))
                    sheet[f"AT{current_row}"] = ProcessHandler.timtrongluong_sanpham(kiemtra) * qty_ord_pcs
            current_row += 1


        kmhoadon = ProcessHandler.find_all_promotions_by_sku_and_time("Hóa Đơn", entry_date,vendor)
        
        

        
        if kmhoadon:
            kiemtra = ProcessHandler.check_value_in_sanpham(kmhoadon) 
            
            
            sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'
            soluongkm = math.floor(tongtien / ProcessHandler.tachtien_khuyenmai(kmhoadon))
            sheet[f"A{current_row}"] = entry_date
            sheet[f"AV{current_row}"] = songayno_MT
            sheet[f"C{current_row}"] = "Chưa thực hiện"
            sheet[f"D{current_row}"] = cancle_date
            sheet[f"E{current_row}"] = shipto
            sheet[f"G{current_row}"] = makhachhang
            sheet[f"L{current_row}"] = diengiai
            sheet[f"Q{current_row}"] = kiemtra[0]
            sheet[f"X{current_row}"] = soluongkm
            sheet[f"V{current_row}"] = kho
            sheet[f"Y{current_row}"] = 0
            sheet[f"Z{current_row}"] = 0
            sheet[f"U{current_row}"] = "Có"
            sheet[f"AJ{current_row}"] = khuvuc
            sheet[f"AE{current_row}"] = 8
            sheet[f"AM{current_row}"] = mien
            sheet[f"T{current_row}"] = "Không"
            sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(kiemtra[0])
            sheet[f"AQ{current_row}"] = kmhoadon
            cachbokem = ProcessHandler.laycachbo_khuyenmai(kmhoadon)
            trongluong += (ProcessHandler.timtrongluong_sanpham(kiemtra[0]) * soluongkm)
            sokienhang += math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(kiemtra[0]))

            sheet[f"AU{current_row}"] = math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(kiemtra[0]))
            sheet[f"AT{current_row}"] = ProcessHandler.timtrongluong_sanpham(kiemtra[0]) * soluongkm
            if cachbokem:
                sheet[f"AO{current_row}"] = cachbokem
            else:
                sheet[f"AO{current_row}"] = f'KM Bó Kèm - Che Barcode'

            self.log_signal.emit(f'🎉 Đơn hàng có chương trình khuyến mãi <b><span style="color: green;">{kmhoadon}</span></b> 🛒')
            self.log_signal.emit(f'Tổng tiền đơn hàng: <b><span style="color: green;">{tongtien} </span></b>\nSố lượng tặng: <b><span style="color: green;">{soluongkm} </span></b> 🛒')
            

        
        

        ProcessHandler.ghi_message(f"tong_tien: {tongtien:,.0f}")
        ProcessHandler.ghi_message(f"tong_trongluong: {ProcessHandler.format_weight_kg(trongluong)}")
        ProcessHandler.ghi_message(f"tong_kienhang: {sokienhang}")
        ProcessHandler.ghi_message(f"sai_gia: {saigia}")
        now2 = datetime.now()
        now_str = now2.strftime("%d/%m/%Y %H:%M:%S")
        ProcessHandler.ghi_message(f"end_time: {now_str}")
        time = now2 - now
        ProcessHandler.ghi_message(f"time: {ProcessHandler.format_timedelta(time)}]")    

        
        sheet[f"L{start_row}"] = f'{diengiai} (Tổng trọng lượng: {ProcessHandler.format_weight_kg(trongluong)})'
        # Lưu lại file Excel
        wb.save(file_path)
        print(f"✅ Đã ghi {len(products)} dòng vào '{file_path}', sheet '{sheet_name}'!")
        
        return saigia
    

    def write_to_dondathang_JIT(self,products, makhachhang, po_number, entry_date, cancle_date,STT_donhang,vendor,shipto,po):

        


        

        #taongaunhien
        """Ghi danh sách sản phẩm vào file dondathang.xlsx"""
        file_path = "dondathang.xlsx"
        sheet_name = "Don dat hang"
        
        
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]

        start_row = sheet.max_row + 1
        current_row = start_row
        # Lấy giá trị G1, đảm bảo nó là số
        
        
        print(f"📍 PO Number {po_number} ")
        print(f"📍 ")
        STT_donhang_str = f"-{po_number}-{shipto}"     
        hethong = 'JIT-CHOICE'
        trongluong = 0
        sokienhang = 0
        
        



        saigia = 0
        tongtien = current_row
        diengiai = f"{hethong} Ngày đổ {po_number} {shipto}" 
        


        if shipto == "WH6_HN":
            khuvuc = "TMĐT_MB"
            kho = "TP_HN_12"
            mien = "HN"
        else:
            kho = "LA_KHOTMDT"
            khuvuc = "TMĐT_MN"
            mien = "LA"




        
            
        for product in products:
            
            sheet[f"A{current_row}"] = entry_date
            sheet[f"AV{current_row}"] = songayno_TMDT

            sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'
            
            sheet[f"C{current_row}"] = "Chưa thực hiện"
            sheet[f"D{current_row}"] = cancle_date
            sheet[f"AO{current_row}"] = po
            sheet[f"E{current_row}"] = shipto
            sheet[f"G{current_row}"] = makhachhang
            sheet[f"L{current_row}"] = diengiai
            sheet[f"Q{current_row}"] = product["Barcode"]
            sheet[f"V{current_row}"] = kho
            sheet[f"AE{current_row}"] = 8
            sheet[f"AJ{current_row}"] = khuvuc
            sheet[f"AM{current_row}"] = mien
            sheet[f"U{current_row}"] = "Không"
            sheet[f"Y{current_row}"] = 0
            sheet[f"T{current_row}"] = "Không"
            sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(product["Barcode"])
            print(f'Mã sản phẩm {product["Barcode"]}')
            print(ProcessHandler.timten_sanpham(product["Barcode"]))

            # ✅ Xử lý "Extended Cost"
            giathuctegoc = ProcessHandler.find_price_by_sku(product["Barcode"],vendor)
            if isinstance(giathuctegoc, str):
                            giathucte = giathuctegoc.replace(",", "").strip()
                            giathucte = float(giathucte) if giathucte else 0


            sheet[f"y{current_row}"] = giathucte
            sheet[f"Z{current_row}"] = f"=Y{current_row}*X{current_row}"

            # ✅ Xử lý "Qty Ord/Pcs"
            qty_ord_pcs = product["Qty Ord/Pcs"]
            if isinstance(qty_ord_pcs, str):  # Nếu là chuỗi, loại bỏ dấu phẩy
                qty_ord_pcs = float(qty_ord_pcs.replace(",", ""))
            qty_ord_pcs = int(qty_ord_pcs) if qty_ord_pcs.is_integer() else qty_ord_pcs
            sheet[f"X{current_row}"] = qty_ord_pcs

            trongluong += (ProcessHandler.timtrongluong_sanpham(product["Barcode"]) * qty_ord_pcs)
            sokienhang += math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(product["Barcode"]))

            sheet[f"AT{current_row}"] = ProcessHandler.timtrongluong_sanpham(product["Barcode"]) * qty_ord_pcs
            sheet[f"AU{current_row}"] = math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(product["Barcode"]))

            current_row += 1

        

        sheet[f"L{start_row}"] = f'{diengiai}'   #f'{diengiai} (Tổng trọng lượng: {ProcessHandler.format_weight_kg(trongluong)})'
        wb.save(file_path)
        print(f"✅ Đã ghi {len(products)} dòng vào '{file_path}', sheet '{sheet_name}'!")
        
        return saigia
    




    
    def write_to_dondathang_Jupviec(self,products, makhachhang, po_number, entry_date, cancle_date,STT_donhang,vendor,shipto,po):

        #taongaunhien
        """Ghi danh sách sản phẩm vào file dondathang.xlsx"""
        file_path = "dondathang.xlsx"
        sheet_name = "Don dat hang"
        
        
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]

        start_row = sheet.max_row + 1
        current_row = start_row
        # Lấy giá trị G1, đảm bảo nó là số
        
        
        print(f"📍 PO Number {po_number} ")
        print(f"📍 ")
        STT_donhang_str = f"-{po_number}"     
        hethong = 'JV-Mart'
        trongluong = 0
        



        saigia = 0
        tongtien = current_row
        diengiai = f"{hethong} {po_number}" 
        


        if shipto == "HN":
            khuvuc = "TMĐT_MB"
            kho = "TP_HN_12"
            mien = "HN"
        else:
            kho = "LA_KHOTMDT"
            khuvuc = "TMĐT_MN"
            mien = "LA"




        
            
        for product in products:
            
            sheet[f"A{current_row}"] = entry_date
            sheet[f"AV{current_row}"] = songayno_TMDT
            sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'
            
            sheet[f"C{current_row}"] = "Chưa thực hiện"
            sheet[f"D{current_row}"] = cancle_date
            sheet[f"AO{current_row}"] = po
            sheet[f"E{current_row}"] = shipto
            sheet[f"G{current_row}"] = makhachhang
            sheet[f"L{current_row}"] = diengiai
            sheet[f"Q{current_row}"] = product["Barcode"]
            sheet[f"V{current_row}"] = kho
            sheet[f"AE{current_row}"] = 8
            sheet[f"AJ{current_row}"] = khuvuc
            sheet[f"AM{current_row}"] = mien
            sheet[f"U{current_row}"] = "Không"
            sheet[f"Y{current_row}"] = 0
            sheet[f"T{current_row}"] = "Không"
            sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(product["Barcode"])
            print(f'Mã sản phẩm {product["Barcode"]}')
            print(ProcessHandler.timten_sanpham(product["Barcode"]))

            # ✅ Xử lý "Extended Cost"
            giathuctegoc = ProcessHandler.find_price_by_sku(product["Barcode"],vendor)
            if isinstance(giathuctegoc, str):
                            giathucte = giathuctegoc.replace(",", "").strip()
                            giathucte = float(giathucte) if giathucte else 0


            sheet[f"y{current_row}"] = giathucte
            sheet[f"Z{current_row}"] = f"=Y{current_row}*X{current_row}"

            # ✅ Xử lý "Qty Ord/Pcs"
            qty_ord_pcs = product["Qty Ord/Pcs"]
            if isinstance(qty_ord_pcs, str):  # Nếu là chuỗi, loại bỏ dấu phẩy
                qty_ord_pcs = float(qty_ord_pcs.replace(",", ""))
            qty_ord_pcs = int(qty_ord_pcs) if qty_ord_pcs.is_integer() else qty_ord_pcs
            sheet[f"X{current_row}"] = qty_ord_pcs

            trongluong += (ProcessHandler.timtrongluong_sanpham(product["Barcode"]) * qty_ord_pcs)
            #sokienhang += math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(product["Barcode"]))
            sheet[f"AT{current_row}"] = ProcessHandler.timtrongluong_sanpham(product["Barcode"]) * qty_ord_pcs
            sheet[f"AU{current_row}"] = math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(product["Barcode"]))

            current_row += 1

        

        sheet[f"L{start_row}"] = f'{diengiai}'   #f'{diengiai} (Tổng trọng lượng: {ProcessHandler.format_weight_kg(trongluong)})'
        wb.save(file_path)
        print(f"✅ Đã ghi {len(products)} dòng vào '{file_path}', sheet '{sheet_name}'!")
        
        return saigia
    




    def write_to_dondathang_KOC(self,products, makhachhang, po_number, entry_date, cancle_date,STT_donhang,vendor,shipto,po,ten_khach,diachi):

        
        """Ghi danh sách sản phẩm vào file dondathang.xlsx"""
        file_path = "dondathang.xlsx"
        sheet_name = "Don dat hang"
        
        
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]

        start_row = sheet.max_row + 1
        current_row = start_row
        # Lấy giá trị G1, đảm bảo nó là số
        
        
        print(f"📍 PO Number {po_number} ")
        print(f"📍 ")
        STT_donhang_str = f"-{po}"     
        hethong = 'Shopee - KOC'
        trongluong = 0
        



        saigia = 0
        tongtien = current_row
        diengiai = f"{hethong} {po_number}" 
        


        if shipto == "HN":
            khuvuc = "TMĐT_MB"
            kho = "TP_HN_12"
            mien = "HN"
        else:
            kho = "LA_KHOTMDT"
            khuvuc = "TMĐT_MN"
            mien = "LA"




        
            
        for product in products:
            
            sheet[f"A{current_row}"] = entry_date
            sheet[f"AV{current_row}"] = songayno_TMDT
            sheet[f"B{current_row}"] = f'ĐĐH{STT_donhang_str}'
            
            sheet[f"C{current_row}"] = "Chưa thực hiện"
            sheet[f"D{current_row}"] = cancle_date
            sheet[f"AO{current_row}"] = po
            sheet[f"E{current_row}"] = diachi
            sheet[f"H{current_row}"] = ten_khach
            sheet[f"I{current_row}"] = diachi
            sheet[f"G{current_row}"] = makhachhang
            sheet[f"L{current_row}"] = diengiai
            sheet[f"Q{current_row}"] = product["Barcode"]
            sheet[f"V{current_row}"] = kho
            sheet[f"AE{current_row}"] = 8
            sheet[f"AJ{current_row}"] = khuvuc
            sheet[f"AM{current_row}"] = mien
            sheet[f"U{current_row}"] = "Không"
            sheet[f"Y{current_row}"] = 0
            sheet[f"T{current_row}"] = "Không"
            sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(product["Barcode"])
            print(f'Mã sản phẩm {product["Barcode"]}')
            print(ProcessHandler.timten_sanpham(product["Barcode"]))

            # ✅ Xử lý "Extended Cost"
            


            sheet[f"y{current_row}"] = 0
            sheet[f"Z{current_row}"] = f"=Y{current_row}*X{current_row}"

            # ✅ Xử lý "Qty Ord/Pcs"
            qty_ord_pcs = product["Qty Ord/Pcs"]
            if isinstance(qty_ord_pcs, str):  # Nếu là chuỗi, loại bỏ dấu phẩy
                qty_ord_pcs = float(qty_ord_pcs.replace(",", ""))
            qty_ord_pcs = int(qty_ord_pcs) if qty_ord_pcs.is_integer() else qty_ord_pcs
            sheet[f"X{current_row}"] = qty_ord_pcs

            
            

            current_row += 1

        

        sheet[f"L{start_row}"] = f'{diengiai}'   #f'{diengiai} (Tổng trọng lượng: {ProcessHandler.format_weight_kg(trongluong)})'
        wb.save(file_path)
        print(f"✅ Đã ghi {len(products)} dòng vào '{file_path}', sheet '{sheet_name}'!")
        
        return saigia




    def write_to_dondathang_JITxls(self,products, makhachhang, po_number, entry_date, cancle_date,STT_donhang,vendor,shipto,po):

        #taongaunhien
        """Ghi danh sách sản phẩm vào file dondathang.xlsx"""
        file_path = "dondathang.xlsx"
        sheet_name = "Don dat hang"
        
        
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]

        start_row = sheet.max_row + 1
        current_row = start_row
        # Lấy giá trị G1, đảm bảo nó là số
        trongluong = 0
        
        
        print(f"📍 PO Number {po_number} ")
        print(f"📍 ")
        STT_donhang_str = f"-{po_number}"     
        hethong = 'JIT-CHOICE'
        



        saigia = 0
        tongtien = current_row
        diengiai = f"{hethong} Ngày đổ {po_number} {shipto}" 
        


        if shipto == "WH6_HN":
            khuvuc = "TMĐT_MB"
            kho = "TP_HN_12"
            mien = "HN"
        else:
            kho = "LA_KHOTMDT"
            khuvuc = "TMĐT_MN"
            mien = "LA"
            




        
            
        for product in products:
            
            sheet[f"A{current_row}"] = entry_date
            sheet[f"AV{current_row}"] = songayno_TMDT
            sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'
                   
            sheet[f"C{current_row}"] = "Chưa thực hiện"
            sheet[f"D{current_row}"] = cancle_date
            sheet[f"AO{current_row}"] = po
            sheet[f"E{current_row}"] = shipto
            sheet[f"G{current_row}"] = makhachhang
            sheet[f"L{current_row}"] = diengiai
            sheet[f"Q{current_row}"] = product["Barcode"]
            sheet[f"V{current_row}"] = kho
            sheet[f"AE{current_row}"] = 8
            sheet[f"AJ{current_row}"] = khuvuc
            sheet[f"AM{current_row}"] = mien
            sheet[f"U{current_row}"] = "Không"
            sheet[f"Y{current_row}"] = 0
            sheet[f"T{current_row}"] = "Không"
            sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(product["Barcode"])

            # ✅ Xử lý "Extended Cost"
            giathuctegoc = ProcessHandler.find_price_by_sku(product["Barcode"],vendor)
            if isinstance(giathuctegoc, str):
                            giathucte = giathuctegoc.replace(",", "").strip()
                            giathucte = float(giathucte) if giathucte else 0


            sheet[f"y{current_row}"] = giathucte
            sheet[f"Z{current_row}"] = f"=Y{current_row}*X{current_row}"

            # ✅ Xử lý "Qty Ord/Pcs"
            qty_ord_pcs = product["Qty Ord/Pcs"]
            if isinstance(qty_ord_pcs, str):  # Nếu là chuỗi, loại bỏ dấu phẩy
                qty_ord_pcs = float(qty_ord_pcs.replace(",", ""))
            qty_ord_pcs = int(qty_ord_pcs) if qty_ord_pcs.is_integer() else qty_ord_pcs
            sheet[f"X{current_row}"] = qty_ord_pcs
            trongluong += (ProcessHandler.timtrongluong_sanpham(product["Barcode"]) * qty_ord_pcs)
            sokienhang += math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(product["Barcode"]))
            sheet[f"AU{current_row}"] = math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(product["Barcode"]))
            sheet[f"AT{current_row}"] = ProcessHandler.timtrongluong_sanpham(product["Barcode"]) * qty_ord_pcs
            current_row += 1

        sheet[f"L{start_row}"] = f'{diengiai} (Tổng trọng lượng: {ProcessHandler.format_weight_kg(trongluong)})'
        wb.save(file_path)
        print(f"✅ Đã ghi {len(products)} dòng vào '{file_path}', sheet '{sheet_name}'!")
        
        return saigia
    




    def write_to_dondathang_TMDT(self, products, makhachhang, entry_date, cancle_date):
        """
        products: list hóa đơn
        mỗi hóa đơn có dong_chi_tiet
        """

        file_path = "dondathang.xlsx"
        sheet_name = "Don dat hang"

        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]

        total_written = 0

        kho_haravan = products[0].get("kho_haravan", "").strip()

        if kho_haravan == "Miền Nam - Kho mặc định":
            khuvuc = "TMĐT_MN"
            kho = "LA_KHOTMDT"
            mien = "LA"
        else:
            khuvuc = "TMĐT_MB"
            kho = "TP_HN_12"
            mien = "HN"

        # ===============================
        # DUYỆT TỪNG HÓA ĐƠN
        # ===============================
        for hoa_don in products:
            kenh_ban = hoa_don.get("kenh_ban", "")
            ma_donhang = hoa_don.get("ma_hoadon", "")
            ma_vandon = hoa_don.get("ma_vandon", "")
            trang_thai = hoa_don.get("trang_thai", "")

            hethong = "TMĐT-TikTok" if ma_donhang.startswith("HDTTS") else "TMĐT-Shopee"
            diengiai = f"{hethong}-{kenh_ban} - {ma_donhang} - Ngày đổ {entry_date} - {mien}"

            self.log_signal.emit(
                f'Bắt đầu xử lý đơn <b><span style="color: green;">{ma_donhang}</span></b>'
            )

            # ===============================
            # DUYỆT TỪNG DÒNG CHI TIẾT
            # ===============================
            for dong in hoa_don.get("dong_chi_tiet", []):

                thanh_tien_dong = dong.get("thanh_tien", 0)
                san_pham = dong.get("san_pham", [])

                if not san_pham:
                    continue

                # 🔹 tổng SL trong DÒNG
                tong_sl_dong = sum(sp.get("sl_tp", 0) for sp in san_pham)

                if tong_sl_dong <= 0:
                    continue

                # 🔹 đơn giá trung bình THEO DÒNG
                don_gia_tb = thanh_tien_dong / tong_sl_dong
                hang_km = "Không"
                
                if don_gia_tb == 0:
                    hang_km = "Có"

                # ===============================
                # DUYỆT SẢN PHẨM TRONG DÒNG
                # ===============================
                for sp in san_pham:
                    ma_tp = sp.get("ma_tp")
                    sl_tp = sp.get("sl_tp")

                    if not ma_tp or not sl_tp:
                        continue

                    current_row = sheet.max_row + 1

                    sheet[f"A{current_row}"] = entry_date
                    sheet[f"AV{current_row}"] = songayno_TMDT
                    sheet[f"B{current_row}"] = f"ĐĐH{hethong}-{ma_donhang}"
                    sheet[f"C{current_row}"] = "Chưa thực hiện"
                    sheet[f"D{current_row}"] = cancle_date
                    sheet[f"E{current_row}"] = mien
                    sheet[f"G{current_row}"] = makhachhang
                    sheet[f"L{current_row}"] = diengiai
                    sheet[f"AO{current_row}"] = ma_donhang
                    sheet[f"Q{current_row}"] = ma_tp
                    sheet[f"V{current_row}"] = kho
                    sheet[f"AE{current_row}"] = 8
                    sheet[f"AJ{current_row}"] = khuvuc
                    sheet[f"AM{current_row}"] = mien
                    sheet[f"AP{current_row}"] = ma_vandon
                    sheet[f"AQ{current_row}"] = trang_thai
                    sheet[f"U{current_row}"] = hang_km
                    sheet[f"T{current_row}"] = "Không"

                    # ✅ Tên sản phẩm
                    try:
                        ten_sp = ProcessHandler.timten_sanpham(ma_tp)
                        sheet[f"S{current_row}"] = ten_sp
                        self.log_signal.emit(
                            f'Tên SP <b><span style="color: green;">{ten_sp}</span></b>'
                        )
                    except Exception:
                        sheet[f"S{current_row}"] = ""

                    # ✅ Số lượng
                    sheet[f"X{current_row}"] = sl_tp

                    # ✅ Đơn giá (chưa VAT)
                    sheet[f"Y{current_row}"] = don_gia_tb / 1.08

                    # ✅ Thành tiền
                    sheet[f"Z{current_row}"] = f"=Y{current_row}*X{current_row}"

                    total_written += 1

        wb.save(file_path)
        print(f"✅ Đã ghi {total_written} dòng vào '{file_path}', sheet '{sheet_name}'")
        return







    


    
    def write_to_dondathang_CNHCM(self,products, makhachhang, po_number, entry_date, cancle_date,STT_donhang,vendor,diachigiaohang):

        #taongaunhien
        """Ghi danh sách sản phẩm vào file dondathang.xlsx"""
        file_path = "dondathang.xlsx"
        sheet_name = "Don dat hang"
        
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]

        start_row = sheet.max_row + 1
        current_row = start_row
        trongluong = 0
        songayno_GT = 30
        STT_donhang_str = f"-{po_number}"
        diengiai = f"{vendor} PO {po_number.replace("-", "/")}"
        if makhachhang[:2] == "MB":
            khuvuc = "Sỉ_MN"
            kho = "TP_HN_12"
            mien = "HN"
        else:
            kho = "LA_TP"
            khuvuc = "Sỉ_MN"
            mien = "LA"


        
        for product in products:

            sheet[f"A{current_row}"] = entry_date
            sheet[f"AV{current_row}"] = songayno_GT

            sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'

            sheet[f"C{current_row}"] = "Chưa thực hiện"
            sheet[f"D{current_row}"] = cancle_date
            sheet[f"G{current_row}"] = makhachhang
            sheet[f"L{current_row}"] = diengiai
            sheet[f"Q{current_row}"] = product["Mã hàng"]
            sheet[f"V{current_row}"] = kho
            sheet[f"AE{current_row}"] = 0
            sheet[f"AJ{current_row}"] = khuvuc
            sheet[f"AM{current_row}"] = mien
            sheet[f"U{current_row}"] = "Không"
            sheet[f"E{current_row}"] = diachigiaohang
            sheet[f"W{current_row}"] = product["Đơn vị tính"]
            sheet[f"T{current_row}"] = "Không"
            sheet[f"S{current_row}"] = product["Tên sản phẩm"]
            sheet[f"Z{current_row}"] = f"=Y{current_row}*X{current_row}"
            sheet[f"X{current_row}"] = product["Số lượng"]
            sheet[f"Y{current_row}"] = 0
            self.log_signal.emit(f'Đã thêm {product["Mã hàng"]} - {product["Tên sản phẩm"]} ')

            

            current_row += 1

        wb.save(file_path)
        print(f"✅ Đã ghi {len(products)} dòng vào '{file_path}', sheet '{sheet_name}'!")
        return 0




    def write_to_dondathang_lotte(self,products, makhachhang, po_number, entry_date, cancle_date,STT_donhang,vendor,diachigiaohang,url):
        now = datetime.now()
        now_str = now.strftime("%d/%m/%Y %H:%M:%S")
        vendor = vendor.upper()
        """Ghi danh sách sản phẩm vào file dondathang.xlsx"""
        file_path = "dondathang.xlsx"
        sheet_name = "Don dat hang"
        
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]

        start_row = sheet.max_row + 1
        current_row = start_row
        # Lấy giá trị G1, đảm bảo nó là số
        trongluong = 0
        
        print(f"📍 PO Number {po_number} ")
        STT_donhang_str = f"-{po_number}"     
        sku_mapping = ProcessHandler.load_sku_mapping()  # Load mapping SKU
        products = ProcessHandler.replace_sku_numbers(products, sku_mapping)  # Thay SKU
        saigia = 0
        tongtien = 0.0
        sokienhang = 0
        diengiai = f"{vendor} PO{po_number}"
        if makhachhang[:2] == "MB":
            khuvuc = "MT_MB"
            kho = "TP_HN_12"
            mien = "HN"
        else:
            kho = "LA_TP"
            khuvuc = "MT_MN"
            mien = "LA"


        ProcessHandler.ghi_message(f"[PO: {po_number}")
        ProcessHandler.ghi_message(f"vendor: {vendor}")
        ProcessHandler.ghi_message(f"store: {diachigiaohang}")
        ProcessHandler.ghi_message(f"Mã Khách hàng: {makhachhang}")
        ProcessHandler.ghi_message(f"start_time: {now_str}")
        if url:
            ProcessHandler.ghi_message(f"url: {url}")




        sheet[f"A{current_row}"] = entry_date
        sheet[f"AV{current_row}"] = songayno_MT
    
        sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'
        
        sheet[f"C{current_row}"] = "Chưa thực hiện"
        sheet[f"D{current_row}"] = cancle_date
        sheet[f"G{current_row}"] = makhachhang
        sheet[f"L{current_row}"] = diengiai
        sheet[f"V{current_row}"] = kho
        sheet[f"AE{current_row}"] = 8
        sheet[f"AJ{current_row}"] = khuvuc
        sheet[f"AM{current_row}"] = mien
        sheet[f"U{current_row}"] = "Không"
        sheet[f"Z{current_row}"] = 0
        sheet[f"S{current_row}"] = f"{vendor} PO{po_number}"
        sheet[f"T{current_row}"] = "Có"
        sheet[f"X{current_row}"] = 0
        sheet[f"Y{current_row}"] = 0
        sheet[f"E{current_row}"] = diachigiaohang
        


        
        current_row += 1

        for product in products:
            
            sheet[f"A{current_row}"] = entry_date
            sheet[f"AV{current_row}"] = songayno_MT

            sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'

            sheet[f"C{current_row}"] = "Chưa thực hiện"
            sheet[f"D{current_row}"] = cancle_date
            sheet[f"G{current_row}"] = makhachhang
            sheet[f"L{current_row}"] = diengiai
            sheet[f"Q{current_row}"] = product["Barcode"]
            sheet[f"V{current_row}"] = kho
            sheet[f"AE{current_row}"] = 8
            sheet[f"AJ{current_row}"] = khuvuc
            sheet[f"AM{current_row}"] = mien
            sheet[f"U{current_row}"] = "Không"
            sheet[f"E{current_row}"] = diachigiaohang

            sheet[f"T{current_row}"] = "Không"
            sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(product["Barcode"])

            # ✅ Xử lý "Extended Cost"
            extended_cost = product["Total Price"]
            if isinstance(extended_cost, str):  # Nếu là chuỗi, loại bỏ dấu phẩy
                extended_cost = float(extended_cost.replace(",", ""))
            extended_cost = int(extended_cost) if extended_cost.is_integer() else extended_cost
            sheet[f"Z{current_row}"] = f"=Y{current_row}*X{current_row}"

            # ✅ Xử lý "Qty Ord/Pcs"
            qty_ord_pcs = product["Qty-Box"] * product["Box Quantity"]
            if isinstance(qty_ord_pcs, str):  # Nếu là chuỗi, loại bỏ dấu phẩy
                qty_ord_pcs = float(qty_ord_pcs.replace(",", ""))
            qty_ord_pcs = int(qty_ord_pcs) if qty_ord_pcs.is_integer() else qty_ord_pcs
            sheet[f"X{current_row}"] = qty_ord_pcs
            trongluong += (ProcessHandler.timtrongluong_sanpham(product["Barcode"]) * qty_ord_pcs)
            sheet[f"AT{current_row}"] = ProcessHandler.timtrongluong_sanpham(product["Barcode"]) * qty_ord_pcs
            sokienhang += math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(product["Barcode"]))

            sheet[f"AU{current_row}"] = math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(product["Barcode"]))


            # ✅ Tính giá hóa đơn
            giahoadon = extended_cost / qty_ord_pcs, 2

           
            giathuctegoc = ProcessHandler.find_price_by_sku(product["Barcode"],vendor)
            giathucte = giathuctegoc
            # 🔍 **Tìm giá trị trong CTKM.xlsx theo SKU & thời gian**
            
            results = ProcessHandler.find_all_promotions_by_sku_and_time(product["Barcode"], entry_date,vendor)
            slkhuyenmai = len(results)
            print(f'Giá thực tế: {giathucte}')
            print(f'🔎 Số lượng CTKM khớp: {slkhuyenmai}')
            khuyenmai = ''



            giakhop = False  # Cờ kiểm tra khớp giá

            for col, value in results:
                print(f'🧪 Thử CTKM: {col} → {value}')
                
                if value:
                    khuyenmai = value
                    discount = ProcessHandler.extract_discount(value)
                    print(f'🔻 Giảm giá: {discount} ({type(discount)})')

                    if discount != 0:
                        if isinstance(discount, str):
                            discount = discount.replace("%", "").strip()
                            discount = float(discount) if discount else 0

                        if isinstance(giathuctegoc, str):
                            giathucte = giathuctegoc.replace(",", "").strip()
                            giathucte = float(giathucte) if giathucte else 0

                        giathucte = giathucte - (giathucte * float(discount) / 100)
                    else:
                        
                        # Nếu không có giảm giá, kiểm tra lại giá trị giathucte
                        if isinstance(giathuctegoc, str):
                            giathucte = giathuctegoc.replace(",", "").strip()
                            giathucte = float(giathucte) if giathucte else 0



                    giahoadon = float(giahoadon[0]) if isinstance(giahoadon, tuple) else float(giahoadon)
                    if isinstance(giathucte, str):
                        giathucte = giathucte.replace(",", "").strip()
                        giathucte = float(giathucte) if giathucte else 0

                    if math.isclose(giahoadon, giathucte, rel_tol=1e-4):
                        sheet[f"Y{current_row}"] = giathucte
                        self.log_signal.emit(
                        f'🎉 Mã hàng <b><span style="color: blue;">{product["Barcode"]}</span></b> '
                        f'có CTKM <b><span style="color: green;">{value}</span></b> từ  🛒')
                        self.log_signal.emit(
                            f'✅ Mã hàng <b><span style="color: blue;">{product["Barcode"]}</span></b> '
                            f'có giá chính xác 🎯: <span style="color: green;"><b>{giahoadon}</b></span>'
                        )
                        
                        
                        giakhop = True
                        break  # ✅ THOÁT VÒNG LẶP khi đã khớp giá
                    

            if not results:
                khuyenmai = ''
                giahoadon = float(giahoadon[0]) if isinstance(giahoadon, tuple) else float(giahoadon)
                if isinstance(giathucte, str):
                    giathucte = giathucte.replace(",", "").strip()
                    giathucte = float(giathucte) if giathucte else 0

                if math.isclose(giahoadon, giathucte, rel_tol=1e-4):
                    sheet[f"Y{current_row}"] = giathucte
                    self.log_signal.emit(
                    f'🎉 Mã hàng <b><span style="color: blue;">{product["Barcode"]}</span></b> '
                    f'có CTKM <b><span style="color: green;">{khuyenmai}</span></b> từ  🛒')
                    self.log_signal.emit(
                        f'✅ Mã hàng <b><span style="color: blue;">{product["Barcode"]}</span></b> '
                        f'có giá chính xác 🎯: <span style="color: green;"><b>{giahoadon}</b></span>'
                    )
                    giakhop = True



            # ❌ Sau khi thử hết, không khớp giá nào:
            if not giakhop:
                sheet[f"Y{current_row}"] = giathucte
                red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                sheet[f"Y{current_row}"].fill = red_fill
                comment_text = f"Kiểm tra lại giá mã này! - Giá hóa đơn: {giahoadon} - Chênh lệch: {giahoadon - giathucte}"
                comment = Comment(comment_text, "System")
                sheet[f"Y{current_row}"].comment = comment

                self.log_signal.emit(
                    f"⚠️ Mã hàng <b>{product["Barcode"]}</b> sai giá! 🛑 "
                    f"Hóa đơn: <b><span style='color: red;'>{giahoadon}</span></b> – "
                    f"Giá hệ thống: <b><span style='color: green;'>{giathucte}</span></b>"
                )

                ProcessHandler.ghi_message(f"text: Mã hàng {product["Barcode"]} - {ProcessHandler.timten_sanpham(product["Barcode"])}, Giá trên PO {giahoadon:,}, Giá đúng: {giathucte:,}. Chênh lệch {(giahoadon - giathucte):,.0f}")

                saigia += 1

            tongtien += giathucte * qty_ord_pcs
                

            # 🔍 Kiểm tra SKU có trong 'SanPham'
            kiemtra = ProcessHandler.check_value_in_sanpham(khuyenmai)
            sheet[f"AQ{current_row}"] = khuyenmai
            match = re.search(r"(\d+)\s*\+\s*1", khuyenmai)  # Tìm số đầu tiên trong biểu thức X+1
            if match:
                x = int(match.group(1))  # Chuyển đổi thành số nguyên
                kiemtra = kiemtra or product["Barcode"]
                
                
                if x >= 2:  # Chỉ chia khi x >= 2, còn 1+1 thì giữ nguyên
                    qty_ord_pcs = math.floor(qty_ord_pcs / x)  # Chia rồi làm tròn xuống

            
               
            if kiemtra:
                if isinstance(kiemtra, list):  # Nếu có nhiều mã SKU, nối thành chuỗi
                    kiemtra = ", ".join(kiemtra)
                cachbokem = ProcessHandler.laycachbo_khuyenmai(value)
                if cachbokem:
                    sheet[f"AO{current_row}"] = cachbokem
                    cachbokem_lower = cachbokem.lower()
                    if "bó kèm" in cachbokem_lower or "quấn kèm" in cachbokem_lower:
                        sheet[f"AP{current_row}"] = f'{ProcessHandler.layduoi_mahang(product["Barcode"])}_{ProcessHandler.layduoi_mahang(kiemtra)}_1'
                        sheet[f"AP{current_row + 1}"] = f'{ProcessHandler.layduoi_mahang(product["Barcode"])}_{ProcessHandler.layduoi_mahang(kiemtra)}_1'
                    
                else:
                    sheet[f"AO{current_row}"] = f'KM Giao Rời - Không Che Barcode'
                current_row += 1
               
                self.log_signal.emit(
                    f'🎁 Mã hàng <b><span style="color: blue;">{product["Barcode"]}</span></b> '
                    f'đã thêm hàng khuyến mãi <b><span style="color: green;">{kiemtra}</span></b> tặng kèm! 🎉'
                )
                sheet[f"A{current_row}"] = entry_date
                sheet[f"AV{current_row}"] = songayno_MT
                sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'

                sheet[f"C{current_row}"] = "Chưa thực hiện"
                sheet[f"D{current_row}"] = cancle_date
                sheet[f"G{current_row}"] = makhachhang
                sheet[f"L{current_row}"] = diengiai
                sheet[f"Q{current_row}"] = kiemtra
                sheet[f"X{current_row}"] = qty_ord_pcs
                sheet[f"V{current_row}"] = kho
                sheet[f"Y{current_row}"] = 0
                sheet[f"Z{current_row}"] = 0
                sheet[f"U{current_row}"] = "Có"
                sheet[f"AJ{current_row}"] = khuvuc
                sheet[f"AE{current_row}"] = 8
                sheet[f"AM{current_row}"] = mien
                sheet[f"T{current_row}"] = "Không"
                sheet[f"E{current_row}"] = diachigiaohang
                sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(kiemtra)
                trongluong += (ProcessHandler.timtrongluong_sanpham(kiemtra) * qty_ord_pcs)
                sokienhang += math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(kiemtra))
                sheet[f"AT{current_row}"] = ProcessHandler.timtrongluong_sanpham(kiemtra) * qty_ord_pcs
                sheet[f"AU{current_row}"] = math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(kiemtra))

            sheet["G1"] = STT_donhang
            current_row += 1
        


        kmhoadon = ProcessHandler.find_all_promotions_by_sku_and_time("Hóa Đơn", entry_date,vendor)

        
        print(kmhoadon)
        if kmhoadon:
            kmhoadon = str(kmhoadon[0][1])
            kiemtra = ProcessHandler.check_value_in_sanpham(kmhoadon) 
            sheet[f"AQ{current_row}"] = kmhoadon

            sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'
            print(tongtien)
            print(ProcessHandler.tachtien_khuyenmai(kmhoadon))
            soluongkm = math.floor(tongtien / ProcessHandler.tachtien_khuyenmai(kmhoadon))
            sheet[f"A{current_row}"] = entry_date
            sheet[f"AV{current_row}"] = songayno_MT
            sheet[f"C{current_row}"] = "Chưa thực hiện"
            sheet[f"D{current_row}"] = cancle_date
            sheet[f"E{current_row}"] = diachigiaohang
            sheet[f"G{current_row}"] = makhachhang
            sheet[f"L{current_row}"] = diengiai
            sheet[f"Q{current_row}"] = kiemtra[0]
            sheet[f"X{current_row}"] = soluongkm
            sheet[f"V{current_row}"] = kho
            sheet[f"Y{current_row}"] = 0
            sheet[f"Z{current_row}"] = 0
            sheet[f"U{current_row}"] = "Có"
            sheet[f"AJ{current_row}"] = khuvuc
            sheet[f"AE{current_row}"] = 8
            sheet[f"AM{current_row}"] = mien
            sheet[f"T{current_row}"] = "Không"
            sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(kiemtra[0])
            trongluong += (ProcessHandler.timtrongluong_sanpham(kiemtra[0]) * soluongkm)
            sheet[f"AU{current_row}"] = math.ceil(soluongkm / ProcessHandler.timquycach_sanpham(kiemtra[0]))
            sokienhang += math.ceil(soluongkm / ProcessHandler.timquycach_sanpham(kiemtra[0]))

            sheet[f"AT{current_row}"] = ProcessHandler.timtrongluong_sanpham(kiemtra[0]) * soluongkm

            cachbokem = ProcessHandler.laycachbo_khuyenmai(kmhoadon)
            if cachbokem:
                sheet[f"AO{current_row}"] = cachbokem
            else:
                sheet[f"AO{current_row}"] = f'KM Bó Kèm - Che Barcode'

            self.log_signal.emit(f'🎉 Đơn hàng có chương trình khuyến mãi <b><span style="color: green;">{kmhoadon}</span></b> 🛒')
            self.log_signal.emit(f'Tổng tiền đơn hàng: <b><span style="color: green;">{tongtien} </span></b>\nSố lượng tặng: <b><span style="color: green;">{soluongkm} </span></b> 🛒')
        
        sheet["G1"] = STT_donhang
        current_row += 1


        ProcessHandler.ghi_message(f"tong_tien: {tongtien:,.0f}")
        ProcessHandler.ghi_message(f"tong_trongluong: {ProcessHandler.format_weight_kg(trongluong)}")
        ProcessHandler.ghi_message(f"tong_kienhang: {sokienhang}")
        ProcessHandler.ghi_message(f"sai_gia: {saigia}")
        now2 = datetime.now()
        now_str = now2.strftime("%d/%m/%Y %H:%M:%S")
        ProcessHandler.ghi_message(f"end_time: {now_str}")
        time = now2 - now
        ProcessHandler.ghi_message(f"time: {ProcessHandler.format_timedelta(time)}]")  

        
        sheet[f"L{start_row}"] = f'{diengiai} (Tổng trọng lượng: {ProcessHandler.format_weight_kg(trongluong)})'
        # Lưu lại file Excel
        wb.save(file_path)
        print(f"✅ Đã ghi {len(products)} dòng vào '{file_path}', sheet '{sheet_name}'!")
        return saigia
    





    

    


    def write_to_dondathang_satra(self,products, makhachhang, po_number, entry_date, cancle_date,STT_donhang,vendor,diachigiao,url):
        now = datetime.now()
        now_str = now.strftime("%d/%m/%Y %H:%M:%S")
        vendor = vendor.upper()
        """Ghi danh sách sản phẩm vào file dondathang.xlsx"""
        file_path = "dondathang.xlsx"
        sheet_name = "Don dat hang"
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]

        start_row = sheet.max_row + 1
        current_row = start_row
        # Lấy giá trị G1, đảm bảo nó là số
        print(f"📍 PO Number {po_number} ")
        print(f"📍 Địa chỉ giao {diachigiao} ")

        STT_donhang_str = f"-{po_number}"     
        saigia = 0
        tongtien = 0.0
        trongluong = 0
        sokienhang = 0

        ProcessHandler.ghi_message(f"[PO: {po_number}")
        ProcessHandler.ghi_message(f"store: {diachigiao}")
        ProcessHandler.ghi_message(f"vendor: {vendor}")
        ProcessHandler.ghi_message(f"Mã Khách hàng: {makhachhang}")
        ProcessHandler.ghi_message(f"start_time: {now_str}")
        if url:
            ProcessHandler.ghi_message(f"url: {url}")

        if makhachhang[:2] == "MB":
            khuvuc = "MT_MB"
            kho = "TP_HN_12"
            mien = "HN"
        else:
            kho = "LA_TP"
            khuvuc = "MT_MN"
            mien = "LA"
        khonggiaothu7 = ''
        if makhachhang == "MN_MT_stph":
            khonggiaothu7 = '- Không giao thứ 7'

        diengiai = f"{vendor} {po_number} {khonggiaothu7}"


        sheet[f"A{current_row}"] = entry_date
        sheet[f"AV{current_row}"] = songayno_MT
        sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'

        sheet[f"C{current_row}"] = "Chưa thực hiện"
        sheet[f"D{current_row}"] = cancle_date
        sheet[f"G{current_row}"] = makhachhang
        sheet[f"L{current_row}"] = diengiai
        sheet[f"V{current_row}"] = kho
        sheet[f"AE{current_row}"] = 8
        sheet[f"AJ{current_row}"] = khuvuc
        sheet[f"AM{current_row}"] = mien
        sheet[f"U{current_row}"] = "Không"
        sheet[f"Z{current_row}"] = 0
        sheet[f"S{current_row}"] = f"{vendor} {po_number}"
        sheet[f"T{current_row}"] = "Có"
        sheet[f"X{current_row}"] = 0
        sheet[f"Y{current_row}"] = 0
        sheet[f"E{current_row}"] = diachigiao
        current_row += 1

        for product in products:
            
            barcode = product["Barcode"]
            soluong = float(product["OU Qty"])
            
            dongia = float(str(product["Total Price"]).replace(',', '.'))
            #dongia = 159200
           
            sku_mapping = ProcessHandler.load_sku_mapping()  # Load mapping SKU
            products = ProcessHandler.replace_sku_numbers(products, sku_mapping)  # Thay SKU
            sheet[f"A{current_row}"] = entry_date
            sheet[f"AV{current_row}"] = songayno_MT

            sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'
            sheet[f"C{current_row}"] = "Chưa thực hiện"
            sheet[f"D{current_row}"] = cancle_date
            sheet[f"G{current_row}"] = makhachhang
            sheet[f"L{current_row}"] = diengiai
            sheet[f"Q{current_row}"] = product["Barcode"]
            sheet[f"V{current_row}"] = kho
            sheet[f"AE{current_row}"] = 8
            sheet[f"AJ{current_row}"] = khuvuc
            sheet[f"AM{current_row}"] = mien
            sheet[f"U{current_row}"] = "Không"
            sheet[f"E{current_row}"] = diachigiao
            sheet[f"T{current_row}"] = "Không"
            sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(product["Barcode"])

            # ✅ Xử lý "Extended Cost"
            
            sheet[f"Z{current_row}"] = f"=Y{current_row}*X{current_row}"

            # ✅ Xử lý "Qty Ord/Pcs"
            qty_ord_pcs = product["OU Qty"]
            if isinstance(qty_ord_pcs, str):  # Nếu là chuỗi, loại bỏ dấu phẩy
                qty_ord_pcs = float(qty_ord_pcs.replace(",", ""))
            qty_ord_pcs = int(qty_ord_pcs) if qty_ord_pcs.is_integer() else qty_ord_pcs
            sheet[f"X{current_row}"] = qty_ord_pcs
            
            print(product["Barcode"])
            trongluong += float(ProcessHandler.timtrongluong_sanpham(product["Barcode"])) * float(qty_ord_pcs)

            sokienhang += math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(product["Barcode"]))
            sheet[f"AU{current_row}"] = math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(product["Barcode"]))

            sheet[f"AT{current_row}"] = ProcessHandler.timtrongluong_sanpham(product["Barcode"]) * qty_ord_pcs


            # ✅ Tính giá hóa đơn
            giahoadon = dongia
            print(f'Giá hóa đơn {giahoadon}')


            giathuctegoc = ProcessHandler.find_price_by_sku(product["Barcode"],vendor)
            giathucte = giathuctegoc
            
            # 🔍 **Tìm giá trị trong CTKM.xlsx theo SKU & thời gian**
            
            results = ProcessHandler.find_all_promotions_by_sku_and_time(product["Barcode"], entry_date,vendor)
            slkhuyenmai = len(results)
            print(f'Giá thực tế: {giathucte}')
            print(f'🔎 Số lượng CTKM khớp: {slkhuyenmai}')
            khuyenmai = ''
            value =''


            giakhop = False  # Cờ kiểm tra khớp giá

            for col, value in results:
                print(f'🧪 Thử CTKM: {col} → {value}')
                
                if value:
                    khuyenmai = value
                    discount = ProcessHandler.extract_discount(value)
                    print(f'🔻 Giảm giá: {discount} ({type(discount)})')

                    if discount != 0:
                        if isinstance(discount, str):
                            discount = discount.replace("%", "").strip()
                            discount = float(discount) if discount else 0

                        if isinstance(giathuctegoc, str):
                            giathucte = giathuctegoc.replace(",", "").strip()
                            giathucte = float(giathucte) if giathucte else 0

                        giathucte = giathucte - (giathucte * float(discount) / 100)
                    else:
                        giathucte = giathucte



                    giahoadon = float(giahoadon[0]) if isinstance(giahoadon, tuple) else float(giahoadon)
                    if isinstance(giathucte, str):
                        giathucte = giathucte.replace(",", "").strip()
                        giathucte = float(giathucte) if giathucte else 0

                    if math.isclose(giahoadon, giathucte, rel_tol=1e-4):
                        sheet[f"Y{current_row}"] = giahoadon
                        self.log_signal.emit(
                        f'🎉 Mã hàng <b><span style="color: blue;">{product["Barcode"]}</span></b> '
                        f'có CTKM <b><span style="color: green;">{value}</span></b>  🛒')
                        self.log_signal.emit(
                            f'✅ Mã hàng <b><span style="color: blue;">{product["Barcode"]}</span></b> '
                            f'có giá chính xác 🎯: <span style="color: green;"><b>{giahoadon}</b></span>'
                        )
                        
                        giakhop = True
                        
                        break  # ✅ THOÁT VÒNG LẶP khi đã khớp giá

            if not results:
                khuyenmai = ''
                giahoadon = float(giahoadon[0]) if isinstance(giahoadon, tuple) else float(giahoadon)




                if isinstance(giathucte, str):
                    giathucte = giathucte.replace(",", "").strip()
                    giathucte = float(giathucte) if giathucte else 0

                
                    if math.isclose(giahoadon, giathucte, rel_tol=1e-4):
                        sheet[f"Y{current_row}"] = giahoadon
                        self.log_signal.emit(
                        f'🎉 Mã hàng <b><span style="color: blue;">{product["Barcode"]}</span></b> '
                        f'có CTKM <b><span style="color: green;">{value}</span></b>  🛒')
                        self.log_signal.emit(
                            f'✅ Mã hàng <b><span style="color: blue;">{product["Barcode"]}</span></b> '
                            f'có giá chính xác 🎯: <span style="color: green;"><b>{giahoadon}</b></span>'
                        )
                        
                        giakhop = True



            # ❌ Sau khi thử hết, không khớp giá nào:
            if not giakhop:
                sheet[f"Y{current_row}"] = giathucte
                red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                sheet[f"Y{current_row}"].fill = red_fill
                comment_text = f"Kiểm tra lại giá mã này! - Giá hóa đơn: {giahoadon} - Chênh lệch: {giahoadon - giathucte}"
                comment = Comment(comment_text, "System")
                sheet[f"Y{current_row}"].comment = comment

                self.log_signal.emit(
                    f"⚠️ Mã hàng <b>{product["Barcode"]}</b> sai giá! 🛑 "
                    f"Hóa đơn: <b><span style='color: red;'>{giahoadon}</span></b> – "
                    f"Giá hệ thống: <b><span style='color: green;'>{giathucte}</span></b>"
                )
                ProcessHandler.ghi_message(f"text: Mã hàng {product["Barcode"]} - {ProcessHandler.timten_sanpham(product["Barcode"])}, Giá trên PO {giahoadon:,}, Giá đúng: {giathucte:,}. Chênh lệch {(giahoadon - giathucte):,.0f}")

                saigia += 1

            tongtien += giathucte * qty_ord_pcs
            sheet[f"AQ{current_row}"] = khuyenmai
            
            nhieuCtkm = khuyenmai.split('|')

            for i, hangkm in enumerate(nhieuCtkm):
                print(hangkm)
                kiemtra = ProcessHandler.check_value_in_sanpham(hangkm)
                match = re.search(r"(\d+)\s*\+\s*1", hangkm)  # Tìm số đầu tiên trong biểu thức X+1
                sheet[f"AQ{current_row}"] = khuyenmai
                
                
                    
                    
                if match:
                    x = int(match.group(1))  # Chuyển đổi thành số nguyên
                    kiemtra = kiemtra or product["Barcode"]
                    
                    if x >= 2:  # Chỉ chia khi x >= 2, còn 1+1 thì giữ nguyên
                        qty_ord_pcs = math.floor(qty_ord_pcs / x)  # Chia rồi làm tròn xuống
                
                if kiemtra:
                    if isinstance(kiemtra, list):  # Nếu có nhiều mã SKU, nối thành chuỗi
                        kiemtra = ", ".join(kiemtra)
                    cachbokem = ProcessHandler.laycachbo_khuyenmai(hangkm) or "KM Bó Kèm - Che Barcode"
                    if cachbokem:
                        if i == 0:
                            sheet[f"AO{current_row}"] = cachbokem
                            



                            cachbokem_lower = cachbokem.lower()
                            if "bó kèm" in cachbokem_lower or "quấn kèm" in cachbokem_lower:
                                sheet[f"AP{current_row}"] = f'{ProcessHandler.layduoi_mahang(product["Barcode"])}_{ProcessHandler.layduoi_mahang(kiemtra)}_1'
                                sheet[f"AP{current_row + 1}"] = f'{ProcessHandler.layduoi_mahang(product["Barcode"])}_{ProcessHandler.layduoi_mahang(kiemtra)}_1'
                        elif i > 0:
                            sheet[f"AO{current_row + 1}"] = cachbokem
                            

                            cachbokem_lower = cachbokem.lower()
                            if "bó kèm" in cachbokem_lower or "quấn kèm" in cachbokem_lower:
                                sheet[f"AP{current_row + 1}"] = f'{ProcessHandler.layduoi_mahang(product["Barcode"])}_{ProcessHandler.layduoi_mahang(kiemtra)}_1'
                             




                    self.log_signal.emit(
        f'🎁 Mã hàng <b><span style="color: blue;">{product["Barcode"]}</span></b> '
        f'đã thêm hàng khuyến mãi <b><span style="color: green;">{kiemtra}</span></b> tặng kèm! 🎉'
        )
                    current_row += 1
                    sheet[f"A{current_row}"] = entry_date
                    sheet[f"AV{current_row}"] = songayno_MT

                    sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'
                    sheet[f"C{current_row}"] = "Chưa thực hiện"
                    sheet[f"D{current_row}"] = cancle_date
                    sheet[f"G{current_row}"] = makhachhang
                    sheet[f"L{current_row}"] = diengiai
                    sheet[f"Q{current_row}"] = kiemtra
                    sheet[f"X{current_row}"] = qty_ord_pcs
                    sheet[f"V{current_row}"] = kho
                    sheet[f"Y{current_row}"] = 0
                    sheet[f"Z{current_row}"] = 0
                    sheet[f"U{current_row}"] = "Có"
                    sheet[f"AJ{current_row}"] = khuvuc
                    sheet[f"AE{current_row}"] = 8
                    sheet[f"AM{current_row}"] = mien
                    sheet[f"T{current_row}"] = "Không"
                    sheet[f"E{current_row}"] = diachigiao
                    sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(kiemtra)
                    trongluong += (ProcessHandler.timtrongluong_sanpham(kiemtra) * qty_ord_pcs)
                    sokienhang += math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(kiemtra))
                    sheet[f"AU{current_row}"] = math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(kiemtra))
                    sheet[f"AT{current_row}"] = ProcessHandler.timtrongluong_sanpham(kiemtra) * qty_ord_pcs
            current_row += 1



        kmhoadon = ProcessHandler.find_all_promotions_by_sku_and_time("Hóa Đơn", entry_date,vendor)
        
       

        if kmhoadon:
            kmhoadon = str(kmhoadon[0][1])
            kiemtra = ProcessHandler.check_value_in_sanpham(kmhoadon) 
            sheet[f"AQ{current_row}"] = kmhoadon
            
            sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'
            soluongkm = math.floor(tongtien / ProcessHandler.tachtien_khuyenmai(kmhoadon))
            sheet[f"A{current_row}"] = entry_date
            sheet[f"AV{current_row}"] = songayno_MT
            sheet[f"C{current_row}"] = "Chưa thực hiện"
            sheet[f"D{current_row}"] = cancle_date
            sheet[f"E{current_row}"] = diachigiao
            sheet[f"G{current_row}"] = makhachhang
            sheet[f"L{current_row}"] = diengiai
            sheet[f"Q{current_row}"] = kiemtra[0]
            sheet[f"X{current_row}"] = soluongkm
            sheet[f"V{current_row}"] = kho
            sheet[f"Y{current_row}"] = 0
            sheet[f"Z{current_row}"] = 0
            sheet[f"U{current_row}"] = "Có"
            sheet[f"AJ{current_row}"] = khuvuc
            sheet[f"AE{current_row}"] = 8
            sheet[f"AM{current_row}"] = mien
            sheet[f"T{current_row}"] = "Không"
            sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(kiemtra[0])
            trongluong += (ProcessHandler.timtrongluong_sanpham(kiemtra[0]) * soluongkm)
            sokienhang += math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(kiemtra[0]))
            sheet[f"AU{current_row}"] = math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(kiemtra[0]))

            sheet[f"AT{current_row}"] = ProcessHandler.timtrongluong_sanpham(kiemtra[0]) * soluongkm

            cachbokem = ProcessHandler.laycachbo_khuyenmai(kmhoadon)
            if cachbokem:
                sheet[f"AO{current_row}"] = cachbokem
            else:
                sheet[f"AO{current_row}"] = f'KM Bó Kèm - Che Barcode'

            self.log_signal.emit(f'🎉 Đơn hàng có chương trình khuyến mãi <b><span style="color: green;">{kmhoadon}</span></b> 🛒')
            self.log_signal.emit(f'Tổng tiền đơn hàng: <b><span style="color: green;">{tongtien} </span></b>\nSố lượng tặng: <b><span style="color: green;">{soluongkm} </span></b> 🛒')
            

        ProcessHandler.ghi_message(f"tong_tien: {tongtien:,.0f}")
        ProcessHandler.ghi_message(f"tong_trongluong: {ProcessHandler.format_weight_kg(trongluong)}")
        ProcessHandler.ghi_message(f"tong_kienhang: {sokienhang}")
        ProcessHandler.ghi_message(f"sai_gia: {saigia}")
        now2 = datetime.now()
        now_str = now2.strftime("%d/%m/%Y %H:%M:%S")
        ProcessHandler.ghi_message(f"end_time: {now_str}")
        time = now2 - now
        ProcessHandler.ghi_message(f"time: {ProcessHandler.format_timedelta(time)}]")  


        sheet[f"L{start_row}"] = f'{diengiai} (Tổng trọng lượng: {ProcessHandler.format_weight_kg(trongluong)})'
        # Lưu lại file Excel
        wb.save(file_path)
        print(f"✅ Đã ghi {len(products)} dòng vào '{file_path}', sheet '{sheet_name}'!")
        return saigia
    




    def write_to_dondathang_fujimart(self,products, makhachhang, po_number, entry_date, cancle_date,STT_donhang,vendor,diachigiao,url):

        now = datetime.now()
        now_str = now.strftime("%d/%m/%Y %H:%M:%S")

        vendor = vendor.upper()
        """Ghi danh sách sản phẩm vào file dondathang.xlsx"""
        file_path = "dondathang.xlsx"
        sheet_name = "Don dat hang"
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]

        start_row = sheet.max_row + 1
        current_row = start_row
        # Lấy giá trị G1, đảm bảo nó là số
        print(f"📍 PO Number {po_number} ")
        STT_donhang_str = f"-{po_number}"     
        saigia = 0
        tongtien = 0
        trongluong = 0
        sokienhang = 0
        if makhachhang[:2] == "MB":
            khuvuc = "MT_MB"
            kho = "TP_HN_12"
            mien = "HN"
        else:
            kho = "LA_KHO2026"
            khuvuc = "MT_MN"
            mien = "LA"



        ProcessHandler.ghi_message(f"[PO: {po_number}")
        ProcessHandler.ghi_message(f"store: {diachigiao}")
        ProcessHandler.ghi_message(f"vendor: {vendor}")
        ProcessHandler.ghi_message(f"Mã Khách hàng: {makhachhang}")
        ProcessHandler.ghi_message(f"start_time: {now_str}")
        if url:
            ProcessHandler.ghi_message(f"url: {url}")


        diengiai = f"{vendor} PO{po_number}"

        sheet[f"A{current_row}"] = entry_date
        sheet[f"AV{current_row}"] = songayno_MT

        sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'
        sheet[f"C{current_row}"] = "Chưa thực hiện"
        sheet[f"D{current_row}"] = cancle_date
        sheet[f"G{current_row}"] = makhachhang
        sheet[f"L{current_row}"] = diengiai
        sheet[f"V{current_row}"] = kho
        sheet[f"AE{current_row}"] = 8
        sheet[f"AJ{current_row}"] = khuvuc
        sheet[f"AM{current_row}"] = mien
        sheet[f"U{current_row}"] = "Không"
        sheet[f"Z{current_row}"] = 0
        sheet[f"S{current_row}"] = diengiai
        sheet[f"T{current_row}"] = "Có"
        sheet[f"X{current_row}"] = 0
        sheet[f"Y{current_row}"] = 0
        sheet[f"E{current_row}"] = diachigiao
        current_row += 1

        for product in products:
            barcode = product["Barcode"]
            soluong = float(product["OU Qty"])
            
            dongia = float(product["Total Price"])
           
            
            sheet[f"A{current_row}"] = entry_date
            sheet[f"AV{current_row}"] = songayno_MT
        
            sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'
            sheet[f"C{current_row}"] = "Chưa thực hiện"
            sheet[f"D{current_row}"] = cancle_date
            sheet[f"G{current_row}"] = makhachhang
            sheet[f"L{current_row}"] = diengiai
            sheet[f"Q{current_row}"] = product["Barcode"]
            sheet[f"V{current_row}"] = kho
            sheet[f"AE{current_row}"] = 8
            sheet[f"AJ{current_row}"] = khuvuc
            sheet[f"AM{current_row}"] = mien
            sheet[f"U{current_row}"] = "Không"
            sheet[f"E{current_row}"] = diachigiao
            sheet[f"T{current_row}"] = "Không"
            sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(product["Barcode"])

            # ✅ Xử lý "Extended Cost"
            extended_cost =  dongia
            if isinstance(extended_cost, str):  # Nếu là chuỗi, loại bỏ dấu phẩy
                extended_cost = float(extended_cost.replace(",", ""))
            extended_cost = int(extended_cost) if extended_cost.is_integer() else extended_cost
            sheet[f"Z{current_row}"] = f"=Y{current_row}*X{current_row}"

            # ✅ Xử lý "Qty Ord/Pcs"
            qty_ord_pcs = product["OU Qty"]
            if isinstance(qty_ord_pcs, str):  # Nếu là chuỗi, loại bỏ dấu phẩy
                qty_ord_pcs = float(qty_ord_pcs.replace(",", ""))
            qty_ord_pcs = int(qty_ord_pcs) if qty_ord_pcs.is_integer() else qty_ord_pcs
            sheet[f"X{current_row}"] = qty_ord_pcs
            trongluong += (ProcessHandler.timtrongluong_sanpham(product["Barcode"]) * qty_ord_pcs)
            sheet[f"AT{current_row}"] = ProcessHandler.timtrongluong_sanpham(product["Barcode"]) * qty_ord_pcs
            sokienhang += math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(product["Barcode"]))
            sheet[f"AU{current_row}"] = math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(product["Barcode"]))


            # ✅ Tính giá hóa đơn
            giahoadon = dongia / qty_ord_pcs
            
            giathuctegoc = ProcessHandler.find_price_by_sku(product["Barcode"],vendor)
            giathucte = giathuctegoc
            # 🔍 **Tìm giá trị trong CTKM.xlsx theo SKU & thời gian**
            
            results = ProcessHandler.find_all_promotions_by_sku_and_time(product["Barcode"], entry_date,vendor)
            slkhuyenmai = len(results)
            print(f'Giá thực tế: {giathucte}')
            print(f'🔎 Số lượng CTKM khớp: {slkhuyenmai}')
            khuyenmai = ''
            col = ''



            giakhop = False  # Cờ kiểm tra khớp giá

            for col, value in results:
                print(f'🧪 Thử CTKM: {col} → {value}')
                
                if value:
                    khuyenmai = value
                    discount = ProcessHandler.extract_discount(value)
                    print(f'🔻 Giảm giá: {discount} ({type(discount)})')

                    if discount != 0:
                        if isinstance(discount, str):
                            discount = discount.replace("%", "").strip()
                            discount = float(discount) if discount else 0

                        if isinstance(giathuctegoc, str):
                            giathucte = giathuctegoc.replace(",", "").strip()
                            giathucte = float(giathucte) if giathucte else 0

                        giathucte = giathucte - (giathucte * float(discount) / 100)
                    else:
                        giathucte = giathucte



                    giahoadon = float(giahoadon[0]) if isinstance(giahoadon, tuple) else float(giahoadon)
                    if isinstance(giathucte, str):
                        giathucte = giathucte.replace(",", "").strip()
                        giathucte = float(giathucte) if giathucte else 0

                    if math.isclose(giahoadon, giathucte, rel_tol=1e-4):
                        sheet[f"Y{current_row}"] = giathucte
                        self.log_signal.emit(
                        f'🎉 Mã hàng <b><span style="color: blue;">{product["Barcode"]}</span></b> '
                        f'có CTKM <b><span style="color: green;">{value}</span></b> từ <i>{col}</i> 🛒')
                        self.log_signal.emit(
                            f'✅ Mã hàng <b><span style="color: blue;">{product["Barcode"]}</span></b> '
                            f'có giá chính xác 🎯: <span style="color: green;"><b>{giahoadon}</b></span>'
                        )
                        
                        giakhop = True
                        break  # ✅ THOÁT VÒNG LẶP khi đã khớp giá

            if not results:
                khuyenmai = ''
                giahoadon = float(giahoadon[0]) if isinstance(giahoadon, tuple) else float(giahoadon)
                if isinstance(giathucte, str):
                    giathucte = giathucte.replace(",", "").strip()
                    giathucte = float(giathucte) if giathucte else 0

                if math.isclose(giahoadon, giathucte, rel_tol=1e-4):
                    sheet[f"Y{current_row}"] = giathucte
                    self.log_signal.emit(
                    f'🎉 Mã hàng <b><span style="color: blue;">{product["Barcode"]}</span></b> '
                    f'có CTKM <b><span style="color: green;">{khuyenmai}</span></b> từ <i>{col}</i> 🛒')
                    self.log_signal.emit(
                        f'✅ Mã hàng <b><span style="color: blue;">{product["Barcode"]}</span></b> '
                        f'có giá chính xác 🎯: <span style="color: green;"><b>{giahoadon}</b></span>'
                    )
                    giakhop = True



            # ❌ Sau khi thử hết, không khớp giá nào:
            if not giakhop:
                sheet[f"Y{current_row}"] = giathucte
                red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                sheet[f"Y{current_row}"].fill = red_fill
                comment_text = f"Kiểm tra lại giá mã này! - Giá hóa đơn: {giahoadon} - Chênh lệch: {giahoadon - giathucte}"
                comment = Comment(comment_text, "System")
                sheet[f"Y{current_row}"].comment = comment

                self.log_signal.emit(
                    f"⚠️ Mã hàng <b>{product["Barcode"]}</b> sai giá! 🛑 "
                    f"Hóa đơn: <b><span style='color: red;'>{giahoadon}</span></b> – "
                    f"Giá hệ thống: <b><span style='color: green;'>{giathucte}</span></b>"
                )
                ProcessHandler.ghi_message(f"text: Mã hàng {product["Barcode"]} - {ProcessHandler.timten_sanpham(product["Barcode"])}, Giá trên PO {giahoadon:,}, Giá đúng: {giathucte:,}. Chênh lệch {(giahoadon - giathucte):,.0f}")

                saigia += 1

            tongtien += giathucte * qty_ord_pcs
                

            # 🔍 Kiểm tra SKU có trong 'SanPham'
            kiemtra = ProcessHandler.check_value_in_sanpham(khuyenmai)
            sheet[f"AQ{current_row}"] = khuyenmai
            match = re.search(r"(\d+)\s*\+\s*1", khuyenmai)  # Tìm số đầu tiên trong biểu thức X+1
            if match:
                x = int(match.group(1))  # Chuyển đổi thành số nguyên
                kiemtra = kiemtra or barcode
                
                if x >= 2:  # Chỉ chia khi x >= 2, còn 1+1 thì giữ nguyên
                    qty_ord_pcs = math.floor(qty_ord_pcs / x)  # Chia rồi làm tròn xuống

            
               
            if kiemtra:
                if isinstance(kiemtra, list):  # Nếu có nhiều mã SKU, nối thành chuỗi
                    kiemtra = ", ".join(kiemtra)
                cachbokem = ProcessHandler.laycachbo_khuyenmai(value)
                if cachbokem:
                    sheet[f"AO{current_row}"] = cachbokem
                    cachbokem_lower = cachbokem.lower()
                    if "bó kèm" in cachbokem_lower or "quấn kèm" in cachbokem_lower:
                        sheet[f"AP{current_row}"] = f'{ProcessHandler.layduoi_mahang(product["Barcode"])}_{ProcessHandler.layduoi_mahang(kiemtra)}_1'
                        sheet[f"AP{current_row + 1}"] = f'{ProcessHandler.layduoi_mahang(product["Barcode"])}_{ProcessHandler.layduoi_mahang(kiemtra)}_1'
                else:
                    sheet[f"AO{current_row}"] = f'KM Bó Kèm - Không Che Barcode'
                    sheet[f"AP{current_row}"] = f'{ProcessHandler.layduoi_mahang(product["Barcode"])}_{ProcessHandler.layduoi_mahang(kiemtra)}_1'
                    sheet[f"AP{current_row + 1}"] = f'{ProcessHandler.layduoi_mahang(product["Barcode"])}_{ProcessHandler.layduoi_mahang(kiemtra)}_1'
                current_row += 1
                self.log_signal.emit(
    f'🎁 Mã hàng <b><span style="color: blue;">{barcode}</span></b> '
    f'đã thêm hàng khuyến mãi <b><span style="color: green;">{kiemtra}</span></b> tặng kèm! 🎉'
)
                sheet[f"A{current_row}"] = entry_date
                sheet[f"AV{current_row}"] = songayno_MT
            
                sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'
                sheet[f"C{current_row}"] = "Chưa thực hiện"
                sheet[f"D{current_row}"] = cancle_date
                sheet[f"G{current_row}"] = makhachhang
                sheet[f"L{current_row}"] = diengiai
                sheet[f"Q{current_row}"] = kiemtra
                sheet[f"X{current_row}"] = qty_ord_pcs
                sheet[f"V{current_row}"] = kho
                sheet[f"Y{current_row}"] = 0
                sheet[f"Z{current_row}"] = 0
                sheet[f"U{current_row}"] = "Có"
                sheet[f"AJ{current_row}"] = khuvuc
                sheet[f"AE{current_row}"] = 8
                sheet[f"AM{current_row}"] = mien
                sheet[f"E{current_row}"] = diachigiao
                sheet[f"T{current_row}"] = "Không"
                sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(kiemtra)
                trongluong += (ProcessHandler.timtrongluong_sanpham(kiemtra) * qty_ord_pcs)
                sokienhang += math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(kiemtra))
                sheet[f"AU{current_row}"] = math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(kiemtra))
                sheet[f"AT{current_row}"] = ProcessHandler.timtrongluong_sanpham(kiemtra) * qty_ord_pcs

            sheet["G1"] = STT_donhang
            current_row += 1


        kmhoadon = ProcessHandler.find_all_promotions_by_sku_and_time("Hóa Đơn", entry_date,vendor)
        if kmhoadon:
            kmhoadon = str(kmhoadon[0][1])
            kiemtra = ProcessHandler.check_value_in_sanpham(kmhoadon) 
            sheet[f"AQ{current_row}"] = kmhoadon
            
            sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'
            soluongkm = math.floor(tongtien / ProcessHandler.tachtien_khuyenmai(kmhoadon))
            sheet[f"A{current_row}"] = entry_date
            sheet[f"AV{current_row}"] = songayno_MT
            sheet[f"C{current_row}"] = "Chưa thực hiện"
            sheet[f"D{current_row}"] = cancle_date
            sheet[f"E{current_row}"] = diachigiao
            sheet[f"G{current_row}"] = makhachhang
            sheet[f"L{current_row}"] = diengiai
            sheet[f"Q{current_row}"] = kiemtra[0]
            sheet[f"X{current_row}"] = soluongkm
            sheet[f"V{current_row}"] = kho
            sheet[f"Y{current_row}"] = 0
            sheet[f"Z{current_row}"] = 0
            sheet[f"U{current_row}"] = "Có"
            sheet[f"AJ{current_row}"] = khuvuc
            sheet[f"AE{current_row}"] = 8
            sheet[f"AM{current_row}"] = mien
            sheet[f"T{current_row}"] = "Không"
            sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(kiemtra[0])
            cachbokem = ProcessHandler.laycachbo_khuyenmai(kmhoadon)
            trongluong += (ProcessHandler.timtrongluong_sanpham(kiemtra[0]) * soluongkm)
            sokienhang += math.ceil(soluongkm / ProcessHandler.timquycach_sanpham(kiemtra[0]))
            sheet[f"AU{current_row}"] = math.ceil(soluongkm / ProcessHandler.timquycach_sanpham(kiemtra[0]))
            sheet[f"AT{current_row}"] = ProcessHandler.timtrongluong_sanpham(kiemtra[0]) * soluongkm
            if cachbokem:
                sheet[f"AO{current_row}"] = cachbokem
            else:
                sheet[f"AO{current_row}"] = f'KM Bó Kèm - Che Barcode'

            self.log_signal.emit(f'🎉 Đơn hàng có chương trình khuyến mãi <b><span style="color: green;">{kmhoadon}</span></b> 🛒')
            self.log_signal.emit(f'Tổng tiền đơn hàng: <b><span style="color: green;">{tongtien} </span></b>\nSố lượng tặng: <b><span style="color: green;">{soluongkm} </span></b> 🛒')
            
        sheet[f"L{start_row}"] = f'{diengiai} (Tổng trọng lượng: {ProcessHandler.format_weight_kg(trongluong)})'


        ProcessHandler.ghi_message(f"tong_tien: {tongtien:,.0f}")
        ProcessHandler.ghi_message(f"tong_trongluong: {ProcessHandler.format_weight_kg(trongluong)}")
        ProcessHandler.ghi_message(f"sai_gia: {saigia}")
        now2 = datetime.now()
        now_str = now2.strftime("%d/%m/%Y %H:%M:%S")
        ProcessHandler.ghi_message(f"end_time: {now_str}")
        time = now2 - now
        ProcessHandler.ghi_message(f"time: {ProcessHandler.format_timedelta(time)}]")  



        # Lưu lại file Excel
        wb.save(file_path)
        print(f"✅ Đã ghi {len(products)} dòng vào '{file_path}', sheet '{sheet_name}'!")
        return saigia
    



    
    def upload_file_to_drive(path, output_name=None):
        """
        path: đường dẫn file gốc
        output_name: tên file muốn đặt trên Drive (KHÔNG cần đuôi)
                    nếu None -> dùng tên gốc
        """

        script_url = "https://script.google.com/macros/s/AKfycbx2ZJhdxEAZq_79ibt3g5UeqccNqLT2ScOtRldnwlgRQB2JdquUPSnSebMQoYESNSv2/exec"

        # Detect mime
        mime, _ = mimetypes.guess_type(path)
        if not mime:
            mime = "application/octet-stream"

        # Lấy đuôi file gốc
        ext = os.path.splitext(path)[1].lstrip(".")

        # Tên file gửi lên (không đuôi)
        if output_name:
            filename = output_name
        else:
            filename = os.path.splitext(os.path.basename(path))[0]

        # Encode file
        with open(path, "rb") as f:
            b64 = base64.b64encode(f.read()).decode("ascii")

        payload = {
            "filename": filename,   # 👈 tên MỚI
            "ext": ext,             # 👈 giữ nguyên đuôi
            "mime": mime,
            "file_b64": b64
        }

        r = requests.post(
            script_url,
            data=json.dumps(payload),
            headers={"Content-Type": "application/json"},
            timeout=60
        )

        r.raise_for_status()
        return r.json()



    def write_to_dondathang_farmer(self,products, makhachhang, po_number, entry_date, cancle_date,STT_donhang,vendor,diachigiao,url):
        now = datetime.now()
        now_str = now.strftime("%d/%m/%Y %H:%M:%S")

        vendor = vendor.upper()
        """Ghi danh sách sản phẩm vào file dondathang.xlsx"""
        file_path = "dondathang.xlsx"
        sheet_name = "Don dat hang"
      
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]

        start_row = sheet.max_row + 1
        current_row = start_row
        # Lấy giá trị G1, đảm bảo nó là số
        print(f"📍 PO Number {po_number} ")
        STT_donhang_str = f"-{po_number}"     

        saigia = 0
        tongtien = 0
        trongluong = 0
        sokienhang = 0

        if makhachhang[:2] == "MB":
            khuvuc = "MT_MB"
            kho = "TP_HN_12"
            mien = "HN"
        else:
            kho = "LA_KHO2026"
            khuvuc = "MT_MN"
            mien = "LA"

        diengiai = f"{vendor} {po_number}"

        ProcessHandler.ghi_message(f"[PO: {po_number}")
        ProcessHandler.ghi_message(f"vendor: {vendor}")
        ProcessHandler.ghi_message(f"store: {diachigiao}")
        ProcessHandler.ghi_message(f"Mã Khách hàng: {makhachhang}")
        ProcessHandler.ghi_message(f"start_time: {now_str}")
        if url:
            ProcessHandler.ghi_message(f"url: {url}")





        sheet[f"A{current_row}"] = entry_date
        sheet[f"AV{current_row}"] = songayno_MT
       
        sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'
        sheet[f"C{current_row}"] = "Chưa thực hiện"
        sheet[f"D{current_row}"] = cancle_date
        sheet[f"G{current_row}"] = makhachhang
        sheet[f"L{current_row}"] = diengiai
        sheet[f"V{current_row}"] = kho
        sheet[f"AE{current_row}"] = 8
        sheet[f"AJ{current_row}"] = khuvuc
        sheet[f"AM{current_row}"] = mien
        sheet[f"U{current_row}"] = "Không"
        sheet[f"Z{current_row}"] = 0
        sheet[f"S{current_row}"] = f"{vendor} {po_number}"
        sheet[f"T{current_row}"] = "Có"
        sheet[f"X{current_row}"] = 0
        sheet[f"Y{current_row}"] = 0
        sheet[f"E{current_row}"] = diachigiao
        current_row += 1

        for product in products:
            barcode = product["Barcode"]
            soluong = float(product["OU Qty"])
            
            dongia = float(product["Total Price"])
           
            
            sheet[f"A{current_row}"] = entry_date
            sheet[f"AV{current_row}"] = songayno_MT
            sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'
            sheet[f"C{current_row}"] = "Chưa thực hiện"
            sheet[f"D{current_row}"] = cancle_date
            sheet[f"G{current_row}"] = makhachhang
            sheet[f"L{current_row}"] = diengiai
            sheet[f"Q{current_row}"] = product["Barcode"]
            sheet[f"V{current_row}"] = kho
            sheet[f"AE{current_row}"] = 8
            sheet[f"AJ{current_row}"] = khuvuc
            sheet[f"AM{current_row}"] = mien
            sheet[f"U{current_row}"] = "Không"
            sheet[f"E{current_row}"] = diachigiao
            sheet[f"T{current_row}"] = "Không"
            sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(product["Barcode"])

            # ✅ Xử lý "Extended Cost"
            extended_cost =  dongia
            if isinstance(extended_cost, str):  # Nếu là chuỗi, loại bỏ dấu phẩy
                extended_cost = float(extended_cost.replace(",", ""))
            extended_cost = int(extended_cost) if extended_cost.is_integer() else extended_cost
            sheet[f"Z{current_row}"] = f"=Y{current_row}*X{current_row}"

            # ✅ Xử lý "Qty Ord/Pcs"
            qty_ord_pcs = product["OU Qty"]
            if isinstance(qty_ord_pcs, str):  # Nếu là chuỗi, loại bỏ dấu phẩy
                qty_ord_pcs = float(qty_ord_pcs.replace(",", ""))
            qty_ord_pcs = int(qty_ord_pcs) if qty_ord_pcs.is_integer() else qty_ord_pcs
            sheet[f"X{current_row}"] = qty_ord_pcs

            trongluong += (ProcessHandler.timtrongluong_sanpham(product["Barcode"]) * qty_ord_pcs)
            sokienhang += math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(product["Barcode"]))
            sheet[f"AU{current_row}"] = math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(product["Barcode"]))
            sheet[f"AT{current_row}"] = ProcessHandler.timtrongluong_sanpham(product["Barcode"]) * qty_ord_pcs

            # ✅ Tính giá hóa đơn
            giahoadon = dongia / qty_ord_pcs
            
            giathuctegoc = ProcessHandler.find_price_by_sku(product["Barcode"],vendor)
            giathucte = giathuctegoc
            # 🔍 **Tìm giá trị trong CTKM.xlsx theo SKU & thời gian**
            
            results = ProcessHandler.find_all_promotions_by_sku_and_time(product["Barcode"], entry_date,vendor)
            slkhuyenmai = len(results)
            print(f'Giá thực tế: {giathucte}')
            print(f'🔎 Số lượng CTKM khớp: {slkhuyenmai}')
            khuyenmai = ''



            giakhop = False  # Cờ kiểm tra khớp giá

            for col, value in results:
                print(f'🧪 Thử CTKM: {col} → {value}')
                
                if value:
                    khuyenmai = value
                    discount = ProcessHandler.extract_discount(value)
                    print(f'🔻 Giảm giá: {discount} ({type(discount)})')

                    if discount != 0:
                        if isinstance(discount, str):
                            discount = discount.replace("%", "").strip()
                            discount = float(discount) if discount else 0

                        if isinstance(giathuctegoc, str):
                            giathucte = giathuctegoc.replace(",", "").strip()
                            giathucte = float(giathucte) if giathucte else 0

                        giathucte = giathucte - (giathucte * float(discount) / 100)
                    else:
                        giathucte = giathucte



                    giahoadon = float(giahoadon[0]) if isinstance(giahoadon, tuple) else float(giahoadon)
                    if isinstance(giathucte, str):
                        giathucte = giathucte.replace(",", "").strip()
                        giathucte = float(giathucte) if giathucte else 0

                    if math.isclose(giahoadon, giathucte, rel_tol=1e-4):
                        sheet[f"Y{current_row}"] = giathucte
                        self.log_signal.emit(
                        f'🎉 Mã hàng <b><span style="color: blue;">{product["Barcode"]}</span></b> '
                        f'có CTKM <b><span style="color: green;">{value}</span></b> từ <i>col</i> 🛒')
                        self.log_signal.emit(
                            f'✅ Mã hàng <b><span style="color: blue;">{product["Barcode"]}</span></b> '
                            f'có giá chính xác 🎯: <span style="color: green;"><b>{giahoadon}</b></span>'
                        )
                        
                        giakhop = True
                        break  # ✅ THOÁT VÒNG LẶP khi đã khớp giá

            if not results:
                khuyenmai = ''
                giahoadon = float(giahoadon[0]) if isinstance(giahoadon, tuple) else float(giahoadon)
                if isinstance(giathucte, str):
                    giathucte = giathucte.replace(",", "").strip()
                    giathucte = float(giathucte) if giathucte else 0

                if math.isclose(giahoadon, giathucte, rel_tol=1e-4):
                    sheet[f"Y{current_row}"] = giathucte
                    self.log_signal.emit(
                    f'🎉 Mã hàng <b><span style="color: blue;">{product["Barcode"]}</span></b> '
                    f'có CTKM <b><span style="color: green;">{khuyenmai}</span></b> từ <i>col</i> 🛒')
                    self.log_signal.emit(
                        f'✅ Mã hàng <b><span style="color: blue;">{product["Barcode"]}</span></b> '
                        f'có giá chính xác 🎯: <span style="color: green;"><b>{giahoadon}</b></span>'
                    )
                    giakhop = True



            # ❌ Sau khi thử hết, không khớp giá nào:
            if not giakhop:
                sheet[f"Y{current_row}"] = giathucte
                red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                sheet[f"Y{current_row}"].fill = red_fill
                comment_text = f"Kiểm tra lại giá mã này! - Giá hóa đơn: {giahoadon} - Chênh lệch: {giahoadon - giathucte}"
                comment = Comment(comment_text, "System")
                sheet[f"Y{current_row}"].comment = comment

                self.log_signal.emit(
                    f"⚠️ Mã hàng <b>{product["Barcode"]}</b> sai giá! 🛑 "
                    f"Hóa đơn: <b><span style='color: red;'>{giahoadon}</span></b> – "
                    f"Giá hệ thống: <b><span style='color: green;'>{giathucte}</span></b>"
                )
                ProcessHandler.ghi_message(f"text: Mã hàng {product["Barcode"]} - {ProcessHandler.timten_sanpham(product["Barcode"])}, Giá trên PO {giahoadon:,}, Giá đúng: {giathucte:,}. Chênh lệch {(giahoadon - giathucte):,.0f}")

                saigia += 1

            tongtien += giathucte * qty_ord_pcs
                

            # 🔍 Kiểm tra SKU có trong 'SanPham'
            kiemtra = ProcessHandler.check_value_in_sanpham(khuyenmai)
            match = re.search(r"(\d+)\s*\+\s*1", khuyenmai)  # Tìm số đầu tiên trong biểu thức X+1
            sheet[f"AQ{current_row}"] = khuyenmai
            if match:
                x = int(match.group(1))  # Chuyển đổi thành số nguyên
                kiemtra = kiemtra or barcode
                
                if x >= 2:  # Chỉ chia khi x >= 2, còn 1+1 thì giữ nguyên
                    qty_ord_pcs = math.floor(qty_ord_pcs / x)  # Chia rồi làm tròn xuống

            
               
            if kiemtra:
                if isinstance(kiemtra, list):  # Nếu có nhiều mã SKU, nối thành chuỗi
                    kiemtra = ", ".join(kiemtra)
                cachbokem = ProcessHandler.laycachbo_khuyenmai(value)
                if cachbokem:
                    sheet[f"AO{current_row}"] = cachbokem
                    cachbokem_lower = cachbokem.lower()
                    if "bó kèm" in cachbokem_lower or "quấn kèm" in cachbokem_lower:
                        sheet[f"AP{current_row}"] = f'{ProcessHandler.layduoi_mahang(product["Barcode"])}_{ProcessHandler.layduoi_mahang(kiemtra)}_1'
                        sheet[f"AP{current_row + 1}"] = f'{ProcessHandler.layduoi_mahang(product["Barcode"])}_{ProcessHandler.layduoi_mahang(kiemtra)}_1'
                else:
                    sheet[f"AO{current_row}"] = f'KM Giao Rời - Không Che Barcode'
                current_row += 1
                self.log_signal.emit(
                    f'🎁 Mã hàng <b><span style="color: blue;">{barcode}</span></b> '
                    f'đã thêm hàng khuyến mãi <b><span style="color: green;">{kiemtra}</span></b> tặng kèm! 🎉'
                )
                sheet[f"A{current_row}"] = entry_date
                sheet[f"AV{current_row}"] = songayno_MT
                
                sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'
                sheet[f"C{current_row}"] = "Chưa thực hiện"
                sheet[f"D{current_row}"] = cancle_date
                sheet[f"G{current_row}"] = makhachhang
                sheet[f"L{current_row}"] = diengiai
                sheet[f"Q{current_row}"] = kiemtra
                sheet[f"X{current_row}"] = qty_ord_pcs
                sheet[f"V{current_row}"] = kho
                sheet[f"Y{current_row}"] = 0
                sheet[f"Z{current_row}"] = 0
                sheet[f"U{current_row}"] = "Có"
                sheet[f"AJ{current_row}"] = khuvuc
                sheet[f"AE{current_row}"] = 8
                sheet[f"AM{current_row}"] = mien
                sheet[f"E{current_row}"] = diachigiao
                sheet[f"T{current_row}"] = "Không"
                sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(kiemtra)
                trongluong += (ProcessHandler.timtrongluong_sanpham(kiemtra) * qty_ord_pcs)
                sokienhang += math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(kiemtra))
                sheet[f"AU{current_row}"] = math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(kiemtra))
                sheet[f"AT{current_row}"] = ProcessHandler.timtrongluong_sanpham(kiemtra) * qty_ord_pcs

            sheet["G1"] = STT_donhang
            current_row += 1

        kmhoadon = ProcessHandler.find_all_promotions_by_sku_and_time("Hóa Đơn", entry_date,vendor)
        if kmhoadon:
            kmhoadon = str(kmhoadon[0][1])
            kiemtra = ProcessHandler.check_value_in_sanpham(kmhoadon)
            sheet[f"AQ{current_row}"] = kmhoadon 
            
            sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'
            soluongkm = math.floor(tongtien / ProcessHandler.tachtien_khuyenmai(kmhoadon))
            sheet[f"A{current_row}"] = entry_date
            sheet[f"AV{current_row}"] = songayno_MT
            sheet[f"C{current_row}"] = "Chưa thực hiện"
            sheet[f"D{current_row}"] = cancle_date
            sheet[f"E{current_row}"] = diachigiao
            sheet[f"G{current_row}"] = makhachhang
            sheet[f"L{current_row}"] = diengiai
            sheet[f"Q{current_row}"] = kiemtra[0]
            sheet[f"X{current_row}"] = soluongkm
            sheet[f"V{current_row}"] = kho
            sheet[f"Y{current_row}"] = 0
            sheet[f"Z{current_row}"] = 0
            sheet[f"U{current_row}"] = "Có"
            sheet[f"AJ{current_row}"] = khuvuc
            sheet[f"AE{current_row}"] = 8
            sheet[f"AM{current_row}"] = mien
            sheet[f"T{current_row}"] = "Không"
            sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(kiemtra[0])

            trongluong += (ProcessHandler.timtrongluong_sanpham(kiemtra[0]) * soluongkm)
            sokienhang += math.ceil(soluongkm / ProcessHandler.timquycach_sanpham(kiemtra[0]))
            sheet[f"AU{current_row}"] = math.ceil(soluongkm / ProcessHandler.timquycach_sanpham(kiemtra[0]))
            sheet[f"AT{current_row}"] = ProcessHandler.timtrongluong_sanpham(kiemtra[0]) * soluongkm
            cachbokem = ProcessHandler.laycachbo_khuyenmai(kmhoadon)
            if cachbokem:
                sheet[f"AO{current_row}"] = cachbokem
            else:
                sheet[f"AO{current_row}"] = f'KM Bó Kèm - Che Barcode'

            self.log_signal.emit(f'🎉 Đơn hàng có chương trình khuyến mãi <b><span style="color: green;">{kmhoadon}</span></b> 🛒')
            self.log_signal.emit(f'Tổng tiền đơn hàng: <b><span style="color: green;">{tongtien} </span></b>\nSố lượng tặng: <b><span style="color: green;">{soluongkm} </span></b> 🛒')
            

        sheet[f"L{start_row}"] = f'{diengiai} (Tổng trọng lượng: {ProcessHandler.format_weight_kg(trongluong)})'
        # Lưu lại file Excel
        

        ProcessHandler.ghi_message(f"tong_tien: {tongtien:,.0f}")
        ProcessHandler.ghi_message(f"tong_trongluong: {ProcessHandler.format_weight_kg(trongluong)}")
        ProcessHandler.ghi_message(f"tong_kienhang: {sokienhang}")
        ProcessHandler.ghi_message(f"sai_gia: {saigia}")
        now2 = datetime.now()
        now_str = now2.strftime("%d/%m/%Y %H:%M:%S")
        ProcessHandler.ghi_message(f"end_time: {now_str}")
        time = now2 - now
        ProcessHandler.ghi_message(f"time: {ProcessHandler.format_timedelta(time)}]")  

        wb.save(file_path)
        print(f"✅ Đã ghi {len(products)} dòng vào '{file_path}', sheet '{sheet_name}'!")
        return saigia
    

    def write_to_dondathang_bhx(self,products, makhachhang, po_number, entry_date, cancle_date,STT_donhang,vendor,diachigiao,url):
        now = datetime.now()
        now_str = now.strftime("%d/%m/%Y %H:%M:%S")

        vendor = vendor.upper()
        """Ghi danh sách sản phẩm vào file dondathang.xlsx"""
        file_path = "dondathang.xlsx"
        sheet_name = "Don dat hang"
        
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]

        start_row = sheet.max_row + 1
        current_row = start_row
        # Lấy giá trị G1, đảm bảo nó là số
        print(f"📍 PO Number {po_number} ")
        STT_donhang_str = f"-{po_number}"     
        saigia = 0
        tongtien = 0
        trongluong = 0
        sokienhang = 0
        if makhachhang[:2] == "MB":
            khuvuc = "MT_MB"
            kho = "TP_HN_12"
            mien = "HN"
        else:
            kho = "LA_KHO2026"
            khuvuc = "MT_MN"
            mien = "LA"

        diengiai = f"{vendor} PO{po_number}"

        ProcessHandler.ghi_message(f"[PO: {po_number}")
        ProcessHandler.ghi_message(f"vendor: {vendor}")
        ProcessHandler.ghi_message(f"store: {diachigiao}")
        ProcessHandler.ghi_message(f"Mã Khách hàng: {makhachhang}")
        ProcessHandler.ghi_message(f"start_time: {now_str}")
        if url:
            ProcessHandler.ghi_message(f"url: {url}")





        sheet[f"A{current_row}"] = entry_date
        sheet[f"AV{current_row}"] = songayno_MT

        sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'
        sheet[f"C{current_row}"] = "Chưa thực hiện"
        sheet[f"D{current_row}"] = cancle_date
        sheet[f"G{current_row}"] = makhachhang
        sheet[f"L{current_row}"] = diengiai
        sheet[f"V{current_row}"] = kho
        sheet[f"AE{current_row}"] = 8
        sheet[f"AJ{current_row}"] = khuvuc
        sheet[f"AM{current_row}"] = mien
        sheet[f"U{current_row}"] = "Không"
        sheet[f"Z{current_row}"] = 0
        sheet[f"S{current_row}"] = f"{vendor} PO{po_number}"
        sheet[f"T{current_row}"] = "Có"
        sheet[f"X{current_row}"] = 0
        sheet[f"Y{current_row}"] = 0
        sheet[f"E{current_row}"] = diachigiao
        current_row += 1

        for product in products:
            barcode = product["Barcode"]
            soluong = float(product["OU Qty"])
            
            dongia = float(product["Total Price"])
           
            
            sheet[f"A{current_row}"] = entry_date
            sheet[f"AV{current_row}"] = songayno_MT
            
            sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'
            sheet[f"C{current_row}"] = "Chưa thực hiện"
            sheet[f"D{current_row}"] = cancle_date
            sheet[f"G{current_row}"] = makhachhang
            sheet[f"L{current_row}"] = diengiai
            sheet[f"Q{current_row}"] = product["Barcode"]
            sheet[f"V{current_row}"] = kho
            sheet[f"AE{current_row}"] = 8
            sheet[f"AJ{current_row}"] = khuvuc
            sheet[f"AM{current_row}"] = mien
            sheet[f"U{current_row}"] = "Không"
            sheet[f"E{current_row}"] = diachigiao
            sheet[f"T{current_row}"] = "Không"
            sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(product["Barcode"])

            # ✅ Xử lý "Extended Cost"
            extended_cost =  dongia
            if isinstance(extended_cost, str):  # Nếu là chuỗi, loại bỏ dấu phẩy
                extended_cost = float(extended_cost.replace(",", ""))
            extended_cost = int(extended_cost) if extended_cost.is_integer() else extended_cost
            sheet[f"Z{current_row}"] = f"=Y{current_row}*X{current_row}"

            # ✅ Xử lý "Qty Ord/Pcs"
            qty_ord_pcs = product["OU Qty"]
            if isinstance(qty_ord_pcs, str):  # Nếu là chuỗi, loại bỏ dấu phẩy
                qty_ord_pcs = float(qty_ord_pcs.replace(",", ""))
            qty_ord_pcs = int(qty_ord_pcs) if qty_ord_pcs.is_integer() else qty_ord_pcs
            sheet[f"X{current_row}"] = qty_ord_pcs
            
            trongluong += (ProcessHandler.timtrongluong_sanpham(product["Barcode"]) * qty_ord_pcs)
            sokienhang += math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(product["Barcode"]))
            sheet[f"AU{current_row}"] = math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(product["Barcode"]))
            sheet[f"AT{current_row}"] = ProcessHandler.timtrongluong_sanpham(product["Barcode"]) * qty_ord_pcs

            # ✅ Tính giá hóa đơn
            giahoadon = dongia / qty_ord_pcs
            
            giathuctegoc = ProcessHandler.find_price_by_sku(product["Barcode"],vendor)
            giathucte = giathuctegoc
            # 🔍 **Tìm giá trị trong CTKM.xlsx theo SKU & thời gian**
            
            results = ProcessHandler.find_all_promotions_by_sku_and_time(product["Barcode"], entry_date,vendor)
            slkhuyenmai = len(results)
            print(f'Giá thực tế: {giathucte}')
            print(f'🔎 Số lượng CTKM khớp: {slkhuyenmai}')
            khuyenmai = ''


            col = ''
            giakhop = False  # Cờ kiểm tra khớp giá

            for col, value in results:
                print(f'🧪 Thử CTKM: {col} → {value}')
                
                if value:
                    khuyenmai = value
                    discount = ProcessHandler.extract_discount(value)
                    print(f'🔻 Giảm giá: {discount} ({type(discount)})')

                    if discount != 0:
                        if isinstance(discount, str):
                            discount = discount.replace("%", "").strip()
                            discount = float(discount) if discount else 0

                        if isinstance(giathuctegoc, str):
                            giathucte = giathuctegoc.replace(",", "").strip()
                            giathucte = float(giathucte) if giathucte else 0

                        giathucte = giathucte - (giathucte * float(discount) / 100)
                    else:
                        giathucte = giathucte



                    giahoadon = float(giahoadon[0]) if isinstance(giahoadon, tuple) else float(giahoadon)
                    if isinstance(giathucte, str):
                        giathucte = giathucte.replace(",", "").strip()
                        giathucte = float(giathucte) if giathucte else 0

                    if math.isclose(giahoadon, giathucte, rel_tol=1e-4):
                        sheet[f"Y{current_row}"] = giathucte
                        self.log_signal.emit(
                        f'🎉 Mã hàng <b><span style="color: blue;">{product["Barcode"]}</span></b> '
                        f'có CTKM <b><span style="color: green;">{value}</span></b> từ <i>{col}</i> 🛒')
                        self.log_signal.emit(
                            f'✅ Mã hàng <b><span style="color: blue;">{product["Barcode"]}</span></b> '
                            f'có giá chính xác 🎯: <span style="color: green;"><b>{giahoadon}</b></span>'
                        )
                        
                        giakhop = True
                        break  # ✅ THOÁT VÒNG LẶP khi đã khớp giá

            if not results:
                khuyenmai = ''
                giahoadon = float(giahoadon[0]) if isinstance(giahoadon, tuple) else float(giahoadon)
                if isinstance(giathucte, str):
                    giathucte = giathucte.replace(",", "").strip()
                    giathucte = float(giathucte) if giathucte else 0

                if math.isclose(giahoadon, giathucte, rel_tol=1e-4):
                    sheet[f"Y{current_row}"] = giathucte
                    self.log_signal.emit(
                    f'🎉 Mã hàng <b><span style="color: blue;">{product["Barcode"]}</span></b> '
                    f'có CTKM <b><span style="color: green;">{khuyenmai}</span></b> từ <i>{col}</i> 🛒')
                    self.log_signal.emit(
                        f'✅ Mã hàng <b><span style="color: blue;">{product["Barcode"]}</span></b> '
                        f'có giá chính xác 🎯: <span style="color: green;"><b>{giahoadon}</b></span>'
                    )
                    giakhop = True



            # ❌ Sau khi thử hết, không khớp giá nào:
            if not giakhop:
                sheet[f"Y{current_row}"] = giathucte
                red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                sheet[f"Y{current_row}"].fill = red_fill
                comment_text = f"Kiểm tra lại giá mã này! - Giá hóa đơn: {giahoadon} - Chênh lệch: {giahoadon - giathucte}"
                comment = Comment(comment_text, "System")
                sheet[f"Y{current_row}"].comment = comment

                self.log_signal.emit(
                    f"⚠️ Mã hàng <b>{product["Barcode"]}</b> sai giá! 🛑 "
                    f"Hóa đơn: <b><span style='color: red;'>{giahoadon}</span></b> – "
                    f"Giá hệ thống: <b><span style='color: green;'>{giathucte}</span></b>"
                )
                ProcessHandler.ghi_message(f"text: Mã hàng {product["Barcode"]} - {ProcessHandler.timten_sanpham(product["Barcode"])}, Giá trên PO {giahoadon:,}, Giá đúng: {giathucte:,}. Chênh lệch {(giahoadon - giathucte):,.0f}")


                saigia += 1

            tongtien += giathucte * qty_ord_pcs
                

            # 🔍 Kiểm tra SKU có trong 'SanPham'
            kiemtra = ProcessHandler.check_value_in_sanpham(khuyenmai)
            sheet[f"AQ{current_row}"] = khuyenmai 
            match = re.search(r"(\d+)\s*\+\s*1", khuyenmai)  # Tìm số đầu tiên trong biểu thức X+1
            if match:
                x = int(match.group(1))  # Chuyển đổi thành số nguyên
                kiemtra = kiemtra or barcode
                
                if x >= 2:  # Chỉ chia khi x >= 2, còn 1+1 thì giữ nguyên
                    qty_ord_pcs = math.floor(qty_ord_pcs / x)  # Chia rồi làm tròn xuống

            
               
            if kiemtra:
                if isinstance(kiemtra, list):  # Nếu có nhiều mã SKU, nối thành chuỗi
                    kiemtra = ", ".join(kiemtra)
                cachbokem = ProcessHandler.laycachbo_khuyenmai(value)
                if cachbokem:
                    sheet[f"AO{current_row}"] = cachbokem
                    cachbokem_lower = cachbokem.lower()
                    if "bó kèm" in cachbokem_lower or "quấn kèm" in cachbokem_lower:
                        sheet[f"AP{current_row}"] = f'{ProcessHandler.layduoi_mahang(product["Barcode"])}_{ProcessHandler.layduoi_mahang(kiemtra)}_1'
                        sheet[f"AP{current_row + 1}"] = f'{ProcessHandler.layduoi_mahang(product["Barcode"])}_{ProcessHandler.layduoi_mahang(kiemtra)}_1'
                else:
                    sheet[f"AO{current_row}"] = f'KM Giao Rời - Không Che Barcode'
                current_row += 1
                self.log_signal.emit(
    f'🎁 Mã hàng <b><span style="color: blue;">{barcode}</span></b> '
    f'đã thêm hàng khuyến mãi <b><span style="color: green;">{kiemtra}</span></b> tặng kèm! 🎉'
)
                sheet[f"A{current_row}"] = entry_date
                sheet[f"AV{current_row}"] = songayno_MT
             
                sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'
                sheet[f"C{current_row}"] = "Chưa thực hiện"
                sheet[f"D{current_row}"] = cancle_date
                sheet[f"G{current_row}"] = makhachhang
                sheet[f"L{current_row}"] = diengiai
                sheet[f"Q{current_row}"] = kiemtra
                sheet[f"X{current_row}"] = qty_ord_pcs
                sheet[f"V{current_row}"] = kho
                sheet[f"Y{current_row}"] = 0
                sheet[f"Z{current_row}"] = 0
                sheet[f"U{current_row}"] = "Có"
                sheet[f"AJ{current_row}"] = khuvuc
                sheet[f"AE{current_row}"] = 8
                sheet[f"AM{current_row}"] = mien
                sheet[f"E{current_row}"] = diachigiao
                sheet[f"T{current_row}"] = "Không"
                sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(kiemtra)
                trongluong += (ProcessHandler.timtrongluong_sanpham(kiemtra) * qty_ord_pcs)
                sokienhang += math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(kiemtra))
                sheet[f"AU{current_row}"] = math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(kiemtra))
                sheet[f"AT{current_row}"] = ProcessHandler.timtrongluong_sanpham(kiemtra) * qty_ord_pcs

            sheet["G1"] = STT_donhang
            current_row += 1


        kmhoadon = ProcessHandler.find_all_promotions_by_sku_and_time("Hóa Đơn", entry_date,vendor)
        if kmhoadon:
            kmhoadon = str(kmhoadon[0][1])
            kiemtra = ProcessHandler.check_value_in_sanpham(kmhoadon) 

            sheet[f"AQ{current_row}"] = kmhoadon 
            
            sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'
            soluongkm = math.floor(tongtien / ProcessHandler.tachtien_khuyenmai(kmhoadon))
            sheet[f"A{current_row}"] = entry_date
            sheet[f"AV{current_row}"] = songayno_MT
            sheet[f"C{current_row}"] = "Chưa thực hiện"
            sheet[f"D{current_row}"] = cancle_date
            sheet[f"E{current_row}"] = diachigiao
            sheet[f"G{current_row}"] = makhachhang
            sheet[f"L{current_row}"] = diengiai
            sheet[f"Q{current_row}"] = kiemtra[0]
            sheet[f"X{current_row}"] = soluongkm
            sheet[f"V{current_row}"] = kho
            sheet[f"Y{current_row}"] = 0
            sheet[f"Z{current_row}"] = 0
            sheet[f"U{current_row}"] = "Có"
            sheet[f"AJ{current_row}"] = khuvuc
            sheet[f"AE{current_row}"] = 8
            sheet[f"AM{current_row}"] = mien
            sheet[f"T{current_row}"] = "Không"
            sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(kiemtra[0])


            trongluong += (ProcessHandler.timtrongluong_sanpham(kiemtra[0]) * qty_ord_pcs)
            sokienhang += math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(kiemtra[0]))
            sheet[f"AU{current_row}"] = math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(kiemtra[0]))
            sheet[f"AT{current_row}"] = ProcessHandler.timtrongluong_sanpham(kiemtra[0]) * qty_ord_pcs

            cachbokem = ProcessHandler.laycachbo_khuyenmai(kmhoadon)
            if cachbokem:
                sheet[f"AO{current_row}"] = cachbokem
            else:
                sheet[f"AO{current_row}"] = f'KM Bó Kèm - Che Barcode'

            self.log_signal.emit(f'🎉 Đơn hàng có chương trình khuyến mãi <b><span style="color: green;">{kmhoadon}</span></b> 🛒')
            self.log_signal.emit(f'Tổng tiền đơn hàng: <b><span style="color: green;">{tongtien} </span></b>\nSố lượng tặng: <b><span style="color: green;">{soluongkm} </span></b> 🛒')
            
        sheet[f"L{start_row}"] = f'{diengiai} (Tổng trọng lượng: {ProcessHandler.format_weight_kg(trongluong)})'
        # Lưu lại file Excel
        wb.save(file_path)
        ProcessHandler.ghi_message(f"tong_tien: {tongtien:,.0f}")
        ProcessHandler.ghi_message(f"tong_trongluong: {ProcessHandler.format_weight_kg(trongluong)}")
        ProcessHandler.ghi_message(f"tong_kienhang: {sokienhang}")
        ProcessHandler.ghi_message(f"sai_gia: {saigia}")
        now2 = datetime.now()
        now_str = now2.strftime("%d/%m/%Y %H:%M:%S")
        ProcessHandler.ghi_message(f"end_time: {now_str}")
        time = now2 - now
        ProcessHandler.ghi_message(f"time: {ProcessHandler.format_timedelta(time)}]")  

        print(f"✅ Đã ghi {len(products)} dòng vào '{file_path}', sheet '{sheet_name}'!")
        return saigia





    def write_to_dondathang_kingfood(self,products, makhachhang, po_number, entry_date, cancle_date,STT_donhang,vendor,delivery,url):
        now = datetime.now()
        now_str = now.strftime("%d/%m/%Y %H:%M:%S")


        vendor = vendor.upper()
        """Ghi danh sách sản phẩm vào file dondathang.xlsx"""
        file_path = "dondathang.xlsx"
        sheet_name = "Don dat hang"

        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]

        start_row = sheet.max_row + 1
        current_row = start_row
        # Lấy giá trị G1, đảm bảo nó là số
        print(f"📍 PO Number {po_number} ")
        delivery = delivery or "KHO SEEDLOG"
        STT_donhang_str = f"-{po_number}"  
        saigia = 0  
        tongtien = 0
        trongluong = 0
        sokienhang = 0
        if makhachhang[:2] == "MB":
            khuvuc = "MT_MB"
            kho = "TP_HN_12"
            mien = "HN"
        else:
            kho = "LA_KHO2026"
            khuvuc = "MT_MN"
            mien = "LA"

        diengiai = f"{vendor} {po_number}"
        sheet[f"A{current_row}"] = entry_date
        sheet[f"AV{current_row}"] = songayno_MT

        ProcessHandler.ghi_message(f"[PO: {po_number}")
        ProcessHandler.ghi_message(f"store: {delivery}")
        ProcessHandler.ghi_message(f"vendor: {vendor}")
        ProcessHandler.ghi_message(f"Mã Khách hàng: {makhachhang}")
        ProcessHandler.ghi_message(f"start_time: {now_str}")
        if url:
            ProcessHandler.ghi_message(f"url: {url}")
      
        sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'
        sheet[f"C{current_row}"] = "Chưa thực hiện"
        sheet[f"D{current_row}"] = cancle_date
        sheet[f"G{current_row}"] = makhachhang
        sheet[f"L{current_row}"] = diengiai
        sheet[f"V{current_row}"] = kho
        sheet[f"AE{current_row}"] = 8
        sheet[f"AJ{current_row}"] = khuvuc
        sheet[f"AM{current_row}"] = mien
        sheet[f"U{current_row}"] = "Không"
        sheet[f"Z{current_row}"] = 0
        sheet[f"S{current_row}"] = f"{vendor} {po_number}"
        sheet[f"T{current_row}"] = "Có"
        sheet[f"X{current_row}"] = 0
        sheet[f"Y{current_row}"] = 0
        
        sheet[f"E{current_row}"] = delivery


        
        current_row += 1


        for product in products:
            barcode = product["Barcode"]
            soluong = float(product["OU Qty"])
            dongia = float(product["Total Price"])
           
            sku_mapping = ProcessHandler.load_sku_mapping()  # Load mapping SKU
            products = ProcessHandler.replace_sku_numbers(products, sku_mapping)  # Thay SKU
            sheet[f"A{current_row}"] = entry_date
            sheet[f"AV{current_row}"] = songayno_MT
            
            sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'
            sheet[f"C{current_row}"] = "Chưa thực hiện"
            sheet[f"D{current_row}"] = cancle_date
            sheet[f"G{current_row}"] = makhachhang
            sheet[f"L{current_row}"] = diengiai
            sheet[f"Q{current_row}"] = product["Barcode"]
            sheet[f"V{current_row}"] = kho
            sheet[f"AE{current_row}"] = 8
            sheet[f"AJ{current_row}"] = khuvuc
            sheet[f"AM{current_row}"] = mien
            sheet[f"U{current_row}"] = "Không"
            sheet[f"E{current_row}"] = delivery
            sheet[f"T{current_row}"] = "Không"
            sheet[f"AR{current_row}"] = product.get("Mahang", '')
            sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(product["Barcode"])

            # ✅ Xử lý "Extended Cost"
            extended_cost = soluong * dongia
            if isinstance(extended_cost, str):  # Nếu là chuỗi, loại bỏ dấu phẩy
                extended_cost = float(extended_cost.replace(",", ""))
            extended_cost = int(extended_cost) if extended_cost.is_integer() else extended_cost
            sheet[f"Z{current_row}"] = f"=Y{current_row}*X{current_row}"

            # ✅ Xử lý "Qty Ord/Pcs"
            qty_ord_pcs = product["OU Qty"]
            if isinstance(qty_ord_pcs, str):  # Nếu là chuỗi, loại bỏ dấu phẩy
                qty_ord_pcs = float(qty_ord_pcs.replace(",", ""))
            qty_ord_pcs = int(qty_ord_pcs) if qty_ord_pcs.is_integer() else qty_ord_pcs
            sheet[f"X{current_row}"] = qty_ord_pcs

            trongluong += (float(ProcessHandler.timtrongluong_sanpham(product["Barcode"])) * qty_ord_pcs)
            
            sokienhang += math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(product["Barcode"]))
            sheet[f"AU{current_row}"] = math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(product["Barcode"]))

            sheet[f"AT{current_row}"] = (float(ProcessHandler.timtrongluong_sanpham(product["Barcode"]) * qty_ord_pcs))

            # ✅ Tính giá hóa đơn
            giahoadon = dongia
            giathuctegoc = ProcessHandler.find_price_by_sku(product["Barcode"],vendor)
            giathucte = giathuctegoc
            # 🔍 **Tìm giá trị trong CTKM.xlsx theo SKU & thời gian**
            
            results = ProcessHandler.find_all_promotions_by_sku_and_time(product["Barcode"], entry_date,vendor)
            slkhuyenmai = len(results)
            print(f'Giá thực tế: {giathucte}')
            print(f'🔎 Số lượng CTKM khớp: {slkhuyenmai}')
            khuyenmai = ''



            giakhop = False  # Cờ kiểm tra khớp giá

            for col, value in results:
                print(f'🧪 Thử CTKM: {col} → {value}')
                
                if value:
                    khuyenmai = value
                    discount = ProcessHandler.extract_discount(value)
                    print(f'🔻 Giảm giá: {discount} ({type(discount)})')

                    if discount != 0:
                        if isinstance(discount, str):
                            discount = discount.replace("%", "").strip()
                            discount = float(discount) if discount else 0

                        if isinstance(giathuctegoc, str):
                            giathucte = giathuctegoc.replace(",", "").strip()
                            giathucte = float(giathucte) if giathucte else 0

                        giathucte = giathucte - (giathucte * float(discount) / 100)
                    else:
                        giathucte = giathucte



                    giahoadon = float(giahoadon[0]) if isinstance(giahoadon, tuple) else float(giahoadon)
                    if isinstance(giathucte, str):
                        giathucte = giathucte.replace(",", "").strip()
                        giathucte = float(giathucte) if giathucte else 0

                    if math.isclose(giahoadon, giathucte, rel_tol=1e-4):
                        sheet[f"Y{current_row}"] = giathucte
                        self.log_signal.emit(
                        f'🎉 Mã hàng <b><span style="color: blue;">{product["Barcode"]}</span></b> '
                        f'có CTKM <b><span style="color: green;">{value}</span></b> từ <i>{col}</i> 🛒')
                        self.log_signal.emit(
                            f'✅ Mã hàng <b><span style="color: blue;">{product["Barcode"]}</span></b> '
                            f'có giá chính xác 🎯: <span style="color: green;"><b>{giahoadon}</b></span>'
                        )
                        
                        giakhop = True
                        break  # ✅ THOÁT VÒNG LẶP khi đã khớp giá

            if not results:
                khuyenmai = ''
                giahoadon = float(giahoadon[0]) if isinstance(giahoadon, tuple) else float(giahoadon)
                if isinstance(giathucte, str):
                    giathucte = giathucte.replace(",", "").strip()
                    giathucte = float(giathucte) if giathucte else 0

                if math.isclose(giahoadon, giathucte, rel_tol=1e-4):
                    sheet[f"Y{current_row}"] = giathucte
                    self.log_signal.emit(
                    f'🎉 Mã hàng <b><span style="color: blue;">{product["Barcode"]}</span></b> '
                    f'có CTKM <b><span style="color: green;">{khuyenmai}</span></b> 🛒')
                    self.log_signal.emit(
                        f'✅ Mã hàng <b><span style="color: blue;">{product["Barcode"]}</span></b> '
                        f'có giá chính xác 🎯: <span style="color: green;"><b>{giahoadon}</b></span>'
                    )
                    giakhop = True



            # ❌ Sau khi thử hết, không khớp giá nào:
            if not giakhop:
                sheet[f"Y{current_row}"] = giathucte
                red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                sheet[f"Y{current_row}"].fill = red_fill
                comment_text = f"Kiểm tra lại giá mã này! - Giá hóa đơn: {giahoadon} - Chênh lệch: {giahoadon - giathucte}"
                comment = Comment(comment_text, "System")
                sheet[f"Y{current_row}"].comment = comment

                self.log_signal.emit(
                    f"⚠️ Mã hàng <b>{product["Barcode"]}</b> sai giá! 🛑 "
                    f"Hóa đơn: <b><span style='color: red;'>{giahoadon}</span></b> – "
                    f"Giá hệ thống: <b><span style='color: green;'>{giathucte}</span></b>"
                )
                ProcessHandler.ghi_message(f"text: Mã hàng {product["Barcode"]} - {ProcessHandler.timten_sanpham(product["Barcode"])}, Giá trên PO {giahoadon:,}, Giá đúng: {giathucte:,}. Chênh lệch {(giahoadon - giathucte):,.0f}")

                saigia += 1

            tongtien += giathucte * qty_ord_pcs
                

            # 🔍 Kiểm tra SKU có trong 'SanPham'
            kiemtra = ProcessHandler.check_value_in_sanpham(khuyenmai)
            sheet[f"AQ{current_row}"] = khuyenmai 
            match = re.search(r"(\d+)\s*\+\s*1", khuyenmai)  # Tìm số đầu tiên trong biểu thức X+1
            if match:
                x = int(match.group(1))  # Chuyển đổi thành số nguyên
                kiemtra = kiemtra or barcode
                
                if x >= 2:  # Chỉ chia khi x >= 2, còn 1+1 thì giữ nguyên
                    qty_ord_pcs = math.floor(qty_ord_pcs / x)  # Chia rồi làm tròn xuống
               
            if kiemtra:
                if isinstance(kiemtra, list):  # Nếu có nhiều mã SKU, nối thành chuỗi
                    kiemtra = ", ".join(kiemtra)
                cachbokem = ProcessHandler.laycachbo_khuyenmai(value)
                if cachbokem:
                    sheet[f"AO{current_row}"] = cachbokem
                    cachbokem_lower = cachbokem.lower()
                    if "bó kèm" in cachbokem_lower or "quấn kèm" in cachbokem_lower:
                        sheet[f"AP{current_row}"] = f'{ProcessHandler.layduoi_mahang(product["Barcode"])}_{ProcessHandler.layduoi_mahang(kiemtra)}_1'
                        sheet[f"AP{current_row + 1}"] = f'{ProcessHandler.layduoi_mahang(product["Barcode"])}_{ProcessHandler.layduoi_mahang(kiemtra)}_1'
                else:
                    sheet[f"AO{current_row}"] = f'KM Giao Rời - Không Che Barcode'
                current_row += 1
                self.log_signal.emit(
    f'🎁 Mã hàng <b><span style="color: blue;">{barcode}</span></b> '
    f'đã thêm hàng khuyến mãi <b><span style="color: green;">{kiemtra}</span></b> tặng kèm! 🎉'
)
                sheet[f"A{current_row}"] = entry_date
                sheet[f"AV{current_row}"] = songayno_MT
                
                sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'
                sheet[f"C{current_row}"] = "Chưa thực hiện"
                sheet[f"D{current_row}"] = cancle_date
                sheet[f"G{current_row}"] = makhachhang
                sheet[f"L{current_row}"] = diengiai
                sheet[f"Q{current_row}"] = kiemtra
                sheet[f"X{current_row}"] = qty_ord_pcs
                sheet[f"V{current_row}"] = kho
                sheet[f"Y{current_row}"] = 0
                sheet[f"Z{current_row}"] = 0
                sheet[f"U{current_row}"] = "Có"
                sheet[f"AJ{current_row}"] = khuvuc
                sheet[f"AE{current_row}"] = 8
                sheet[f"AM{current_row}"] = mien
                sheet[f"T{current_row}"] = "Không"
                sheet[f"E{current_row}"] = delivery
                sheet[f"AR{current_row}"] = product.get("Mahang", '')
                sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(kiemtra)

                trongluong += (ProcessHandler.timtrongluong_sanpham(kiemtra) * qty_ord_pcs)
                
                sokienhang += math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(kiemtra))
                sheet[f"AU{current_row}"] = math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(kiemtra))
                sheet[f"AT{current_row}"] = ProcessHandler.timtrongluong_sanpham(kiemtra) * qty_ord_pcs
        
            current_row += 1
        kmhoadon = ProcessHandler.find_all_promotions_by_sku_and_time("Hóa Đơn", entry_date)
        if kmhoadon:
            kmhoadon = str(kmhoadon[0][1])
            kiemtra = ProcessHandler.check_value_in_sanpham(kmhoadon) 
            sheet[f"AQ{current_row}"] = kmhoadon
            
            sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'
            soluongkm = math.floor(tongtien / ProcessHandler.tachtien_khuyenmai(kmhoadon))
            sheet[f"A{current_row}"] = entry_date
            sheet[f"AV{current_row}"] = songayno_MT
            sheet[f"C{current_row}"] = "Chưa thực hiện"
            sheet[f"D{current_row}"] = cancle_date

            sheet[f"E{current_row}"] = delivery
            sheet[f"G{current_row}"] = makhachhang
            sheet[f"L{current_row}"] = diengiai
            sheet[f"Q{current_row}"] = kiemtra[0]
            sheet[f"X{current_row}"] = soluongkm
            sheet[f"V{current_row}"] = kho
            sheet[f"Y{current_row}"] = 0
            sheet[f"Z{current_row}"] = 0
            sheet[f"U{current_row}"] = "Có"
            sheet[f"AJ{current_row}"] = khuvuc
            sheet[f"AE{current_row}"] = 8
            sheet[f"AM{current_row}"] = mien
            sheet[f"T{current_row}"] = "Không"
            sheet[f"AR{current_row}"] = product.get("Mahang", '')
            sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(kiemtra[0])


            trongluong += (ProcessHandler.timtrongluong_sanpham( kiemtra[0]) * soluongkm)
            
            sokienhang += math.ceil(soluongkm / ProcessHandler.timquycach_sanpham(kiemtra[0]))
            sheet[f"AU{current_row}"] = math.ceil(soluongkm / ProcessHandler.timquycach_sanpham(kiemtra[0]))
            sheet[f"AT{current_row}"] = ProcessHandler.timtrongluong_sanpham( kiemtra[0]) * soluongkm

            cachbokem = ProcessHandler.laycachbo_khuyenmai(kmhoadon)
            if cachbokem:
                sheet[f"AO{current_row}"] = cachbokem
            else:
                sheet[f"AO{current_row}"] = f'KM Bó Kèm - Che Barcode'

            self.log_signal.emit(f'🎉 Đơn hàng có chương trình khuyến mãi <b><span style="color: green;">{kmhoadon}</span></b> 🛒')
            self.log_signal.emit(f'Tổng tiền đơn hàng: <b><span style="color: green;">{tongtien} </span></b>\nSố lượng tặng: <b><span style="color: green;">{soluongkm} </span></b> 🛒')
            



            sheet["G1"] = STT_donhang
            current_row += 1
        sheet[f"L{start_row}"] = f'{diengiai} (Tổng trọng lượng: {ProcessHandler.format_weight_kg(trongluong)})'
        # Lưu lại file Excel
        wb.save(file_path)
        
        ProcessHandler.ghi_message(f"tong_tien: {tongtien:,.0f}")
        ProcessHandler.ghi_message(f"tong_trongluong: {ProcessHandler.format_weight_kg(trongluong)}")
        ProcessHandler.ghi_message(f"sai_gia: {saigia}")
        ProcessHandler.ghi_message(f"tong_kienhang: {sokienhang}")
        now2 = datetime.now()
        now_str = now2.strftime("%d/%m/%Y %H:%M:%S")
        ProcessHandler.ghi_message(f"end_time: {now_str}")
        time = now2 - now
        ProcessHandler.ghi_message(f"time: {ProcessHandler.format_timedelta(time)}]")  

        print(f"✅ Đã ghi {len(products)} dòng vào '{file_path}', sheet '{sheet_name}'!")
        return saigia






    def write_to_dondathang_winmart(self,products, makhachhang, po_number, entry_date, cancle_date,STT_donhang,vendor,diachigiaohang,ghichu,file_url):
        now = datetime.now()
        now_str = now.strftime("%d/%m/%Y %H:%M:%S")
        vendor = vendor.upper()
        """Ghi danh sách sản phẩm vào file dondathang.xlsx"""
        file_path = "dondathang.xlsx"
        sheet_name = "Don dat hang"
        
        
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]

        start_row = sheet.max_row + 1
        current_row = start_row
        # Lấy giá trị G1, đảm bảo nó là số
        print(f"📍 PO Number {po_number} ")
        STT_donhang_str = f"-{po_number}"   
        saigia = 0  
        tongtien = 0
        trongluong = 0
        sokienhang = 0
        print(makhachhang)
        
        if makhachhang[:2] == "MB":
            khuvuc = "MT_MB"
            kho = "TP_HN_12"
            mien = "HN"
        else:
            kho = "LA_KHO2026"
            khuvuc = "MT_MN"
            mien = "LA"

        if makhachhang == 'MN_MT_WIN1326':
            kho = "TP_DN_1"
            khuvuc = "MT_MN"
            mien = "DN"

        
        ProcessHandler.ghi_message(f"[PO: {po_number}")
        ProcessHandler.ghi_message(f"store: {diachigiaohang}")
        ProcessHandler.ghi_message(f"vendor: {vendor}")
        ProcessHandler.ghi_message(f"Mã Khách hàng: {makhachhang}")
        ProcessHandler.ghi_message(f"start_time: {now_str}")
        if file_url:
            ProcessHandler.ghi_message(f"url: {file_url}")
        
    



        if ghichu:
            po_number = f'{po_number} - {ghichu}'

        diengiai = f"{vendor} PO{po_number}"


        sheet[f"A{current_row}"] = entry_date
        sheet[f"AV{current_row}"] = songayno_MT
       
        sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'
        sheet[f"C{current_row}"] = "Chưa thực hiện"
        sheet[f"D{current_row}"] = cancle_date
        sheet[f"G{current_row}"] = makhachhang
        sheet[f"L{current_row}"] = diengiai
        sheet[f"V{current_row}"] = kho
        sheet[f"AE{current_row}"] = 8
        sheet[f"AJ{current_row}"] = khuvuc
        sheet[f"AM{current_row}"] = mien
        sheet[f"U{current_row}"] = "Không"
        sheet[f"Z{current_row}"] = 0
        sheet[f"S{current_row}"] = f"{vendor} PO{po_number.split('-')[0]}"
        sheet[f"T{current_row}"] = "Có"
        sheet[f"X{current_row}"] = 0
        sheet[f"Y{current_row}"] = 0
        sheet[f"E{current_row}"] = diachigiaohang
        current_row += 1


        for product in products:


            '''"Barcode": match.group("barcode"),  # Dùng mã vạch làm mã hàng
                    "Total Price": match.group("amount").replace(",", ""),
                    "OU Qty": match.group("quantity").replace(",", "")'''
            
            


            barcode = product["Barcode"]
            soluong = float(product["OU Qty"])
            dongia = float(product["Total Price"])



            ########################################## KHUYẾN MÃI WINMART #######################################
            if dongia == 0 and current_row - 2 >= 9:
                sheet[f"AO{current_row - 2}"] = 'KM Giao Rời - Không Che' #Có km trên PO
                sheet[f"AP{current_row - 2}"] = ''
                sheet[f"AP{current_row - 1}"] = ''
                continue
           
            sku_mapping = ProcessHandler.load_sku_mapping()  # Load mapping SKU
            products = ProcessHandler.replace_sku_numbers(products, sku_mapping)  # Thay SKU
            sheet[f"A{current_row}"] = entry_date
            sheet[f"AV{current_row}"] = songayno_MT
 
            sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'
            sheet[f"C{current_row}"] = "Chưa thực hiện"
            sheet[f"D{current_row}"] = cancle_date
            sheet[f"G{current_row}"] = makhachhang
            sheet[f"L{current_row}"] = diengiai
            sheet[f"Q{current_row}"] = product["Barcode"]
            sheet[f"V{current_row}"] = kho
            sheet[f"AE{current_row}"] = 8
            sheet[f"AJ{current_row}"] = khuvuc
            sheet[f"AM{current_row}"] = mien
            sheet[f"U{current_row}"] = "Không"
            sheet[f"T{current_row}"] = "Không"
            sheet[f"E{current_row}"] = diachigiaohang
            sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(product["Barcode"])

            # ✅ Xử lý "Extended Cost"
            extended_cost = product["Total Price"]
            if isinstance(extended_cost, str):  # Nếu là chuỗi, loại bỏ dấu phẩy
                extended_cost = float(extended_cost.replace(",", ""))
            extended_cost = int(extended_cost) if extended_cost.is_integer() else extended_cost
            sheet[f"Z{current_row}"] = f"=Y{current_row}*X{current_row}"

            # ✅ Xử lý "Qty Ord/Pcs"
            qty_ord_pcs = product["OU Qty"]
            if isinstance(qty_ord_pcs, str):  # Nếu là chuỗi, loại bỏ dấu phẩy
                qty_ord_pcs = float(qty_ord_pcs.replace(",", ""))
            qty_ord_pcs = int(qty_ord_pcs) if qty_ord_pcs.is_integer() else qty_ord_pcs
            sheet[f"X{current_row}"] = qty_ord_pcs

            trongluong += (ProcessHandler.timtrongluong_sanpham(product["Barcode"]) * qty_ord_pcs)

            sokienhang += math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(product["Barcode"]))
            sheet[f"AU{current_row}"] = math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(product["Barcode"]))

            sheet[f"AT{current_row}"] = ProcessHandler.timtrongluong_sanpham(product["Barcode"]) * qty_ord_pcs

            # ✅ Tính giá hóa đơn
            if vendor == 'WINMART':
                giahoadon = extended_cost / qty_ord_pcs
            elif vendor == 'BC MART':
                giahoadon = extended_cost
            
            
            giathuctegoc = ProcessHandler.find_price_by_sku(product["Barcode"],vendor)
            giathucte = giathuctegoc
            # 🔍 **Tìm giá trị trong CTKM.xlsx theo SKU & thời gian**
            
            results  = ProcessHandler.find_all_promotions_by_sku_and_time(product["Barcode"], entry_date,vendor)
            slkhuyenmai = len(results)
            print(f'Giá thực tế: {giathucte}')
            print(f'🔎 Số lượng CTKM khớp: {slkhuyenmai}')
            khuyenmai = ''



            giakhop = False  # Cờ kiểm tra khớp giá

            for col, value in results:
                print(f'🧪 Thử CTKM: {col} → {value}')
                
                if value:
                    khuyenmai = value
                    discount = ProcessHandler.extract_discount(value)
                    print(f'🔻 Giảm giá: {discount} ({type(discount)})')

                    if discount != 0:
                        if isinstance(discount, str):
                            discount = discount.replace("%", "").strip()
                            discount = float(discount) if discount else 0

                        if isinstance(giathuctegoc, str):
                            giathucte = giathuctegoc.replace(",", "").strip()
                            giathucte = float(giathucte) if giathucte else 0

                        giathucte = giathucte - (giathucte * float(discount) / 100)
                    else:
                        giathucte = giathucte



                    giahoadon = float(giahoadon[0]) if isinstance(giahoadon, tuple) else float(giahoadon)
                    if isinstance(giathucte, str):
                        giathucte = giathucte.replace(",", "").strip()
                        giathucte = float(giathucte) if giathucte else 0

                    if math.isclose(giahoadon, giathucte, rel_tol=1e-4):
                        sheet[f"Y{current_row}"] = giathucte
                        self.log_signal.emit(
                        f'🎉 Mã hàng <b><span style="color: blue;">{product["Barcode"]}</span></b> '
                        f'có CTKM <b><span style="color: green;">{value}</span></b> từ <i>{col}</i> 🛒')
                        self.log_signal.emit(
                            f'✅ Mã hàng <b><span style="color: blue;">{product["Barcode"]}</span></b> '
                            f'có giá chính xác 🎯: <span style="color: green;"><b>{giahoadon}</b></span>'
                        )
                        
                        giakhop = True
                        break  # ✅ THOÁT VÒNG LẶP khi đã khớp giá
                if giakhop == True: 
                    break

            if not results:
                khuyenmai = ''
                giahoadon = float(giahoadon[0]) if isinstance(giahoadon, tuple) else float(giahoadon)
                if isinstance(giathucte, str):
                    giathucte = giathucte.replace(",", "").strip()
                    giathucte = float(giathucte) if giathucte else 0

                if math.isclose(giahoadon, giathucte, rel_tol=1e-4):
                    sheet[f"Y{current_row}"] = giathucte
                    self.log_signal.emit(
                    f'🎉 Mã hàng <b><span style="color: blue;">{product["Barcode"]}</span></b> '
                    f'có CTKM <b><span style="color: green;">{khuyenmai}</span></b>  🛒')
                    self.log_signal.emit(
                        f'✅ Mã hàng <b><span style="color: blue;">{product["Barcode"]}</span></b> '
                        f'có giá chính xác 🎯: <span style="color: green;"><b>{giahoadon}</b></span>'
                    )
                    giakhop = True



            # ❌ Sau khi thử hết, không khớp giá nào:
            if not giakhop:
                sheet[f"Y{current_row}"] = giathucte
                red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                sheet[f"Y{current_row}"].fill = red_fill
                comment_text = f"Kiểm tra lại giá mã này! - Giá hóa đơn: {giahoadon} - Chênh lệch: {giahoadon - giathucte}"
                comment = Comment(comment_text, "System")
                sheet[f"Y{current_row}"].comment = comment

                self.log_signal.emit(
                    f"⚠️ Mã hàng <b>{product["Barcode"]}</b> sai giá! 🛑 "
                    f"Hóa đơn: <b><span style='color: red;'>{giahoadon}</span></b> – "
                    f"Giá hệ thống: <b><span style='color: green;'>{giathucte}</span></b>"
                )
                
                ProcessHandler.ghi_message(f"text: Mã hàng {product["Barcode"]} - {ProcessHandler.timten_sanpham(product["Barcode"])}, Giá trên PO {giahoadon:,}, Giá đúng: {giathucte:,}. Chênh lệch {(giahoadon - giathucte):,.0f}")


                saigia += 1

            tongtien += giathucte * qty_ord_pcs
                

            # 🔍 Kiểm tra SKU có trong 'SanPham'
            kiemtra = ProcessHandler.check_value_in_sanpham(khuyenmai)
            sheet[f"AQ{current_row}"] = khuyenmai
            match = re.search(r"(\d+)\s*\+\s*1", khuyenmai)  # Tìm số đầu tiên trong biểu thức X+1
            if match:
                x = int(match.group(1))  # Chuyển đổi thành số nguyên
                kiemtra = kiemtra or barcode
                
                if x >= 2:  # Chỉ chia khi x >= 2, còn 1+1 thì giữ nguyên
                    qty_ord_pcs = math.floor(qty_ord_pcs / x)  # Chia rồi làm tròn xuống
               
            if kiemtra:
                if isinstance(kiemtra, list):  # Nếu có nhiều mã SKU, nối thành chuỗi
                    kiemtra = ", ".join(kiemtra)


                cachbokem = ProcessHandler.laycachbo_khuyenmai(value)
                if cachbokem:
                    sheet[f"AO{current_row}"] = cachbokem
                    cachbokem_lower = cachbokem.lower()
                    if "bó kèm" in cachbokem_lower or "quấn kèm" in cachbokem_lower:
                        sheet[f"AP{current_row}"] = f'{ProcessHandler.layduoi_mahang(product["Barcode"])}_{ProcessHandler.layduoi_mahang(kiemtra)}_1'
                        sheet[f"AP{current_row + 1}"] = f'{ProcessHandler.layduoi_mahang(product["Barcode"])}_{ProcessHandler.layduoi_mahang(kiemtra)}_1'
                else:
                    sheet[f"AO{current_row}"] = f'KM Giao Rời - Không Che Barcode'
                current_row += 1
                self.log_signal.emit(
                    f'🎁 Mã hàng <b><span style="color: blue;">{barcode}</span></b> '
                    f'đã thêm hàng khuyến mãi <b><span style="color: green;">{kiemtra}</span></b> tặng kèm! 🎉'
                )
                sheet[f"A{current_row}"] = entry_date
                sheet[f"AV{current_row}"] = songayno_MT
             
                sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'
                sheet[f"C{current_row}"] = "Chưa thực hiện"
                sheet[f"D{current_row}"] = cancle_date
                sheet[f"G{current_row}"] = makhachhang
                sheet[f"L{current_row}"] = diengiai
                sheet[f"Q{current_row}"] = kiemtra
                sheet[f"X{current_row}"] = qty_ord_pcs
                sheet[f"V{current_row}"] = kho
                sheet[f"Y{current_row}"] = 0
                sheet[f"Z{current_row}"] = 0
                sheet[f"U{current_row}"] = "Có"
                sheet[f"AJ{current_row}"] = khuvuc
                sheet[f"AE{current_row}"] = 8
                sheet[f"AM{current_row}"] = mien
                sheet[f"T{current_row}"] = "Không"
                sheet[f"E{current_row}"] = diachigiaohang
                sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(kiemtra)

                trongluong += (ProcessHandler.timtrongluong_sanpham(kiemtra) * qty_ord_pcs)
                sokienhang += math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(kiemtra))
                sheet[f"AU{current_row}"] = math.ceil(qty_ord_pcs / ProcessHandler.timquycach_sanpham(kiemtra))

                sheet[f"AT{current_row}"] = ProcessHandler.timtrongluong_sanpham(kiemtra) * qty_ord_pcs

            sheet["G1"] = STT_donhang
            current_row += 1

        kmhoadon = ProcessHandler.find_all_promotions_by_sku_and_time("Hóa Đơn", entry_date,vendor)
        print(kmhoadon)
        if kmhoadon:
            kmhoadon = str(kmhoadon[0][1])
            kiemtra = ProcessHandler.check_value_in_sanpham(kmhoadon) 
            sheet[f"AQ{current_row}"] = kmhoadon
      
            sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'
            soluongkm = math.floor(tongtien / ProcessHandler.tachtien_khuyenmai(kmhoadon))
            sheet[f"A{current_row}"] = entry_date
            sheet[f"AV{current_row}"] = songayno_MT
            sheet[f"C{current_row}"] = "Chưa thực hiện"
            sheet[f"D{current_row}"] = cancle_date
            sheet[f"E{current_row}"] = diachigiaohang
            sheet[f"G{current_row}"] = makhachhang
            sheet[f"L{current_row}"] = diengiai
            sheet[f"Q{current_row}"] = kiemtra[0]
            sheet[f"X{current_row}"] = soluongkm
            sheet[f"V{current_row}"] = kho
            sheet[f"Y{current_row}"] = 0
            sheet[f"Z{current_row}"] = 0
            sheet[f"U{current_row}"] = "Có"
            sheet[f"AJ{current_row}"] = khuvuc
            sheet[f"AE{current_row}"] = 8
            sheet[f"AM{current_row}"] = mien
            sheet[f"T{current_row}"] = "Không"
            sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(kiemtra[0])

            trongluong += (ProcessHandler.timtrongluong_sanpham(kiemtra[0]) * soluongkm)
            sokienhang += math.ceil(soluongkm / ProcessHandler.timquycach_sanpham(kiemtra[0]))
            sheet[f"AU{current_row}"] = math.ceil(soluongkm / ProcessHandler.timquycach_sanpham(kiemtra[0]))
            sheet[f"AT{current_row}"] = ProcessHandler.timtrongluong_sanpham(kiemtra[0]) * soluongkm
            
            cachbokem = ProcessHandler.laycachbo_khuyenmai(kmhoadon)
            if cachbokem:
                sheet[f"AO{current_row}"] = cachbokem
            else:
                sheet[f"AO{current_row}"] = f'KM Bó Kèm - Che Barcode'

            self.log_signal.emit(f'🎉 Đơn hàng có chương trình khuyến mãi <b><span style="color: green;">{kmhoadon}</span></b> 🛒')
            self.log_signal.emit(f'Tổng tiền đơn hàng: <b><span style="color: green;">{tongtien} </span></b>\nSố lượng tặng: <b><span style="color: green;">{soluongkm} </span></b> 🛒')
            

        sheet[f"L{start_row}"] = f'{diengiai} (Tổng trọng lượng: {ProcessHandler.format_weight_kg(trongluong)})'
        # Lưu lại file Excel
        wb.save(file_path)

        ProcessHandler.ghi_message(f"tong_tien: {tongtien:,.0f}")
        ProcessHandler.ghi_message(f"tong_trongluong: {ProcessHandler.format_weight_kg(trongluong)}")
        
        ProcessHandler.ghi_message(f"tong_kienhang: {sokienhang}")
        ProcessHandler.ghi_message(f"sai_gia: {saigia}")
        now2 = datetime.now()
        now_str = now2.strftime("%d/%m/%Y %H:%M:%S")
        ProcessHandler.ghi_message(f"end_time: {now_str}")
        time = now2 - now
        ProcessHandler.ghi_message(f"time: {ProcessHandler.format_timedelta(time)}]")  
        print(f"✅ Đã ghi {len(products)} dòng vào '{file_path}', sheet '{sheet_name}'!")
        return saigia
        



        

    def tim_gia_tri_congtrinh(congtrinh):
        file_path = "data.xlsx"
        sheet_name = "MaKH"

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb[sheet_name]

            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=3, values_only=True):
                ma_ct, gia_tri = row  # Cột B và C
                if ma_ct == congtrinh:
                    return gia_tri  # Trả về giá trị cột C

            return congtrinh.replace(" ", "")  # Không tìm thấy

        except Exception as e:
            print(f"Lỗi khi mở file: {e}")
            return None

    



    def write_to_dondathang_bigc(self,products,items, po_number, entry_date, cancle_date, congtrinh,STT_donhang ,makhachhang, vendor,page_num,diachigiao,bat_dau,url):

        vendor = vendor.upper()

        """Ghi danh sách sản phẩm vào file dondathang.xlsx"""
        file_path = "dondathang.xlsx"
        sheet_name = "Don dat hang"
       
        
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]

        start_row = sheet.max_row + 1
        current_row = start_row
        ten_store = congtrinh
        congtrinh = ProcessHandler.tim_gia_tri_congtrinh(congtrinh)


        STT_donhang_str = f"-{po_number}"    

        items = ProcessHandler.ghepgia_donhangbigc(items,products) 

        sku_mapping = ProcessHandler.load_sku_mapping()  # Load mapping SKU
        items = ProcessHandler.replace_sku_numbers(items, sku_mapping)  # Thay SKU
        saigia = 0
        tongtien = 0
        trongluong = 0
        if makhachhang[:2] == "MB":
            khuvuc = "MT_MB"
            kho = "TP_HN_12"
            mien = "HN"
        else:
            kho = "LA_KHO2026"
            khuvuc = "MT_MN"
            mien = "LA"
        if page_num == 1:
            sheet[f"A{current_row}"] = entry_date
            sheet[f"AV{current_row}"] = songayno_MT
            sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'
            sheet[f"C{current_row}"] = "Chưa thực hiện"
            sheet[f"D{current_row}"] = cancle_date
            sheet[f"G{current_row}"] = makhachhang
            sheet[f"L{current_row}"] = f"{vendor} PO{po_number}"
            sheet[f"V{current_row}"] = kho
            sheet[f"AE{current_row}"] = 8
            sheet[f"AJ{current_row}"] = khuvuc
            sheet[f"AM{current_row}"] = mien
            sheet[f"U{current_row}"] = "Không"
            sheet[f"Z{current_row}"] = 0
            sheet[f"S{current_row}"] = f"{vendor} PO{po_number}"
            sheet[f"T{current_row}"] = "Có"
            sheet[f"X{current_row}"] = 0
            sheet[f"Y{current_row}"] = 0
            sheet[f"AN{current_row}"] = "BIGCANLAC"
            sheet[f"E{current_row}"] = diachigiao
            current_row += 1

        for item in items:
            if ProcessHandler.timten_sanpham(item["Barcode"]) == "Không thấy tên sản phẩm":
                continue
                

            sheet[f"A{current_row}"] = entry_date
            sheet[f"AV{current_row}"] = songayno_MT
          
            sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'
            sheet[f"C{current_row}"] = "Chưa thực hiện"
            sheet[f"D{current_row}"] = cancle_date
            sheet[f"L{current_row}"] = f"{vendor} PO{po_number}"
            sheet[f"G{current_row}"] = makhachhang
            sheet[f"Q{current_row}"] = item["Barcode"]
            sheet[f"V{current_row}"] = kho
            sheet[f"AE{current_row}"] = 8
            sheet[f"AJ{current_row}"] = khuvuc
            sheet[f"AM{current_row}"] = mien
            sheet[f"U{current_row}"] = "Không"
            sheet[f"AN{current_row}"] = congtrinh #viết mã công trình
            sheet[f"T{current_row}"] = "Không"
            sheet[f"E{current_row}"] = diachigiao
            sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(item["Barcode"])
            



            sku_ou = item["SKU/OU"]
            if isinstance(sku_ou, str):  # Nếu là chuỗi, loại bỏ dấu phẩy
                sku_ou = float(sku_ou.replace(",", ""))
            sku_ou = int(sku_ou) if sku_ou.is_integer() else sku_ou

            ou_qty = item["OU Qty"]
            if isinstance(ou_qty, str):  # Nếu là chuỗi, loại bỏ dấu phẩy
                ou_qty = float(ou_qty.replace(",", ""))
            ou_qty = int(ou_qty) if ou_qty.is_integer() else ou_qty

            qty_ord_pcs = ou_qty * sku_ou

            trongluong += (ProcessHandler.timtrongluong_sanpham(item["Barcode"]) * qty_ord_pcs)
            sheet[f"AT{current_row}"] = ProcessHandler.timtrongluong_sanpham(item["Barcode"]) * qty_ord_pcs

            sheet[f"X{current_row}"] = qty_ord_pcs #ghi số lượng
            sheet[f"Z{current_row}"] = f"=Y{current_row}*X{current_row}"


            
            giahoadon = float(item["Total Price"])

            giathuctegoc = ProcessHandler.find_price_by_sku(item["Barcode"],vendor)
            giathucte = giathuctegoc
            # 🔍 **Tìm giá trị trong CTKM.xlsx theo SKU & thời gian**
            
            results  = ProcessHandler.find_all_promotions_by_sku_and_time(item["Barcode"], entry_date,vendor)
            slkhuyenmai = len(results)
            print(f'Giá thực tế: {giathucte}')
            print(f'🔎 Số lượng CTKM khớp: {slkhuyenmai}')
            khuyenmai = ''



            giakhop = False  # Cờ kiểm tra khớp giá

            for col, value in results:
                print(f'🧪 Thử CTKM: {col} → {value}')
                
                if value:
                    khuyenmai = value
                    discount = ProcessHandler.extract_discount(value)
                    print(f'🔻 Giảm giá: {discount} ({type(discount)})')

                    if discount != 0:
                        if isinstance(discount, str):
                            discount = discount.replace("%", "").strip()
                            discount = float(discount) if discount else 0

                        if isinstance(giathuctegoc, str):
                            giathucte = giathuctegoc.replace(",", "").strip()
                            giathucte = float(giathucte) if giathucte else 0

                        giathucte = giathucte - (giathucte * float(discount) / 100)
                    else:
                        giathucte = giathucte



                    giahoadon = float(giahoadon[0]) if isinstance(giahoadon, tuple) else float(giahoadon)
                    if isinstance(giathucte, str):
                        giathucte = giathucte.replace(",", "").strip()
                        giathucte = float(giathucte) if giathucte else 0

                    if math.isclose(giahoadon, giathucte, rel_tol=1e-4):
                        sheet[f"Y{current_row}"] = giathucte
                        self.log_signal.emit(
                        f'🎉 Mã hàng <b><span style="color: blue;">{item["Barcode"]}</span></b> '
                        f'có CTKM <b><span style="color: green;">{value}</span></b> từ <i>{col}</i> 🛒')
                        self.log_signal.emit(
                            f'✅ Mã hàng <b><span style="color: blue;">{item["Barcode"]}</span></b> '
                            f'có giá chính xác 🎯: <span style="color: green;"><b>{giahoadon}</b></span>'
                        )
                        
                        giakhop = True
                        break  # ✅ THOÁT VÒNG LẶP khi đã khớp giá
            

            if not results:
                khuyenmai = ''
                giahoadon = float(giahoadon[0]) if isinstance(giahoadon, tuple) else float(giahoadon)
                if isinstance(giathucte, str):
                    giathucte = giathucte.replace(",", "").strip()
                    giathucte = float(giathucte) if giathucte else 0

                if math.isclose(giahoadon, giathucte, rel_tol=1e-4):
                    sheet[f"Y{current_row}"] = giathucte
                    self.log_signal.emit(
                    f'🎉 Mã hàng <b><span style="color: blue;">{item["Barcode"]}</span></b> '
                    f'có CTKM <b><span style="color: green;">{khuyenmai}</span></b> 🛒')
                    self.log_signal.emit(
                        f'✅ Mã hàng <b><span style="color: blue;">{item["Barcode"]}</span></b> '
                        f'có giá chính xác 🎯: <span style="color: green;"><b>{giahoadon}</b></span>'
                    )
                    giakhop = True



            # ❌ Sau khi thử hết, không khớp giá nào:
            if not giakhop:
                sheet[f"Y{current_row}"] = giathucte
                red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                sheet[f"Y{current_row}"].fill = red_fill
                comment_text = f"Kiểm tra lại giá mã này! - Giá hóa đơn: {giahoadon} - Chênh lệch: {giahoadon - giathucte}"
                comment = Comment(comment_text, "System")
                sheet[f"Y{current_row}"].comment = comment

                self.log_signal.emit(
                    f"⚠️ Mã hàng <b>{item["Barcode"]}</b> sai giá! 🛑 "
                    f"Hóa đơn: <b><span style='color: red;'>{giahoadon}</span></b> – "
                    f"Giá hệ thống: <b><span style='color: green;'>{giathucte}</span></b>"
                )
                saigia += 1
                
                ProcessHandler.ghi_message(f"text: Store: {ten_store} Mã hàng {item["Barcode"]} - {ProcessHandler.timten_sanpham(item["Barcode"])}, Giá trên PO {giahoadon:,}, Giá đúng: {giathucte:,}. Chênh lệch {(giahoadon - giathucte):,.0f}")


            tongtien += giathucte * qty_ord_pcs
            print(tongtien)
            
                

            # 🔍 Kiểm tra SKU có trong 'SanPham'
            kiemtra = ProcessHandler.check_value_in_sanpham(khuyenmai)
            match = re.search(r"(\d+)\s*\+\s*1", khuyenmai)  # Tìm số đầu tiên trong biểu thức X+1
            sheet[f"AQ{current_row}"] = khuyenmai
            if match:
                x = int(match.group(1))  # Chuyển đổi thành số nguyên
                kiemtra = kiemtra or item["Barcode"]
                
                if x >= 2:  # Chỉ chia khi x >= 2, còn 1+1 thì giữ nguyên
                    qty_ord_pcs = math.floor(qty_ord_pcs / x)  # Chia rồi làm tròn xuống
            
            if kiemtra:
                if isinstance(kiemtra, list):  # Nếu có nhiều mã SKU, nối thành chuỗi
                    kiemtra = ", ".join(kiemtra)

                cachbokem = ProcessHandler.laycachbo_khuyenmai(value)
                if cachbokem:
                    sheet[f"AO{current_row}"] = cachbokem
                    cachbokem_lower = cachbokem.lower()
                    if "bó kèm" in cachbokem_lower or "quấn kèm" in cachbokem_lower:
                        sheet[f"AP{current_row}"] = f'{ProcessHandler.layduoi_mahang(item["Barcode"])}_{ProcessHandler.layduoi_mahang(kiemtra)}_1'
                        sheet[f"AP{current_row + 1}"] = f'{ProcessHandler.layduoi_mahang(item["Barcode"])}_{ProcessHandler.layduoi_mahang(kiemtra)}_1'
                else:
                    sheet[f"AO{current_row}"] = f'KM Rời - Không Che Barcode'
                current_row += 1
                self.log_signal.emit(
f'🎁 Mã hàng <b><span style="color: blue;">{item["Barcode"]}</span></b> '
f'đã thêm hàng khuyến mãi <b><span style="color: green;">{kiemtra}</span></b> tặng kèm! 🎉'
)
                sheet[f"A{current_row}"] = entry_date
                sheet[f"AV{current_row}"] = songayno_MT
                
                sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'
                sheet[f"C{current_row}"] = "Chưa thực hiện"
                sheet[f"D{current_row}"] = cancle_date
                sheet[f"G{current_row}"] = makhachhang
                sheet[f"L{current_row}"] = f"{vendor} PO{po_number}"
                sheet[f"Q{current_row}"] = kiemtra
                sheet[f"X{current_row}"] = qty_ord_pcs
                sheet[f"V{current_row}"] = kho
                sheet[f"Y{current_row}"] = 0
                sheet[f"Z{current_row}"] = 0
                sheet[f"U{current_row}"] = "Có"
                sheet[f"AJ{current_row}"] = khuvuc
                sheet[f"AE{current_row}"] = 8
                sheet[f"AM{current_row}"] = mien
                sheet[f"AN{current_row}"] = congtrinh
                sheet[f"T{current_row}"] = "Không"
                sheet[f"E{current_row}"] = diachigiao
                sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(kiemtra)

                trongluong += (ProcessHandler.timtrongluong_sanpham(kiemtra) * qty_ord_pcs)
                sheet[f"AT{current_row}"] = ProcessHandler.timtrongluong_sanpham(kiemtra) * qty_ord_pcs
            
            current_row += 1
        sheet["G1"] = STT_donhang

        kmhoadon = ProcessHandler.find_all_promotions_by_sku_and_time("Hóa Đơn", entry_date,vendor)
        if kmhoadon:
            kmhoadon = str(kmhoadon[0][1])
            kiemtra = ProcessHandler.check_value_in_sanpham(kmhoadon) 
            sheet[f"AQ{current_row}"] = kmhoadon
           
            sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'
            soluongkm = math.floor(tongtien / ProcessHandler.tachtien_khuyenmai(kmhoadon))
            sheet[f"A{current_row}"] = entry_date
            sheet[f"AV{current_row}"] = songayno_MT
            sheet[f"C{current_row}"] = "Chưa thực hiện"
            sheet[f"D{current_row}"] = cancle_date
            sheet[f"E{current_row}"] = diachigiao
            sheet[f"G{current_row}"] = makhachhang
            sheet[f"L{current_row}"] = f"{vendor} PO{po_number}"
            sheet[f"Q{current_row}"] = kiemtra[0]
            sheet[f"X{current_row}"] = soluongkm
            sheet[f"V{current_row}"] = kho
            sheet[f"Y{current_row}"] = 0
            sheet[f"Z{current_row}"] = 0
            sheet[f"U{current_row}"] = "Có"
            sheet[f"AJ{current_row}"] = khuvuc
            sheet[f"AE{current_row}"] = 8
            sheet[f"AM{current_row}"] = mien
            sheet[f"T{current_row}"] = "Không"
            sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(kiemtra[0])

            trongluong += (ProcessHandler.timtrongluong_sanpham(kiemtra[0]) * soluongkm)
            sheet[f"AT{current_row}"] = ProcessHandler.timtrongluong_sanpham(kiemtra[0]) * soluongkm
            
            cachbokem = ProcessHandler.laycachbo_khuyenmai(kmhoadon)
            if cachbokem:
                sheet[f"AO{current_row}"] = cachbokem
            else:
                sheet[f"AO{current_row}"] = f'KM Bó Kèm - Che Barcode'

            self.log_signal.emit(f'🎉 Đơn hàng có chương trình khuyến mãi <b><span style="color: green;">{kmhoadon}</span></b> 🛒')
            self.log_signal.emit(f'Tổng tiền đơn hàng: <b><span style="color: green;">{tongtien} </span></b>\nSố lượng tặng: <b><span style="color: green;">{soluongkm} </span></b> 🛒')
        
        sheet[f"L{start_row}"] = f'{vendor} PO{po_number} (Tổng trọng lượng: {ProcessHandler.format_weight_kg(trongluong)})'

        if bat_dau:
            if url:
                ProcessHandler.ghi_message(f"url: {url}")
            col_AT = 46
            col_Z  = 26

            total_AT = 0
            total_Z  = 0

            for r in range(bat_dau, current_row):

                # ─── Tổng AT (giá trị số) ───
                v_AT = sheet.cell(row=r, column=col_AT).value
                if isinstance(v_AT, (int, float)):
                    total_AT += v_AT

                # ─── Tổng Z (công thức Y*X) ───
                formula = sheet.cell(row=r, column=col_Z).value
                if isinstance(formula, str) and "*" in formula:
                    left, right = formula.replace("=", "").split("*")

                    v1 = sheet[left.strip()].value
                    v2 = sheet[right.strip()].value

                    if isinstance(v1, (int, float)) and isinstance(v2, (int, float)):
                        total_Z += v1 * v2



            
            ProcessHandler.ghi_message(f"tong_tien: {total_Z:,.0f}")

            print("Tổng AT:", total_AT)


            
            
            ProcessHandler.ghi_message(f"tong_trongluong: {ProcessHandler.format_weight_kg(total_AT)}")
           


            sheet[f"L{bat_dau}"] = f'{vendor} PO{po_number} (Tổng trọng lượng: {ProcessHandler.format_weight_kg(total_AT)})'
        # Lưu lại file Excel
        wb.save(file_path)
        print(f"✅ Đã ghi {len(products)} dòng vào '{file_path}', sheet '{sheet_name}'!")
        return saigia,tongtien

    def write_to_dondathang_emart(self,items, po_number, entry_date, cancle_date, congtrinh,STT_donhang ,makhachhang="MN_MT_KH0032", vendor="Emart", url = None):

        now = datetime.now()
        now_str = now.strftime("%d/%m/%Y %H:%M:%S")
        vendor = vendor.upper()
        """Ghi danh sách sản phẩm vào file dondathang.xlsx"""
        file_path = "dondathang.xlsx"
        sheet_name = "Don dat hang"
        
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]
        diachigiaohang = congtrinh

        start_row = sheet.max_row + 1
        current_row = start_row

        mapping = {
    "EMART GO VAP": "PVT",
    "EMART PHI": "PHI",
    "EMART SALA": "SALA"
}
        
        congtrinh = mapping.get(congtrinh, congtrinh)



        STT_donhang_str = f"-{po_number}"    

        sku_mapping = ProcessHandler.load_sku_mapping()  # Load mapping SKU
        items = ProcessHandler.replace_sku_numbers(items, sku_mapping)  # Thay SKU
        saigia = 0
        tongtien = 0
        trongluong = 0
        diengiai = f"{vendor} PO{po_number}"
        if makhachhang[:2] == "MB":
            khuvuc = "MT_MB"
            kho = "TP_HN_12"
            mien = "HN"
        else:
            kho = "LA_KHO2026"
            khuvuc = "MT_MN"
            mien = "LA"

        ProcessHandler.ghi_message(f"[PO: {po_number}")
        ProcessHandler.ghi_message(f"store: {diachigiaohang}")
        ProcessHandler.ghi_message(f"vendor: {vendor}")
        ProcessHandler.ghi_message(f"Mã Khách hàng: {makhachhang}")
        ProcessHandler.ghi_message(f"start_time: {now_str}")
        if url:
            ProcessHandler.ghi_message(f"url: {url}")


        sheet[f"A{current_row}"] = entry_date
        sheet[f"AV{current_row}"] = songayno_MT
        sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'
        sheet[f"C{current_row}"] = "Chưa thực hiện"
        sheet[f"D{current_row}"] = cancle_date
        sheet[f"G{current_row}"] = makhachhang
        sheet[f"L{current_row}"] = diengiai
        sheet[f"V{current_row}"] = kho
        sheet[f"AE{current_row}"] = 8
        sheet[f"AJ{current_row}"] = khuvuc
        sheet[f"AM{current_row}"] = mien
        sheet[f"U{current_row}"] = "Không"
        sheet[f"Z{current_row}"] = 0
        sheet[f"S{current_row}"] = f"{vendor} PO{po_number}"
        sheet[f"T{current_row}"] = "Có"
        sheet[f"X{current_row}"] = 0
        sheet[f"Y{current_row}"] = 0
        sheet[f"E{current_row}"] = diachigiaohang
        if congtrinh == "PVT":
            sheet[f"K{current_row}"] = "SIÊU THỊ EMART PHAN VĂN TRỊ"
        elif congtrinh == "SALA":
            sheet[f"K{current_row}"] = "SIÊU THỊ EMART SALA"
        elif congtrinh == "PHI":
            sheet[f"K{current_row}"] = "SIÊU THỊ EMART PHAN HUY ÍCH"

        
        sheet[f"AN{current_row}"] = 'PVT' #viết mã công trình
        current_row += 1

        for item in items:
            
            sheet[f"A{current_row}"] = entry_date
            sheet[f"AV{current_row}"] = songayno_MT
            
            sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'
            sheet[f"C{current_row}"] = "Chưa thực hiện"
            sheet[f"D{current_row}"] = cancle_date
            sheet[f"L{current_row}"] = diengiai
            sheet[f"G{current_row}"] = makhachhang
            sheet[f"Q{current_row}"] = item["Barcode"]
            sheet[f"V{current_row}"] = kho
            sheet[f"AE{current_row}"] = 8
            sheet[f"AJ{current_row}"] = khuvuc
            sheet[f"AM{current_row}"] = mien
            sheet[f"U{current_row}"] = "Không"
            sheet[f"E{current_row}"] = diachigiaohang

            sheet[f"AN{current_row}"] = congtrinh #viết mã công trình
            sheet[f"T{current_row}"] = "Không"
            sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(item["Barcode"])
            sheet[f"Z{current_row}"] = f"=Y{current_row}*X{current_row}"
            

            ou_qty = item["OU Qty"]
            if isinstance(ou_qty, str):  # Nếu là chuỗi, loại bỏ dấu phẩy
                ou_qty = float(ou_qty.replace(",", ""))
            ou_qty = int(ou_qty) if ou_qty.is_integer() else ou_qty

            qty_ord_pcs = ou_qty 

            sheet[f"X{current_row}"] = qty_ord_pcs #ghi số lượng


            trongluong += (ProcessHandler.timtrongluong_sanpham(item["Barcode"]) * qty_ord_pcs)
            sheet[f"AT{current_row}"] = ProcessHandler.timtrongluong_sanpham(item["Barcode"]) * qty_ord_pcs


            giahoadon = float(item["Total Price"])

            giathuctegoc = ProcessHandler.find_price_by_sku(item["Barcode"],vendor)
            giathucte = giathuctegoc
            # 🔍 **Tìm giá trị trong CTKM.xlsx theo SKU & thời gian**
            
            results  = ProcessHandler.find_all_promotions_by_sku_and_time(item["Barcode"], entry_date,vendor)
            
            slkhuyenmai = len(results)
            print(f'Giá thực tế: {giathucte}')
            print(f'🔎 Số lượng CTKM khớp: {slkhuyenmai}')
            khuyenmai = ''




            giakhop = False  # Cờ kiểm tra khớp giá

            for col, value in results:
                print(f'🧪 Thử CTKM: {col} → {value}')

                if value:
                    khuyenmai = value
                    giathucte = giathuctegoc  # Reset về giá gốc trước mỗi lần thử CTKM
                    discount = ProcessHandler.extract_discount(value)
                    print(f'🔻 Giảm giá: {discount} ({type(discount)})')

                    if discount != 0:
                        if isinstance(discount, str):
                            discount = discount.replace("%", "").strip()
                            discount = float(discount) if discount else 0

                        if isinstance(giathucte, str):
                            giathucte = giathucte.replace(",", "").strip()
                            giathucte = float(giathucte) if giathucte else 0

                        giathucte = giathucte - (giathucte * float(discount) / 100)

                    giahoadon = float(giahoadon[0]) if isinstance(giahoadon, tuple) else float(giahoadon)
                    if isinstance(giathucte, str):
                        giathucte = giathucte.replace(",", "").strip()
                        giathucte = float(giathucte) if giathucte else 0

                    if math.isclose(giahoadon, giathucte, rel_tol=1e-4):
                        sheet[f"Y{current_row}"] = giathucte
                        self.log_signal.emit(
                        f'🎉 Mã hàng <b><span style="color: blue;">{item["Barcode"]}</span></b> '
                        f'có CTKM <b><span style="color: green;">{value}</span></b> từ <i>{col}</i> 🛒')
                        self.log_signal.emit(
                            f'✅ Mã hàng <b><span style="color: blue;">{item["Barcode"]}</span></b> '
                            f'có giá chính xác 🎯: <span style="color: green;"><b>{giahoadon}</b></span>'
                        )

                        giakhop = True
                        break  # ✅ THOÁT VÒNG LẶP khi đã khớp giá

            if not results:
                khuyenmai = ''
                giahoadon = float(giahoadon[0]) if isinstance(giahoadon, tuple) else float(giahoadon)
                if isinstance(giathucte, str):
                    giathucte = giathucte.replace(",", "").strip()
                    giathucte = float(giathucte) if giathucte else 0

                if math.isclose(giahoadon, giathucte, rel_tol=1e-4):
                    sheet[f"Y{current_row}"] = giathucte
                    self.log_signal.emit(
                    f'🎉 Mã hàng <b><span style="color: blue;">{item["Barcode"]}</span></b> '
                    f'có CTKM <b><span style="color: green;">{khuyenmai}</span></b> 🛒')
                    self.log_signal.emit(
                        f'✅ Mã hàng <b><span style="color: blue;">{item["Barcode"]}</span></b> '
                        f'có giá chính xác 🎯: <span style="color: green;"><b>{giahoadon}</b></span>'
                    )
                    giakhop = True



            # ❌ Sau khi thử hết, không khớp giá nào:
            if not giakhop:
                sheet[f"Y{current_row}"] = giathucte
                red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                sheet[f"Y{current_row}"].fill = red_fill
                comment_text = f"Kiểm tra lại giá mã này! - Giá hóa đơn: {giahoadon} - Chênh lệch: {giahoadon - giathucte}"
                comment = Comment(comment_text, "System")
                sheet[f"Y{current_row}"].comment = comment

                self.log_signal.emit(
                    f"⚠️ Mã hàng <b>{item["Barcode"]}</b> sai giá! 🛑 "
                    f"Hóa đơn: <b><span style='color: red;'>{giahoadon}</span></b> – "
                    f"Giá hệ thống: <b><span style='color: green;'>{giathucte}</span></b>"
                )

                ProcessHandler.ghi_message(f"text: Mã hàng {item["Barcode"]} - {ProcessHandler.timten_sanpham(item["Barcode"])}, Giá trên PO {giahoadon:,}, Giá đúng: {giathucte:,}. Chênh lệch {(giahoadon - giathucte):,.0f}")

                saigia += 1

            tongtien += giathucte * qty_ord_pcs


            # 🔍 Kiểm tra SKU có trong 'SanPham'


            nhieuCtkm = khuyenmai.split('|')

            for i, hangkm in enumerate(nhieuCtkm):
                kiemtra = ProcessHandler.check_value_in_sanpham(hangkm)
                match = re.search(r"(\d+)\s*\+\s*1", hangkm)  # Tìm số đầu tiên trong biểu thức X+1
                sheet[f"AQ{current_row}"] = khuyenmai
                    
                    
                if match:
                    x = int(match.group(1))  # Chuyển đổi thành số nguyên
                    kiemtra = kiemtra or item["Barcode"]
                    
                    if x >= 2:  # Chỉ chia khi x >= 2, còn 1+1 thì giữ nguyên
                        qty_ord_pcs = math.floor(qty_ord_pcs / x)  # Chia rồi làm tròn xuống
                
                if kiemtra:
                    if isinstance(kiemtra, list):  # Nếu có nhiều mã SKU, nối thành chuỗi
                        kiemtra = ", ".join(kiemtra)
                    cachbokem = ProcessHandler.laycachbo_khuyenmai(hangkm)
                    if i > 0:
                        if cachbokem:
                            sheet[f"AO{current_row + 1}"] = cachbokem
                            cachbokem_lower = cachbokem.lower()
                            if "bó kèm" in cachbokem_lower or "quấn kèm" in cachbokem_lower:
                                sheet[f"AP{current_row + 1}"] = f'{ProcessHandler.layduoi_mahang(item["Barcode"])}_{ProcessHandler.layduoi_mahang(kiemtra)}_1'
                                sheet[f"AP{current_row + 2}"] = f'{ProcessHandler.layduoi_mahang(item["Barcode"])}_{ProcessHandler.layduoi_mahang(kiemtra)}_1'
                        else:
                            sheet[f"AO{current_row + 1}"] = f'KM Rời - Không Che Barcode'
                    else:
                        if cachbokem:
                            sheet[f"AO{current_row}"] = cachbokem
                            sheet[f"AO{current_row + 1}"] = cachbokem
                            cachbokem_lower = cachbokem.lower()
                            if "bó kèm" in cachbokem_lower or "quấn kèm" in cachbokem_lower:
                                sheet[f"AP{current_row}"] = f'{ProcessHandler.layduoi_mahang(item["Barcode"])}_{ProcessHandler.layduoi_mahang(kiemtra)}_1'
                                sheet[f"AP{current_row + 1}"] = f'{ProcessHandler.layduoi_mahang(item["Barcode"])}_{ProcessHandler.layduoi_mahang(kiemtra)}_1'
                        else:
                            sheet[f"AO{current_row}"] = f'KM Rời - Không Che Barcode'

                    self.log_signal.emit(
        f'🎁 Mã hàng <b><span style="color: blue;">{item["Barcode"]}</span></b> '
        f'đã thêm hàng khuyến mãi <b><span style="color: green;">{kiemtra}</span></b> tặng kèm! 🎉'
        )
                    current_row += 1
                    sheet[f"A{current_row}"] = entry_date
                    sheet[f"AV{current_row}"] = songayno_MT
                   
                    sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'
                    sheet[f"C{current_row}"] = "Chưa thực hiện"
                    sheet[f"D{current_row}"] = cancle_date
                    sheet[f"G{current_row}"] = makhachhang
                    sheet[f"L{current_row}"] = diengiai
                    sheet[f"Q{current_row}"] = kiemtra
                    sheet[f"X{current_row}"] = qty_ord_pcs
                    sheet[f"V{current_row}"] = kho
                    sheet[f"Y{current_row}"] = 0
                    sheet[f"Z{current_row}"] = 0
                    sheet[f"U{current_row}"] = "Có"
                    sheet[f"AJ{current_row}"] = khuvuc
                    sheet[f"AE{current_row}"] = 8
                    sheet[f"AM{current_row}"] = mien
                    sheet[f"AN{current_row}"] = congtrinh
                    sheet[f"T{current_row}"] = "Không"
                    sheet[f"E{current_row}"] = diachigiaohang
                    sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(kiemtra)

                    trongluong += (ProcessHandler.timtrongluong_sanpham(kiemtra) * qty_ord_pcs)
                    sheet[f"AT{current_row}"] = ProcessHandler.timtrongluong_sanpham(kiemtra) * qty_ord_pcs

            current_row += 1
        sheet["G1"] = STT_donhang
        kmhoadon = ProcessHandler.find_all_promotions_by_sku_and_time("Hóa Đơn", entry_date, vendor)
        if kmhoadon:
            kmhoadon = str(kmhoadon[0][1])
            kiemtra = ProcessHandler.check_value_in_sanpham(kmhoadon)
            sheet[f"AQ{current_row}"] = kmhoadon


            sheet[f"B{current_row}"] = f'ĐĐH{vendor}{STT_donhang_str}'
            soluongkm = math.floor(tongtien / ProcessHandler.tachtien_khuyenmai(kmhoadon))
            sheet[f"A{current_row}"] = entry_date
            sheet[f"AV{current_row}"] = songayno_MT
            sheet[f"C{current_row}"] = "Chưa thực hiện"
            sheet[f"D{current_row}"] = cancle_date
            sheet[f"E{current_row}"] = diachigiaohang
            sheet[f"G{current_row}"] = makhachhang
            sheet[f"L{current_row}"] = diengiai
            sheet[f"Q{current_row}"] = kiemtra[0]
            sheet[f"X{current_row}"] = soluongkm
            sheet[f"V{current_row}"] = kho
            sheet[f"Y{current_row}"] = 0
            sheet[f"Z{current_row}"] = 0
            sheet[f"U{current_row}"] = "Có"
            sheet[f"AJ{current_row}"] = khuvuc
            sheet[f"AE{current_row}"] = 8
            sheet[f"AM{current_row}"] = mien
            sheet[f"T{current_row}"] = "Không"
            sheet[f"S{current_row}"] = ProcessHandler.timten_sanpham(kiemtra[0])


            trongluong += (ProcessHandler.timtrongluong_sanpham(kiemtra[0]) * soluongkm)
            sheet[f"AT{current_row}"] = ProcessHandler.timtrongluong_sanpham(kiemtra[0]) * soluongkm

            

            cachbokem = ProcessHandler.laycachbo_khuyenmai(kmhoadon)
            if cachbokem:
                sheet[f"AO{current_row}"] = cachbokem
            else:
                sheet[f"AO{current_row}"] = f'KM Bó Kèm - Che Barcode'

            self.log_signal.emit(f'🎉 Đơn hàng có chương trình khuyến mãi <b><span style="color: green;">{kmhoadon}</span></b> 🛒')
            self.log_signal.emit(f'Tổng tiền đơn hàng: <b><span style="color: green;">{tongtien} </span></b>\nSố lượng tặng: <b><span style="color: green;">{soluongkm} </span></b> 🛒')
            
        sheet[f"L{start_row}"] = f'{diengiai} (Tổng trọng lượng: {ProcessHandler.format_weight_kg(trongluong)})'
        # Lưu lại file Excel
        wb.save(file_path)

        ProcessHandler.ghi_message(f"tong_tien: {tongtien:,.0f}")
        ProcessHandler.ghi_message(f"tong_trongluong: {ProcessHandler.format_weight_kg(trongluong)}")
        ProcessHandler.ghi_message(f"sai_gia: {saigia}")
        now2 = datetime.now()
        now_str = now2.strftime("%d/%m/%Y %H:%M:%S")
        ProcessHandler.ghi_message(f"end_time: {now_str}")
        time = now2 - now
        ProcessHandler.ghi_message(f"time: {ProcessHandler.format_timedelta(time)}]") 
        print(f"✅ Đã ghi {len(items)} dòng vào '{file_path}', sheet '{sheet_name}'!")
        return saigia
    

    def laymahang_JIT(text):
        # Tìm tất cả chuỗi bắt đầu bằng CH và theo sau là 1 hoặc nhiều số
        matches = re.findall(r'\bCH\d+\b', text)
        return matches if matches else text

    def xoa_du_lieu_don_dat_hang():
        """
        Xóa toàn bộ dòng từ dòng 9 đến hết trong file 'dondathang.xlsx', sheet 'Don dat hang'.
        """
        file_path = "dondathang.xlsx"
        sheet_name = "Don dat hang"

        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]

        max_row = sheet.max_row

        if max_row >= 9:
            sheet.delete_rows(9, max_row - 8)  # Xóa từ dòng 9 đến hết

        wb.save(file_path)
        wb.close()
        print(f"Đã xóa toàn bộ dòng từ 9 đến {max_row} trong sheet '{sheet_name}'.")

    @staticmethod
    def clean_sku_number(sku):
        if isinstance(sku, list):  
            # Nếu sku là list thì lấy phần tử đầu tiên (hoặc join lại)
            sku = sku[0] if sku else ""
        elif sku is None:
            return ""

        match = re.search(r"(\d{7})-\d", str(sku))
        if match:
            return match.group(1)
        return str(sku)


    def load_sku_mapping():
        """Tạo mapping SKU từ sheet 'SanPham' (C → cuối, bỏ toàn bộ khoảng trắng)"""
        file_path = "data.xlsx"
        sheet_name = "SanPham"

        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb[sheet_name]

        sku_mapping = {}

        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Cột A: SKU chuẩn (giữ nguyên, chỉ strip)
            sku_code = str(row[0]).strip() if row[0] else None
            if not sku_code:
                continue

            # Cột C → cuối: xóa TOÀN BỘ khoảng trắng
            for cell in row[2:]:
                if not cell:
                    continue

                value = re.sub(r"\s+", "", str(cell))

                if value:
                    sku_mapping[value] = sku_code

        return sku_mapping



    def replace_sku_numbers(products, sku_mapping):
        
        """Thay thế SKU Number hoặc Article nếu tìm thấy trong mapping"""
        for product in products:
            for key in ["Barcode", "Barcode", "Barcode"]:  # Kiểm tra cả 2 trường
                if key in product:
                    clean_sku = ProcessHandler.clean_sku_number(product[key])  # Làm sạch mã
                    product[key] = ProcessHandler.clean_sku_number(product[key])
                    if clean_sku in sku_mapping:
                        product[key] = sku_mapping[clean_sku]  # Thay bằng mã mới

        return products


    def convert_entry_date(entry_date):
        """Chuyển đổi Entry Date từ dd/mm/yy sang dd/mm/yyyy với yy thuộc 2000+"""
        match = re.match(r"(\d{2})/(\d{2})/(\d{2})", entry_date)
        if match:
            day, month, year = match.groups()
            year = f"20{year}"  # Luôn giả định năm là từ 2000 trở đi
            return f"{day}/{month}/{year}"
        return entry_date  # Trả về nguyên nếu không đúng định dạng
    

    @staticmethod
    def demsodonhang1trang_coop(text):
        # Chuẩn hóa: thay \xa0 bằng khoảng trắng thông thường
        text = text.replace("\xa0", " ").replace("pom343", "POM343").replace("pom346", "POM346")
        
        pom34_pattern = r'POM34[36]\b'  # Tìm cả POM343 và POM346
        sub_total_pattern = r'\bSub\s*(?:Total|Tot\s*al)\b'

        
        pom34_count = len(re.findall(pom34_pattern, text))
        sub_total_count = len(re.findall(sub_total_pattern, text))
        
        return {"POM343": pom34_count, "Sub Total": sub_total_count}



   


    def process_coop_invoice(self,text,stt,path,page_label,doc):
        file_name = os.path.basename(path)

        """Xử lý nếu Vendor là Coop"""
        print("\n🔹 Vendor xác nhận là Coop. Tiến hành xử lý...")

        
        # 📍 Lấy thông tin từ extract_info()
        info = ProcessHandler.extract_info(text)

        # 🟢 3. Xóa cụm từ không mong muốn
        # Danh sách các cụm từ cần loại bỏ (không phân biệt hoa thường, khoảng trắng tùy ý)
        patterns = [
            r"X\s*i\s*n\s*v\s*u\s*i\s*l\s*o\s*n\s*g\s*k\s*e\s*m\s*D\s*D\s*H\s*k\s*h\s*i\s*g\s*i\s*a\s*o\s*h\s*a\s*n\s*g",
            r"\*\s*=\s*T\s*h\s*i\s*s\s*S\s*K\s*U\s*D\s*i\s*s\s*c\s*o\s*u\s*n\s*t\s*e\s*d",
            r"M\s*o\s*t\s*H\s*o\s*a\s*D\s*o\s*n\s*c\s*h\s*i\s*x\s*u\s*a\s*t\s*c\s*h\s*o\s*m\s*o\s*t\s*P\s*O",
            r"m\s*u\s*a\s*1\s*T\s*A\s*N\s*G\s*1\s*C\s*U\s*N\s*G\s*L\s*O\s*A\s*I",
            r"1\s*T\s*A\s*N\s*G\s*1\s*C\s*U\s*N\s*G\s*L\s*O\s*A\s*I"
        ]

        # Tìm nội dung giữa "Notes" (hoặc "No tes") và "FOB"
        match = re.search(r"N\s*o\s*t\s*e\s*s\s*-\s*(.*?)\s*FOB", text, re.DOTALL | re.IGNORECASE)
        ghichu = match.group(1).strip() if match else ""

        # Loại bỏ các cụm từ không mong muốn
        for pattern in patterns:
            ghichu = re.sub(pattern, "", ghichu, flags=re.IGNORECASE)

        # Xóa khoảng trắng dư thừa
        ghichu = re.sub(r"\s+", " ", ghichu).strip()

        # Loại bỏ trùng lặp từ nhưng vẫn giữ thứ tự gốc
        words = ghichu.split()
        ghichu = " ".join(dict.fromkeys(words))

        print(f"Ghi chú: {ghichu}")
        po_number = info.get("P/O Number", "Không tìm thấy")
        po_location = info.get("P/O Location", "Không tìm thấy")


        shipto = text
        shipto = re.sub(r"\s+", " ", shipto)  # Thay thế nhiều khoảng trắng bằng một space


        # Regex để tìm "Ship To"
        pattern = r"S\s*h\s*i\s*p\s*T\s*o\s*:\s*S\s*t\s*a\s*t\s*u\s*s\s*-\s*\d+\s*R\s*E\s*L\s*E\s*A\s*S\s*E\s*D\s*(.*?)\s*C\s*o\s*n\s*t\s*a\s*c\s*t\s*-"

        match = re.search(pattern, shipto, re.DOTALL | re.IGNORECASE)
        
        if match:
            shipto = match.group(1).strip()  # Lấy nội dung và loại bỏ khoảng trắng đầu/cuối
            print("Ship To:", shipto)
        else:
            pattern = r"S\s*t\s*o\s*r\s*e\s*-\s*(.*?)\s*V\s*e\s*n\s*d\s*o\s*r"
            match = re.search(pattern, shipto, re.DOTALL | re.IGNORECASE)
            
            if match:
                shipto = match.group(1).strip()  # Lấy nội dung và loại bỏ khoảng trắng đầu/cuối
                print("Ship To:", shipto)
            else:
                print("Không tìm thấy thông tin Ship To.")
                shipto = ""

        

        entry_date_str = info.get("Entry Date", "Không tìm thấy")
        cancle_date_str = info.get("Cancel Date", "Không tìm thấy")

        def convert_date_format(date_str):
            """ Chuyển ngày từ 'dd/MM/yy' thành 'dd/MM/yyyy' """
            if date_str and date_str != "Không tìm thấy":
                try:
                    # Đọc theo định dạng "dd/MM/yy"
                    date_obj = datetime.strptime(date_str, "%d/%m/%y")
                    # Chuyển sang định dạng "dd/MM/yyyy"
                    return date_obj.strftime("%d/%m/%Y")
                except ValueError:
                    return "Không hợp lệ"  # Tránh lỗi khi format sai
            return "Không tìm thấy"

        # Chuyển đổi ngày
        entry_date = convert_date_format(entry_date_str)
        cancle_date = convert_date_format(cancle_date_str)

        # Nếu cancle_date không có, cộng 65 ngày vào entry_date
        if cancle_date == "Không tìm thấy" and entry_date != "Không tìm thấy":
            entry_date_obj = datetime.strptime(entry_date, "%d/%m/%Y")
            cancle_date = (entry_date_obj + timedelta(days=65)).strftime("%d/%m/%Y")


        



        print(f"📍 PO Number: {po_number}")
        print(f"📍 P/O Location: {po_location}")
        print(f"📅 Entry Date: {entry_date}")
        print(f"⏳ Cancel Date: {cancle_date}")
    
        self.log_signal.emit(f"<b>📦 Số PO:</b> <span style='color:blue;'>{po_number}</span>")
        self.log_signal.emit(f"<b>🗓️ Ngày đặt hàng:</b> <span style='color:green;'>{entry_date}</span>")
        self.log_signal.emit(f"<b>⏳ Hạn đơn hàng:</b> <span style='color:red;'>{cancle_date}</span>")

        print(f"📅 Cancle Date: {cancle_date}")
        # 🏢 Lấy Mã Khách Hàng
        makhachhang = "Không tìm thấy"
        if po_location:
            makhachhang = ProcessHandler.get_makhachhang(po_location)
            if makhachhang == "Không tìm thấy":
                # Nếu không tìm thấy, thử lại với nửa đầu của po_location
                half_location = po_location[:len(po_location) // 2].strip()
                makhachhang = ProcessHandler.get_makhachhang(half_location)
                if not makhachhang:
                    makhachhang = "Không tìm thấy"

        print(f"🏢 Mã Khách Hàng: {makhachhang}")

        # 📦 Trích xuất danh sách sản phẩm
        products = ProcessHandler.extract_products(text)
        soluongsanpham = len(products)
        self.log_signal.emit(f"<b>📊 Tổng số lượng sản phẩm:</b> <span style='color:purple;'>{soluongsanpham}</span>")

        # 🔽 Thay thế SKU Number từ Excel
        if products:
            sku_mapping = ProcessHandler.load_sku_mapping()  # Load mapping SKU
            products = ProcessHandler.replace_sku_numbers(products, sku_mapping)  # Thay SKU
            
            # Ghi vào dondathang.xls (truyền thêm `entry_date`)
            hethong = ProcessHandler.layhethong_COOP(makhachhang)
            

            
            if page_label == '1/1':
                print('')
                result  = ProcessHandler.upload_file_to_drive(path, po_number)
                print(result )
                file_url = result.get("url")
            else:
                print(page_label)
                po_name = f"{po_number}.pdf"
                ProcessHandler.cat_trang_hien_tai(doc,po_name,page_label)
                
                result  = ProcessHandler.upload_file_to_drive(po_name, po_number)
                file_url = result.get("url")

                if os.path.exists(po_name):
                    os.remove(po_name)

            saigia = ProcessHandler.write_to_dondathang(self,products, makhachhang, po_number, entry_date,cancle_date,stt,"COOP",ghichu,shipto,file_url)
            if not makhachhang: 
                makhachhang = "Không xác định"

            
            if saigia > 0 or makhachhang == "Không xác định":
                if saigia == 0:
                    saigia = "Đúng giá"
                else:
                    saigia = f'Có {saigia} mã sai giá'
                hoanthanh = "⚠️Hoàn Thành"
            else:
                saigia = "Đúng giá"
                hoanthanh = "✅Hoàn Thành"

            
            
            self.table_signal.emit(file_name, page_label,hethong, makhachhang ,po_number, saigia, hoanthanh)  # Phát tín hiệu cập nhật bảng
            
        stt = stt + 1

        print("\n📦 Danh sách sản phẩm:")
        for product in products:
            print(product)
        
    def parse_date(date_str):
        """Chuyển đổi chuỗi ngày tháng về định dạng chuẩn dd/mm/yyyy"""
        if not date_str or date_str in ["Không tìm thấy", ""]:  
            return None  # Trả về None nếu dữ liệu không hợp lệ
        
        if isinstance(date_str, datetime):
            return date_str  # Nếu đã là datetime, không cần xử lý

        if isinstance(date_str, str):
            parts = date_str.split('/')
            if len(parts) == 3:
                day, month, year = parts
                if not (day.isdigit() and month.isdigit() and year.isdigit()):
                    return None  # Nếu có ký tự không phải số -> bỏ qua
                
                if len(year) == 2:  # Nếu năm chỉ có 2 chữ số, thêm '20' phía trước
                    year = "20" + year

                date_str = f"{day.zfill(2)}/{month.zfill(2)}/{year}"

                try:
                    return datetime.strptime(date_str, "%d/%m/%Y")
                except ValueError:
                    return None  # Ngăn lỗi nếu dữ liệu không hợp lệ

        return None  # Nếu không thuộc kiểu string hoặc datetime
    
    def get_gid(sheet_name):
        """ Truy xuất GID từ tên sheet trong thẻ <gid>...</gid> của setting.ini """
        with open(CONFIG_FILE, "r", encoding="utf-8") as file:
            content = file.read()
        
        # Tìm nội dung bên trong thẻ <gid>...</gid>
        match = re.search(r"<gid>(.*?)</gid>", content, re.DOTALL)
        if not match:
            return None  # Không tìm thấy thẻ <gid>

        gid_section = match.group(1)  # Lấy nội dung bên trong <gid>...</gid>
        
        # Chuyển đổi thành dictionary
        gid_dict = {}
        for line in gid_section.strip().split("\n"):
            if "=" in line:
                key, value = line.strip().split("=")
                gid_dict[key.strip()] = value.strip()
        
        return gid_dict.get(sheet_name, None)  # Trả về GID theo tên sheet (hoặc None nếu không có)
    

    def find_price_by_sku(sku_number, sheet_name="COOP"):
        sheet_id = '1yvxE_SPYXKhofcZdhv1CSKAyiwdY1Mf4pFlsiMbtOr4'
        gid = ProcessHandler.get_gid(sheet_name)
        print(f"Mã hàng: {sku_number}")
        sku_number = re.sub(r"\s+", "", sku_number)
        print(f"Sheet Name: {sheet_name}")
        print(f"Sheet ID: {gid}")

        # Tạo URL tải dữ liệu từ Google Sheets
        sheet_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?tqx=out:csv&gid={gid}"

        # Tải dữ liệu từ Google Sheets, chỉ giữ 4 cột đầu
        df = pd.read_csv(sheet_url, dtype=str, usecols=[0, 1, 2, 3])  
        df.fillna("", inplace=True)  # Thay NaN bằng chuỗi rỗng

        # Tìm hàng có SKU trong cột thứ 2 (index = 1)
        matched_row = df[df.iloc[:, 1] == sku_number]

        if not matched_row.empty:
            price = matched_row.iloc[0, 3]  # Cột thứ 4 (index = 3)

            price = price.replace(".", "")  # Xóa dấu "." và ","
            
            print(price if price.strip() else "Không có giá trị")
            return price if price.strip() else "Không có giá trị"
        else:
            return "Không tìm thấy SKU"
        




    def laygiathucte_CNHCM(sku_number, sheet_name="CN-HCM",thung = 0):
        sheet_id = '1yvxE_SPYXKhofcZdhv1CSKAyiwdY1Mf4pFlsiMbtOr4'
        gid = ProcessHandler.get_gid(sheet_name)
        print(f"Mã hàng: {sku_number}")
        print(f"Sheet Name: {sheet_name}")

        # Tạo URL tải dữ liệu từ Google Sheets
        sheet_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?tqx=out:csv&gid={gid}"

        # Tải dữ liệu từ Google Sheets, chỉ giữ 4 cột đầu
        df = pd.read_csv(sheet_url, dtype=str, usecols=[0, 1, 2, 8])  
        df.fillna("", inplace=True)  # Thay NaN bằng chuỗi rỗng

        # Tìm hàng có SKU trong cột thứ 2 (index = 1)
        matched_row = df[df.iloc[:, 1] == sku_number]


        if not matched_row.empty:
            if thung == 1:
                price = matched_row.iloc[0, 3]  # Cột thứ 4 (index = 3)
            else:
                price = matched_row.iloc[0, 3]  # Cột thứ 4 (index = 3)

            price = price.replace(".", "")  # Xóa dấu "." và ","
            
            print(price if price.strip() else "Không có giá trị")
            return price if price.strip() else "Không có giá trị"
        else:
            return "Không tìm thấy SKU"



    def find_all_promotions_by_sku_and_time(sku_code, time_to_check, sheet_name="Coop"):
 

        sheet_id = '1yvxE_SPYXKhofcZdhv1CSKAyiwdY1Mf4pFlsiMbtOr4'
        gid = ProcessHandler.get_gid(sheet_name)

        #print(f"✅ Sheet name: {sheet_name}, GID: {gid}")
        if not gid:
            return []

        sheet_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?tqx=out:csv&gid={gid}"
        #print(f"📥 Tải dữ liệu từ: {sheet_url}")

        try:
            df_raw = pd.read_csv(sheet_url, dtype=str, header=None)
            df_raw.fillna("", inplace=True)
        except Exception as e:
           # print(f"❌ Lỗi tải Google Sheet: {e}")
            return []

        # ✅ Dùng dòng thứ 2 làm tiêu đề (index = 1), dữ liệu từ dòng thứ 3 trở đi
        df = df_raw.iloc[1:].copy()
        df.columns = [col.strip().replace("\n", " ").replace("\r", "") for col in df_raw.iloc[0]]

        #print(f"📑 Các cột hiện có: {df.columns.tolist()}")

        # 🔍 Tìm cột "Mã hàng"
        sku_col = next((col for col in df.columns if "Mã hàng" in col), None)
        if not sku_col:
            return []

        sku_rows = df[df[sku_col] == sku_code]
        if sku_rows.empty:
            return []

        sku_row_idx = sku_rows.index[0]
       # print(f"✅ SKU {sku_code} nằm ở dòng index: {sku_row_idx}")

        def normalize_date(date_str):
            day, month = date_str.split("/")
            return f"{int(day):02}/{int(month):02}"

        def is_within_date_range(time_to_check, column_name):
            match = re.search(r"(\d{1,2}/\d{1,2})-(\d{1,2}/\d{1,2})", column_name)
            if match:
                start_date, end_date = match.groups()
                try:
                    # Xóa phần năm, xử lý cả M/D/YYYY (6/22/2026) lẫn D/M/YYYY (22/6/2026)
                    year_match = re.search(r"/(\d{4})$", time_to_check)
                    time_to_check = re.sub(r"/\d{4}$", "", time_to_check)
                    parts = time_to_check.split("/")
                    if len(parts) == 2:
                        p1, p2 = int(parts[0]), int(parts[1])
                        # Nếu phần đầu > 12 thì là D/M, nếu phần sau > 12 thì là M/D → đổi lại
                        if p1 <= 12 and p2 > 12:
                            time_to_check = f"{p2}/{p1}"
                    start_date = normalize_date(start_date)
                    end_date = normalize_date(end_date)
                    current_year = datetime.now().year

                    time_to_check_dt = datetime.strptime(f"{time_to_check}/{current_year}", "%d/%m/%Y")
                    start_date_dt = datetime.strptime(f"{start_date}/{current_year}", "%d/%m/%Y")
                    end_date_dt = datetime.strptime(f"{end_date}/{current_year}", "%d/%m/%Y")

                    return start_date_dt <= time_to_check_dt <= end_date_dt

                except Exception as e:
                    print(f"⚠️ Lỗi khi parse ngày tháng '{column_name}': {e}")
            return False

        matching_cols = [col for col in df.columns if is_within_date_range(time_to_check, col)]
        print(f"📆 Các cột khớp thời gian: {matching_cols}")

        matched_promotions = []
        for col in matching_cols:
            value = df.at[sku_row_idx, col]
           # print(f"🔍 {col}: {repr(value)}")
            print(f"Type of value: {type(value)}")
            print(f"Value content: {value}")
            if value.strip():
                matched_promotions.append((col, value))

        print(f"🎯 Tìm thấy {len(matched_promotions)} khuyến mãi hợp lệ.")
        return matched_promotions












    def kiemtrakhuyenmaihoadon( time_to_check, sheet_name="Coop"):
        sheet_id = '1yvxE_SPYXKhofcZdhv1CSKAyiwdY1Mf4pFlsiMbtOr4'
        gid = ProcessHandler.get_gid(sheet_name)

        # 1️⃣ Tải dữ liệu từ Google Sheets
        sheet_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?tqx=out:csv&gid={gid}"
        df = pd.read_csv(sheet_url, dtype=str)  # Đọc dữ liệu, giữ nguyên định dạng chuỗi
        df.fillna("", inplace=True)  # Thay NaN bằng chuỗi rỗng để tránh lỗi

        # Chuẩn hóa tiêu đề cột
        df.columns = [col.strip().replace("\n", " ") for col in df.columns]

        # Tìm cột "Mã hàng"
        sku_col = next((col for col in df.columns if "TÊN SẢN PHẨM" in col), None)
        if not sku_col:
            return "Không có cột TÊN SẢN PHẨM"

        # Kiểm tra SKU có tồn tại không
        sku_rows = df[df[sku_col] == "Khuyến mãi trên 1 Hoá đơn"]
        if sku_rows.empty:
            return "Không có giá trị"

        sku_row_idx = sku_rows.index[0]  # Lấy chỉ mục dòng chứa SKU

        def normalize_date(date_str):
            """Chuyển ngày về dạng dd/mm"""
            day, month = date_str.split("/")
            return f"{int(day):02}/{int(month):02}"  # Đảm bảo 2 chữ số

        def is_within_date_range(time_to_check, column_name):
            match = re.search(r"(\d{1,2}/\d{1,2})-(\d{1,2}/\d{1,2})", column_name)
            if match:
                start_date, end_date = match.groups()
                time_to_check = re.sub(r"/\d{4}$", "", time_to_check)  # Xóa /2025 nếu có

                # Chuẩn hóa ngày
                start_date, end_date = normalize_date(start_date), normalize_date(end_date)
                current_year = datetime.now().year  # Lấy năm hiện tại

                time_to_check_dt = datetime.strptime(f"{time_to_check}/{current_year}", "%d/%m/%Y")
                start_date_dt = datetime.strptime(f"{start_date}/{current_year}", "%d/%m/%Y")
                end_date_dt = datetime.strptime(f"{end_date}/{current_year}", "%d/%m/%Y")

                return start_date_dt <= time_to_check_dt <= end_date_dt
            return False

        # Lưu danh sách các cột khớp thời gian
        matching_cols = [col for col in df.columns if is_within_date_range(time_to_check, col)]
        
        # Kiểm tra và lấy giá trị từ cột đầu tiên có dữ liệu
        for col in matching_cols:
            value = df.at[sku_row_idx, col]
            if value.strip():  # Nếu có dữ liệu, trả về ngay
                print(f"✅ Tìm thấy giá trị '{value}' tại Khuyến mãi trên 1 Hoá đơn, cột {col}")
                return value
        
        # Nếu tất cả cột khớp thời gian đều rỗng
        return "Không có giá trị"



        



    def to_date_only(val):
        """Chuyển datetime -> date, bỏ giờ."""
        if isinstance(val, datetime):
            return val.date()
        if isinstance(val, date):
            return val
        if val is None:
            return None
        s = str(val).strip()
        fmts = ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%Y-%m-%d", "%d/%m/%Y")
        for f in fmts:
            try:
                return datetime.strptime(s, f).date()
            except Exception:
                pass
        try:
            return datetime.fromisoformat(s).date()
        except Exception:
            return s
        


    def laydanhsachsanpham_bigc(text):
        """
        Trích xuất danh sách sản phẩm từ nội dung văn bản BigC (bỏ qua Article Description).
        """


            # 1️⃣ Xác định vị trí danh sách sản phẩm (từ dòng chứa "Barcode")
        match_start = re.search(r"\bArticle\b", text)
    
        if not match_start:
            print("Không tìm thấy tiêu đề danh sách sản phẩm!")
            return []
        text = text[match_start.start():].strip()
            

        pattern = re.compile(
            r"(\d{13})\s+.+?\s+Pack\s+\d+\s+(\d+)\s+(\d+)\s+\d+\s+([\d,]+)\s+\w+\s+([\d,]+)",
            re.DOTALL
        )

        matches = pattern.findall(text)

        product_list = []
        for match in matches:
            if len(match) != 5:
                print(f"LỖI: Không đủ dữ liệu trong dòng: {match}")
                continue  # Bỏ qua dòng bị thiếu dữ liệu

            article, sku_ou, ou_qty, net_price, total_price = match

            # Chuẩn hóa dữ liệu (bỏ dấu phẩy và chuyển thành số)
            net_price = int(net_price.replace(",", ""))
            total_price = int(total_price.replace(",", ""))

            product_list.append({
                "Barcode": article.strip(),
                "SKU/OU": sku_ou.strip(),
                "OU Qty": ou_qty.strip(),
                "Total Price": net_price,
                "Total Net Purchase Price": total_price
            })

        return product_list




    def lay_ten_store(text):
        # Tìm vị trí của kho "FM LOGISTIC VSIP 2"
        match = re.search(r"(FM LOGISTIC VSIP 2|LINFOX WAREHOUSE \(802\)).*?Vietnam\s*\n(.*?)\n", text, re.DOTALL)
        
        if match:
            return match.group(2).strip()  # Trả về dòng ngay sau đó (tên cửa hàng)
        return None



    def ghepgia_donhangbigc(items, products):
        # Chuyển đổi products thành dict để tra cứu nhanh hơn
        product_dict = {p["Barcode"]: p["Total Price"] for p in products}

        # Thêm giá vào mỗi item
        for item in items:
            article = item["Barcode"]
            item["Total Price"] = product_dict.get(article, 0)  # Nếu không có thì báo lỗi

        return items


    def trichxuatdanhsachforstore_bigc(text):
        # Regex để tìm các dòng chứa thông tin sản phẩm
        pattern = r"(?<=\n)(\d{13})\s*\n(.*?)\s*\nPack\s*\n\d+\s*\n(\d+)\s*\n(\d+)"
        matches = re.findall(pattern, text, re.DOTALL)

        # Chuyển kết quả thành danh sách từ điển
        data = [{"Barcode": m[0], "SKU/OU": m[2], "OU Qty": m[3]} for m in matches]
        return data
    # --------- Chạy chương trình ---------
    def read_text_file(file_path):
        """Đọc file TXT với khả năng xử lý nhiều kiểu mã hóa"""
        encodings = ["utf-8", "utf-16", "utf-16-le", "utf-16-be", "latin-1"]

        for encoding in encodings:
            try:
                with open(file_path, "r", encoding=encoding) as f:
                    return f.read()
            except UnicodeDecodeError:
                continue  # Thử mã hóa tiếp theo

        raise UnicodeDecodeError("Không thể đọc file TXT với các bộ mã hóa phổ biến!")




    def ghi_gia_tri_G1(stt):
        """
        Ghi giá trị 10 vào ô G1 trong file 'dondathang.xlsx', sheet 'Don dat hang'.
        """
        file_path = "dondathang.xlsx"
        sheet_name = "Don dat hang"

        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]

        sheet["G1"].value = stt  # Ghi giá trị 10 vào ô G1

        wb.save(file_path)
        wb.close()


    def trichxuatinfo_donbigc(text):
        """Trích xuất PO Number, Entry Date và Cancel Date từ văn bản"""

        # Chuẩn hóa văn bản (xoá khoảng trắng thừa)
        cleaned_text = re.sub(r"\s+", " ", text).strip()

        # 1️⃣ Tìm PO Number & Entry Date
        match_po = re.search(r"(\d{13,})\s+(\d{2}/\d{2}/\d{2})", cleaned_text)
        po_number = match_po.group(1) if match_po else None
        entry_date = match_po.group(2) if match_po else None

        # 2️⃣ Tìm tất cả vị trí "Total Net Purchase Price"
        matches = list(re.finditer(r"Total Net Purchase Price", cleaned_text))

        # 3️⃣ Nếu tìm thấy, lấy ngày sau lần xuất hiện CUỐI CÙNG
        if matches:
            last_match = matches[-1]  # Lấy lần xuất hiện cuối
            search_region = cleaned_text[last_match.end():]  # Lấy phần sau đó
            match_cancel = re.search(r"(\d{2}/\d{2}/\d{2})", search_region)
            cancel_date = match_cancel.group(1) if match_cancel else None
        else:
            cancel_date = None

        # Nếu không tìm thấy cancel_date, đặt cancel_date = entry_date + 5 ngày
        if cancel_date is None and entry_date:
            try:
                entry_dt = datetime.strptime(entry_date, "%d/%m/%y")
                cancel_dt = entry_dt + timedelta(days=5)
                cancel_date = cancel_dt.strftime("%d/%m/%y")  # Chuyển lại thành chuỗi
            except ValueError:
                cancel_date = None  # Nếu entry_date không hợp lệ, giữ cancel_date là None

        return po_number, ProcessHandler.convert_entry_date(entry_date), ProcessHandler.convert_entry_date(cancel_date)




    def catdonra_nhieutrang(text):
        segments = []
        
        # Tách đoạn đầu tiên từ đầu đến "Sub Total" đầu tiên
        parts = text.split("Sub Total", 1)
        if len(parts) > 1:
            segments.append(parts[0])  # Đoạn trước "Sub Total"
            text = parts[-1]  # Giữ lại phần còn lại để tiếp tục xử lý
        
        # Tách các đoạn bắt đầu từ "POM343" hoặc "POM346"
        while any(keyword in text.lower() for keyword in ["pom343", "pom346"]) and "Sub Total" in text:
            for keyword in ["pom343", "pom346"]:
                if keyword in text.lower():
                    text = text.split(keyword, 1)[-1]  # Cắt bỏ phần trước từ khóa
                    parts = text.split("Sub Total", 1)  # Cắt đến "Sub Total" tiếp theo
                    if len(parts) > 1:
                        segments.append(keyword.upper() + parts[0])  # Giữ lại phần có từ khóa
                        text = parts[-1]  # Tiếp tục với phần còn lại
                    else:
                        segments.append(keyword.upper() + text)  # Nếu không còn "Sub Total", lấy hết và kết thúc
                        return segments
        
        return segments



    def get_grouped_data_xlsx(file_path: str, sheet_name: str = "Sheet1"):
        wb = openpyxl.load_workbook(filename=file_path, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active  # Chọn sheet

        products = []
        po_location = po_number = entry_date = cancel_date = None

        first_po_number = None  # Biến lưu P/O Number đầu tiên để so sánh
        rows_to_delete = []  # Danh sách lưu index các dòng cần xóa

        for idx, row in enumerate(ws.iter_rows(min_row=2, max_col=8, values_only=True), start=2):
            if row[1] is None:  # Bỏ qua dòng không có P/O Number
                continue

            if first_po_number is None:  # Lấy P/O Number đầu tiên
                first_po_number = row[1]
                po_location = row[0]  
                po_number = row[1]  
                entry_date = row[2]  
                cancel_date = row[3]  

            if row[1] != first_po_number:  # Nếu gặp P/O Number mới, dừng lấy dữ liệu
                break

            sku = row[4]  
            net_buy_cost = row[5]  
            qty_ord_pcs = row[6]  
            extended_cost = row[7]  

            products.append({
                "Barcode": sku,
                "Net Buy Cost": net_buy_cost,
                "Qty Ord/Pcs": qty_ord_pcs,
                "Extended Cost": extended_cost
            })

            rows_to_delete.append(idx)  # Đánh dấu dòng cần xóa

        # Xóa các dòng đã lấy từ dưới lên để tránh lỗi chỉ mục
        for row_idx in reversed(rows_to_delete):
            ws.delete_rows(row_idx)

        # Lưu lại file sau khi xóa dữ liệu
        wb.save(file_path)

        return po_location, po_number, entry_date, cancel_date, products
    
    def tachcancledate_lotte(text, po_number):
        start_pattern = re.escape(po_number)  # Tìm dòng bắt đầu
        end_marker = "00:00"  # Dòng kết thúc
        date_pattern = r"\d{1,2}/\d{1,2}/\d{4}"  # Định dạng ngày (dd/mm/yyyy)

        lines = text.split("\n")
        start_index = None
        end_index = None

        for i, line in enumerate(lines):
            if start_index is None and re.match(start_pattern, line):  # Tìm dòng đầu tiên
                start_index = i
            if line.strip() == end_marker:  # Tìm dòng cuối cùng
                end_index = i
                break  # Dừng ngay khi tìm thấy dòng cuối

        if start_index is not None and end_index is not None and start_index < end_index:
            filtered_dates = [line.strip() for line in lines[start_index + 1:end_index] if re.search(date_pattern, line)]
            return "\n".join(filtered_dates)  # Trả về chỉ các dòng có ngày tháng

        return None  # Không tìm thấy hai dòng
    

    def tachsanpham_lotte(text):
        text = ProcessHandler.lamsachdonhang_lotte(text)
        pattern = r'(\d{1,2}-\d{6}-\d{3})\s+(\d{12,13})[\s\S]*?(\d+)\s+BOX\s+(\d+)\s+(\d+)\s+[\d,]+\s+([\d,]+)'
        matches = re.findall(pattern, text)
        
        product_data = []
        for match in matches:
            product_code, barcode, unit_count, box_qty, loose_qty, total_price = match
            product_data.append({
                "Product Code": product_code,
                "Barcode": barcode,
                "Qty-Box": int(unit_count),
                "Box Quantity": int(box_qty),
                "Loose Quantity": int(loose_qty),
                "Total Price": float(total_price.replace(',', ''))
            })
        
        return product_data
    


    def cat_giua_theo_dong(text: str,
                       dau_line: str = "Mã vận đơn:",
                       cuoi_line: str = "Top Value",
                       ignore_case: bool = False) -> str:
        """
        Lấy nguyên vẹn các dòng nằm giữa một dòng bắt đầu bằng 'dau_line'
        và dòng bằng 'cuoi_line' (dòng đơn, sau khi strip).
        Không đụng vào định dạng các dòng ở giữa.
        """
        lines = text.splitlines()

        def norm(s: str) -> str:
            return s.casefold() if ignore_case else s

        start = None
        end = None

        for i, line in enumerate(lines):
            if start is None and norm(line.strip()).startswith(norm(dau_line)):
                start = i + 1  # lấy sau dòng "Mã vận đơn:"
                continue
            if start is not None and norm(line.strip()) == norm(cuoi_line):
                end = i        # dừng trước dòng "Top Value"
                break

        if start is None or end is None or end <= start:
            return ""

        return "\n".join(lines[start:end]).rstrip()

    def merge_CH_code(text: str) -> str:
        # Ghép liền khi là "CHx\nx" hoặc "CHxx\nxx" (x là chữ số)
        text = re.sub(r"\b(CH\d{1,2})\s*\n\s*(\d{1,2})\b", r"\1\2", text)
        # Các xuống dòng khác -> đổi thành 1 khoảng trắng
        text = re.sub(r"\s*\n\s*", " ", text)
        return text
    



    def tach_sanpham_Clevy_tiktok(text):
        products = []
        lines = [l.strip() for l in text.splitlines() if l.strip()]

        # 1) Loại bỏ dòng như "Qty Total: 1"
        cleaned = []
        for line in lines:
            if line.lower().startswith("qty total"):
                continue
            cleaned.append(line)

        lines = cleaned

        # 2) Gom block kết thúc bằng dòng chỉ có số (qty)
        blocks = []
        current = []

        for line in lines:
            if line.isdigit():     # DÒNG QTY
                current.append(line)
                blocks.append(current)
                current = []
            else:
                current.append(line)

        # 3) Xử lý từng block
        for block in blocks:
            if not block:
                continue

            # dòng cuối = qty
            qty = int(block[-1])

            # các dòng trước = tên sản phẩm (giữ nguyên tất cả)
            name_parts = block[:-1]

            # GHÉP TÊN
            product_name = " ".join(name_parts).strip()

            products.append({
                "product_name": product_name,
                "Qty Ord/Pcs": qty,
                "Barcode":  ProcessHandler.laymahang_JIT(re.sub(r"\s+", "", product_name)),
                "don_gia": 0,
                "thanh_tien": 0
            })

        return products



    def tach_sanpham_JIT_airway(text: str):
        """
        Phân tích danh sách sản phẩm trong text, mỗi sản phẩm kết thúc bằng 'SL: <số>'.
        Chịu được trường hợp 'SL:' bị tách dòng thành 'S\\nL:'.
        """
        # 🔹 B1: Gộp đặc thù (của bạn) rồi thêm chuẩn hoá 'S\\nL:' -> 'SL:'
        merged = ProcessHandler.merge_CH_code(text)
        merged = re.sub(r"S\s*[\r\n]+\s*L\s*:", "SL:", merged, flags=re.I)

        # 🔹 B2: Regex khoan dung: cho phép khoảng trắng giữa S và L
        SL_TAG = r"S\s*L\s*:"
        
        # Tách block: từ số thứ tự '1. ' ... đến 'SL: <số>'
        pattern_block = rf"(?s)(\d+\.\s.*?{SL_TAG}\s*\d+)"
        blocks = re.findall(pattern_block, merged)

        products = []
        for block in blocks:
            block_clean = block.strip()

            # 🧩 Tên sản phẩm: sau '1.' đến trước 'SL:'
            name_match = re.search(rf"\d+\.\s*(.*?)\s*,?\s*{SL_TAG}", block_clean, flags=re.S)
            product_name = name_match.group(1).strip() if name_match else None

            # 🔢 Số lượng
            
            qty_match = re.search(rf"{SL_TAG}\s*(\d+)", block_clean)
            qty = int(qty_match.group(1)) if qty_match else None

            if product_name:
                # Làm sạch tên do PDF xuống dòng ngẫu nhiên: gom khoảng trắng liền kề
                product_name = re.sub(r"\s+", " ", product_name).strip(" ,")

                products.append({
                    "product_name": product_name,
                    "Qty Ord/Pcs": qty,
                    "Barcode": ProcessHandler.laymahang_JIT(re.sub(r"\s+", "",product_name)),
                    "don_gia": 0,
                    "thanh_tien": 0
                })

        return products




    def tach_sanpham_clevy_viettelpost(text: str):
        """
        Nhận text nhiều dòng, tìm các dòng có dạng 'số x tên sản phẩm'
        Trả về list: [{'Qty Ord/Pcs': qty, 'Barcode': barcode}, ...]
        """
        results = []
        lines = [l.strip() for l in text.splitlines() if l.strip()]

        for line in reversed(lines):      # duyệt từ dưới lên
            if "x" not in line:
                continue

            parts = line.split("x", 1)    # chỉ tách 1 lần
            qty = parts[0].strip()
            barcode = parts[1].strip()

            # đảm bảo qty là số
            if qty.isdigit():
                results.append({
                    "Qty Ord/Pcs": qty,
                    "Barcode": barcode
                })

        return results

    
    def tach_sanpham_KOC(text):
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        products = []
        buf = ""

        for line in lines:
            # gặp sản phẩm mới
            if re.match(r"^\d+\.", line):
                if buf:
                    m = re.search(r"SL\s*:\s*(\d+)", buf)
                    products.append({
                        "Barcode": re.sub(r"^\d+\.\s*", "", re.sub(r",?\s*SL\s*:\s*\d+.*", "", buf)).strip(),
                        "Qty Ord/Pcs": int(m.group(1)) if m else 0
                    })
                buf = line
            else:
                buf += " " + line

        # sản phẩm cuối
        if buf:
            m = re.search(r"SL\s*:\s*(\d+)", buf)
            products.append({
                "Barcode": re.sub(r"^\d+\.\s*", "", re.sub(r",?\s*SL\s*:\s*\d+.*", "", buf)).strip(),
                "Qty Ord/Pcs": int(m.group(1)) if m else 0
            })

        return products


    def tach_san_pham_JITpdf(text: str):
        def to_int(s: str):
            """Convert chuỗi số tiền/qty hợp lệ thành int, else None"""
            if re.match(r"^\d{1,3}(,\d{3})*$", s) or re.match(r"^\d+$", s):
                return int(s.replace(",", ""))
            return None

        def tach_qty_tu_line(line: str):
            """Tách qty nếu dính ở cuối dòng text, ví dụ 'Đậu xanh - 2.1L 2' -> ('Đậu xanh - 2.1L', 2)"""
            parts = line.strip().rsplit(" ", 1)
            if len(parts) == 2 and parts[1].isdigit():
                return parts[0], int(parts[1])
            return line, None

        lines = [line.strip() for line in text.split("\n") if line.strip()]
        products = []
        ma_don_hang = ""

        # --- Lấy mã đơn hàng ---
        for i, line in enumerate(lines):
            if "Mã đơn hàng:" in line:
                if len(line.split("Mã đơn hàng:")) > 1 and line.split("Mã đơn hàng:")[1].strip():
                    ma_don_hang = line.split("Mã đơn hàng:")[1].strip().split(" ")[0]
                elif i + 1 < len(lines):
                    ma_don_hang = lines[i + 1].split(" ")[0]
                break

        # --- Bỏ tất cả từ đầu đến dòng "Thành tiền" ---
        start_index = 0
        for i, line in enumerate(lines):
            if "Thành tiền" in line:
                start_index = i + 1
                break
        
        # Lấy phần còn lại sau header
        remaining_lines = lines[start_index:]
        
        if not remaining_lines:
            return ma_don_hang, products

        # --- Tìm các block sản phẩm ---
        i = 0
        while i < len(remaining_lines):
            found_product = False
            
            for j in range(i, len(remaining_lines) - 2):
                line1 = remaining_lines[j]
                line2 = remaining_lines[j + 1] 
                line3 = remaining_lines[j + 2]
                
                num1 = to_int(line1)
                num2 = to_int(line2)
                num3 = to_int(line3)
                
                # Case chuẩn: qty * đơn giá = thành tiền
                if num1 is not None and num2 is not None and num3 is not None:
                    if abs(num1 * num2 - num3) < 1:
                        product_block = remaining_lines[i:j+3]
                        product_name_lines = product_block[:-3]
                        qty_line = product_block[-3]
                        price_line = product_block[-2]
                        total_line = product_block[-1]

                        product_name = " ".join(product_name_lines).strip()
                        if product_name and product_name.split()[0].isdigit():
                            product_name = " ".join(product_name.split()[1:])

                        product = {
                            "product_name": product_name,
                            "Barcode": "",
                            "Qty Ord/Pcs": qty_line,
                            "don_gia": price_line,
                            "thanh_tien": total_line,
                        }
                        if "ProcessHandler" in globals():
                            product["Barcode"] =  ProcessHandler.laymahang_JIT(re.sub(r"\s+", "", product_name))

                        products.append(product)
                        i = j + 3
                        found_product = True
                        break
                
                # Case lệch cột: qty dính chung với dòng trước
                elif num2 is not None and num3 is not None:
                    prev_line = remaining_lines[j]
                    text_part, qty_found = tach_qty_tu_line(prev_line)
                    if qty_found is not None and abs(qty_found * num2 - num3) < 1:
                        product_block = remaining_lines[i:j+2]  # qty nằm trong prev_line
                        product_name_lines = product_block[:-1]  # bỏ dòng cuối chứa qty dính
                        product_name = " ".join(product_name_lines).strip()
                        if product_name and product_name.split()[0].isdigit():
                            product_name = " ".join(product_name.split()[1:])

                        product = {
                            "product_name": text_part.strip() if text_part else product_name,
                            "Barcode": "",
                            "Qty Ord/Pcs": str(qty_found),
                            "don_gia": line2,
                            "thanh_tien": line3,
                        }
                        if "ProcessHandler" in globals():
                            product["Barcode"] = ProcessHandler.laymahang_JIT(re.sub(r"\s+", "", product_name))

                        products.append(product)
                        i = j + 2
                        found_product = True
                        break

            if not found_product:
                i += 1

        return ma_don_hang, products






    def lamsachdonhang_lotte(text):
        start_pattern = "Sply qty"  # Tìm dòng bắt đầu
        end_marker = "Tot add tax"  # Dòng kết thúc

        lines = text.split("\n")
        start_index = None
        end_index = None

        for i, line in enumerate(lines):
            if start_index is None and re.match(start_pattern, line):  # Tìm dòng đầu tiên
                start_index = i
            if line.strip() == end_marker:  # Tìm dòng cuối cùng
                end_index = i
                break  # Dừng ngay khi tìm thấy dòng cuối

        if start_index is not None and end_index is not None and start_index < end_index:
            return "\n".join(lines[start_index + 1:end_index])  # Lấy nội dung giữa hai dòng

        return None  # Không tìm thấy hai dòng
    
    




    def tachdonhang_shopeechoice(text):
        lines = [line.strip() for line in text.split("\n") if line.strip()]
        groups = []
        i = 0

        while i < len(lines):
            if lines[i].isdigit():  # Số thứ tự
                stt = int(lines[i])
                i += 1

                # Gom tên sản phẩm tới khi gặp barcode (barcode có "_")
                name_parts = []
                while i < len(lines) and "_" not in lines[i]:
                    name_parts.append(lines[i])
                    i += 1
                name = " ".join(name_parts)

                # Nếu hết dữ liệu thì dừng
                if i >= len(lines):
                    break

                barcode = lines[i]
                qty = int(lines[i+1])
                price = float(lines[i+2])
                total = float(lines[i+3])

                groups.append({
                    "stt": stt,
                    "ten_san_pham": name,
                    "Barcode": barcode,
                    "OU Qty": qty,
                    "Total Price": price,
                    "tong_tien": total
                })

                i += 4
            else:
                i += 1


        return groups

    def lamsachdonhang(text,start_pattern,end_marker):

        lines = text.split("\n")
        start_index = None
        end_index = None

        for i, line in enumerate(lines):
            if start_index is None and re.match(start_pattern, line):  # Tìm dòng đầu tiên
                start_index = i
            if line.strip() == end_marker:  # Tìm dòng cuối cùng
                end_index = i
                break  # Dừng ngay khi tìm thấy dòng cuối

        if start_index is not None and end_index is not None and start_index < end_index:
            return "\n".join(lines[start_index + 1:end_index])  # Lấy nội dung giữa hai dòng

        return None  # Không tìm thấy hai dòng
    


    def trichxuatsanpham_satra(text):
        text = re.split(r"\bTổng cộng\b", text, maxsplit=1)[0].strip()
        pattern = re.compile(r"^\s*(\d+)\s+\d+\s*\n\s*(\d{13})", re.VERBOSE | re.MULTILINE)
        so_luong_pattern = re.compile(r"\b(\d{1,3}),000\b")  # Tìm số có dạng "x,000", "xx,000", "xxx,000"

        matches = list(pattern.finditer(text))
        if not matches:
            print("⚠️ Không tìm thấy dữ liệu phù hợp với regex!")
            return []

        positions = [(m.start(), m.group(2)) for m in matches]
        positions.append((len(text), None))  

        nhom_sanpham = []
        for i in range(len(positions) - 1):
            start, barcode = positions[i]
            end, _ = positions[i + 1]
            lines = [line.replace(" ", "") for line in text[start:end].strip().split("\n") if line.strip()]

            so_luong_index = next((i for i, line in enumerate(lines) if so_luong_pattern.search(line)), None)
            if so_luong_index is not None:
                so_luong = so_luong_pattern.sub(r"\1", lines[so_luong_index])
                thanh_tien = lines[so_luong_index + 1] if so_luong_index + 1 < len(lines) else ""  
                thanh_tien = re.sub(r",00$", "", thanh_tien)  # Bỏ phần ",00" nếu có
                # 👉 Chuyển về số và kiểm tra
                try:
                    gia = float(thanh_tien.replace(",", ""))
                    if gia == 0:
                        continue  # ✅ Bỏ qua nếu giá trị bằng 0
                except:
                    continue  # Bỏ qua nếu không chuyển được thành số
                nhom_sanpham.append({
                "Barcode": barcode,
                "OU Qty": so_luong,
                "Total Price": thanh_tien
            })

        return nhom_sanpham 


    def laydanhsanpham_emart(text):
        # Regex sửa đổi để khớp chính xác hơn
        product_pattern = re.finditer(r"""
            (?P<article_code>\d{7})\s*        # Article Code (7 chữ số)
            (?P<barcode>\d{12,13})\s*        # Barcode (12-13 chữ số)
            \s*(?P<description>.+?)\s+        # Description (kết thúc trước unit, ít nhất 1 khoảng trắng)
            (?P<unit>[A-Z]{2,})\s+           # Unit (ít nhất 2 chữ cái, như EA, BOX)
            \s*(?P<qty_in_box>\d+)\s+        # Qty in Box (yêu cầu khoảng trắng sau)
            \s*(?P<quantity>\d+)\s+          # PO Quantity (yêu cầu khoảng trắng sau)
            \s*(?P<purchase_price>[\d\.,]+)  # Purchase Price
        """, text.strip(), re.VERBOSE | re.DOTALL)

        results = []
        for match in product_pattern:
            purchase_price = match.group("purchase_price").replace(".", "")  # Xóa dấu chấm
            purchase_price_value = float(purchase_price.replace(",", "."))  # Chuyển đổi thành số
            
            if purchase_price_value == 0:
                continue  # Bỏ qua nếu giá trị "Total Price" bằng 0

            extracted_data = {
                "Barcode": match.group("barcode"),
                "OU Qty": int(match.group("quantity")),  # Ép kiểu thành int
                "Total Price": purchase_price
            }
            results.append(extracted_data)

        if not results:
            print("No valid matches found. Please check input format.")

        return results


    def laytenstore_lotte(text, po_number):
        start_pattern = "DOAN TUAN ANH"  # Dòng bắt đầu
        end_marker = po_number  # Dòng kết thúc

        lines = text.split("\n")
        start_index = None
        end_index = None

        for i, line in enumerate(lines):
            if start_index is None and re.match(start_pattern, line):  # Tìm dòng đầu tiên
                start_index = i
            if line.strip() == end_marker:  # Tìm dòng cuối cùng
                end_index = i
                break  # Dừng ngay khi tìm thấy dòng cuối

        if start_index is not None and end_index is not None and start_index < end_index:
            last_line = lines[end_index - 1].strip()  # Lấy dòng cuối cùng trước `end_marker`
            return last_line if last_line else None  # Trả về dòng cuối cùng (không rỗng)

        return None  # Không tìm thấy hai dòng
    




    def lamsachdonhang_kingfood(text: str):
        # Cho phép "Ghi", "Ghi ", "Ghi\n", hoặc "Ghi   \n  Chú"
        start_pattern = re.compile(r"Khu vực", re.I)
        end_marker = "TỔNG CỘNG"

        lines = text.split("\n")
        start_index = None
        end_index = None

        # 🔹 Ghép lại text để tìm vị trí "Ghi Chú" chính xác
        full_text = "\n".join(lines)
        start_match = start_pattern.search(full_text)

        if start_match:
            start_pos = start_match.end()  # kết thúc của cụm "Ghi Chú"
            # Lấy phần sau đó
            after_start = full_text[start_pos:]
            # Tìm vị trí "TỔNG CỘNG"
            end_match = re.search(rf"(?<=\n){end_marker}", after_start)
            if end_match:
                end_pos = end_match.start()
                return after_start[:end_pos].strip()

        return "Không có sản phẩm"

    
    def laydanhsachsanpham_kingfood(text):

        text = ProcessHandler.lamsachdonhang_kingfood(text)
        print('bắt đầu')
        print(text)
        print('kết thúc')
        # Pattern linh hoạt hơn cho tên sản phẩm nhiều dòng
        

        pattern = re.compile(r"""
    (?P<stt>\d+)\s*\n
    (?P<barcode>\d{13})\s*\n

    (?P<name>(?:.+\n)+?)

    (?P<unit>HỘP|TÚI|CHAI|LON|GÓI)\s*\n

    (?P<quantity>[\d.]+)\s*\n     # <-- cho phép 1.464

    \d+\s*\n
    .+\s*\n

    (?:.*\n){4}

    (?P<price>[0-9.,]+)

    """, re.VERBOSE)

        products = []

        matches = pattern.finditer(text)

        for i, match in enumerate(matches, 1):

            barcode = match.group("barcode")

            name = re.sub(r'\s+', ' ', match.group("name")).strip()

            unit = match.group("unit")

            quantity = match.group("quantity").replace('.', '')

            price_str = match.group("price")

            try:

                price = float(price_str.replace('.', '').replace(',', '.'))

                print(f"[MATCH {i}] Barcode: {barcode}")
                print(f"Name: {name}")
                print(f"Qty: {quantity}")
                print(f"Price: {price}")
                print("-" * 50)

                products.append({
                    "Barcode": barcode,
                    "Product Name": name,
                    "Unit": unit,
                    "OU Qty": int(quantity),
                    "Total Price": price
                })

            except Exception as e:
                print("Lỗi:", e)

        print(f"\n✅ Tổng số sản phẩm: {len(products)}")

        return products
    


    




    def trichxuatsanpham_winmart(text):
        # Gộp toàn bộ text lại thành một block
        if isinstance(text, list):
            text = "\n".join(text)

        # Chuẩn hóa: loại bỏ nhiều khoảng trắng thừa
        text = re.sub(r"[ \t]+", " ", text)

        # Regex nhận diện 1 sản phẩm Winmart (đa trang OK)
        pattern = re.compile(r"""
            (?P<no>\d+)\s*\n                          # Stt
            (?P<article>\d+)\s*\n                     # Mã hàng
            (?P<barcode>\d+)\s*\n                     # Mã vạch
            (?P<qty>[\d,]+)\s*\n                      # Số lượng
            [A-Z0-9]{2,4}\s*\n                        # ĐVT (G1, CHA…)
            (?P<unit_price>[\d,]+)\s*\n              # Đơn giá
            (?P<amount>[\d,]+)                       # Thành tiền
        """, re.VERBOSE)

        ket_qua = []

        for m in pattern.finditer(text):
            qty = m.group("qty").replace(",", "")
            amount = m.group("amount").replace(",", "")

            ket_qua.append({
                "Barcode": m.group("barcode"),
                "OU Qty": qty,
                "Total Price": amount
            })

        return ket_qua
    




    
    def cat_trang_hien_tai(src_doc, output_path, page=None):
        """
        page có thể là:
        - int            → page_index (0-based)
        - '3/10'         → page_label
        - '3'            → page số 3 (1-based)
        """

        page_index = None

        # 1️⃣ Tự nhận diện kiểu page
        if isinstance(page, int):
            page_index = page

        elif isinstance(page, str):
            page = page.strip()

            # Trường hợp 'x/y'
            if "/" in page:
                try:
                    page_num = int(page.split("/")[0])
                    page_index = page_num - 1
                except Exception:
                    raise ValueError(f"page_label không hợp lệ: {page}")

            # Trường hợp '3'
            else:
                if page.isdigit():
                    page_index = int(page) - 1
                else:
                    raise ValueError(f"page không hợp lệ: {page}")

        else:
            raise ValueError("page phải là int hoặc str")

        # 2️⃣ Kiểm tra phạm vi
        if page_index < 0 or page_index >= len(src_doc):
            raise IndexError(f"Trang không tồn tại: {page_index + 1}")

        # 3️⃣ Cắt trang
        dst = fitz.open()
        dst.insert_pdf(src_doc, from_page=page_index, to_page=page_index)
        dst.save(output_path)
        dst.close()



    def tach_san_pham_BHX(text: str) -> list:
        lines = [line.strip() for line in text.replace('\r', '\n').split('\n') if line.strip()]
        product_indices = [
            i for i in range(len(lines) - 1)
            if re.fullmatch(r"\d{1,3}", lines[i]) and re.fullmatch(r"\d{10,15}", lines[i + 1])
        ]

        ket_qua = []

        for idx, start in enumerate(product_indices):
            end = product_indices[idx + 1] if idx + 1 < len(product_indices) else len(lines)
            block = lines[start:end]

            # Tìm các dòng số trong block
            all_numbers = [line for line in block if re.fullmatch(r"[\d,]+\.\d{2}", line)]

            if len(all_numbers) < 3:
                print("❌ Không đủ số lượng để tách:", block)
                continue

            try:
                article = block[2]  # dòng thứ 3 là barcode
                quantity = all_numbers[0]  # thường là dòng chứa OU Qty
                amount = all_numbers[-1]  # dòng cuối là thành tiền

                ket_qua.append({
                    "Barcode": article,
                    "OU Qty": quantity.replace(",", ""),
                    "Total Price": amount.replace(",", "")
                })
            except Exception as e:
                print("❌ Lỗi xử lý block:", block, "→", e)

        return ket_qua

    
    def tach_san_pham_DIY(text):
        lines = text.strip().split('\n')
        ket_qua = []
        
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            
            # Tìm dòng có pattern "số)" để xác định sản phẩm
            if ')' in line and line.split(')')[0].isdigit():
                # Lấy OU Qty (2 dòng trước đó)
                if i >= 2:
                    ou_qty = lines[i-2].strip()
                    mahang = lines[i-5].strip()
                else:
                    ou_qty = ""
                    mahang = ""
                # Lấy Article/Barcode (dòng tiếp theo)
                if i + 1 < len(lines):
                    article = lines[i+1].strip()
                else:
                    article = ""
                
                # Lấy Total Price (dòng sau Article)
                if i + 2 < len(lines):
                    total_price = lines[i+2].strip()
                else:
                    total_price = ""
                
                # Thêm vào kết quả
                ket_qua.append({
                    "Barcode": article,
                    "Mahang": mahang,
                    "GTIN": article,
                    "OU Qty": ou_qty.replace(",", ""),
                    "Total Price": total_price.replace(",", "")
                })
            
            i += 1
        
        return ket_qua




    def tachsanpham_JMart(text: str):

        text = re.sub(r'(\d+,\d+)\.(\d)\s*\n\s*([0-9]{2})', r'\1.\2\3', text)
        text = re.sub(r'(\d+,\d+)\.\s*\n\s*([0-9]+)', r'\1.\2', text)

        lines = [l.strip() for l in text.splitlines()]
        result = []

        price_pattern = r'\d{1,3}(?:,\d{3})+\.\d{3}'

        for idx, line in enumerate(lines):

            if not re.fullmatch(r'\d{13}', line):
                continue

            barcode = line
            ou_qty = None
            total_price = None

            # ===== tìm OU Qty đúng block =====
            for i in range(idx - 1, max(idx - 20, 0), -1):

                if lines[i] == "1.00":
                    if re.fullmatch(r'[1-9]\d*\.000', lines[i - 1]):
                        ou_qty = lines[i - 1].split('.')[0]
                    break

            # ===== tìm giá tiền =====
            for j in range(idx - 1, -1, -1):
                if re.fullmatch(price_pattern, lines[j]):
                    total_price = lines[j].replace(",", "")
                    break

            result.append({
                "Barcode": barcode,
                "OU Qty": ou_qty,
                "Total Price": total_price
            })

        return result
    

    def debug_structure(text: str):
        # Gom số tiền bị tách dòng
        text = re.sub(r'(\d+,\d+)\.\s*\n\s*(\d+)', r'\1.\2', text)
        
        # Tìm tất cả barcode
        barcodes = re.findall(r'\d{13}', text)
        print(f"Tìm thấy {len(barcodes)} barcode")
        
        # Split và xem text trước mỗi barcode
        parts = re.split(r'(\d{13})', text)
        
        for i in range(1, len(parts), 2):
            barcode = parts[i]
            text_truoc = parts[i-1]
            text_sau = parts[i+1] if i+1 < len(parts) else ""
            
            print(f"\n{'='*50}")
            print(f"BARCODE: {barcode}")
            print(f"\nText TRƯỚC (50 ký tự cuối):")
            print(repr(text_truoc[-100:]))
            print(f"\nText SAU (100 ký tự đầu):")
            print(repr(text_sau[:100]))


    

    def format_date(value):
        if isinstance(value, datetime):
            return value.strftime("%d/%m/%Y")
        elif value is None:
            return None
        else:
            return str(value)  # fallback nếu là text sẵn


    def tach_san_pham_Fujimart(text: str) -> list:
        
        # Pattern để tìm từng sản phẩm
        pattern = re.compile(r"""
            (\d+)\n                     # STT (1, 2, 3...)
            ([\d.]+)\n                  # Quantity 
            ([\d,]+)\n                  # Total Amount
            ([A-Z]+)\n                  # Unit (TUI, HOP...)
            ([\d,]+)\n                  # Unit Price
            (.+?)\n                     # Product Name
            (\d+|[A-Z0-9]+)             # Barcode hoặc Product Code
        """, re.VERBOSE | re.MULTILINE)
        
        ket_qua = []
        
        for match in pattern.finditer(text):
            stt = match.group(1)
            quantity = match.group(2)
            total_amount = match.group(3)
            unit = match.group(4)
            unit_price = match.group(5)
            product_name = match.group(6)
            barcode = match.group(7)
            
            print(f"Match found - STT: {stt}, Barcode: {barcode}, Name: {product_name}")
            
            ket_qua.append({
                "Barcode": barcode,
                "Product Name": product_name.strip(),
                "Unit": unit,
                "Unit Price": unit_price.replace(",", ""),
                "Total Price": total_amount.replace(",", ""),
                "OU Qty": quantity.replace(",", "")
            })
        
        print(f"\n✅ Regex - Tổng số sản phẩm tìm được: {len(ket_qua)}")
        return ket_qua
    


    
    def tach_san_pham_BCMart(text: str,  debug: bool = True):
        # 1) Chuẩn hóa & bỏ dòng trống
        lines = [l.strip() for l in text.splitlines() if l.strip()]

        start_indices = []

        # 2) Xác định vị trí bắt đầu mỗi sản phẩm:
        #    STT (số) + mã 8 số + barcode 10–14 số
        for i in range(len(lines) - 2):
            stt_line = lines[i]
            code_line = lines[i + 1]
            barcode_line = lines[i + 2]

            if (re.fullmatch(r"\d+", stt_line) and          # STT
                re.fullmatch(r"\d{8}", code_line) and       # mã hàng 8 số
                re.fullmatch(r"\d{10,14}", barcode_line)):  # barcode 10–14 số
                start_indices.append(i)

        if debug:
            print("== DEBUG: start_indices =", start_indices)

        # 3) Cắt thành từng block sản phẩm
        blocks = []
        for idx, start in enumerate(start_indices):
            end = start_indices[idx + 1] if idx + 1 < len(start_indices) else len(lines)
            block = lines[start:end]
            blocks.append(block)

            if debug:
                print(f"\n=== BLOCK {idx + 1} (lines {start} -> {end - 1}) ===")
                for j, ln in enumerate(block):
                    print(f"{j:2d}: {ln}")

        # 4) Parse từng block
        ket_qua = []
        money_re = re.compile(r"^\d{1,3}(?:\.\d{3})*,\d{2}$")  # dạng 95.153,00

        for idx, b in enumerate(blocks):
            barcode = ""
            ou_qty = ""
            total_price = ""

            # Đảm bảo block đủ dài theo format:
            # 0: STT
            # 1: mã nội bộ
            # 2: barcode
            # 3: tên hàng
            # 4: đơn vị (Túi/Can/...)
            # 5: OU Qty
            # 6: đơn giá (95.153,00)
            if len(b) >= 7:
                barcode = b[2]
                ou_qty  = b[5]

                # Tìm dòng tiền đầu tiên sau qty
                for ln in b[6:]:
                    if money_re.match(ln):
                        total_price = ln
                        break

            if debug:
                print(f"\n>>> PARSED BLOCK {idx + 1}:")
                print("    Barcode    =", barcode)
                print("    OU Qty     =", ou_qty)
                print("    TotalPrice =", total_price)

            ket_qua.append({
                "Barcode": barcode or "",
                "OU Qty": ou_qty.replace(",", "") if ou_qty else "",
                "Total Price": total_price.replace(",00", "").replace(".", "") if total_price else ""
            })

        return ket_qua


    def tach_san_pham_Farmer(text: str) -> list:
        def parse_number(value: str) -> float:
            """Chuyển '765.419,00' → 765419.00"""
            value = value.strip()
            if ',' in value:
                value = value.replace('.', '').replace(',', '.')
            return float(value)

        lines = [line.strip() for line in text.strip().splitlines() if line.strip()]
        start_index = next((i for i, line in enumerate(lines) if line.strip() == "1"), None)
        if start_index is None:
            print("❌ Không tìm thấy dòng bắt đầu là '1'")
            return []

        lines = lines[start_index:]
        ket_qua = []
        i = 0

        while i < len(lines):
            line = lines[i]
            barcode_match = re.match(r"^\d{13}\b", line)
            if barcode_match:
                barcode = barcode_match.group()
                block = [line]
                j = i + 1
                while j < len(lines) and len(block) < 9:
                    block.append(lines[j])
                    j += 1

                try:
                    # Tìm số lượng: số đầu tiên có dạng xxx,00 sau barcode
                    quantity = None
                    for line in block:
                        numbers = re.findall(r"\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{2})", line)
                        for num in numbers:
                            val = parse_number(num)
                            if val.is_integer() and val < 1000:  # Số lượng thường nhỏ hơn 1000
                                quantity = int(val)
                                break
                        if quantity:
                            break

                    # Tìm amount: số tiền cuối cùng
                    all_numbers = re.findall(r"\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{2})", " ".join(block))
                    amount = parse_number(all_numbers[-1]) if all_numbers else None

                    if not quantity or not amount:
                        raise ValueError("Không tìm được đầy đủ dữ liệu.")

                    ket_qua.append({
                        "Barcode": barcode,
                        "Total Price": f"{amount:.2f}",
                        "OU Qty": f"{quantity}"
                    })

                except Exception as e:
                    print(f"⚠️ Lỗi block: {block} → {e}")

                i = j
            else:
                i += 1

        return ket_qua
    
    @staticmethod
    def taongaunhien(length: int) -> str:
        """
        Sinh chuỗi ngẫu nhiên từ tập ký tự A–Z, a–z, 0–9.
        
        :param length: Độ dài chuỗi cần sinh
        :return: Chuỗi ngẫu nhiên
        """
        chars = string.ascii_letters + string.digits
        return ''.join(random.choice(chars) for _ in range(length))

        

    def fix_vni_to_unicode(text):
        vni_to_unicode = {
        "¸": "á", "µ": "à", "¶": "ả", "·": "ã", "¹": "ạ",
        "¨": "â", "Ç": "ầ", "©": "ấ", "Ê": "ẩ", "È": "ẫ", "É": "ậ",
        "¢": "ă", "¤": "ằ", "¡": "ắ", "£": "ẳ", "¥": "ẵ", "¦": "ặ",

        "è": "è", "é": "é", "ê": "ẻ", "ë": "ẽ", "ç": "ẹ",
        "ª": "ê", "Ì": "ề", "Ñ": "ế", "Í": "ể", "Î": "ễ", "Ï": "ệ",

        "ï": "ì", "Ý": "í", "ó": "ỉ", "ò": "ĩ", "î": "ị",

        "ó": "ó", "ò": "ò", "ô": "ỏ", "õ": "õ", "ö": "ọ",
        "¯": "ô", "×": "ố", "Ø": "ồ", "Ù": "ổ", "Ú": "ỗ", "Ü": "ộ",
        "­": "ơ", "¬": "ờ", "½": "ớ", "¾": "ở", "¿": "ỡ", "À": "ợ",

        "ú": "ú", "ù": "ù", "û": "ủ", "ü": "ũ", "þ": "ụ",
        "®": "ư", "Å": "ứ", "Ä": "ừ", "Æ": "ử", "Ç": "ữ", "È": "ự",

        "ý": "ý", "ÿ": "ỷ", "×": "ỹ", "Ø": "ỳ", "Þ": "ỵ",

        "§": "đ",
    }

        for vni_char, unicode_char in vni_to_unicode.items():
            text = text.replace(vni_char, unicode_char)
        return text


    def cat_donhang_KOC(text: str):
        lines = [l.rstrip() for l in text.splitlines()]

        start_idx = None
        end_idx = None

        # regex dòng bắt đầu: Nội dung hàng (Tổng SL sản phẩm: X)
        start_pattern = re.compile(r"Nội dung hàng\s*\(Tổng SL sản phẩm:\s*\d+\)", re.IGNORECASE)

        end_text = "*** Người nhận đồng ý việc SPX thu thập bằng chứng giao hàng"

        for i, line in enumerate(lines):
            if start_idx is None and start_pattern.search(line):
                start_idx = i
                continue

            if start_idx is not None and end_text in line:
                end_idx = i
                break

        if start_idx is None or end_idx is None or end_idx <= start_idx:
            return None

        # Cắt đoạn (loại bỏ dòng tiêu đề nếu muốn thì +1)
        content_lines = lines[start_idx + 1 : end_idx]

        return "\n".join(l for l in content_lines if l.strip())

    def process_file(self, file_path,stt = 1):





        """Xử lý PDF, TXT hoặc XLSX"""
        self.log_signal.emit(f"🔄 Đang xử lý file: {file_path}")
        file_name = os.path.basename(file_path)
        fileloi = ''
        

        #try: 

        
                    


        if file_path.lower().endswith(".pdf"):
            doc = fitz.open(file_path)
            kiemtrapo = ""
            
            full_text = "\n".join(page.get_text("text") for page in doc)
        

            for page_num in range(len(doc)):
                text = doc[page_num].get_text("text")
                
                
                
                
                vendor = ProcessHandler.identify_vendor(text)


                if vendor == "Unknown" and (
    file_name.startswith("package_list_") or file_name.startswith("air_waybill_")
):
                    vendor = "JIT"
                    # Xử lý khi thỏa 2 điều kiện
                    page_label = f'{page_num + 1}/{len(doc)}'  # đổi tên biến này


                    
                    match = re.match(r"package_list_(.+)_(\d+)\.pdf", file_name)
                    if match:
                        wh = match.group(1)   # WH6_HN
                        od_raw = match.group(2)   # 19082025
                       
                        od = datetime.strptime(od_raw, "%d%m%Y").strftime("%d/%m/%Y")
                    
                        products = ProcessHandler.tach_san_pham_JITpdf(text)[1]
                        po_number = ProcessHandler.tach_san_pham_JITpdf(text)[0]


                        now = datetime.now(ZoneInfo("Asia/Ho_Chi_Minh"))
                        khung_gio = 'sáng' if now.hour < 12 else 'chiều'

                        dien_giai = f'{od} ({khung_gio})'


                        print(po_number)
                        print(products)
                        # Load và thay SKU theo mapping
                        sku_mapping = ProcessHandler.load_sku_mapping()
                        print(f"[DEBUG] SKU mapping size: {len(sku_mapping)}")


                        



                        products = ProcessHandler.replace_sku_numbers(products, sku_mapping)


                        # Gửi từng nhóm sang JIT
                        ProcessHandler.write_to_dondathang_JIT(
                            self,
                            products,
                            'MN_JIT_01512',
                            dien_giai,
                            od,  # order_date
                            od,  # order_date
                            stt,
                            'JIT',
                            wh,
                            po_number # warehouse
                        )

                        self.table_signal.emit(file_name, page_label, 'JIT-CHOICE', wh, po_number, '', "✅Hoàn Thành")

                    #JIT air_waybill
                    match = re.match(r"air_waybill_(.+?)_(\d+)\.pdf", file_name)
                    if match:
                        wh = match.group(1)   # WH6_HN
                        od_raw = match.group(2)   # 19082025

                        if page_num == 0:
                            now_xuly = datetime.now()
                            now_str = now_xuly.strftime("%d/%m/%Y %H:%M:%S")
                            
                            
                            ProcessHandler.ghi_message(f"[vendor: {vendor}")
                            ProcessHandler.ghi_message(f"start_time: {now_str}")
                            ProcessHandler.ghi_message(f"khu_vuc: {wh}")
                            ProcessHandler.ghi_message(f"Mã Khách hàng: MN_JIT_01512")
                            now = datetime.now(ZoneInfo("Asia/Ho_Chi_Minh"))
                            khung_gio = 'sáng' if now.hour < 12 else 'chiều'
                            
                            ProcessHandler.ghi_message(f"khung_gio: {khung_gio}")
                            


                        

                        
                       
                       
                        od = datetime.strptime(od_raw, "%d%m%Y").strftime("%d/%m/%Y")

                        
                        products = ProcessHandler.cat_giua_theo_dong(text, "Mã vận đơn:", "Top Value")
                        
                        
                        
                        products = ProcessHandler.tach_sanpham_JIT_airway(products)
                        
                        if not products:
                            products = ProcessHandler.cat_giua_theo_dong(text, "Top Value", "Mã đơn hàng:")
                        
                        
                        
                            products = ProcessHandler.tach_sanpham_JIT_airway(products)


                        try:
                            po_number = text.splitlines()[text.splitlines().index(next(l for l in text.splitlines() if "Mã đơn hàng:" in l)) + 2].strip()
                        except StopIteration:
                            # xử lý riêng file thiếu chữ "Mã đơn hàng:"
                            for line in text.splitlines():
                                if line.strip() == "2603157B6CWHCS":
                                    po_number = line.strip()
                                    break                                           
                        # ✅ Kiểm tra 6 ký tự đầu có phải số không
                        if not po_number[:6].isdigit():
                            po_number = text.splitlines()[text.splitlines().index(next(l for l in text.splitlines() if "Mã đơn hàng:" in l)) - 1].strip()
                            if not po_number[:6].isdigit():
                                po_number = text.splitlines()[text.splitlines().index(next(l for l in text.splitlines() if "Mã vận đơn:" in l)) - 1].strip()


                        
                        
                        dien_giai = f'{od} ({khung_gio})'


                        print(po_number)
                        print(products)
                        # Load và thay SKU theo mapping
                        sku_mapping = ProcessHandler.load_sku_mapping()
                        products = ProcessHandler.replace_sku_numbers(products, sku_mapping)
                        
                        


                        # Gửi từng nhóm sang JIT
                        ProcessHandler.write_to_dondathang_JIT(
                            self,
                            products,
                            'MN_JIT_01512',
                            dien_giai,
                            od,  # order_date
                            od,  # order_date
                            stt,
                            'JIT',
                            wh,
                            po_number # warehouse
                        )

                        self.table_signal.emit(file_name, page_label, 'JIT-CHOICE', wh, po_number, '', "✅Hoàn Thành")
                        
                        if page_num == len(doc) - 1:
                            print('Trang cuối')
                            ProcessHandler.ghi_message(f"tong_don: {len(doc)}")
                            now2 = datetime.now()
                            now_str = now2.strftime("%d/%m/%Y %H:%M:%S")
                            ProcessHandler.ghi_message(f"end_time: {now_str}")
                            time = now2 - now_xuly
                            ProcessHandler.ghi_message(f"time: {ProcessHandler.format_timedelta(time)}]")

                
                    

                    




                    





                
                if vendor != "Unknown":
                    kiemtrapo = vendor
                    #print(text)
                page_label = f'{page_num + 1}/{len(doc)}'  # đổi tên biến này



                



                
                
                
                
                
                if text == "":
                    # Chuyển trang PDF thành ảnh
                    page = doc.load_page(page_num)  # thêm dòng này để lấy đúng trang PDF
                    # Dùng dpi cao để ảnh rõ
                    pix = page.get_pixmap(dpi=500)
                    img = Image.open(io.BytesIO(pix.tobytes("png")))

                    # Tiền xử lý ảnh: grayscale + tăng tương phản
                    img = img.convert("L")  # grayscale
                    img = ImageEnhance.Contrast(img).enhance(2.5)

                    # OCR tiếng Việt với cấu hình tốt
                    text = pytesseract.image_to_string(img, lang='vie', config='--psm 11')

                    print(f"\n🧾 Trang {page_num + 1}/{len(doc)}:\n{text}\n{'='*80}")

                    

                
                
                
                
            
                if page_num == 0:
                    self.log_signal.emit(f"<b style='color:blue;'>📌 Phát hiện Đơn Hàng:</b> <span style='color:green;'>{vendor}</span>")
                    self.log_signal.emit(f"<b style='color:purple;'>📄 Tổng số trang:</b> <span style='color:red;'>{len(doc)}</span>")
                    self.log_signal.emit(f"<b style='color:orange;'>📝 Đang xử lý trang:</b> <span style='color:brown;'>{page_num + 1}</span>")

                

                if vendor == "Coop":
                    sodon = process_handler.demsodonhang1trang_coop(text)
                    
                    pom343 = sodon.get("POM343")
                    sub_total = sodon.get("Sub Total")
                    
                
                
                    print(f"Trang {page_num + 1}")
                    print(f"POM343: {pom343}, Sub Total: {sub_total}")
                    
                    if pom343 == 1 and sub_total == 1:
                        ProcessHandler.process_coop_invoice(self,text,stt,file_path,page_label,doc)
                        stt = stt +1
                    elif pom343 > 1 and sub_total > 1:





                        if pom343 == sub_total:
                            
                            catdon = ProcessHandler.catdonra_nhieutrang(text)
                            for idx, segment in enumerate(catdon):
                                ProcessHandler.process_coop_invoice(self,segment,stt,file_path,page_label)
                                stt = stt + 1
                        else:
                            print("khác nhau")



                



                elif vendor == "Clevy-ViettelPost":
                    print(text)
                    lines = text.strip().splitlines()
                    last2 = lines[-1].split(' ') if lines else ""
                    entry_date = last2[0]
                    gio_dat = last2[1]

                    print(entry_date)
                    print(gio_dat)

                    ma_donhang = lines[2].strip() if len(lines) >= 3 else ""
                    print(ma_donhang)
                    products = ProcessHandler.cat_giua_theo_dong(text, f"{ma_donhang} /", "CÔNG TY CỔ PHẦN EDUCI GROUP")
                    print('bắt đầu')
                    print(products)
                    print('Kết thúc')
                    products = ProcessHandler.tach_sanpham_clevy_viettelpost(products)
                    khuvuc = 'HCM'
                    print(products)
                    
                    sku_mapping = ProcessHandler.load_sku_mapping()
                    products = ProcessHandler.replace_sku_numbers(products, sku_mapping)
                    print(products)
                    url = "https://script.google.com/macros/s/AKfycbwgWp1NIGDRIDbDint3YUArofepaFCGW6wzInkYh7h1NhRuSVqtWLU9JuGdQK8kvMecjQ/exec"

                    for p in products:
                        payload = {
                            "he_thong": "Sàn",
                            "data": {
                                "ngay_dathang": entry_date,
                                "ma_vandon": ma_donhang,
                                "ma_sanpham": p.get('Barcode', ''),
                                "ten_sanpham": ProcessHandler.timten_sanpham(p.get('Barcode', '')), 
                                "so_luong": p.get('Qty Ord/Pcs', 0),
                                "dvvc": "Viettelpost",
                                "tinh_trang": "Chờ xử lý",
                                "ghi_chu": f'Mã Đơn hàng: {ma_donhang} - Ngày giờ đặt: {entry_date} {gio_dat}'
                            }
                        }

                        r = requests.post(url, json=payload)
                        print(r.text)
                        self.table_signal.emit(file_name, page_label, 'Clevy', khuvuc, ma_donhang, '', "✅Hoàn Thành")





                
                elif vendor == "Clevy":
                    print(text)
                    dong_dau = text.strip().split('\n')[0]

                    if not dong_dau.startswith('V'):
                        print("Shopee")
                        if not dong_dau.startswith('Từ:'):
                            lines = text.splitlines()
                            for i, line in enumerate(lines):
                                if "Mã đơn hàng:" in line:
                                    ma_vandon = lines[i+1].strip()
                                    ma_donhang = lines[i+2].strip()
                                    print(ma_vandon, ma_donhang)


                            products = ProcessHandler.cat_giua_theo_dong(text, "Mã vận đơn:", "CLEVY VIỆT NAM")
                            print(products)
                            lines = products.splitlines()
                            last2 = lines[-3:-1] 
                            entry_date = last2[0]
                            gio_dat = last2[1]


                            print(entry_date)
                            print(gio_dat)
                            products = ProcessHandler.tach_sanpham_JIT_airway(products)
                            khuvuc = "HCM" if "SĐT: 84335617279" in text else "HN"
                            print(khuvuc)
                            if not products:
                                products = ProcessHandler.cat_giua_theo_dong(text, "CLEVY VIỆT NAM", "Mã đơn hàng:")
                
                                products = ProcessHandler.tach_sanpham_JIT_airway(products)

                            # Load và thay SKU theo mapping
                            sku_mapping = ProcessHandler.load_sku_mapping()
                            products = ProcessHandler.replace_sku_numbers(products, sku_mapping)


                            print(products)
                            url = "https://script.google.com/macros/s/AKfycbwgWp1NIGDRIDbDint3YUArofepaFCGW6wzInkYh7h1NhRuSVqtWLU9JuGdQK8kvMecjQ/exec"

                            for p in products:
                                payload = {
                                    "he_thong": "TMĐT",
                                    "data": {
                                        "ngay_dathang": entry_date,
                                        "ma_vandon": ma_vandon,
                                        "ma_sanpham": p.get('Barcode', ''),
                                        "ten_sanpham": ProcessHandler.timten_sanpham(p.get('Barcode', '')), 
                                        "so_luong": p.get('Qty Ord/Pcs', 0),
                                        "dvvc": "",
                                        "tinh_trang": "",
                                        "ghi_chu": f'Shopee - Mã Đơn hàng: {ma_donhang} - Ngày giờ đặt: {entry_date} {gio_dat}'
                                    }
                                }

                                r = requests.post(url, json=payload)
                                print(r.text)
                                self.table_signal.emit(file_name, page_label, 'Clevy', khuvuc, ma_donhang, '', "✅Hoàn Thành")
                        else:
                            lines = text.splitlines()
                            for i, line in enumerate(lines):
                                if "Mã đơn hàng:" in line:
                                    ma_donhang = lines[i-1].strip()
                            for i, line in enumerate(lines):
                                if "Mã vận đơn:" in line:
                                    ma_vandon = lines[i+1].strip()
                                    

                            print(ma_vandon, ma_donhang)
                            ngay_dathang = ProcessHandler.cat_giua_theo_dong(text, "Được đồng kiểm", "Ngày đặt hàng:")
                            

                            ngay_dathang = ngay_dathang.splitlines()

                            entry_date = ngay_dathang[0]
                            gio_dat = ngay_dathang[1]
                            print(entry_date)
                            print(gio_dat)
                            khuvuc = "HCM" if "SĐT: 84335617279" in text else "HN"
                            products = ProcessHandler.cat_giua_theo_dong(text, "Mã vận đơn:", "CLEVY VIỆT NAM")

                            products = ProcessHandler.tach_sanpham_JIT_airway(products)

                            print(khuvuc)
                            if not products:
                                products = ProcessHandler.cat_giua_theo_dong(text, "CLEVY VIỆT NAM", "Mã đơn hàng:")
                
                                products = ProcessHandler.tach_sanpham_JIT_airway(products)

                            # Load và thay SKU theo mapping
                            sku_mapping = ProcessHandler.load_sku_mapping()
                            products = ProcessHandler.replace_sku_numbers(products, sku_mapping)


                            print(products)
                            url = "https://script.google.com/macros/s/AKfycbwgWp1NIGDRIDbDint3YUArofepaFCGW6wzInkYh7h1NhRuSVqtWLU9JuGdQK8kvMecjQ/exec"

                            for p in products:
                                payload = {
                                    "he_thong": "TMĐT",
                                    "data": {
                                        "ngay_dathang": entry_date,
                                        "ma_vandon": ma_vandon,
                                        "ma_sanpham": p.get('Barcode', ''),
                                        "ten_sanpham": ProcessHandler.timten_sanpham(p.get('Barcode', '')), 
                                        "so_luong": p.get('Qty Ord/Pcs', 0),
                                        "dvvc": "",
                                        "tinh_trang": "",
                                        "ghi_chu": f'Shopee - Mã Đơn hàng: {ma_donhang} - Ngày giờ đặt: {entry_date} {gio_dat}'
                                    }
                                }

                                r = requests.post(url, json=payload)
                                print(r.text)
                                self.table_signal.emit(file_name, page_label, 'Clevy', khuvuc, ma_donhang, '', "✅Hoàn Thành")










                    else:
                        print("Tiktok")
                        lines = text.splitlines()
                        for i, line in enumerate(lines):
                            if "CLEVY VIỆT NAM" in line:
                                if i >= 3:
                                    ma_vandon = lines[i - 3]
                                    print(ma_vandon)
                                break

                        for i, line in enumerate(lines):
                            if "Thời gian đặt hàng:" in line:
                                if i >= 3:
                                    thoi_gian = lines[i - 1]
                                    print(thoi_gian)
                                break
                        entry_date = thoi_gian.split()[0]
                        gio_dat = thoi_gian.split()[1]
                        print(entry_date)
                        print(gio_dat)
                        
                        for i, line in enumerate(lines):
                            if "Order ID:" in line:
                                ma_donhang = lines[i+1] if i+1 < len(lines) else ""
                                print(ma_donhang)
                                break

                        products = ProcessHandler.cat_giua_theo_dong(text, "Product Name", f"Order ID: {ma_donhang}")
                        
                        products = '\n'.join(products.splitlines()[3:])
                        print(products)
                        
                        products = ProcessHandler.tach_sanpham_Clevy_tiktok(products)
                        
                        sku_mapping = ProcessHandler.load_sku_mapping()
                        products = ProcessHandler.replace_sku_numbers(products, sku_mapping)
                        khuvuc = 'HCM'


                        print(products)
                        url = "https://script.google.com/macros/s/AKfycbwgWp1NIGDRIDbDint3YUArofepaFCGW6wzInkYh7h1NhRuSVqtWLU9JuGdQK8kvMecjQ/exec"

                        for p in products:
                            payload = {
                                "he_thong": "TMĐT",
                                "data": {
                                    "ngay_dathang": entry_date,
                                    "ma_vandon": ma_vandon,
                                    "ma_sanpham": p.get('Barcode', ''),
                                    "ten_sanpham": ProcessHandler.timten_sanpham(p.get('Barcode', '')), 
                                    "so_luong": p.get('Qty Ord/Pcs', 0),
                                    "dvvc": "",
                                    "tinh_trang": "",
                                    "ghi_chu": f'Tiktok - Mã Đơn hàng: {ma_donhang} - Ngày giờ đặt: {entry_date} {gio_dat} '
                                }
                            }

                            r = requests.post(url, json=payload)
                            print(r.text)
                            self.table_signal.emit(file_name, page_label, 'Clevy', khuvuc, ma_donhang, '', "✅Hoàn Thành")
                        

                        





                        






                
                elif vendor == 'Clevy-Care':
                    print("Tiktok")
                    lines = text.splitlines()
                    for i, line in enumerate(lines):
                        if "Clevy VN – Care Plus" in line:
                            if i >= 3:
                                ma_vandon = lines[i - 3]
                                print(ma_vandon)
                            break

                    for i, line in enumerate(lines):
                        if "Thời gian đặt hàng:" in line:
                            if i >= 3:
                                thoi_gian = lines[i - 1]
                                print(thoi_gian)
                            break
                    entry_date = thoi_gian.split()[0]
                    gio_dat = thoi_gian.split()[1]
                    print(entry_date)
                    print(gio_dat)
                    
                    for i, line in enumerate(lines):
                        if "Order ID:" in line:
                            ma_donhang = lines[i+1] if i+1 < len(lines) else ""
                            print(ma_donhang)
                            break

                    products = ProcessHandler.cat_giua_theo_dong(text, "Product Name", f"Order ID: {ma_donhang}")
                    
                    products = '\n'.join(products.splitlines()[3:])
                    print(products)
                    
                    products = ProcessHandler.tach_sanpham_Clevy_tiktok(products)
                    
                    sku_mapping = ProcessHandler.load_sku_mapping()
                    products = ProcessHandler.replace_sku_numbers(products, sku_mapping)
                    khuvuc = 'HCM'


                    print(products)
                    url = "https://script.google.com/macros/s/AKfycbwgWp1NIGDRIDbDint3YUArofepaFCGW6wzInkYh7h1NhRuSVqtWLU9JuGdQK8kvMecjQ/exec"

                    for p in products:
                        payload = {
                            "he_thong": "TMĐT",
                            "data": {
                                "ngay_dathang": entry_date,
                                "ma_vandon": ma_vandon,
                                "ma_sanpham": p.get('Barcode', ''),
                                "ten_sanpham": ProcessHandler.timten_sanpham(p.get('Barcode', '')), 
                                "so_luong": p.get('Qty Ord/Pcs', 0),
                                "dvvc": "",
                                "tinh_trang": "",
                                "ghi_chu": f'Tiktok - Mã Đơn hàng: {ma_donhang} - Ngày giờ đặt: {entry_date} {gio_dat}'
                            }
                        }

                        r = requests.post(url, json=payload)
                        print(r.text)
                        self.table_signal.emit(file_name, page_label, 'Clevy', khuvuc, ma_donhang, '', "✅Hoàn Thành")
                

                elif vendor == "JV-Mart":
                    if page_num == 0:
                        now = datetime.now()
                        now_str = now.strftime("%d/%m/%Y %H:%M:%S")
                        
                        
                        ProcessHandler.ghi_message(f"[vendor: {vendor}")
                        ProcessHandler.ghi_message(f"start_time: {now_str}")
                        khuvuc = "HCM" if "SĐT: 84909764354" in text else "HN"
                        ProcessHandler.ghi_message(f"khu_vuc: {khuvuc}")
                        ProcessHandler.ghi_message(f"Mã Khách hàng: MB_SHOPEEGIUPVIEC")
                    



                    print(text)
                    lines = text.splitlines()
                    for i, line in enumerate(lines):
                        if "Mã đơn hàng:" in line:
                            ma_vandon = lines[i+1].strip()
                            ma_donhang = lines[i+2].strip()
                            print(ma_vandon, ma_donhang)


                    products = ProcessHandler.cat_giua_theo_dong(text, "Mã vận đơn:", "JV Mart – Siêu Thị Đồ Vệ Sinh")
                    print(products)

                    lines = products.splitlines()
                    last2 = lines[-3:-1] 
                    entry_date = last2[0]
                    gio_dat = last2[1]
                    print(entry_date)
                    print(gio_dat)
                    products = ProcessHandler.tach_sanpham_JIT_airway(products)
                    khuvuc = "HCM" if "SĐT: 84335617279" in text else "HN"
                    print(khuvuc)
                    if not products:
                        products = ProcessHandler.cat_giua_theo_dong(text, "JV Mart – Siêu Thị Đồ Vệ Sinh", "Mã đơn hàng:")
          
                        products = ProcessHandler.tach_sanpham_JIT_airway(products)


                    

                    dien_giai = f'{ma_donhang} {ma_vandon}-{khuvuc}'
                    ProcessHandler.ghi_message(f"text: {ma_donhang}")


          
                    # Load và thay SKU theo mapping
                    sku_mapping = ProcessHandler.load_sku_mapping()
                    products = ProcessHandler.replace_sku_numbers(products, sku_mapping)


                    # Gửi từng nhóm sang JIT
                    ProcessHandler.write_to_dondathang_Jupviec(
                        self,
                        products,
                        'MB_SHOPEEGIUPVIEC',
                        dien_giai,
                        entry_date,  # order_date
                        entry_date,  # order_date
                        stt,
                        'JV-Mart',
                        khuvuc,
                        ma_donhang # warehouse
                    )

                    self.table_signal.emit(file_name, page_label, 'JV-Mart', khuvuc, ma_donhang, '', "✅Hoàn Thành")
                    if page_num == len(doc) - 1:
                        print('Trang cuối')
                        ProcessHandler.ghi_message(f"tong_don: {len(doc)}")
                        now2 = datetime.now()
                        now_str = now2.strftime("%d/%m/%Y %H:%M:%S")
                        ProcessHandler.ghi_message(f"end_time: {now_str}")
                        time = now2 - now
                        ProcessHandler.ghi_message(f"time: {ProcessHandler.format_timedelta(time)}]")  

                    
                
            
                



                elif vendor == "JV-Mart-2":
                    if page_num == 0:
                        now = datetime.now()
                        now_str = now.strftime("%d/%m/%Y %H:%M:%S")
                        
                        
                        ProcessHandler.ghi_message(f"[vendor: JV-Mart")
                        ProcessHandler.ghi_message(f"start_time: {now_str}")
                        khuvuc = "HCM" if "SĐT: 84909764354" in text else "HN"
                        ProcessHandler.ghi_message(f"khu_vuc: {khuvuc}")
                        ProcessHandler.ghi_message(f"Mã Khách hàng: MB_SHOPEEGIUPVIEC")

                    print(text)
                    lines = text.splitlines()
                    for i, line in enumerate(lines):
                        if "Mã đơn hàng:" in line:
                            ma_vandon = lines[i+1].strip()
                            ma_donhang = lines[i+2].strip()
                            print(ma_vandon, ma_donhang)


                    products = ProcessHandler.cat_giua_theo_dong(text, "Mã vận đơn:", "JV Mart - Chăm sóc Nhà cửa")
                    print(products)

                    lines = products.splitlines()
                    last2 = lines[-3:-1] 
                    entry_date = last2[0]
                    gio_dat = last2[1]
                    print(entry_date)
                    print(gio_dat)
                    products = ProcessHandler.tach_sanpham_JIT_airway(products)
                    khuvuc = "HCM" if "SĐT: 84909764354" in text else "HN"
                    print(khuvuc)
                    if not products:
                        products = ProcessHandler.cat_giua_theo_dong(text, "JV Mart - Chăm sóc Nhà cửa", "Mã đơn hàng:")
          
                        products = ProcessHandler.tach_sanpham_JIT_airway(products)


                    

                    dien_giai = f'{ma_donhang} {ma_vandon}-{khuvuc}'
                    ProcessHandler.ghi_message(f"text: {ma_donhang}")


          
                    # Load và thay SKU theo mapping
                    sku_mapping = ProcessHandler.load_sku_mapping()
                    products = ProcessHandler.replace_sku_numbers(products, sku_mapping)


                    # Gửi từng nhóm sang JIT
                    ProcessHandler.write_to_dondathang_Jupviec(
                        self,
                        products,
                        'MB_SHOPEEGIUPVIEC',
                        dien_giai,
                        entry_date,  # order_date
                        entry_date,  # order_date
                        stt,
                        'JV-Mart',
                        khuvuc,
                        ma_donhang # warehouse
                    )

                    self.table_signal.emit(file_name, page_label, 'JV-Mart', khuvuc, ma_donhang, '', "✅Hoàn Thành")
                    if page_num == len(doc) - 1:
                        print('Trang cuối')
                        ProcessHandler.ghi_message(f"tong_don: {len(doc)}")
                        now2 = datetime.now()
                        now_str = now2.strftime("%d/%m/%Y %H:%M:%S")
                        ProcessHandler.ghi_message(f"end_time: {now_str}")
                        time = now2 - now
                        ProcessHandler.ghi_message(f"time: {ProcessHandler.format_timedelta(time)}]")  







                elif vendor == "SHOPEE-CHOICE":

                    lines = [line.strip() for line in text.split('\n')]
    
                    po_number = None
                    entry_date = None
                    cancel_date = None
                    makhachhang = 'MN_TMDT_00013'
                    delivery_address = 'Khách hàng qua kho Long An lấy hàng trực tiếp'
                    
                    
                    # Tìm dòng chứa "OUR REF NO :"
                    for i, line in enumerate(lines):
                        if "Mã Đơn đặt hàng" in line:
                            # Lấy dòng tiếp theo làm po_number
                            if i + 1 < len(lines):
                                po_number = lines[i + 1].strip()
                            
                            # Lấy dòng sau po_number làm entry_date  
                            if i + 2 < len(lines):
                                entry_date = lines[i + 2].strip()
                                entry_date = entry_date.split(' ')[0]
                            
                            break

                    for line in lines:
                        if line.startswith("Ngày nhập hàng dự kiến"):
                            match = re.search(r'\d{4}-\d{2}-\d{2}', line)
                            if match:
                                cancle_date = match.group(0)
                               


                    print(po_number)
                    print(entry_date)
                    print(cancle_date)


                    products = ProcessHandler.lamsachdonhang(text, r"Tổng tiền sau thuế \(VND\)", "Đại diện Nhà cung cấp")

                    products = ProcessHandler.tachdonhang_shopeechoice(products)
                    if products:
                        sku_mapping = ProcessHandler.load_sku_mapping()  # Load mapping SKU
                        products = ProcessHandler.replace_sku_numbers(products, sku_mapping)  # Thay SKU
                        soluongsanphamstore = len(products)

                        
                        


                        self.log_signal.emit(f"<b>🏬 Đơn hàng Store:</b> <span style='color:orange;'>{vendor}</span>")
                        self.log_signal.emit(f"<b>📦 Số PO:</b> <span style='color:blue;'>{po_number}</span>")
                        self.log_signal.emit(f"<b>🗓️ Ngày đặt hàng:</b> <span style='color:green;'>{entry_date}</span>")
                        self.log_signal.emit(f"<b>⏳ Hạn đơn hàng:</b> <span style='color:red;'>{cancle_date}</span>")
                        self.log_signal.emit(f"<b>🛒 Số lượng sản phẩm đặt của Store:</b> <span style='color:brown;'>{soluongsanphamstore}</span>")

                        
                        if page_label == '1/1':
                            print('')
                            result  = ProcessHandler.upload_file_to_drive(file_path, po_number)
                            print(result )
                            file_url = result.get("url")
                        else:
                            print(page_label)
                            print(page_num)
                            po_name = f"{po_number}.pdf"

                            ProcessHandler.cat_trang_hien_tai(doc,po_name,page_num)
                            result  = ProcessHandler.upload_file_to_drive(po_name, po_number)
                            file_url = result.get("url")
                            if os.path.exists(po_name):
                                os.remove(po_name)


                        saigia = ProcessHandler.write_to_dondathang_bhx(self,products,makhachhang,po_number,entry_date,cancel_date,stt,vendor,delivery_address,file_url)
                        
                        if not makhachhang: 
                            makhachhang = "Không xác định"
                            
                        if saigia > 0 or makhachhang == "Không xác định":
                            
                            if saigia == 0:
                                saigia = "Đúng giá"
                            else:
                                saigia = f'Có {saigia} mã sai giá'
                            hoanthanh = "⚠️Hoàn Thành"
                        else:
                            saigia = "Đúng giá"
                            hoanthanh = "✅Hoàn Thành"


                        self.table_signal.emit(file_name, page_label,vendor, makhachhang ,po_number, saigia, hoanthanh)  # Phát tín hiệu cập nhật bảng
                    


                    

                    





                elif vendor == "JMart":
                    makhachhang = 'MN_MT_JM0001'
                    print(text)
                    entry_date = re.search(r"Ngày in\s*:\s*(\d{1,2}/\d{1,2}/\d{4})", text).group(1)
                    print(entry_date)
                    cancel_date = entry_date
                    po_number = re.search(r"Số phiếu đặt\s*:\s*([A-Z0-9]+)", text).group(1)
                    print(po_number)
                    m = re.search(r"Địa chỉ giao hàng\s*:\s*(.+?)\s*SĐT nhận hàng\s*:", text, re.S)
                    delivery_address = m.group(1).strip() if m else None
                    print(delivery_address)
                    print("Tách sản phẩm")
                    products = ProcessHandler.cat_giua_theo_dong(text,"Mã vật tư","Tổng:")
                    print(products)
                    products = ProcessHandler.tachsanpham_JMart(products)
                    print("--------------")
                    #print(ProcessHandler.debug_structure(products))
                    print(products)
                    if products:
                        sku_mapping = ProcessHandler.load_sku_mapping()  # Load mapping SKU
                        products = ProcessHandler.replace_sku_numbers(products, sku_mapping)  # Thay SKU
                        soluongsanphamstore = len(products)
                        print(products)
                        
                        

                        self.log_signal.emit(f"<b>🏬 Đơn hàng Store:</b> <span style='color:orange;'>{delivery_address}</span>")
                        self.log_signal.emit(f"<b>📦 Số PO:</b> <span style='color:blue;'>{po_number}</span>")
                        self.log_signal.emit(f"<b>🗓️ Ngày đặt hàng:</b> <span style='color:green;'>{entry_date}</span>")
                        self.log_signal.emit(f"<b>⏳ Hạn đơn hàng:</b> <span style='color:red;'>{cancel_date}</span>")
                        self.log_signal.emit(f"<b>🛒 Số lượng sản phẩm đặt của Store:</b> <span style='color:brown;'>{soluongsanphamstore}</span>")

                        
                        if page_label == '1/1':
                            print('')
                            result  = ProcessHandler.upload_file_to_drive(file_path, po_number)
                            print(result )
                            file_url = result.get("url")
                        else:
                            print(page_label)
                            print(page_num)
                            po_name = f"{po_number}.pdf"

                            ProcessHandler.cat_trang_hien_tai(doc,po_name,page_num)
                            result  = ProcessHandler.upload_file_to_drive(po_name, po_number)
                            file_url = result.get("url")
                            if os.path.exists(po_name):
                                os.remove(po_name)
                        
                        saigia = ProcessHandler.write_to_dondathang_kingfood(self,products,makhachhang,po_number,entry_date,cancel_date,stt,vendor,delivery_address,file_url)
                        
                        if not makhachhang: 
                            makhachhang = "Không xác định"
                            
                        if saigia > 0 or makhachhang == "Không xác định":
                            
                            if saigia == 0:
                                saigia = "Đúng giá"
                            else:
                                saigia = f'Có {saigia} mã sai giá'
                            hoanthanh = "⚠️Hoàn Thành"
                        else:
                            saigia = "Đúng giá"
                            hoanthanh = "✅Hoàn Thành"


                        self.table_signal.emit(file_name, page_label,vendor, makhachhang ,po_number, saigia, hoanthanh)  # Phát tín hiệu cập nhật bảng








                elif vendor == "MR.DIY":
                

                    lines = [line.strip() for line in text.split('\n')]
    
                    po_number = None
                    entry_date = None
                    cancel_date = None
                    makhachhang = 'MN_MT_MRDIY'
                    print(text)
                    
                    # Tìm dòng chứa "OUR REF NO :"
                    for i, line in enumerate(lines):
                        if "OUR REF NO :" in line:
                            # Lấy dòng tiếp theo làm po_number
                            if i + 1 < len(lines):
                                po_number = lines[i + 1].strip()
                            
                            # Lấy dòng sau po_number làm entry_date  
                            if i + 2 < len(lines):
                                entry_date = lines[i + 2].strip()
                            
                            break

                    for i, line in enumerate(lines):
                        if "DELIVERY DATE" in line:
                            # Lấy dòng tiếp theo làm po_number
                            if i + 1 < len(lines):
                                cancel_date = lines[i + 1].strip()
                            
                            
                            break



                    products = ProcessHandler.lamsachdonhang(text,"VAT/GST No. 1102018142","Subtotal VATABLE  :")

                    products = ProcessHandler.tach_san_pham_DIY(products)
                    delivery_address = "14 Phan Đăng Lưu, Khu phố 7, Biên Hòa, Đồng Nai"

                    if products:
                        sku_mapping = ProcessHandler.load_sku_mapping()  # Load mapping SKU
                        products = ProcessHandler.replace_sku_numbers(products, sku_mapping)  # Thay SKU
                        soluongsanphamstore = len(products)
                        print(products)
                        
                        

                        self.log_signal.emit(f"<b>🏬 Đơn hàng Store:</b> <span style='color:orange;'>{delivery_address}</span>")
                        self.log_signal.emit(f"<b>📦 Số PO:</b> <span style='color:blue;'>{po_number}</span>")
                        self.log_signal.emit(f"<b>🗓️ Ngày đặt hàng:</b> <span style='color:green;'>{entry_date}</span>")
                        self.log_signal.emit(f"<b>⏳ Hạn đơn hàng:</b> <span style='color:red;'>{cancel_date}</span>")
                        self.log_signal.emit(f"<b>🛒 Số lượng sản phẩm đặt của Store:</b> <span style='color:brown;'>{soluongsanphamstore}</span>")

                        
                        if page_label == '1/1':
                            print('')
                            result  = ProcessHandler.upload_file_to_drive(file_path, po_number)
                            print(result )
                            file_url = result.get("url")
                        else:
                            print(page_label)
                            print(page_num)
                            po_name = f"{po_number}.pdf"

                            ProcessHandler.cat_trang_hien_tai(doc,po_name,page_num)
                            result  = ProcessHandler.upload_file_to_drive(po_name, po_number)
                            file_url = result.get("url")
                            if os.path.exists(po_name):
                                os.remove(po_name)
                        
                        saigia = ProcessHandler.write_to_dondathang_kingfood(self,products,makhachhang,po_number,entry_date,cancel_date,stt,vendor,delivery_address,file_url)
                        
                        if not makhachhang: 
                            makhachhang = "Không xác định"
                            
                        if saigia > 0 or makhachhang == "Không xác định":
                            
                            if saigia == 0:
                                saigia = "Đúng giá"
                            else:
                                saigia = f'Có {saigia} mã sai giá'
                            hoanthanh = "⚠️Hoàn Thành"
                        else:
                            saigia = "Đúng giá"
                            hoanthanh = "✅Hoàn Thành"


                        self.table_signal.emit(file_name, page_label,vendor, makhachhang ,po_number, saigia, hoanthanh)  # Phát tín hiệu cập nhật bảng



                    
                    
                    
                








                elif vendor == "CN-HCM":
                    match = re.search(r"Ngày:\s*(\d{2}/\d{2}/\d{4})", text)
                    if match:
                        entry_date = match.group(1)
                        print("Ngày:", entry_date)

                    match = re.search(r"Số:\s*(\S+)", text)
                    if match:
                        po_number = match.group(1).replace('/','-')
                        print("Số đơn:", po_number)

                    
                    dia_chi = "103/7 đường Ao Đôi, Phường Bình Trị Đông A, Quận Bình Tân, Thành phố Hồ Chí Minh, Việt Nam"
                    makhachhang = "MN_SI_CT07"


                    entry_date_obj = datetime.strptime(entry_date, "%d/%m/%Y")
                    cancel_date = (entry_date_obj + timedelta(days=2)).strftime("%d/%m/%Y")
                    print(cancel_date)


                    products = ProcessHandler.trichdon_CNHCM(text)

                    if not products:
                        products = ProcessHandler.trichdon_CNHCM2(text)

                    

                    


                    





                    if products:
                        
                        saigia = ProcessHandler.write_to_dondathang_CNHCM(self,products,makhachhang,po_number,entry_date,cancel_date,stt,vendor,dia_chi)
                            
                        if not makhachhang: 
                            makhachhang = "Không xác định"
                            
                        if saigia > 0 or makhachhang == "Không xác định":
                            
                            if saigia == 0:
                                saigia = "Đúng giá"
                            else:
                                saigia = f'Có {saigia} mã sai giá'
                            hoanthanh = "⚠️Hoàn Thành"
                        else:
                            saigia = "Đúng giá"
                            hoanthanh = "✅Hoàn Thành"


                        self.table_signal.emit(file_name, page_label,vendor, makhachhang ,po_number, saigia, hoanthanh)  # Phát tín hiệu cập nhật bảng










                elif vendor == "Tiktok":
                    print(text)
                    sohoadon = ""
                    ngayhoadon = ""
                    mail = ""
                    tencongty = ""
                    dongia = ""
                    thue = ""
                    tongtien = ""

                    
                    lines = text.strip().split('\n')  # Cắt chuỗi thành từng dòng


                    for i, line in enumerate(lines):
                        if "Billing Email" in line and i + 1 < len(lines):
                            mail = lines[i + 1]
                            break

                    
                    for i, line in enumerate(lines):
                        if "Subtotal (excluding Tax)" in line and i + 1 < len(lines):
                            dongia = lines[i + 1]
                            break
                        
                    for i, line in enumerate(lines):
                        if "Total Tax @5%" in line and i + 1 < len(lines):
                            thue = lines[i + 1]
                            break

                    for i, line in enumerate(lines):
                        if "Total Amount Due" in line and i + 1 < len(lines):
                            tongtien = lines[i + 1]
                            break

                    for i, line in enumerate(lines):
                        if "Client Name" in line and i + 1 < len(lines):
                            tencongty = lines[i + 1]
                            break

                    for i, line in enumerate(lines):
                        if "Invoice No." in line and i + 1 < len(lines):
                            sohoadon = lines[i + 1]
                            break

                    for i, line in enumerate(lines):
                        if "Invoice Date" in line and i + 1 < len(lines):
                            ngayhoadon = lines[i + 1]
                            break


                    file_path = "dondathang.xlsx"
                    sheet_name = "Don dat hang"
                    
                    
                    wb = openpyxl.load_workbook(file_path)
                    sheet = wb[sheet_name]

                    start_row = sheet.max_row + 1
                    current_row = start_row

                    sheet[f"A{current_row}"] = sohoadon
                    sheet[f"B{current_row}"] = ngayhoadon
                    sheet[f"C{current_row}"] = tencongty
                    sheet[f"D{current_row}"] = mail
                    sheet[f"E{current_row}"] = dongia
                    sheet[f"F{current_row}"] = thue
                    sheet[f"G{current_row}"] = tongtien
                    wb.save(file_path)

                






                    




                elif vendor == "BC Mart":
                    makhachhang = 'MB_MT_8292'
                    
                    po_number = entry_date = cancel_date = None

                    lines = [l.strip() for l in text.splitlines()]

                    for i, line in enumerate(lines):
                        if line.startswith("Số: PO"):
                            po_number  = line.split(":",1)[1].strip()
                            entry_date = lines[i+1].split(":",1)[1].strip()
                            cancel_date= lines[i+2].split(":",1)[1].strip()
                            break

                    print(po_number, entry_date, cancel_date)


                    delivery_address = None
                    ghi_chu = None

                    lines = [l.strip() for l in text.splitlines()]

                    for i, line in enumerate(lines):

                        if "Địa chỉ giao hàng" in line:
                            delivery_address = lines[i+1].strip()

                        if line.startswith("- Nội dung") or "Nội dung" in line:
                            ghi_chu = lines[i+1].strip()

                    print("delivery_address =", delivery_address)
                    print("ghi_chu =", ghi_chu)
                    self.log_signal.emit(f"<b>🏬 Đơn hàng Store:</b> <span style='color:orange;'>{delivery_address}</span>")
                    self.log_signal.emit(f"<b>📦 Số PO:</b> <span style='color:blue;'>{po_number}</span>")
                    self.log_signal.emit(f"<b>🗓️ Ngày đặt hàng:</b> <span style='color:green;'>{entry_date}</span>")
                    self.log_signal.emit(f"<b>⏳ Hạn đơn hàng:</b> <span style='color:red;'>{cancel_date}</span>")
                    #self.log_signal.emit(f"<b>🛒 Số lượng sản phẩm đặt của Store:</b> <span style='color:brown;'>{soluongsanphamstore}</span>")
                    products = ProcessHandler.cat_giua_theo_dong(text,'KM','TỔNG CỘNG')
                    products = ProcessHandler.tach_san_pham_BCMart(products)
                    print(products)
                    if products:
                        sku_mapping = ProcessHandler.load_sku_mapping()  # Load mapping SKU
                        products = ProcessHandler.replace_sku_numbers(products, sku_mapping)  # Thay SKU
                        soluongsanphamstore = len(products)
                        self.log_signal.emit(f"<b>🛒 Số lượng sản phẩm đặt của Store:</b> <span style='color:brown;'>{soluongsanphamstore}</span>")

                        if page_label == '1/1':
                            print('')
                            result  = ProcessHandler.upload_file_to_drive(file_path, po_number)
                            print(result )
                            file_url = result.get("url")
                        else:
                            print(page_label)
                            print(page_num)
                            po_name = f"{po_number}.pdf"

                            ProcessHandler.cat_trang_hien_tai(doc,po_name,page_num)
                            result  = ProcessHandler.upload_file_to_drive(po_name, po_number)
                            file_url = result.get("url")
                            if os.path.exists(po_name):
                                os.remove(po_name)
                        saigia = ProcessHandler.write_to_dondathang_winmart(self,products,makhachhang,po_number,entry_date,cancel_date,stt,vendor,delivery_address,ghi_chu,file_url)
                        
                        if not makhachhang: 
                            makhachhang = "Không xác định"
                            
                        if saigia > 0 or makhachhang == "Không xác định":
                            
                            if saigia == 0:
                                saigia = "Đúng giá"
                            else:
                                saigia = f'Có {saigia} mã sai giá'
                            hoanthanh = "⚠️Hoàn Thành"
                        else:
                            saigia = "Đúng giá"
                            hoanthanh = "✅Hoàn Thành"


                        self.table_signal.emit(file_name, page_label,vendor, makhachhang ,po_number, saigia, hoanthanh)  # Phát tín hiệu cập nhật bảng







                elif vendor == "BHX":
                    
                    lines = text.split("\n")
                    idx = next((i for i, line in enumerate(lines) if "Mã đơn hàng:" in line), -1)
                    po_number = lines[idx + 1].strip() if idx != -1 and idx + 1 < len(lines) else None
                    po_number = po_number.replace('.','/')
                    print(po_number)


                    # Tìm địa chỉ giao hàng
                    delivery_idx = next((i for i, line in enumerate(lines) if "Kho NCC giao" in line), -1)
                    delivery_address = lines[delivery_idx + 2].strip() if delivery_idx != -1 and delivery_idx + 2 < len(lines) else None
                    print("Địa chỉ giao hàng:", delivery_address)


                    # Tìm dòng chứa "Ngày đặt hàng:"
                    entry_idx = next((i for i, line in enumerate(lines) if "Ngày đặt hàng:" in line), -1)
                    entry_date = None

                    if entry_idx != -1:
                        parts = lines[entry_idx].split(":")
                        if len(parts) > 1:
                            date_time_str = parts[1].strip()
                            entry_date = date_time_str.split(" ")[0]  # Lấy phần ngày

                    print("entry_date:", entry_date)


                    # Tìm dòng chứa "Ngày NCC giao hàng:"
                    cancel_idx = next((i for i, line in enumerate(lines) if "Ngày NCC giao hàng" in line), -1)
                    cancel_date = None

                    if cancel_idx != -1 and cancel_idx + 1 < len(lines):
                        date_range = lines[cancel_idx + 1].strip()
                        if " - " in date_range:
                            cancel_date = date_range.split(" - ")[1].strip()

                    print("cancel_date:", cancel_date)



                    

                    start = text.rfind("Thành tiền")  # tìm vị trí xuất hiện cuối cùng
                    end = text.find("Tổng")

                    product_block = text[start + len("Thành tiền"):end].strip()
                    lines = product_block.splitlines()
                    result = "\n".join(lines[1:])  # bỏ dòng đầu
                    print(result)
                    products = ProcessHandler.tach_san_pham_BHX(result)
                    print(products)
                    if products:
                        sku_mapping = ProcessHandler.load_sku_mapping()  # Load mapping SKU
                        products = ProcessHandler.replace_sku_numbers(products, sku_mapping)  # Thay SKU
                        soluongsanphamstore = len(products)
                        
                        socuahang = delivery_address.split()[0]
                        makhachhang = ProcessHandler.laymakhachhang_STF(f'BHX{socuahang}')


                        self.log_signal.emit(f"<b>🏬 Đơn hàng Store:</b> <span style='color:orange;'>{delivery_address}</span>")
                        self.log_signal.emit(f"<b>📦 Số PO:</b> <span style='color:blue;'>{po_number}</span>")
                        self.log_signal.emit(f"<b>🗓️ Ngày đặt hàng:</b> <span style='color:green;'>{entry_date}</span>")
                        self.log_signal.emit(f"<b>⏳ Hạn đơn hàng:</b> <span style='color:red;'>{cancel_date}</span>")
                        self.log_signal.emit(f"<b>🛒 Số lượng sản phẩm đặt của Store:</b> <span style='color:brown;'>{soluongsanphamstore}</span>")
                        
                        if page_label == '1/1':
                            print('')
                            result  = ProcessHandler.upload_file_to_drive(file_path, po_number)
                            print(result )
                            file_url = result.get("url")
                        else:
                            print(page_label)
                            print(page_num)
                            po_name = f"{po_number}.pdf"

                            ProcessHandler.cat_trang_hien_tai(doc,po_name,page_num)
                            result  = ProcessHandler.upload_file_to_drive(po_name, po_number)
                            file_url = result.get("url")
                            if os.path.exists(po_name):
                                os.remove(po_name)
                        saigia = ProcessHandler.write_to_dondathang_bhx(self,products,makhachhang,po_number,entry_date,cancel_date,stt,vendor,delivery_address)
                        
                        if not makhachhang: 
                            makhachhang = "Không xác định"
                            
                        if saigia > 0 or makhachhang == "Không xác định":
                            
                            if saigia == 0:
                                saigia = "Đúng giá"
                            else:
                                saigia = f'Có {saigia} mã sai giá'
                            hoanthanh = "⚠️Hoàn Thành"
                        else:
                            saigia = "Đúng giá"
                            hoanthanh = "✅Hoàn Thành"


                        self.table_signal.emit(file_name, page_label,vendor, makhachhang ,po_number, saigia, hoanthanh)  # Phát tín hiệu cập nhật bảng

                        





                elif vendor == "Farmer":
                    print(text)
                    
                    po_number = None  # Khởi tạo mặc định
                    match = re.search(r"Số hiệu chứng từ\s*:\s*(PO-\w+)", text)
                    if match:
                        po_number = match.group(1)

                    print(po_number)
                    entry_date = None
                    match = re.search(r"Ngày chứng từ\s*:\s*(\d{1,2}/\d{1,2}/\d{4})", text)
                    if match:
                        entry_date = match.group(1)
                    print(entry_date)

                    tenstore = None
                    match = re.search(r"Chi nhánh yêu cầu\s*:\s*(.+)", text)
                    if match:
                        tenstore = match.group(1).strip()

                    


                    cancel_date = None
                    match = re.search(r"Ngày giao hàng\s*:\s*(\d{1,2}/\d{1,2}/\d{4})", text)
                    if match:
                        cancel_date = match.group(1)
                    print(cancel_date)

                    start = text.rfind("(VND)")  # tìm vị trí xuất hiện cuối cùng
                    end = text.find("TỔNG CỘNG TRƯỚC THUẾ VAT")

                    product_block = text[start + len("(VND)"):end].strip()
                    result = "\n".join(product_block.splitlines())
                    print(result)

                    products = ProcessHandler.tach_san_pham_Farmer(result)

                    print(products)

                    if products:
                        sku_mapping = ProcessHandler.load_sku_mapping()  # Load mapping SKU
                        products = ProcessHandler.replace_sku_numbers(products, sku_mapping)  # Thay SKU
                        soluongsanphamstore = len(products)
                        
                        
                        makhachhang = "MN_MT_FMDC01"
                        

                        self.log_signal.emit(f"<b>🏬 Đơn hàng Store:</b> <span style='color:orange;'>{tenstore}</span>")
                        self.log_signal.emit(f"<b>📦 Số PO:</b> <span style='color:blue;'>{po_number}</span>")
                        self.log_signal.emit(f"<b>🗓️ Ngày đặt hàng:</b> <span style='color:green;'>{entry_date}</span>")
                        self.log_signal.emit(f"<b>⏳ Hạn đơn hàng:</b> <span style='color:red;'>{cancel_date}</span>")
                        self.log_signal.emit(f"<b>🛒 Số lượng sản phẩm đặt của Store:</b> <span style='color:brown;'>{soluongsanphamstore}</span>")
                        
                        
                        
                        if page_label == '1/1':
                            print('')
                            result  = ProcessHandler.upload_file_to_drive(file_path, po_number)
                            print(result )
                            file_url = result.get("url")
                        else:
                            print(page_label)
                            print(page_num)
                            po_name = f"{po_number}.pdf"

                            ProcessHandler.cat_trang_hien_tai(doc,po_name,page_num)
                            result  = ProcessHandler.upload_file_to_drive(po_name, po_number)
                            file_url = result.get("url")
                            if os.path.exists(po_name):
                                os.remove(po_name)

                        saigia = ProcessHandler.write_to_dondathang_farmer(self,products,makhachhang,po_number,entry_date,cancel_date,stt,vendor,tenstore,file_url)
                        
                        if not makhachhang: 
                            makhachhang = "Không xác định"
                            
                        if saigia > 0 or makhachhang == "Không xác định":
                            
                            if saigia == 0:
                                saigia = "Đúng giá"
                            else:
                                saigia = f'Có {saigia} mã sai giá'
                            hoanthanh = "⚠️Hoàn Thành"
                        else:
                            saigia = "Đúng giá"
                            hoanthanh = "✅Hoàn Thành"

                        self.table_signal.emit(file_name, page_label,vendor, makhachhang ,po_number, saigia, hoanthanh)  # Phát tín hiệu cập nhật bảng

                        







                elif vendor == "KOC":
                    print(text)
                    lines = [l.strip() for l in text.splitlines() if l.strip()]
                    ma_vandon = None
                    entry_date = None
                    for i, line in enumerate(lines):
                        if "Mã vận đơn:" in line:
                            entry_date = lines[i + 1] if i + 1 < len(lines) else None
                            ma_vandon = lines[i + 3] if i + 3 < len(lines) else None
                            break
                    print(entry_date)
                    print(ma_vandon)

                    lines = [l.strip() for l in text.splitlines() if l.strip()]

                    ten_nguoinhan = None
                    diachi_giaohang = []

                    for i, line in enumerate(lines):
                        if line == "Gọi người gửi nếu giao không được;":
                            # 1 dòng sau là tên người nhận
                            if i + 1 < len(lines):
                                ten_nguoinhan = lines[i + 1]

                            # từ sau tên đến trước "Đến:"
                            for j in range(i + 2, len(lines)):
                                if lines[j].startswith("Đến:"):
                                    break
                                diachi_giaohang.append(lines[j])

                            break
                           
                    diachi_giaohang = "\n".join(diachi_giaohang)                
                    print(ten_nguoinhan)
                    print(diachi_giaohang)
                    products = ProcessHandler.cat_donhang_KOC(text)
                
                    products = ProcessHandler.tach_sanpham_KOC(products)
                    khuvuc = "HCM" if "SĐT: 84335617279" in text else "HN"

                    print(products)
                    if products:
                        sku_mapping = ProcessHandler.load_sku_mapping()  # Load mapping SKU
                        products = ProcessHandler.replace_sku_numbers(products, sku_mapping)  # Thay SKU
                        dien_giai = f'xuất hàng mẫu KOC ngày {entry_date} - {ten_nguoinhan} - {ma_vandon}'

                        ProcessHandler.write_to_dondathang_KOC(
                        self,
                        products,
                        'MB_LE_00217',
                        dien_giai,
                        entry_date,  # order_date
                        entry_date,  # order_date
                        stt,
                        'Shopee_KOC',
                        khuvuc,
                        ma_vandon, # warehouse,
                        ten_nguoinhan,
                        diachi_giaohang
                    )

                    self.table_signal.emit(file_name, page_label, 'Shopee KOC', khuvuc, ma_vandon, '', "✅Hoàn Thành")






                elif vendor == "FujiMart":

                    
                    
                    # Chuyển trang PDF thành ảnh
                    page = doc.load_page(page_num)  # thêm dòng này để lấy đúng trang PDF
                    # Dùng dpi cao để ảnh rõ
                    pix = page.get_pixmap(dpi=500)
                    img = Image.open(io.BytesIO(pix.tobytes("png")))

                    # Tiền xử lý ảnh: grayscale + tăng tương phản
                    img = img.convert("L")  # grayscale
                    img = ImageEnhance.Contrast(img).enhance(2.5)

                    # OCR tiếng Việt với cấu hình tốt
                    info = pytesseract.image_to_string(img, lang='vie')
                    
                    
                    
                    
                    #print(text)
                    
                    lines = [line.strip() for line in text.strip().splitlines() if line.strip()]
                    for i, line in enumerate(lines):
                        if "Sè §¬n:" in line and i >= 3:
                            entry_date = lines[i - 3]
                    print(entry_date)

                    cancel_date = next((line.split("Ngµy giao:")[1].strip() for line in text.splitlines() if "Ngµy giao:" in line), "")
                    
                    # Kiểm tra ngày giao hợp lệ
                    if not re.match(r"\d{2}/\d{2}/\d{4}$", cancel_date):
                        
                        cancel_date = "Không tìm thấy"

                        # Chỉ cộng thêm nếu entry_date hợp lệ
                        if cancel_date == "Không tìm thấy" and re.match(r"\d{2}/\d{2}/\d{4}$", entry_date):
                            entry_date_obj = datetime.strptime(entry_date, "%d/%m/%Y")
                            cancel_date = (entry_date_obj + timedelta(days=2)).strftime("%d/%m/%Y")

                    


                    
                    # Kiểm tra ngày giao hợp lệ
                    if not re.match(r"\d{2}/\d{2}/\d{4}$", entry_date):
                        
                        entry_date = "Không tìm thấy"

                        # Chỉ cộng thêm nếu entry_date hợp lệ
                        if entry_date == "Không tìm thấy" and re.match(r"\d{2}/\d{2}/\d{4}$", cancel_date):
                            cancel_date_obj = datetime.strptime(cancel_date, "%d/%m/%Y")
                            entry_date = (cancel_date_obj - timedelta(days=2)).strftime("%d/%m/%Y")

                    m = re.search(rf'^{entry_date}\s*\n(.+)', text, re.MULTILINE)
                    
                    po_number = m.group(1) if m else None
                    
                    print(po_number)
                    
                    

                    

                    tenstore = ""
                    match = re.search(r"N\s*ơ\s*i\s*[\s]*n\s*h\s*ậ\s*n\s*:\s*(.+?)(?=\n|$)", info, re.IGNORECASE)

                    if match:
                        tenstore = match.group(1)


                    
                    start = text.rfind("§Þa chØ:")  # tìm vị trí xuất hiện cuối cùng
                    end = text.find("VAT")

                    product_block = text[start + len("§Þa chØ:"):end].strip()
                    result = "\n".join(product_block.splitlines())
                    
                    products = ProcessHandler.tach_san_pham_Fujimart(result)
                    
                    if products:
                        sku_mapping = ProcessHandler.load_sku_mapping()  # Load mapping SKU
                        products = ProcessHandler.replace_sku_numbers(products, sku_mapping)  # Thay SKU
                        soluongsanphamstore = len(products)

                        
                        socuahang = tenstore.split()[0]
                        
                        makhachhang = 'MB_MT_FUJI'
                        print(f'SỐ cửa hàng: FJ{socuahang}')
                        

                        self.log_signal.emit(f"<b>🏬 Đơn hàng Store:</b> <span style='color:orange;'>{tenstore}</span>")
                        self.log_signal.emit(f"<b>📦 Số PO:</b> <span style='color:blue;'>{po_number}</span>")
                        self.log_signal.emit(f"<b>🗓️ Ngày đặt hàng:</b> <span style='color:green;'>{entry_date}</span>")
                        self.log_signal.emit(f"<b>⏳ Hạn đơn hàng:</b> <span style='color:red;'>{cancel_date}</span>")
                        self.log_signal.emit(f"<b>🛒 Số lượng sản phẩm đặt của Store:</b> <span style='color:brown;'>{soluongsanphamstore}</span>")
                        
                        
                        
                        if page_label == '1/1':
                            print('')
                            result  = ProcessHandler.upload_file_to_drive(file_path, po_number)
                            print(result )
                            file_url = result.get("url")
                        else:
                            print(page_label)
                            print(page_num)
                            po_name = f"{po_number}.pdf"

                            ProcessHandler.cat_trang_hien_tai(doc,po_name,page_num)
                            result  = ProcessHandler.upload_file_to_drive(po_name, po_number)
                            file_url = result.get("url")
                            if os.path.exists(po_name):
                                os.remove(po_name)

                        saigia = ProcessHandler.write_to_dondathang_fujimart(self,products,makhachhang,po_number,entry_date,cancel_date,stt,vendor,tenstore,file_url)
                        
                        if not makhachhang: 
                            makhachhang = "Không xác định"
                            
                        if saigia > 0 or makhachhang == "Không xác định":
                            
                            if saigia == 0:
                                saigia = "Đúng giá"
                            else:
                                saigia = f'Có {saigia} mã sai giá'
                            hoanthanh = "⚠️Hoàn Thành"
                        else:
                            saigia = "Đúng giá"
                            hoanthanh = "✅Hoàn Thành"


                        self.table_signal.emit(file_name, page_label,vendor, makhachhang ,po_number, saigia, hoanthanh)  # Phát tín hiệu cập nhật bảng

                        

                    




                    
                    
                    
                    





                

                elif vendor == "Winmart":
                    print(text)

                    lines = text.split("\n")

                    idx = next((i for i, line in enumerate(lines) if "Ngày đặt hàng (PO date)" in line), -1)
                    entry_date = lines[idx + 1].strip() if idx != -1 and idx + 1 < len(lines) else None
                    entry_date = entry_date.replace('.', '/')


                    ghichu = "\n".join(
                        text.split("Ghi chú")[1]
                        .split("Nhà cung cấp (Supplier): 0002011398")[0]
                        .strip()
                        .splitlines()[:-1]
                    )
                    ghichu = ghichu.replace('\n', ' ')


                    idx = next((i for i, line in enumerate(lines) if "Số đơn hàng (PO No.)" in line), -1)
                    po_number = lines[idx + 1].strip() if idx != -1 and idx + 1 < len(lines) else None

                    idx = next((i for i, line in enumerate(lines) if "Ngày giao (Delivery Date)" in line), -1)
                    cancel_date = lines[idx + 1].strip() if idx != -1 and idx + 1 < len(lines) else None
                    cancel_date = cancel_date.replace('.', '/')


                    # ===== ĐỊA CHỈ GIAO HÀNG =====

                    idx = next(
                        (i for i, line in enumerate(lines)
                        if "Địa chỉ giao hàng (Delivery Address)" in line),
                        -1
                    )

                    if idx != -1:
                        ma_kho = lines[idx + 1].strip()

                        # lấy phần từ sau mã kho đến trước "Thông tin đơn hàng"
                        address_lines = []

                        for line in lines[idx + 2:]:
                            if "Thông tin đơn hàng (Information)" in line:
                                break

                            line = line.strip()

                            # bỏ dòng trùng kiểu: 6863 - WM+ HCM 60 Liên khu 10-11
                            if "WM+" in line:
                                continue

                            if line:
                                address_lines.append(line)

                        diachi = " ".join(address_lines)

                        diachigiaohang = f"{ma_kho} - {diachi}"

                    else:
                        diachigiaohang = None

                    print(diachigiaohang)
                    
                    
                    
                
                    
                    self.log_signal.emit(f"<b>📦 Số PO:</b> <span style='color:blue;'>{po_number}</span>")
                    self.log_signal.emit(f"<b>🗓️ Ngày đặt hàng:</b> <span style='color:green;'>{entry_date}</span>")

                    self.log_signal.emit(f"<b>⏳ Hạn đơn hàng:</b> <span style='color:red;'>{cancel_date}</span>")
                    

                    # Ghép các dòng lại để tránh trường hợp bị xuống dòng mất chữ

                    
                    

                    idx = -1
                    for i in range(len(lines) - 1):  # -1 để tránh vượt chỉ số khi kiểm tra dòng tiếp theo
                        line_lower = lines[i].lower()  # Chuyển dòng hiện tại về chữ thường
                        next_line_lower = lines[i + 1].lower()  # Chuyển dòng tiếp theo về chữ thường
                        
                        # Kiểm tra "TỔNG HỢP" và "WINCOMMERCE" ở hai dòng liên tiếp
                        if "tổng hợp" in line_lower and "wincommerce" in next_line_lower:
                            idx = i  # Lấy dòng đầu tiên làm mốc
                            break
                        # Kiểm tra "WINCOMMERCE" hoặc "TỔNG HỢP WINCOMMERCE" trong cùng một dòng
                        elif "wincommerce" in line_lower:
                            idx = i
                            break

                    if idx != -1:
                        diachi = []
                        for i in range(idx + 1, len(lines)):
                            # Dừng lại khi gặp "MST" hoặc "Địa chỉ giao hàng", không phân biệt hoa/thường
                            line_lower = lines[i].lower()
                            if "mst" in line_lower or "địa chỉ giao hàng" in line_lower:
                                break
                            diachi.append(lines[i].strip())
                        
                        diachi = " ".join(diachi)
                    else:
                        diachi = None

                    print("Địa chỉ giao hàng:", diachigiaohang)
                    
                    

                    


                    products = ProcessHandler.trichxuatsanpham_winmart(text)
                    print(full_text)
                    #if products:
                       # products = ProcessHandler.trichxuatsanpham_winmart(full_text)

                    print(products)
                    if products:
                        sku_mapping = ProcessHandler.load_sku_mapping()  # Load mapping SKU
                        products = ProcessHandler.replace_sku_numbers(products, sku_mapping)  # Thay SKU
                        soluongsanphamstore = len(products)
                        self.log_signal.emit(f"<b>🛒 Số lượng sản phẩm đặt của Store:</b> <span style='color:brown;'>{soluongsanphamstore}</span>")
                        

                        makhachhang = ProcessHandler.laymakhachhang_satra(diachi,"WINMART")
                        print(diachi)

                        if page_label == '1/1':
                            print('')
                            result  = ProcessHandler.upload_file_to_drive(file_path, po_number)
                            print(result )
                            file_url =  result.get("url")
                        else:
                            print(page_label)
                            print(page_num)
                            po_name = f"{po_number}.pdf"
                            ProcessHandler.cat_trang_hien_tai(doc,po_name,page_num)
                            
                            result  = ProcessHandler.upload_file_to_drive(po_name, po_number)
                            file_url = result.get("url")

                            if os.path.exists(po_name):
                                os.remove(po_name)
                            


                        saigia = ProcessHandler.write_to_dondathang_winmart(self,products,makhachhang,po_number,entry_date,cancel_date,stt,vendor,diachigiaohang,ghichu,file_url)

                        if not makhachhang: 
                            makhachhang = "Không xác định"
                        if saigia > 0 or makhachhang == "Không xác định":
                            if saigia == 0:
                                saigia = "Đúng giá"
                            else:
                                saigia = f'Có {saigia} mã sai giá'
                            hoanthanh = "⚠️Hoàn Thành"
                        else:
                            saigia = "Đúng giá"
                            hoanthanh = "✅Hoàn Thành"
                        self.table_signal.emit(file_name, page_label,vendor, makhachhang ,po_number, saigia, hoanthanh)  # Phát tín hiệu cập nhật bảng



                        

                        










                elif vendor == "Lotte":
                    
                    lines = text.splitlines()
                    po_number = lines[1] if len(lines) > 1 else ""
                    if len(po_number) >= 7:
                        po_number = po_number[:6] + "-" + po_number[6:]
                    if len(po_number) >= 12:  # 7 + 1 (dấu -) + 4 = 12
                        po_number = po_number[:12] + "-" + po_number[12:]

                    
                    time_part, store_code, order_number = po_number.split("-")

                    # Chuyển đổi thời gian từ yyMMdd sang dd/MM/yyyy
                    entry_date = datetime.strptime(time_part, "%y%m%d").strftime("%d/%m/%Y")

                    cancel_date = ProcessHandler.tachcancledate_lotte(text,po_number)
                    tenstore = ProcessHandler.laytenstore_lotte(text,po_number)
                    self.log_signal.emit(f"<b>🏬 Đơn hàng Store:</b> <span style='color:orange;'>{tenstore}</span>")
                    self.log_signal.emit(f"<b>📦 Số PO:</b> <span style='color:blue;'>{po_number}</span>")
                    self.log_signal.emit(f"<b>🗓️ Ngày đặt hàng:</b> <span style='color:green;'>{entry_date}</span>")
                    self.log_signal.emit(f"<b>⏳ Hạn đơn hàng:</b> <span style='color:red;'>{cancel_date}</span>")
                
                    diachigiaohang = "Lotte " + tenstore

                    
                    product_details = ProcessHandler.tachsanpham_lotte(text)
                
                
                    soluongsanpham = len(product_details)
                    self.log_signal.emit(f"<b>📊 Tổng số lượng sản phẩm:</b> <span style='color:purple;'>{soluongsanpham}</span>")
                    store_code = ProcessHandler.get_makhachhang_lotte(store_code[1:])
                    
                    if page_label == '1/1':
                        print('')
                        result  = ProcessHandler.upload_file_to_drive(file_path, po_number)
                        print(result )
                        file_url = result.get("url")
                    else:
                        print(page_label)
                        print(page_num)
                        po_name = f"{po_number}.pdf"

                        ProcessHandler.cat_trang_hien_tai(doc,po_name,page_num)
                        result  = ProcessHandler.upload_file_to_drive(po_name, po_number)
                        file_url = result.get("url")
                        if os.path.exists(po_name):
                            os.remove(po_name)
                    saigia = ProcessHandler.write_to_dondathang_lotte(self,product_details,store_code,po_number,entry_date,cancel_date,stt,vendor,diachigiaohang,file_url)
                    
                    if not store_code: 
                        store_code = "Không xác định"
                    if saigia > 0 or store_code == "Không xác định":
                        if saigia == 0:
                            saigia = "Đúng giá"
                        else:
                            saigia = f'Có {saigia} mã sai giá'
                        hoanthanh = "⚠️Hoàn Thành"
                    else:
                        saigia = "Đúng giá"
                        hoanthanh = "✅Hoàn Thành"

                    self.table_signal.emit(file_name, page_label,vendor, store_code ,po_number, saigia, hoanthanh)

                    

                    
                


                elif vendor == "Kingfood":
                    
                    


                    if text:
                        tranggoc = doc[0].get_text("text")
                        

                        po_number = re.search(r"PO Number:\s*\n([^\n]*\n)?([^\n]*)", tranggoc)
                        po_number = po_number.group(1).strip() if po_number else "Không tìm thấy PO Number"
                        self.log_signal.emit(f"<b>📦 Số PO:</b> <span style='color:blue;'>{po_number}</span>")

                        entry_date = re.search(r"Ngày Đặt Hàng:\s*\n([^\n]*\n)?([^\n]*)", tranggoc)
                        entry_date = entry_date.group(1).replace("-","/").strip() if entry_date else "Không tìm thấy ngày đặt hàng"
                        entry_date = datetime.strptime(entry_date, "%d/%m/%Y")  # Chuyển sang datetime
                        entry_date = entry_date.strftime("%d/%m/%Y")
                        self.log_signal.emit(f"<b>🗓️ Ngày đặt hàng:</b> <span style='color:green;'>{entry_date}</span>")

                        cancel_date = re.search(
    r"Ngày\s*Giao\s*Hàng\s*NCC\s*Xác\s*Nhận:\s*\n*([^\n]*\n)?([^\n]*)",
    tranggoc,
    re.IGNORECASE
)



                        cancel_date = cancel_date.group(1).replace("-","/").strip() if cancel_date else "Không tìm thấy ngày giao hàng"
                        
                        cancel_date = datetime.strptime(cancel_date, "%d/%m/%Y")  # Chuyển sang datetime
                       
                        cancel_date = cancel_date.strftime("%d/%m/%Y")
                        self.log_signal.emit(f"<b>⏳ Hạn đơn hàng:</b> <span style='color:red;'>{cancel_date}</span>")

                        products = ProcessHandler.laydanhsachsanpham_kingfood(text)
                        print(products)
                        
                        
                        if products:
                            sku_mapping = ProcessHandler.load_sku_mapping()  # Load mapping SKU
                            products = ProcessHandler.replace_sku_numbers(products, sku_mapping)  # Thay SKU
                            soluongsanphamstore = len(products)
                            self.log_signal.emit(f"<b>🛒 Số lượng sản phẩm đặt của Store:</b> <span style='color:brown;'>{soluongsanphamstore}</span>")
                            store_code = "MN_MT_KFMSL"  
                            delivery = 'Số 324, đường ĐT743A, Phường Đông Hoà, Thành phố Hồ Chí Minh'
                            
                            if page_label == '1/1':
                                print('')
                                result  = ProcessHandler.upload_file_to_drive(file_path, po_number)
                                print(result )
                                file_url = result.get("url")
                            else:
                                print(page_label)
                                print(page_num)
                                po_name = f"{po_number}.pdf"

                                ProcessHandler.cat_trang_hien_tai(doc,po_name,page_num)
                                result  = ProcessHandler.upload_file_to_drive(po_name, po_number)
                                file_url = result.get("url")
                                if os.path.exists(po_name):
                                    os.remove(po_name)
                        
                        
                            saigia = ProcessHandler.write_to_dondathang_kingfood(self,products,store_code,po_number,entry_date,cancel_date,stt,vendor,delivery,file_url)
                            if not store_code: 
                                store_code = "Không xác định"
                            if saigia > 0 or store_code == "Không xác định":
                                if saigia == 0:
                                    saigia = "Đúng giá"
                                else:
                                    saigia = f'Có {saigia} mã sai giá'
                                    hoanthanh = "⚠️Hoàn Thành"
                            else:
                                saigia = "Đúng giá"
                                hoanthanh = "✅Hoàn Thành"

                            self.table_signal.emit(file_name, page_label,vendor, store_code ,po_number, saigia, hoanthanh)  # Phát tín hiệu cập nhật bảng

                            





                elif vendor == "Emart":
                    
                    po_number = re.search(r"PO No\.\s*\n\s*:? ?([^\n]+)", text)

                    if po_number:
                        po_number = po_number.group(1).strip()  # Loại bỏ dấu ":" và trim khoảng trắng
                        self.log_signal.emit(f"<b>📦 Số PO:</b> <span style='color:blue;'>{po_number}</span>")
                        
                    entry_date = re.search(r"Order By / Date\s*\n\s*:? ?([^\n]+)", text)
                    if entry_date:
                        entry_date = entry_date.group(1).strip()  # Loại bỏ dấu ":" và trim khoảng trắng
                        entry_date = entry_date[:10].replace(".", "/")
                        self.log_signal.emit(f"<b>🗓️ Ngày đặt hàng:</b> <span style='color:green;'>{entry_date}</span>")
                    cancel_date = re.search(r"Delivery Date\s*\n\s*:? ?([^\n]+)", text)
                    if cancel_date:
                        cancel_date = cancel_date.group(1).strip()  # Loại bỏ dấu ":" và trim khoảng trắng
                        cancel_date = cancel_date[:10].replace(".", "/")
                        self.log_signal.emit(f"<b>⏳ Hạn đơn hàng:</b> <span style='color:red;'>{cancel_date}</span>")

                    tenstore = re.search(r"^Delivery to :\s*(.+)", text, re.MULTILINE)
                    if tenstore:
                        tenstore = tenstore.group(1).split("   ")[0]
                    else:
                        print("Không tìm thấy thông tin giao hàng.")
                    self.log_signal.emit(f"<b>🏬 Đơn hàng Store:</b> <span style='color:orange;'>{tenstore}</span>")
                    text =  re.search(r"Article Code\s*(.*?)\s*Total Amount\(without VAT\) :", text, re.DOTALL)
                    text = text.group(1).strip()
                    products = ProcessHandler.laydanhsanpham_emart(text)
                    if products:
                        sku_mapping = ProcessHandler.load_sku_mapping()  # Load mapping SKU
                        products = ProcessHandler.replace_sku_numbers(products, sku_mapping)  # Thay SKU
                    soluongsanphamstore = len(products)
                    self.log_signal.emit(f"<b>🛒 Số lượng sản phẩm đặt của Store:</b> <span style='color:brown;'>{soluongsanphamstore}</span>")
                    
                    if page_label == '1/1':
                        print('')
                        result  = ProcessHandler.upload_file_to_drive(file_path, po_number)
                        print(result )
                        file_url = result.get("url")
                    else:
                        print(page_label)
                        print(page_num)
                        po_name = f"{po_number}.pdf"

                        ProcessHandler.cat_trang_hien_tai(doc,po_name,page_num)
                        result  = ProcessHandler.upload_file_to_drive(po_name, po_number)
                        file_url = result.get("url")
                        if os.path.exists(po_name):
                            os.remove(po_name)
                    saigia = ProcessHandler.write_to_dondathang_emart(self,products,po_number,entry_date,cancel_date,tenstore,stt,"MN_MT_KH0032",vendor ,file_url)
                    if not tenstore: 
                        tenstore = "Không xác định"
                        
                    if saigia > 0 or tenstore == "Không xác định":
                        if saigia == 0:
                            saigia = "Đúng giá"
                        else:
                            saigia = f'Có {saigia} mã sai giá'
                        hoanthanh = "⚠️Hoàn Thành"

                    else:
            
                        saigia = "Đúng giá"
                        hoanthanh ="✅Hoàn Thành"

                    self.table_signal.emit(file_name, page_label,vendor, tenstore ,po_number, saigia, hoanthanh)  # Phát tín hiệu cập nhật bảng

                    
                    


                elif vendor == "Satra":


                    print(text)
                    
                    
                    po_number =  re.search(r"\*P-[^*]+\*", text)
                    po_number =  po_number.group(0)[1:-1]
                    self.log_signal.emit(f"<b>📦 Số PO:</b> <span style='color:blue;'>{po_number}</span>")
                    match = re.search(r"Địa chỉ giao hàng:\s*((?:.*\n)+?)Địa chỉ thanh toán:", text)
                    if match:
                        diachi = match.group(1).strip().replace("\n", " ").replace("  ", " ")
                        print("✅ Địa chỉ:", diachi)
                        makhachhang = ProcessHandler.laymakhachhang_satra(diachi,"SATRA")
                        if makhachhang is None:
                            # In ra thông báo nếu không tìm thấy
                            print(f"⚠️ Không tìm thấy mã khách hàng cho địa chỉ: {diachi}")
                            # hoặc nếu bạn muốn vào log GUI:
                            self.log_signal.emit(f"⚠️ Không tìm thấy mã KH cho địa chỉ: {diachi}")
                            self.table_signal.emit(file_name, '',vendor,diachi, po_number, 'Lỗi', '❌Thất bại')  # Phát tín hiệu cập nhật bảng

                    
                    
                    entry_date = re.search(r"(.*?)\nNgày đặt hàng:", text, re.DOTALL)
                    if entry_date:
                        entry_date = entry_date.group(1).split("\n")[-1]  # Lấy dòng cuối cùng trước "Ngày đặt hàng:"
                        entry_date = datetime.strptime(entry_date, "%m/%d/%Y")
                        entry_date = entry_date.strftime("%d/%m/%Y")
                        if entry_date == '01/01/0001':
                            entry_date = re.search(r"(.*?)\nNgày in:", text, re.DOTALL)
                            if entry_date:
                                entry_date = entry_date.group(1).split("\n")[-1]  # Lấy dòng cuối cùng trước "Ngày đặt hàng:"
                                entry_date = datetime.strptime(entry_date, "%m/%d/%Y")
                                entry_date = entry_date.strftime("%d/%m/%Y")
                    self.log_signal.emit(f"<b>🗓️ Ngày đặt hàng:</b> <span style='color:green;'>{entry_date}</span>")
                    
                    cancel_date = re.search(r"Ngày giao hàng:\s*(.*?)\s*Địa chỉ giao hàng:", text, re.DOTALL)
                    if cancel_date:
                        cancel_date = cancel_date.group(1).strip()
                        pattern = r"(\d{1,2}/\d{1,2}/\d{4})"
                        for line in cancel_date.split("\n"):
                            if re.search(pattern, line):
                                cancel_date = line
                                cancel_date = datetime.strptime(cancel_date, "%m/%d/%Y")
                                cancel_date = cancel_date.strftime("%d/%m/%Y")
                    self.log_signal.emit(f"<b>⏳ Hạn đơn hàng:</b> <span style='color:red;'>{cancel_date}</span>")
                    text =  re.search(r"STT\s*(.*?)\s*Hàng phục vụ cho:", text, re.DOTALL)
                    text = text.group(1).strip()
                
                    products = ProcessHandler.trichxuatsanpham_satra(text)
                    if products:
                        sku_mapping = ProcessHandler.load_sku_mapping()  # Load mapping SKU
                        products = ProcessHandler.replace_sku_numbers(products, sku_mapping)  # Thay SKU
                    soluongsanphamstore = len(products)
                    self.log_signal.emit(f"<b>🛒 Số lượng sản phẩm đặt của Store:</b> <span style='color:brown;'>{soluongsanphamstore}</span>")
                    
                    if page_label == '1/1':
                        print('')
                        result  = ProcessHandler.upload_file_to_drive(file_path, po_number)
                        print(result )
                        file_url = result.get("url")

                    
                    if page_label == '1/1':
                        print('')
                        result  = ProcessHandler.upload_file_to_drive(file_path, po_number)
                        print(result )
                        file_url = result.get("url")
                    else:
                        print(page_label)
                        print(page_num)
                        po_name = f"{po_number}.pdf"

                        ProcessHandler.cat_trang_hien_tai(doc,po_name,page_num)
                        result  = ProcessHandler.upload_file_to_drive(po_name, po_number)
                        file_url = result.get("url")
                        if os.path.exists(po_name):
                            os.remove(po_name)

                    
                    saigia = ProcessHandler.write_to_dondathang_satra(self,products,makhachhang,po_number,entry_date,cancel_date,stt,vendor,diachi,file_url)
                    
                    if not makhachhang: 
                        makhachhang = "Không xác định"
                        
                    if saigia > 0 or makhachhang == "Không xác định":
                        
                        if saigia == 0:
                            saigia = "Đúng giá"
                        else:
                            saigia = f'Có {saigia} mã sai giá'
                        hoanthanh = "⚠️Hoàn Thành"
                    else:
                        saigia = "Đúng giá"
                        hoanthanh = "✅Hoàn Thành"

                    self.table_signal.emit(file_name, page_label,vendor, makhachhang ,po_number, saigia, hoanthanh)  # Phát tín hiệu cập nhật bảng

                    
                    
                    
                elif vendor == "BigC":
                    
                    

                    
                    
                    if page_num == 0:
                        now = datetime.now()
                        now_str = now.strftime("%d/%m/%Y %H:%M:%S")
                        saigia_num = 0
                        
                        

                        trangdaubigc = text
                        if "LINFOX WAREHOUSE (802)" in trangdaubigc:
                            makhachhang = "MB_MT_BIGC"
                            diachigiao = "LINFOX WAREHOUSE (802)"
                        else:
                            makhachhang = "MN_MT_BIGCAC"
                            diachigiao = "FM LOGISTIC VSIP 2 (806)"
                            
                    
                        po_number, entry_date, cancel_date = ProcessHandler.trichxuatinfo_donbigc(trangdaubigc)
                        self.log_signal.emit(f"<b>📦 Số PO:</b> <span style='color:blue;'>{po_number}</span>")
                        self.log_signal.emit(f"<b>🗓️ Ngày đặt hàng:</b> <span style='color:green;'>{entry_date}</span>")
                        self.log_signal.emit(f"<b>⏳ Hạn đơn hàng:</b> <span style='color:red;'>{cancel_date}</span>")
                        products = ProcessHandler.laydanhsachsanpham_bigc(trangdaubigc)
                        soluongsanpham = len(products)
                        self.log_signal.emit(f"<b>📊 Tổng số lượng sản phẩm:</b> <span style='color:purple;'>{soluongsanpham}</span>")
                        file_pdf = file_path
                        file_path = "dondathang.xlsx"
                        sheet_name = "Don dat hang"

                        wb = openpyxl.load_workbook(file_path, data_only=True)
                        sheet = wb[sheet_name]

                        last_row = sheet.max_row
                        while last_row > 0 and all(c.value in (None, "", " ") for c in sheet[last_row]):
                            last_row -= 1

                        wb.close()
                        start_row = last_row + 1
                        print(last_row)
                        ProcessHandler.ghi_message(f"[PO: {po_number}")
                        ProcessHandler.ghi_message(f"store: {diachigiao}")
                        ProcessHandler.ghi_message(f"vendor: {vendor}")
                        ProcessHandler.ghi_message(f"Mã Khách hàng: {makhachhang}")
                        ProcessHandler.ghi_message(f"start_time: {now_str}")
                        tongtien = 0




                    elif page_num > 0 and page_num < len(doc) - 1:
                        tenstore = ProcessHandler.lay_ten_store(text)
                    
                        items = ProcessHandler.trichxuatdanhsachforstore_bigc(text)
                        self.log_signal.emit(f"<b>🏬 Đơn hàng Store:</b> <span style='color:orange;'>{tenstore}</span>")
                        soluongsanphamstore = len(items)
                        self.log_signal.emit(f"<b>🛒 Số lượng sản phẩm đặt của Store:</b> <span style='color:brown;'>{soluongsanphamstore}</span>")
                        

                        saigia,tongtien = ProcessHandler.write_to_dondathang_bigc(self,products,items,po_number,entry_date,cancel_date,tenstore,stt,makhachhang,vendor,page_num,diachigiao,False,None)
                        
                        saigia_num = saigia_num + saigia 
                        if not tenstore: 
                            tenstore = "Không xác định"
                            
                        if saigia > 0 or tenstore == "Không xác định":
                            if saigia == 0:
                                saigia = "Đúng giá"
                            else:
                                saigia = f'Có {saigia} mã sai giá'
                            hoanthanh = "⚠️Hoàn Thành"
                        else:
                            saigia = "Đúng giá"
                            hoanthanh = "✅Hoàn Thành"

                        self.table_signal.emit(file_name, page_label,vendor, tenstore ,po_number, saigia, hoanthanh)  # Phát tín hiệu cập nhật bảng
                    
                    elif page_num == len(doc) - 1:
                        'Trang cuối'
                        tenstore = ProcessHandler.lay_ten_store(text)
                    
                        items = ProcessHandler.trichxuatdanhsachforstore_bigc(text)
                        self.log_signal.emit(f"<b>🏬 Đơn hàng Store:</b> <span style='color:orange;'>{tenstore}</span>")
                        soluongsanphamstore = len(items)
                        self.log_signal.emit(f"<b>🛒 Số lượng sản phẩm đặt của Store:</b> <span style='color:brown;'>{soluongsanphamstore}</span>")
                        result  = ProcessHandler.upload_file_to_drive(file_pdf, po_number)
                        print(result )
                        file_url = result.get("url")

                        saigia,tongtien = ProcessHandler.write_to_dondathang_bigc(self,products,items,po_number,entry_date,cancel_date,tenstore,stt,makhachhang,vendor,page_num,diachigiao, start_row,file_url )
                        
                        saigia_num = saigia_num + saigia 

                        if not tenstore: 
                            tenstore = "Không xác định"
                            
                        if saigia > 0 or tenstore == "Không xác định":
                            if saigia == 0:
                                saigia = "Đúng giá"
                            else:
                                saigia = f'Có {saigia} mã sai giá'
                            hoanthanh = "⚠️Hoàn Thành"
                        else:
                            saigia = "Đúng giá"
                            hoanthanh = "✅Hoàn Thành"

                        self.table_signal.emit(file_name, page_label,vendor, tenstore ,po_number, saigia, hoanthanh)  # Phát tín hiệu cập nhật bảng
                        
                        

                        ProcessHandler.ghi_message(f"sai_gia: {saigia_num}")
                        now2 = datetime.now()
                        now_str = now2.strftime("%d/%m/%Y %H:%M:%S")
                        ProcessHandler.ghi_message(f"end_time: {now_str}")
                        time = now2 - now
                        ProcessHandler.ghi_message(f"time: {ProcessHandler.format_timedelta(time)}]") 
                        
                        # (Xử lý BigC ở đây)

                else:
                    
                    if kiemtrapo =="" or kiemtrapo == "Unknown":
                        print(text)
                        self.log_signal.emit("❌ Vendor không hỗ trợ hoặc không xác định!")
                        self.table_signal.emit(file_name, page_label,"Không xác định", "", "","", "❌Thất bại")  # Phát tín hiệu cập nhật bảng


            self.log_signal.emit("✅ Xử lý PDF hoàn tất!")


        elif file_path.lower().endswith(".txt"):
            try:
                text = ProcessHandler.read_text_file(file_path)
                
                vendor = ProcessHandler.identify_vendor(text)

                if vendor == "Coop":
                    ProcessHandler.process_coop_invoice(self,text,stt,file_path,"1/1",None)
                    self.log_signal.emit("✅ Xử lý TXT hoàn tất!")
                else:
                    self.log_signal.emit("❌ Vendor không hỗ trợ hoặc không xác định!")
            except UnicodeDecodeError as e:
                self.log_signal.emit(f"⚠ Lỗi đọc file TXT: {e}")

        elif file_path.lower().endswith(".xlsx"):
            self.log_signal.emit("📊 Đang xử lý file Excel...")




            if file_name == 'đơn hàng tay.xlsx':
                print(file_name)

                wb = openpyxl.load_workbook(filename=file_path, data_only=True)
                ws = wb["Sheet1"]
                last_row = ws.max_row  # dòng cuối cùng có dữ liệu

                # Bỏ qua dòng tiêu đề (giả sử tiêu đề ở dòng 1)
                data = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    he_thong, makhachhang, po_number, entry_date, cancel_date, barcode, ou_qty, total_price = row
                    data.append({
                        "he_thong": he_thong,
                        "makhachhang": makhachhang,
                        "po_number": po_number,
                        "entry_date": entry_date,
                        "cancel_date": cancel_date,
                        "Barcode": barcode,
                        "OU Qty": ou_qty,
                        "Total Price": total_price
                    })

                # Gom nhóm theo po_number
                grouped = defaultdict(lambda: {"products": []})

                for row in data:
                    po = row["po_number"]
                    if "he_thong" not in grouped[po]:
                        grouped[po].update({
                            "he_thong": row["he_thong"],
                            "makhachhang": row["makhachhang"],
                            "po_number": po,
                            "entry_date": row["entry_date"],
                            "cancel_date": row["cancel_date"]
                        })
                    grouped[po]["products"].append({
                        "Barcode": row["Barcode"],
                        "OU Qty": row["OU Qty"],
                        "Total Price": row["Total Price"]
                    })


                result = list(grouped.values())


                for po in result:
                    print("====== PO:", po["po_number"], "======")
                    print("Hệ thống:", po["he_thong"])
                    print("Mã khách hàng:", po["makhachhang"])
                    print("Ngày nhập:", po["entry_date"])
                    print("Ngày hủy:", po["cancel_date"])

                    print("Sản phẩm:")
                    for sp in po["products"]:
                        print(f" - Barcode: {sp['Barcode']}, Qty: {sp['OU Qty']}, Price: {sp['Total Price']}")
                    
                    print("\n")  # xuống dòng ngăn cách

                    vendor = po["he_thong"]
                    po_number = po["po_number"]
                    entry_date = ProcessHandler.format_date(po["entry_date"])
                    cancel_date = ProcessHandler.format_date(po["cancel_date"])
                    tenstore = ''
                    diachi = po["makhachhang"]
                    products = po["products"]
                    





                    self.log_signal.emit(f"<b style='color:blue;'>📌 Phát hiện Đơn Hàng:</b> <span style='color:green;'>{vendor}</span>")
                    self.log_signal.emit(f"<b>📦 Số PO:</b> <span style='color:blue;'>{po_number}</span>")
                    if tenstore:
                        self.log_signal.emit(f"<b>🏬 Đơn hàng Store:</b> <span style='color:orange;'>{tenstore}</span>")
                    self.log_signal.emit(f"<b>🗓️ Ngày đặt hàng:</b> <span style='color:green;'>{entry_date}</span>")
                    self.log_signal.emit(f"<b>⏳ Hạn đơn hàng:</b> <span style='color:red;'>{cancel_date}</span>")

                    if vendor == "SATRA":
                        makhachhang = ProcessHandler.laymakhachhang_satra(diachi,"SATRA")
                        if makhachhang is None:
                            # In ra thông báo nếu không tìm thấy
                            print(f"⚠️ Không tìm thấy mã khách hàng cho địa chỉ: {diachi}")
                            # hoặc nếu bạn muốn vào log GUI:
                            self.log_signal.emit(f"⚠️ Không tìm thấy mã KH cho địa chỉ: {diachi}")
                            self.table_signal.emit(file_name, '',vendor,diachi, po_number, 'Lỗi', '❌Thất bại')  # Phát tín hiệu cập nhật bảng
                        if products:
                            sku_mapping = ProcessHandler.load_sku_mapping()  # Load mapping SKU
                            products = ProcessHandler.replace_sku_numbers(products, sku_mapping)  # Thay SKU
                        soluongsanphamstore = len(products)
                        self.log_signal.emit(f"<b>🛒 Số lượng sản phẩm đặt của Store:</b> <span style='color:brown;'>{soluongsanphamstore}</span>")
                        saigia = ProcessHandler.write_to_dondathang_satra(self,products,makhachhang,po_number,entry_date,cancel_date,stt,vendor,diachi,None)
                        
                        if not makhachhang: 
                            makhachhang = "Không xác định"
                            
                        if saigia > 0 or makhachhang == "Không xác định":
                            
                            if saigia == 0:
                                saigia = "Đúng giá"
                            else:
                                saigia = f'Có {saigia} mã sai giá'
                            hoanthanh = "⚠️Hoàn Thành"
                        else:
                            saigia = "Đúng giá"
                            hoanthanh = "✅Hoàn Thành"

                        self.table_signal.emit(file_name, '1/1',vendor, makhachhang ,po_number, saigia, hoanthanh)  # Phát tín hiệu cập nhật bảng

                    elif vendor == "COOP":
                        makhachhang = po["makhachhang"]
                        
                        if products:
                            sku_mapping = ProcessHandler.load_sku_mapping()  # Load mapping SKU
                            products = ProcessHandler.replace_sku_numbers(products, sku_mapping)  # Thay SKU
                        soluongsanphamstore = len(products)
                        self.log_signal.emit(f"<b>🛒 Số lượng sản phẩm đặt của Store:</b> <span style='color:brown;'>{soluongsanphamstore}</span>")
                        saigia = ProcessHandler.write_to_dondathang_satra(self,products,makhachhang,po_number,entry_date,cancel_date,stt,vendor,diachi,None)
                        
                        if not makhachhang: 
                            makhachhang = "Không xác định"
                            
                        if saigia > 0 or makhachhang == "Không xác định":
                            
                            if saigia == 0:
                                saigia = "Đúng giá"
                            else:
                                saigia = f'Có {saigia} mã sai giá'
                            hoanthanh = "⚠️Hoàn Thành"
                        else:
                            saigia = "Đúng giá"
                            hoanthanh = "✅Hoàn Thành"

                        self.table_signal.emit(file_name, '1/1',vendor, makhachhang ,po_number, saigia, hoanthanh)  # Phát tín hiệu cập nhật bảng


                    elif vendor == "KINGFOOD":
                        makhachhang = diachi
                        
                        if products:
                            sku_mapping = ProcessHandler.load_sku_mapping()  # Load mapping SKU
                            products = ProcessHandler.replace_sku_numbers(products, sku_mapping)  # Thay SKU
                        soluongsanphamstore = len(products)
                        self.log_signal.emit(f"<b>🛒 Số lượng sản phẩm đặt của Store:</b> <span style='color:brown;'>{soluongsanphamstore}</span>")
                        
                        saigia = ProcessHandler.write_to_dondathang_satra(self,products,makhachhang,po_number,entry_date,cancel_date,stt,vendor,diachi)
                        
                        if not makhachhang: 
                            makhachhang = "Không xác định"
                            
                        if saigia > 0 or makhachhang == "Không xác định":
                            
                            if saigia == 0:
                                saigia = "Đúng giá"
                            else:
                                saigia = f'Có {saigia} mã sai giá'
                            hoanthanh = "⚠️Hoàn Thành"
                        else:
                            saigia = "Đúng giá"
                            hoanthanh = "✅Hoàn Thành"

                        self.table_signal.emit(file_name, '1/1',vendor, makhachhang ,po_number, saigia, hoanthanh)  # Phát tín hiệu cập nhật bảng








                    







                return



           
            if file_name.startswith("[MII]"):
                
                groups = defaultdict(list)

                wb = openpyxl.load_workbook(filename=file_path, data_only=True)
                ws = wb["Danh sách đơn hàng_WH6_HN_HCM"]
                last_row = ws.max_row
                print(f"[INFO] Dòng cuối: {last_row}")

                for row in ws.iter_rows(min_row=4, max_row=last_row, values_only=True):
                    order_date = ProcessHandler.to_date_only(row[7])  # H
                    po_number = row[6].split('*')[0] 
                    warehouse = row[2]                 # C
                    product_name = row[12]             # M
                    quantity = row[17]                 # R
                    buyer_id = row[18]                 # S

                    # Debug product
                    print(f"\n[DEBUG] Đang xử lý sản phẩm: {product_name}")
                    raw_barcode = ProcessHandler. ProcessHandler.laymahang_JIT(re.sub(r"\s+", "", product_name))
                    barcode = raw_barcode[0] if isinstance(raw_barcode, list) and raw_barcode else raw_barcode
                    if not barcode:
                        print(f"[WARNING] Barcode trống cho sản phẩm: {product_name}")

                    # Nhóm theo order_date, warehouse và po_number
                    groups[(order_date, warehouse, po_number)].append({
                        "product_name": product_name,
                        "Barcode": barcode,
                        "Qty Ord/Pcs": quantity,
                        "buyer_id": buyer_id,
                        "po_number": po_number.split('*')[0]
                    })

                # --- Xử lý từng nhóm (mỗi po_number riêng)
                for (od, wh, po), items in groups.items():
                    products = [
                        {
                            "Barcode": str(item["Barcode"] or ""),
                            "Qty Ord/Pcs": item["Qty Ord/Pcs"],
                            "product_name": item["product_name"],
                            "Ghi_Chu": f'Mã đơn hàng: {item["po_number"]}, id người mua: {item["buyer_id"]}'
                        }
                        for item in items
                    ]

                    print(f"\n[GROUP] {od} - {wh} - {po}: {products}")

                    now = datetime.now(ZoneInfo("Asia/Ho_Chi_Minh"))
                    khung_gio = 'sáng' if now.hour < 12 else 'chiều'

                    # Load và thay SKU theo mapping
                    sku_mapping = ProcessHandler.load_sku_mapping()
                    print(f"[DEBUG] SKU mapping size: {len(sku_mapping)}")
                    products = ProcessHandler.replace_sku_numbers(products, sku_mapping)

                    # Gửi từng nhóm sang JIT (mỗi po_number riêng)
                    ProcessHandler.write_to_dondathang_JITxls(
                        self,
                        products,
                        'MN_JIT_01512',
                        po,
                        od,  # order_date
                        od,  # order_date
                        stt,
                        'JIT',
                        wh,  # warehouse
                        po
                    )

                    self.table_signal.emit(file_name, '1/1', 'CHOICE', wh, po, '', "✅Hoàn Thành")


            elif file_name.startswith("XUẤT HÀNG "):  

                # 🔹 Regex cho phép có khoảng trắng giữa ngày và tháng
                match = re.search(r'(\d{1,2})\s*-\s*(\d{1,2})', file_name)

                ngay_string = match.group(0) if match else None

                print("📅 Giá trị ngày:", ngay_string)
                  
             
                
                wb = openpyxl.load_workbook(filename=file_path, data_only=True)
                ws = wb["DanhSachChiTietHoaDon_KV1110202"]

                # --- 2️⃣ TÌM DÒNG CUỐI CÓ DỮ LIỆU Ở CỘT A ---
                last_row = ws.max_row
                while last_row > 0:
                    value = ws.cell(row=last_row, column=1).value
                    if value is not None and str(value).strip() != "":
                        break
                    last_row -= 1

                print("👉 Dòng cuối có dữ liệu:", last_row)

                # --- 3️⃣ LẤY HEADER ---
                headers = {}
                for idx, cell in enumerate(ws[1], start=1):
                    if cell.value:
                        headers[cell.value] = idx

                # --- 4️⃣ DANH SÁCH CỘT CẦN ---
                needed_cols = [
                    "Mã Misa", "Mã hóa đơn",
                    "Mã vận đơn", "Trạng thái giao hàng",
                    "Kênh bán", "Thành tiền",
                    "MÃ TP 1", "SLTP1",
                    "MÃ TP 2", "SLTP2",
                    "MÃ TP 3", "SLTP3",
                    "MÃ TP 4", "SLTP4",
                    "Kho dò trên haravan",
                ]

                # --- 5️⃣ GOM DATA ---
                # data = { Mã Misa → { Kho → { Mã hóa đơn → {...} } } }
                data = {}

                for r in range(2, last_row + 1):
                    ma_misa = ws.cell(row=r, column=headers["Mã Misa"]).value
                    ma_hoadon = ws.cell(row=r, column=headers["Mã hóa đơn"]).value

                    if ma_misa in (None, "", "#N/A") or ma_hoadon in (None, "", "#N/A"):
                        continue

                    kenh_ban = ws.cell(row=r, column=headers["Kênh bán"]).value
                    thanh_tien = ws.cell(row=r, column=headers["Thành tiền"]).value
                    kho_haravan = ws.cell(row=r, column=headers["Kho dò trên haravan"]).value

                    ma_vandon = ws.cell(row=r, column=headers["Mã vận đơn"]).value
                    trang_thai = ws.cell(row=r, column=headers["Trạng thái giao hàng"]).value

                    ma_misa = str(ma_misa).strip()
                    ma_hoadon = str(ma_hoadon).strip()
                    kho_haravan = str(kho_haravan).strip() if kho_haravan not in (None, "", "#N/A") else "KHONG_XAC_DINH"

                    ma_vandon = str(ma_vandon).strip() if ma_vandon not in (None, "", "#N/A") else ""
                    trang_thai = str(trang_thai).strip() if trang_thai not in (None, "", "#N/A") else ""

                    # --- KHỞI TẠO MISA ---
                    if ma_misa not in data:
                        data[ma_misa] = {}

                    # --- KHỞI TẠO KHO ---
                    if kho_haravan not in data[ma_misa]:
                        data[ma_misa][kho_haravan] = {}

                    # --- KHỞI TẠO HÓA ĐƠN ---
                    if ma_hoadon not in data[ma_misa][kho_haravan]:
                        data[ma_misa][kho_haravan][ma_hoadon] = {
                            "Mã vận đơn": ma_vandon,
                            "Trạng thái giao hàng": trang_thai,
                            "Kênh bán": str(kenh_ban).strip() if kenh_ban else "",
                            "dong_chi_tiet": []   # ✨ QUAN TRỌNG
                        }
                    else:
                        if ma_vandon:
                            data[ma_misa][kho_haravan][ma_hoadon]["Mã vận đơn"] = ma_vandon
                        if trang_thai:
                            data[ma_misa][kho_haravan][ma_hoadon]["Trạng thái giao hàng"] = trang_thai

                    # --- LẤY SẢN PHẨM THEO DÒNG ---
                    dong_san_pham = []

                    for i in range(1, 5):
                        ma_tp = ws.cell(row=r, column=headers.get(f"MÃ TP {i}", 0)).value
                        sl_tp = ws.cell(row=r, column=headers.get(f"SLTP{i}", 0)).value

                        if ma_tp in (None, "", "#N/A") or sl_tp in (None, "", "#N/A"):
                            continue

                        dong_san_pham.append({
                            "ma_tp": str(ma_tp).strip(),
                            "sl_tp": int(sl_tp)
                        })

                    # --- GHI 1 DÒNG CHI TIẾT ---
                    if dong_san_pham:
                        data[ma_misa][kho_haravan][ma_hoadon]["dong_chi_tiet"].append({
                            "thanh_tien": thanh_tien if isinstance(thanh_tien, (int, float)) else 0,
                            "san_pham": dong_san_pham
                        })

                print("✅ Đã gom dữ liệu – Thành tiền tách theo từng dòng")

                # --- 6️⃣ XỬ LÝ THEO KHO ---
                for ma_misa, kho_dict in data.items():
                    print(f"\n🧾 Mã Misa: {ma_misa}")

                    for kho_haravan, hoadon_dict in kho_dict.items():
                        print(f"📦 Kho Haravan: {kho_haravan}")

                        hoa_don_list = []

                        for ma_hoadon, record in hoadon_dict.items():
                            dong_list = record["dong_chi_tiet"]
                            if not dong_list:
                                continue

                            hoa_don_list.append({
                                "ma_hoadon": ma_hoadon,
                                "ma_vandon": record.get("Mã vận đơn", ""),
                                "trang_thai": record.get("Trạng thái giao hàng", ""),
                                "kenh_ban": record["Kênh bán"],
                                "kho_haravan": kho_haravan,
                                "dong_chi_tiet": dong_list
                            })

                        print(f"👉 Tổng hóa đơn trong kho {kho_haravan}: {len(hoa_don_list)}")



                        # --- GỌI XỬ LÝ ---
                        ProcessHandler.write_to_dondathang_TMDT(
                            self,
                            hoa_don_list,
                            ma_misa,      # makhachhang
                            ngay_string,
                            ngay_string
                        )


                    
                

                




                

               

                







            else:


            

                wb = openpyxl.load_workbook(filename=file_path, data_only=True)
                ws = wb["HÀ THÀNH"]

                
                # Lấy mã sản phẩm từ dòng tiêu đề (E5 và F5)
                article_E = str(ws["E2"].value).strip()
                article_F = str(ws["F2"].value).strip()

                danhsach = []

                for row in range(5, 144):
                    code_store = ws[f"B{row}"].value
                    dia_chi =  ws[f"D{row}"].value
                    qty_E = ws[f"E{row}"].value
                    qty_F = ws[f"F{row}"].value

                    print(f"[DEBUG] Row {row} → code_store: {code_store}, qty_E: {qty_E}, qty_F: {qty_F}")  # 👈 Debug
                    giasanpham =  102143 

                    if code_store:
                        products = []
                        if isinstance(qty_E, (int, float)) and qty_E > 0:
                            products.append({"Barcode": article_E, "OU Qty": int(qty_E), "Total Price": giasanpham})
                        if isinstance(qty_F, (int, float)) and qty_F > 0:
                            products.append({"Barcode": article_F, "OU Qty": int(qty_F), "Total Price": giasanpham})

                        

                        if products:
                            entry = {
                                "code_store": str(code_store).strip(),
                                "dia_chi": str(dia_chi).strip(),
                                "products": products
                            }
                            print(f"✅ ĐÃ THÊM: {entry}")  # 👈 Debug dòng hợp lệ
                            danhsach.append(entry)
                        else:
                            print("⚠️ Không có sản phẩm nào được thêm.")
                    else:
                        print("⛔ Bỏ qua dòng không có mã cửa hàng.")

                # In toàn bộ kết quả cuối cùng
                print("\n📦 TOÀN BỘ KẾT QUẢ CUỐI:")
                print(f"\n📦 Tổng cộng: {len(danhsach)} cửa hàng có sản phẩm.")
                sku_mapping = ProcessHandler.load_sku_mapping()  # Load mapping SKU
                products = ProcessHandler.replace_sku_numbers(products, sku_mapping)  # Thay SKU


                for item in danhsach:
                    makhachhang = ProcessHandler.laymakhachhang_STF(f'STF{item["code_store"]}')
                    po_number = f'P-{item["code_store"]}'
                    entry_date = '05/11/2025'
                    cancel_date = '12/11/2025'
                    vendor = 'Satra'
                    diachi = item["dia_chi"]


                    
                    print(f"🏪 Mã cửa hàng: {code_store}")
                    products = item["products"]
                    
                    
                    saigia = ProcessHandler.write_to_dondathang_satra(self,products,makhachhang,po_number,entry_date,cancel_date,stt,vendor,diachi)
                        
                    if not makhachhang: 
                        makhachhang = "Không xác định"
                        
                    if saigia > 0 or makhachhang == "Không xác định":
                        
                        if saigia == 0:
                            saigia = "Đúng giá"
                        else:
                            saigia = f'Có {saigia} mã sai giá'
                        hoanthanh = "⚠️Hoàn Thành"
                    else:
                        saigia = "Đúng giá"
                        hoanthanh = "✅Hoàn Thành"
                

                






            







                # Xử lý file Excel
            self.log_signal.emit("✅ Xử lý Excel hoàn tất!")

        
        
        
            
            
            
            
            
            #while True:
            """
                # Lấy dữ liệu từ file
            po_location, po_number, entry_date, cancel_date, products = ProcessHandler.get_grouped_data_xlsx(file_path)

                # Kiểm tra nếu không còn dữ liệu để xử lý, thì dừng vòng lặp
            wb = openpyxl.load_workbook(filename=file_path, data_only=True)
            ws = wb["Sheet1"]
    
            makhachhang = ProcessHandler.get_makhachhang(po_location) if po_location else "Không tìm thấy"

            if products:
                sku_mapping = ProcessHandler.load_sku_mapping()  # Load mapping SKU
                products = ProcessHandler.replace_sku_numbers(products, sku_mapping)  # Thay SKU

                # Ghi vào file "dondathang.xlsx"
                saigia = ProcessHandler.write_to_dondathang(self,products, makhachhang, po_number, entry_date, cancel_date,stt,"COOP","")
                if not makhachhang: 
                    makhachhang = "Không xác định"
                    
                if saigia > 0 or makhachhang == "Không xác định":
                    if saigia == 0:
                        saigia = "Đúng giá"
                    else:
                        saigia = f'Có {saigia} mã sai giá'
                    hoanthanh = "⚠️Hoàn Thành"
                else:
                    saigia = "Đúng giá"
                    hoanthanh = "✅Hoàn Thành"

                hethong = ProcessHandler.layhethong_COOP(makhachhang)
                self.table_signal.emit(file_name, "1/1",hethong,makhachhang, po_number, saigia, hoanthanh)  # Phát tín hiệu cập nhật bảng
            if ws.max_row == 1:  # Chỉ còn lại dòng tiêu đề
                print("✅ Đã xử lý xong toàn bộ dữ liệu! Dừng lại.")
                return
            print("✅ Đã xử lý xong một đơn hàng. Tiếp tục...")
                """
                

        else:
            self.log_signal.emit("⚠ Định dạng file không được hỗ trợ!")
        if fileloi:

            self.log_signal.emit(fileloi.strip())
        self.log_signal.emit("✅ Xử lý hoàn tất!")

        #except Exception as e:
            #fileloi = f"{fileloi}\n❌ Lỗi xử lý file: {e}\n{file_name}"
            #self.log_signal.emit(f"❌ Lỗi xử lý file: {e}\n{file_name}")
        

 







        
    


    




process_handler = ProcessHandler()
def lay_gia_tri_G1():
    """
    Lấy giá trị ô G1 từ file 'dondathang.xlsx', sheet 'Don dat hang'.
    Trả về giá trị của ô G1 hoặc None nếu ô trống.
    """
    file_path = "dondathang.xlsx"
    sheet_name = "Don dat hang"

    wb = openpyxl.load_workbook(file_path, data_only=True)  # Đọc giá trị thực (không phải công thức)
    sheet = wb[sheet_name]

    value_G1 = sheet["G1"].value  # Lấy giá trị ô G1

    wb.close()
    return value_G1


def write_headers(file_path='dondathang.xlsx', sheet_name="Don dat hang", row=8):
    try:
        headers = [
            "Ngày đơn hàng (*)",  # A
            "Số đơn hàng (*)",    # B
            "Trạng thái",         # C
            "Ngày giao hàng",     # D
            "Địa điểm giao hàng", # E
            "Tính giá thành",     # F
            "Mã khách hàng",      # G
            "Tên khách hàng",     # H
            "Địa chỉ",            # I
            "Mã số thuế",         # J
            "Người nhận hàng",    # K
            "Diễn giải",          # L
            "Mã nhân viên bán hàng", # M
            "Là đơn đặt hàng phát sinh trước khi sử dụng phần mềm", # N
            "Loại tiền",          # O
            "Tỷ giá",             # P
            "Mã hàng (*)",        # Q
            "Thuộc combo",        # R
            "Tên hàng",           # S
            "Là dòng ghi chú",    # T
            "Hàng khuyến mại",    # U
            "Mã kho",             # V
            "ĐVT",                # W
            "Số lượng",           # X
            "Đơn giá",            # Y
            "Thành tiền",         # Z
            "Thành tiền quy đổi", # AA
            "Tỷ lệ CK (%)",       # AB
            "Tiền chiết khấu",    # AC
            "Tiền chiết khấu quy đổi", # AD
            "% thuế GTGT",        # AE
            "% thuế suất KHAC",   # AF
            "Tiền thuế GTGT",     # AG
            "Tiền thuế GTGT quy đổi", # AH
            "Mã khoản mục chi phí",   # AI
            "Mã đơn vị",          # AJ
            "Mã đối tượng THCP",  # AK
            "Số hợp đồng bán",    # AL
            "Mã thống kê",        # AM
            "Công Trình",         # AN
            "Ghi Chú",            # AO
            "Mã hàng co",         # AP
            "Nội dung khuyến mãi", # AQ
            "Mã barcode DIY",     # AR
            "Mã barcode G TIN"    # AS
        ]

        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]

        # Gán lần lượt vào cột A→AS, dòng row
        for col_idx, value in enumerate(headers, start=1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            sheet[f"{col_letter}{row}"] = value

        wb.save(file_path)
        print(f"✅ Headers đã được ghi vào sheet '{sheet_name}', dòng {row}.")

    except PermissionError:
        print(f"⚠️ File {file_path} đang mở trong Excel. Hãy đóng nó rồi chạy lại.")

        return file_path
